#!/usr/bin/env python3
"""
Generate Trinity-CAM Risk Register from BD template.

Project:          Trinity-CAM
Seller:           Johnson Controls International (JCI)
Buyer:            Robert Bosch GmbH
Business:         Aircondition business
Carve-out model:  Integration (JCI -> Merger Zone -> Bosch)
PMO Lead:         KPMG
Sites:            48 worldwide
IT Users:         12,000
Applications:     1,800+ including massive SAP
IT Delivery:      Infosys (Merger Zone build, operation, migrations)
GoLive:           2028-01-01
"""

import os
import sys
from pathlib import Path
from datetime import date

sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

HERE = Path(__file__).parent
TEMPLATE = HERE / "Reference" / "BD_Risk-Register_template_en_V1.0_Dec2023.xlsx"
OUTPUT   = HERE / "active-projects" / "Trinity-CAM" / "Trinity-CAM_Risk_Register.xlsx"
OUTPUT.parent.mkdir(parents=True, exist_ok=True)

TODAY    = "04.04.2026"
REPORT   = "04.04.2026"

# =============================================================================
# RISK DATA  (25 project-specific risks generated fresh for Trinity-CAM)
#
# Each tuple:
#   (risk_id, category, cause, event, effect, event_date, owner, source,
#    impact, probability, p_or_t, qualitative_desc, eur_cy, eur_3y,
#    strategy, measure, due_date, status, notes)
#
# Categories (template taxonomy):
#   Technology, R&D | Engineering | Manufacturing | Quality |
#   Strategy & Portfolio | Budget | Schedule | Resources | Supply Chain |
#   Market & Competitors | Customers | Raw Materials |
#   Stakeholder Relations & Public Affairs | Intellectual Property |
#   Legal & Compliance | Ecosystems & Ethics | Security & Data Protection
#
# Impact:       Very Low / Low / Moderate / High / Very High
# Probability:  10% / 30% / 50% / 70% / 90%
# p_or_t:       threat / opportunity
# Strategy:     Avoid / Transfer / Mitigate / Accept / Exploit / Enhance / Share
# Status:       not started / in progress / on hold / implemented / cancelled
# =============================================================================

RISKS = [

    # R001 - SAP landscape complexity (Technology)
    (
        "R001", "Technology, R&D",
        "JCI Aircon SAP landscape includes >1,800 applications with deep SAP integration; "
        "over 500 custom Z-programs identified; multiple country-specific payroll and financial "
        "configurations; SAP system copy from JCI to Merger Zone at this scale has not been "
        "performed previously within JCI; Infosys SAP team has limited JCI environment context "
        "at project start.",
        "SAP system copy from JCI to Merger Zone takes longer than planned or fails initial "
        "technical validation; SAP client separation reveals unexpected shared data dependencies "
        "in Finance, HR, and Logistics modules requiring re-work; QG2&3 milestone (2027-07-31) "
        "at risk; SAP Mock Cutover 1 delayed past Oct 2027 target.",
        "Phase 3 Integration Testing (Aug 2027) cannot start on schedule; SAP production "
        "readiness not confirmed by QG4 (2027-12-08); GoLive (2028-01-01) postponed; "
        "JCI business plan impact and extended MZ dual-run costs EUR 500K/month.",
        "2027-07-31", "KPMG SAP Architect", "Formal Risk Review",
        "Very High", "70%", "threat",
        "SAP build and system copy are on the critical path; any delay cascades to UAT, "
        "mock cutovers, and GoLive; business disruption and cost overrun risk is high.",
        "3000000", "1500000",
        "Mitigate",
        "Conduct SAP technical pre-assessment and Z-program inventory by 2026-11-01; "
        "engage SAP SE expert services for system copy support; plan 2 rehearsal copies "
        "before production copy; include 4-week buffer in SAP build schedule; "
        "weekly SAP progress KPI dashboard from Phase 2 start.",
        "2026-11-01", "not started",
        "Top Priority - on critical path to GoLive"
    ),

    # R002 - Infosys Merger Zone delivery delay (Technology)
    (
        "R002", "Technology, R&D",
        "Infosys is the sole responsible vendor for Merger Zone DC/cloud setup, 48-site "
        "network connectivity, and IT service platform build; complex multi-cloud architecture "
        "across global regions; Infosys resource mobilisation at required scale (>50 FTE) "
        "in Phase 2 is subject to bench availability and competing Infosys engagements; "
        "no alternative MZ vendor is contractually identified.",
        "Merger Zone DC environment not ready by 2027-02-01 target (Phase 2 start); "
        "network connectivity to key manufacturing and office sites delayed beyond Sep 2027; "
        "core IT services (AD, M365, ITSM) not available for Phase 3 user migration waves; "
        "QG2&3 gate (2027-07-31) fails.",
        "Phase 3 testing cannot start; user migration waves delayed; SAP system copy "
        "blocked by absence of MZ target environment; programme milestone slip of "
        "3-6 months; extended Infosys contract cost and JCI TSA extension costs.",
        "2027-07-31", "Infosys Programme Manager", "Formal Risk Review",
        "High", "50%", "threat",
        "Single-vendor dependency on Infosys for MZ delivery creates a critical programme "
        "risk; escalation path and contractual remedies must be defined pre-Phase 2.",
        "2000000", "500000",
        "Mitigate",
        "Establish weekly Infosys delivery KPI dashboard from Phase 2 start; "
        "link milestone payments to MZ DC readiness by 2027-02-28; "
        "identify fallback cloud hosting (Azure/AWS) option approved by QG1 (2026-10-01); "
        "KPMG infrastructure architect embedded in Infosys weekly build review.",
        "2026-10-01", "not started",
        "Contractual remedies to be agreed by QG1"
    ),

    # R003 - SAP mock cutover defects vs tight GoLive window (Schedule)
    (
        "R003", "Schedule",
        "SAP Mock Cutover 1 planned end Oct 2027; Mock Cutover 2 planned end Nov 2027 "
        "(target 2027-11-21); QG4 gate is 2027-12-08; recovery window between Mock 2 "
        "end and QG4 is only 17 calendar days; any critical SAP cutover defects found "
        "in Mock 2 (GL posting failures, interface errors, data load timing overruns) "
        "require immediate resolution within this window.",
        "Critical SAP cutover defects discovered in Mock 2 cannot be fully resolved "
        "by QG4 (2027-12-08); QG4 gate deferred; GoLive postponed beyond 2028-01-01; "
        "JCI TSA contractually expired; no JCI IT support coverage available; "
        "Aircon business runs without stable IT systems for extended period.",
        "GoLive delay of 4-8 weeks; extended MZ dual-run costs EUR 500K/month; "
        "JCI TSA emergency extension required at premium rate; business operations "
        "disrupted; Bosch integration timeline delayed; board-level escalation.",
        "2027-12-08", "KPMG PMO Lead", "Formal Risk Review",
        "Very High", "50%", "threat",
        "17-day recovery window between Mock 2 and QG4 is insufficient for complex "
        "SAP cutover defects; pre-defining escalation and rollback criteria is essential.",
        "3000000", "1000000",
        "Mitigate",
        "Conduct Mock Cutover 1 with strict issue timer: all P1 issues resolved within 5 days; "
        "pre-define QG4 go/no-go criteria (zero P1, less than 5 P2 SAP defects); "
        "negotiate JCI TSA emergency extension clause (30 days) in SPA; "
        "identify GoLive postponement date no later than 2028-02-01 as fallback; "
        "reserve post-Mock-2 defect sprint team (Infosys SAP lead x3).",
        "2027-09-01", "not started",
        "TSA extension clause must be in SPA signed before deals close"
    ),

    # R004 - Infosys key personnel loss (Resources)
    (
        "R004", "Resources",
        "Infosys delivery model relies on named SAP Architects, Infrastructure Leads, and "
        "Data Migration Leads; high M&A market demand for Infosys SAP and cloud talent; "
        "18-month engagement creates staff rotation and attrition risk; Infosys HR model "
        "rotates staff every 12-15 months; no contractual named-resource commitment at "
        "project start.",
        "Key Infosys SAP Architect or MZ Infrastructure Lead rotates off project during "
        "Phase 2 (Feb-Jul 2027) or Phase 3 (Aug-Dec 2027); replacement requires 6-8 week "
        "onboarding; critical knowledge of JCI SAP environment and MZ configuration is lost; "
        "SAP build and MZ integration timeline impacted by 4-6 weeks.",
        "QG2&3 gate delayed; SAP integration testing quality reduced; defect rates in "
        "Phase 3 increase; programme PMO must escalate to Infosys leadership; "
        "additional cost for accelerated onboarding.",
        "2027-07-31", "KPMG PMO Lead", "Formal Risk Review",
        "High", "50%", "threat",
        "Resource continuity on 18-month complex programme is a standard delivery risk; "
        "mitigation must be built into Infosys contractual terms.",
        "500000", "0",
        "Mitigate",
        "Contract Infosys for named resource commitment with 90-day minimum notice and "
        "replacement approval by KPMG; require living knowledge base (Confluence/SharePoint) "
        "updated weekly by each Infosys workstream lead; designate KPMG shadow for each "
        "Infosys lead role; conduct mid-project (Phase 2 end) resource continuity review.",
        "2026-08-01", "not started",
        "Named-resource clause to be in Infosys SOW"
    ),

    # R005 - TUPE / employment law across 48 jurisdictions (Legal)
    (
        "R005", "Legal & Compliance",
        "12,000 JCI Aircon employees in 48 jurisdictions are subject to differing TUPE and "
        "transfer of undertaking legislation; Germany and France require 30-90 day employee "
        "consultation periods before IT system changes affecting roles; China requires "
        "local labour authority notification; some APAC jurisdictions have data localisation "
        "laws affecting how user data is migrated.",
        "Employment law non-compliance in key jurisdictions deletes user migration schedule "
        "for Wave 1 or Wave 2 sites; injunctions or regulatory actions in Germany or France "
        "block planned migration execution in Aug-Sep 2027; affected 3,000+ users cannot "
        "be migrated to Merger Zone until legal clearance obtained.",
        "User migration wave schedule disrupted; 3,000+ users cannot access MZ services "
        "on schedule; customer-facing systems at risk of outage; GoLive (2028-01-01) "
        "with partial user base; compliance fines and legal costs EUR 1-3M.",
        "2027-08-15", "JCI Legal Counsel", "Formal Risk Review",
        "Very High", "30%", "threat",
        "Multi-jurisdiction employment law is a high-impact risk for a 12,000-user "
        "carve-out across 48 sites; must be addressed in Phase 1 before migration planning "
        "is finalised.",
        "1500000", "500000",
        "Avoid",
        "Engage local legal counsel in top 10 user-count jurisdictions by QG1 (2026-10-01); "
        "map TUPE triggers per country against wave migration schedule; integrate legal "
        "clearance gates into wave migration plan; begin works council/union consultation "
        "in Germany, France, Netherlands by 2026-11-01; obtain KPMG legal sign-off "
        "on migration schedule before Phase 3 start.",
        "2026-10-01", "not started",
        "German works council consultation is on critical path for Wave 1"
    ),

    # R006 - GDPR / data breach during migration (Security)
    (
        "R006", "Security & Data Protection",
        "12,000 users personal data (PII) and Aircon business IP are migrating across "
        "JCI infrastructure to Infosys-managed Merger Zone across 18 months; data-in-transit "
        "security depends on end-to-end encryption and DLP controls not all implemented "
        "at Phase 2 start; Infosys holds privileged elevated access to both JCI and MZ "
        "environments during dual-run period; 48 cross-border data flows.",
        "Unauthorized access or data exfiltration event during migration window; "
        "GDPR breach notification required within 72 hours of discovery to DPA; "
        "regulatory fines up to EUR 20M or 4% of global turnover; EU regulator "
        "may impose processing stop affecting all EU user migrations; "
        "reputational damage affecting deal valuation.",
        "Deal at risk if data breach during carve-out triggers regulatory investigation; "
        "migration halted for EU 15+ sites; remediation cost EUR 5-20M; "
        "3-6 month programme pause while security audit completed.",
        "2028-01-01", "Infosys Security Lead", "Formal Risk Review",
        "Very High", "30%", "threat",
        "GDPR breach during a high-profile cross-border M&A migration would attract "
        "maximum regulatory scrutiny; risk must be managed through architectural controls "
        "and insurance.",
        "5000000", "2000000",
        "Transfer",
        "Mandate ISO 27001 certification for MZ environment before Phase 3; "
        "require Infosys cyber liability insurance covering data breach during migration; "
        "implement CASB and DLP on all migration data flows from Phase 2 start; "
        "conduct penetration test before Phase 3 (target Aug 2027); "
        "appoint independent GDPR Data Protection Officer; "
        "encrypt all data at rest and in transit; "
        "DPA notification procedure rehearsed before GoLive.",
        "2027-02-01", "not started",
        "Infosys cyber liability insurance must be in place before Phase 2"
    ),

    # R007 - Application incompatibility with Merger Zone (Technology)
    (
        "R007", "Technology, R&D",
        "1,800+ applications built for JCI domain, Active Directory, and on-premise hosting; "
        "Wave 2 (~800 apps) and Wave 3 (~600 apps) include legacy, custom-coded, and "
        "third-party applications; re-pointing to Merger Zone AD, M365, and ITSM requires "
        "application-by-application compatibility testing; some apps require code changes "
        "for MZ compatibility; legacy apps may have no vendor support.",
        "Wave 2 and Wave 3 applications show higher-than-forecast incompatibilities with "
        "MZ environment during Phase 3 (Aug-Dec 2027); remediation backlog grows beyond "
        "capacity; Wave 2 and Wave 3 migrations slip past QG4 deadline (2027-12-08); "
        "GoLive proceeds with incomplete application estate; user productivity impacted.",
        "12,000 users have degraded productivity post-GoLive due to missing applications; "
        "business operations partially dependent on JCI systems post-GoLive requiring "
        "emergency TSA extension; cost of application remediation over budget.",
        "2027-11-30", "Infosys Application Lead", "Formal Risk Review",
        "High", "50%", "threat",
        "1,800+ app migration carries inherent compatibility risk; Wave 2/3 are the "
        "highest-risk waves due to legacy and specialist applications.",
        "2000000", "500000",
        "Mitigate",
        "Complete application compatibility pre-assessment for top 250 apps by QG1; "
        "build 15% schedule buffer into Wave 2 and Wave 3 plans; "
        "establish 3-tier remediation: quick-config (target 2 days), "
        "re-configure (target 2 weeks), retire/replace (plan in Phase 1); "
        "Wave 1 dry run to validate MZ compatibility approach before Wave 2 starts.",
        "2026-10-01", "not started",
        "App pre-assessment results must feed into Wave 2/3 schedule at QG1"
    ),

    # R008 - Infosys cost overrun on MZ delivery (Budget)
    (
        "R008", "Budget",
        "Infosys MZ delivery contract may have underestimated scope; 48-site connectivity, "
        "1,800+ app migration, and SAP system copy at this scale carry high uncertainty; "
        "scope creep from Bosch IT change requests or JCI data quality issues could inflate "
        "MZ contract value; change control process not yet established at project start.",
        "Infosys raises change requests in Phase 2/3 totalling more than 15% of contracted "
        "MZ delivery cost; steering board approval for additional spend takes 4-6 weeks; "
        "Infosys slows delivery pending approval; Phase 2/3 delivery impacted; "
        "total programme budget overrun.",
        "Programme cost overrun of EUR 3-5M; Bosch procurement approval required; "
        "Phase 2/3 delivery paused during approval process; schedule impact of "
        "4-6 weeks per unresolved change request.",
        "2027-06-30", "Bosch Programme Sponsor", "Formal Risk Review",
        "High", "50%", "threat",
        "Fixed-price IT delivery contracts at this scale frequently experience cost overrun; "
        "proactive change control governance is the primary mitigation.",
        "3000000", "0",
        "Mitigate",
        "Establish change control board (KPMG + Bosch + Infosys) with 10-business-day "
        "SLA on all change requests; ring-fence 15% budget contingency for MZ changes; "
        "conduct QG1 scope validation with Infosys before full contract signature; "
        "monthly spend-to-plan dashboard reviewed by steering from Phase 2 start.",
        "2026-10-01", "not started",
        "Change control board charter to be agreed with Infosys at QG1"
    ),

    # R009 - User migration wave delays (Schedule)
    (
        "R009", "Schedule",
        "12,000 users across 4 migration waves covering 48 sites must complete by QG4 "
        "(2027-12-08); Wave 3 (Sites 25-36) and Wave 4 (Sites 37-48) include challenging "
        "APAC sites (China, India) and LATAM sites (Mexico, Brazil) with ISP lead times "
        "of 8-16 weeks; local IT field support availability at tail sites is limited; "
        "network circuit delivery is on the critical path for Wave 3/4 start.",
        "Wave 3 (target Aug-Oct 2027) or Wave 4 (target Oct-Nov 2027) delayed by network "
        "circuit failures or local IT resource unavailability; 3,000+ users for those "
        "waves not migrated by QG4 (2027-12-08); GoLive scope incomplete; "
        "JCI emergency TSA extension required.",
        "GoLive with partial user migration; business disruption for unmigrated users; "
        "TSA emergency extension cost EUR 200-500K/month; programme steering board "
        "escalation; GoLive scope exception needed.",
        "2027-12-08", "Infosys Service Delivery Lead", "Status Meeting",
        "Moderate", "50%", "threat",
        "APAC and LATAM network circuits are on the critical path for Wave 3/4; "
        "early ISP engagement is the key mitigation action.",
        "500000", "0",
        "Mitigate",
        "Pre-qualify ISPs at APAC and LATAM sites during Phase 1 (Oct 2026); "
        "issue ISP purchase orders for long-lead APAC sites before Phase 2 start; "
        "contract local IT field support at tail sites by 2027-03-01; "
        "build 2-week buffer between Wave 4 end target and QG4; "
        "define GoLive-with-exceptions policy for fewer than 50 residual users.",
        "2026-10-01", "not started",
        "ISP RFQs for APAC sites must be issued in Phase 1 discovery phase"
    ),

    # R010 - SAP data quality from JCI system (Quality)
    (
        "R010", "Quality",
        "JCI SAP system has 20+ years of accumulated data with known quality issues: "
        "duplicate vendor master records, inactive cost center allocations, unmapped "
        "GL accounts from legacy acquisitions, and inconsistent material master data "
        "across countries; non-selective SAP copy approach will migrate dirty data "
        "to Merger Zone; data clean-up requires active JCI business team cooperation "
        "during Phase 1 and 2.",
        "SAP data migration dry runs (Phase 2/3, Apr-Aug 2027) reveal significant data "
        "quality failures in financial and logistics data; business validation team rejects "
        "migration output; additional data cleansing sprints of 6-8 weeks required; "
        "SAP Mock Cutover 1 and 2 delayed; QG4 target (2027-12-08) at risk.",
        "SAP production readiness cannot be confirmed by QG4; GoLive delayed; "
        "post-GoLive financial reporting accuracy at risk; statutory reporting in "
        "key countries (Germany, US) delayed; audit qualification risk.",
        "2027-10-31", "JCI SAP Owner", "Formal Risk Review",
        "High", "50%", "threat",
        "SAP data quality is a known risk in all JCI-scale separations; "
        "early data assessment and business engagement is essential.",
        "1000000", "500000",
        "Mitigate",
        "Commission SAP data quality pre-assessment by Phase 1 end (2027-01-31); "
        "establish JCI-led data cleansing sprint Oct-Dec 2026 with weekly KPMG review; "
        "define data acceptance criteria per module before SAP copy in Phase 2; "
        "KPMG data architect validates all extraction outputs; "
        "reject-and-retry mechanism built into SAP copy run book.",
        "2026-11-01", "not started",
        "JCI finance team engagement critical for data quality sign-off"
    ),

    # R011 - JCI stakeholder cooperation decline (Stakeholder)
    (
        "R011", "Stakeholder Relations & Public Affairs",
        "JCI is divesting the Aircon business; after deal financial close, JCI has "
        "reduced commercial incentive to prioritize IT carve-out support; JCI IT staff "
        "assigned to the programme may be affected by JCI corporate restructuring "
        "or redundancy programmes post-deal; JCI programme sponsor focus may shift "
        "to other JCI business units.",
        "JCI IT cooperation declines in Phase 2 (Feb 2027) or Phase 3 (Aug 2027); "
        "access to JCI production SAP system for copy and monitoring is restricted; "
        "JCI data quality clean-up tasks de-prioritized; Infosys blocked from progressing "
        "without JCI system credentials and context; KPMG PMO cannot enforce deliverables.",
        "SAP system copy blocked; application inventory gaps discovered late; "
        "Wave 1 user migration delayed; overall programme milestone slip of 4-8 weeks; "
        "KPMG must escalate to JCI board level.",
        "2027-06-30", "Bosch Programme Sponsor", "Formal Risk Review",
        "High", "30%", "threat",
        "JCI co-operation is a mandatory dependency for all six workstreams; "
        "contractual obligations and financial incentives are the primary mitigations.",
        "1500000", "0",
        "Mitigate",
        "Embed JCI IT cooperation obligations with named deliverable commitments in SPA "
        "and TSA before financial close; KPMG PMO tracks JCI deliverable completion weekly "
        "with RAG status; escalate to joint steering if JCI deliverable SLA breached; "
        "negotiate JCI IT retention bonus linked to programme milestones (QG1, QG2, GoLive); "
        "KPMG shadow of key JCI roles from Phase 1 for knowledge capture.",
        "2026-08-01", "not started",
        "JCI deliverable obligations must be contractually binding in SPA"
    ),

    # R012 - Network connectivity delays 48 sites (Technology)
    (
        "R012", "Technology, R&D",
        "48 sites in multiple regions require SD-WAN/WAN connectivity to Merger Zone; "
        "ISP lead times: China 12-16 weeks, LATAM (Mexico, Brazil) 8-12 weeks, "
        "Eastern Europe 8-10 weeks, Middle East 6-8 weeks; Infosys network team "
        "starts formal procurement in Phase 2 (Feb 2027) which may be too late "
        "for APAC circuits needed for Wave 3 (Aug 2027).",
        "Network circuits for 8-12 APAC and LATAM sites not activated by Phase 3 "
        "user migration start (Aug 2027); 2,000+ users at affected sites cannot be "
        "migrated to Merger Zone; wave schedule disrupted; GoLive with sites still "
        "on JCI network.",
        "User migration Waves 3/4 scope reduced; JCI TSA extension required for "
        "unmigrated sites; GoLive with partial estate; ongoing cost of maintaining "
        "split IT environment between JCI and MZ.",
        "2027-08-01", "Infosys Network Lead", "Formal Risk Review",
        "High", "30%", "threat",
        "Network connectivity is on the critical path for all user migration waves; "
        "APAC and LATAM ISP lead times require pre-Phase 2 procurement actions.",
        "500000", "200000",
        "Mitigate",
        "Initiate ISP RFQs for all APAC and LATAM sites during Phase 1 (Oct-Dec 2026) "
        "before official Phase 2 start; identify China-specific ISP immediately; "
        "evaluate SD-WAN overlay as interim connectivity for tail sites during ISP delay; "
        "network circuit delivery tracked in weekly MZ build dashboard; "
        "all APAC POs issued by 2026-12-01.",
        "2026-10-01", "not started",
        "China ISP selection is on critical path - start in Phase 1"
    ),

    # R013 - Data sovereignty multi-jurisdiction (Legal)
    (
        "R013", "Legal & Compliance",
        "Merger Zone potentially hosted in data centres outside country of origin for "
        "JCI Aircon data; 12,000 users across 48 countries include EU (GDPR, Schrems II), "
        "China (PIPL, data localisation law), and US (state privacy laws); "
        "Aircon product IP and R&D data may have export control restrictions (ITAR/EAR) "
        "under US law; MZ architecture design in Phase 1 must address these requirements.",
        "Data sovereignty analysis in Phase 1 (Oct-Dec 2026) reveals MZ hosting design "
        "is non-compliant for China (PIPL requires local data storage) or EU users "
        "(Schrems II invalidation of data transfers); Infosys MZ architecture requires "
        "redesign to add regional nodes; Phase 2 start delayed 4-8 weeks; "
        "additional MZ hosting cost EUR 1-2M.",
        "Phase 2 delayed; MZ architecture change impacts SAP copy timing; "
        "overall programme schedule pushed 6-8 weeks; regulatory sanctions if "
        "data sovereignty non-compliance proceeds to production.",
        "2027-02-01", "JCI Legal Counsel", "Formal Risk Review",
        "High", "30%", "threat",
        "Data sovereignty is a prerequisite for MZ architecture; must be resolved "
        "in Phase 1 before Infosys finalises MZ blueprint.",
        "1000000", "300000",
        "Avoid",
        "Conduct comprehensive data sovereignty assessment in Phase 1 (target 2026-11-01); "
        "require Infosys MZ blueprint to include China-local and EU-local hosting nodes; "
        "engage Chinese legal counsel for PIPL compliance before MZ design finalised; "
        "review US export control (ITAR/EAR) applicability for Aircon IP data flows; "
        "MZ architecture approved by KPMG legal and privacy team before QG1.",
        "2026-11-01", "not started",
        "China PIPL assessment is priority - must inform MZ blueprint"
    ),

    # R014 - QG4 failure (Schedule)
    (
        "R014", "Schedule",
        "QG4 gate (2027-12-08) requires zero blocking defects across all workstreams: "
        "SAP production readiness, 12,000 user migration completion, 1,800+ app Wave 2/3 "
        "completion, network connectivity all 48 sites, infrastructure readiness, "
        "and DR test passed; six workstreams must all achieve green status simultaneously "
        "in a 6-day pre-gate check window; any late-breaking critical defect causes gate failure.",
        "QG4 gate assessment fails on 2027-12-08 due to open P1/P2 defects in SAP, "
        "network connectivity gaps at tail sites, or user migration Wave 4 incomplete; "
        "gate deferred to 2028-01-15; GoLive pushed from 2028-01-01 to 2028-02-01; "
        "JCI business plan Aircon separation delayed.",
        "GoLive delayed 4-6 weeks; Bosch acquisition integration delayed; "
        "MZ extended operational cost EUR 500K/month; JCI TSA extension at premium; "
        "board-level visibility required.",
        "2027-12-08", "KPMG PMO Lead", "Formal Risk Review",
        "Very High", "30%", "threat",
        "QG4 is the single gate controlling GoLive; all six workstreams must achieve "
        "simultaneous green status; rolling defect management from Phase 3 is critical.",
        "3000000", "1000000",
        "Mitigate",
        "Implement rolling defect triage from UAT week 1 (Aug 2027); "
        "track defect burn-down daily from Oct 2027; "
        "pre-define QG4 go/no-go criteria explicitly: "
        "zero P1 defects, fewer than 5 P2 defects, all Wave 2/3 apps live, "
        "all 48 sites network-connected, 12,000 users migrated; "
        "hold bi-weekly QG4 readiness review from 2027-09-01; "
        "reserve 30-day GoLive buffer (max GoLive 2028-02-01) with steering approval.",
        "2027-09-01", "not started",
        "QG4 go/no-go criteria must be approved by steering before Phase 3 start"
    ),

    # R015 - Extended MZ operational cost (Budget)
    (
        "R015", "Budget",
        "Merger Zone is a transient environment running operational costs including "
        "Infosys management fees, cloud hosting, software licences, and support; "
        "MZ operational burn rate estimated EUR 400-600K/month during Phase 2/3/4 "
        "dual-run period; any schedule delay directly extends dual-run costs; "
        "exact MZ annual budget not yet approved beyond Phase 2 start.",
        "Programme schedule slips 3+ months due to SAP, network, or QG4 issues; "
        "MZ dual-run operational spend increases by EUR 1.2-1.8M per 3-month delay; "
        "Bosch budget authority required for additional spend; "
        "steering board escalation required; project month-end close delayed.",
        "Budget overrun of EUR 1.5-3M impacting Bosch acquisition business case; "
        "procurement escalation process delays delivery further; "
        "Bosch CFO visibility required.",
        "2027-12-31", "Bosch Programme Sponsor", "Formal Risk Review",
        "High", "30%", "threat",
        "MZ dual-run costs compound with any schedule delay; "
        "schedule adherence is the most cost-effective risk control.",
        "1500000", "0",
        "Mitigate",
        "Pre-approve 3-month MZ extension budget contingency at QG1 board decision; "
        "review schedule risk and MZ cost exposure monthly from Phase 2; "
        "phase-down MZ non-essential services within 30 days of GoLive; "
        "Infosys contract includes demobilisation schedule triggered by GoLive confirmation.",
        "2026-10-01", "not started",
        "3-month extension budget approval to be included in QG1 financial decision"
    ),

    # R016 - Ransomware/cyber attack on Merger Zone (Security)
    (
        "R016", "Security & Data Protection",
        "Merger Zone is a new, complex IT environment operated by Infosys with privileged "
        "access to JCI production systems during dual-run; Phase 2 build period (Feb-Jul 2027) "
        "is highest risk window as security controls are not fully hardened; "
        "Aircon business is a high-value target: 12,000 user credentials, SAP financial data, "
        "product IP, and M&A transaction details are accessible via MZ infrastructure; "
        "Infosys supply chain access increases attack surface.",
        "Ransomware attack targets Merger Zone infrastructure in Phase 2/3; "
        "SAP data or app migration data encrypted; DR test environment must be activated "
        "under incident conditions; SAP build timeline set back 4-8 weeks; "
        "personal data of 12,000 users potentially exfiltrated; GDPR breach notification triggered.",
        "Programme delay of 4-8 weeks; SAP data re-extraction required; "
        "GDPR breach notification to DPA with potential EUR 20M fine; "
        "reputational damage to JCI, Bosch, and Infosys; "
        "cyber incident response cost EUR 2-5M.",
        "2027-07-31", "Infosys Security Lead", "Formal Risk Review",
        "Very High", "10%", "threat",
        "Low probability but very high impact; zero-trust and immutable backup architecture "
        "are essential pre-conditions for MZ build.",
        "5000000", "0",
        "Transfer",
        "Implement zero-trust network architecture in MZ from Phase 2 start (Feb 2027); "
        "mandatory 24/7 SOC monitoring by Infosys security team; "
        "cyber insurance covering MZ environment and data breach response; "
        "immutable backup strategy for all SAP data with air-gapped copy; "
        "penetration test of MZ before any JCI production data migrated; "
        "incident response playbook rehearsed before Phase 3.",
        "2027-02-01", "not started",
        "Infosys cyber insurance must be confirmed with KPMG before Phase 2 start"
    ),

    # R017 - JCI IT staff attrition (Resources)
    (
        "R017", "Resources",
        "JCI Aircon IT staff (SAP Basis team, network architects, application owners) "
        "are aware of carve-out and potential redundancy; JCI has limited financial "
        "incentive to retain staff beyond deal close; key individuals (JCI SAP Basis lead, "
        "JCI IT Manager) carry critical tacit knowledge of JCI systems configuration, "
        "SAP authorizations, and network topology across 48 sites.",
        "Key JCI SAP Basis lead departs during Phase 2 (Feb-Jul 2027); SAP system "
        "access credentials and environment knowledge unavailable; Infosys and KPMG "
        "cannot proceed SAP copy without JCI Basis team; SAP build timeline extends "
        "4-6 weeks while JCI provides a replacement.",
        "SAP copy and interface rewiring delayed; QG2&3 gate at risk; "
        "KPMG PMO must escalate to JCI board to force resource replacement; "
        "programme cost of delay EUR 500K.",
        "2027-06-30", "JCI Programme Sponsor", "Status Meeting",
        "High", "30%", "threat",
        "JCI key staff retention is a standard risk in distressed-seller carve-outs; "
        "retention financial incentives and knowledge documentation are key controls.",
        "500000", "0",
        "Mitigate",
        "Negotiate JCI IT staff retention bonus pool covering SAP Basis, "
        "Network Architect, and IT Manager roles through GoLive (2028-01-01); "
        "KPMG shadow programme for all key JCI roles from Oct 2026; "
        "document JCI IT system access procedures and credentials in secure KPMG vault; "
        "contingency plan for each key JCI role identified and approved by QG1.",
        "2026-10-01", "not started",
        "JCI retention terms to be in SPA before deal financial close"
    ),

    # R018 - Manufacturing OT/SCADA systems (Manufacturing)
    (
        "R018", "Manufacturing",
        "JCI Aircon business operates manufacturing facilities for air conditioning "
        "equipment; OT/SCADA/MES systems at these facilities are integrated with "
        "JCI IT applications and SAP (production orders, quality management); "
        "OT systems have long revalidation timelines (6-12 months post-change) and "
        "regulatory certification requirements; OT environments not yet included "
        "in the 1,800 app inventory at project start.",
        "Wave 3 specialist applications (target Oct-Nov 2027) include OT/SCADA-connected "
        "systems not compatible with Merger Zone without factory revalidation; "
        "factory revalidation timelines push Wave 3 completion past QG4 (2027-12-08); "
        "manufacturing continuity risk during OT-IT separation window.",
        "Wave 3 OT applications not migrated by GoLive; factories dependent on split "
        "IT/OT environment post-GoLive; risk of production disruption; "
        "regulatory certification gap in manufacturing quality systems; "
        "GoLive exception required with OT remediation plan.",
        "2027-11-30", "JCI IT Manager", "Formal Risk Review",
        "High", "30%", "threat",
        "OT/SCADA systems are often forgotten in IT carve-outs until too late; "
        "early identification and separate migration track are essential.",
        "1000000", "500000",
        "Mitigate",
        "Identify all OT-connected applications in Phase 1 app inventory by 2026-11-01; "
        "commission OT-IT separation technical assessment by KPMG/Infosys by 2027-01-31; "
        "plan OT systems on a separate migration track with factory approval gates; "
        "engage Bosch OT team (as future owner) in MZ architecture design from Phase 1; "
        "Wave 3 OT migrations scheduled before QG4 with 4-week factory validation buffer.",
        "2026-11-01", "not started",
        "OT inventory must be completed in Phase 1 app assessment by Nov 2026"
    ),

    # R019 - Merger Zone performance under 12k users (Quality)
    (
        "R019", "Quality",
        "Merger Zone is an Infosys-designed environment; performance benchmarks are based "
        "on architecture estimates and comparable projects; actual 12,000-user concurrent "
        "load during performance testing (Phase 3, Aug-Sep 2027) may reveal bottlenecks "
        "in MZ compute, storage, or SAP application server capacity; "
        "Infosys contracts performance SLAs but may not have designed for peak M&A "
        "workloads across 48 sites.",
        "Performance test results in Aug-Sep 2027 show MZ cannot sustain 12,000 concurrent "
        "users at required response times (target <3 sec for SAP transactions); "
        "infrastructure scaling requires additional hardware procurement of 6-8 weeks; "
        "retest required; Phase 3 schedule impacted; QG4 at risk.",
        "Phase 3 bottleneck resolves after hardware remediation but consumes the "
        "QG4 buffer; performance defects carry into QG4 check; "
        "Infosys remediation cost EUR 500K+ not budgeted.",
        "2027-09-30", "Infosys Infrastructure Architect", "Formal Risk Review",
        "High", "30%", "threat",
        "Performance issues discovered late in Phase 3 are expensive and schedule-critical; "
        "contractual SLAs and early incremental testing are key controls.",
        "500000", "0",
        "Mitigate",
        "Define performance SLAs contractually with Infosys pre-Phase 2 (Feb 2027): "
        "SAP response time <3 sec for standard transactions at 12,000 users; "
        "conduct incremental load testing starting at 2,000, then 6,000, then 12,000 users; "
        "Infosys obligated to remediate performance failures within 4-week SLA at no cost; "
        "performance test plan approved by KPMG architect before testing starts.",
        "2026-10-01", "not started",
        "Performance SLAs must be in Infosys contract before Phase 2"
    ),

    # R020 - Bosch integration scope creep (Strategy)
    (
        "R020", "Strategy & Portfolio",
        "Post-GoLive integration of MZ services into Bosch IT is an Integration model "
        "deliverable not fully scoped at project start; Bosch IT may request additional "
        "data migrations, system integrations, architecture changes, or Bosch-standard "
        "tooling that was not anticipated in the original carve-out scope; "
        "Bosch IT organisation is not a primary stakeholder during Phase 2/3 build.",
        "Bosch IT requests significant MZ re-architecture or additional integration "
        "deliverables during hypercare (Jan-Apr 2028) not covered in programme scope; "
        "Infosys and KPMG scope extensions not budgeted; QG5 closure gate (2028-04-01) "
        "delayed; programme extends beyond planned completion.",
        "Programme continues past QG5 date; post-GoLive integration costs not budgeted; "
        "Infosys contract extension required; KPMG engagement extended; "
        "Bosch IT leadership alignment required.",
        "2028-04-01", "Bosch Programme Sponsor", "Status Meeting",
        "Moderate", "30%", "threat",
        "Scope creep in the integration phase is common in Integration-model carve-outs; "
        "defining a clear handover boundary is essential.",
        "1000000", "0",
        "Avoid",
        "Define Bosch IT integration scope in transition agreement by QG1; "
        "scope freeze: all Bosch post-GoLive integration work requiring new delivery "
        "against separate Bosch IT budget from Day 91 of hypercare; "
        "hypercare scope strictly limited to stabilisation and hotfixes (no new features); "
        "run off-programme strategy session with Bosch IT in Phase 1 to identify "
        "integration requirements and budget accordingly.",
        "2026-10-01", "not started",
        "Bosch integration scope definition must be tabled at QG1"
    ),

    # R021 - Hardware procurement lead times (Supply Chain)
    (
        "R021", "Supply Chain",
        "MZ infrastructure requires significant server, storage, and network hardware; "
        "global hardware supply chain disruptions and semiconductor shortages create "
        "16-20 week procurement lead times; Infosys formal hardware procurement "
        "starting Phase 2 (Feb 2027) may be too late for some long-lead components "
        "needed for MZ DC readiness by Apr 2027.",
        "Critical MZ hardware (compute servers, storage arrays, network switches) "
        "not delivered by MZ DC setup target (2027-04-02); Infosys cannot complete "
        "MZ environment build on schedule; SAP system copy start delayed from Jun 2027; "
        "QG2&3 gate (2027-07-31) at risk.",
        "Phase 2 milestone delayed 4-8 weeks; SAP build and app setup on MZ deferred; "
        "cascade to Phase 3 start delay; programme milestone slip of 2-3 months.",
        "2027-04-15", "Infosys Infrastructure Architect", "Formal Risk Review",
        "High", "30%", "threat",
        "Hardware procurement timelines are fixed by market conditions; "
        "early procurement decision before Phase 2 is the only reliable mitigation.",
        "300000", "0",
        "Mitigate",
        "Trigger hardware procurement decisions at QG1 (2026-10-01) with Bosch budget approval; "
        "issue Infosys POs for all long-lead hardware items by 2026-11-01; "
        "evaluate and prefer Azure/AWS cloud-first (IaaS) MZ design to eliminate "
        "physical hardware dependency; Infosys hardware tracker reviewed weekly "
        "from QG1 with KPMG oversight.",
        "2026-10-01", "not started",
        "Hardware or cloud decision at QG1 is a critical scheduling dependency"
    ),

    # R022 - Customer-facing system outage during migration (Customers)
    (
        "R022", "Customers",
        "JCI Aircon serves external customers (building contractors, HVAC systems buyers, "
        "field service engineers) through customer portals, order management, and field "
        "service applications that are included in the 1,800+ app scope; "
        "these customer-facing applications are mission-critical and must maintain "
        "24/7 uptime during migration; Wave 1 includes most critical customer-facing apps.",
        "Customer-facing Wave 1 applications (order portal, service portal, "
        "dealer portal) experience outages or degraded performance during migration "
        "cutover window (Phase 2/3, May-Oct 2027); Aircon customers unable to place orders "
        "or access service management; SLA breach for field service contracts; "
        "customer NPS impact.",
        "Aircon revenue loss estimated EUR 500K-2M per week of customer portal outage; "
        "field service SLA penalties; customer attrition; reputational damage "
        "during M&A transition period; contractual penalty clauses triggered.",
        "2027-08-31", "JCI IT Manager", "Formal Risk Review",
        "High", "30%", "threat",
        "Customer-facing systems require zero-downtime migration planning; "
        "standard wave approach must be enhanced for customer portal stack.",
        "1500000", "500000",
        "Mitigate",
        "Flag all customer-facing applications in Phase 1 app inventory by 2026-11-01; "
        "require zero-downtime migration plan (blue-green deployment) for top 10 "
        "customer-critical apps; conduct extended parallel-run of 2 weeks for "
        "customer portals before DNS cut; test customer portals with 500-user pilot "
        "before general cutover; customer impact communication plan approved by JCI "
        "commercial team before Wave 1.",
        "2026-11-01", "not started",
        "Zero-downtime requirement must be flagged to Infosys app team before Wave 1 design"
    ),

    # R023 - IP segregation completeness (Intellectual Property)
    (
        "R023", "Intellectual Property",
        "JCI Aircon has proprietary HVAC designs, patents, and product roadmap data "
        "managed in PLM (Product Lifecycle Management) and document management systems "
        "integrated with JCI corporate infrastructure; post-deal IP ownership transfers "
        "to Bosch but JCI may retain inadvertent read-access if segregation is incomplete; "
        "PLM system not yet confirmed as in Wave 1 scope.",
        "IP segregation in PLM and document management systems incomplete at GoLive "
        "(2028-01-01); JCI retains read-access to Bosch-owned Aircon IP in shared systems "
        "post-deal; Bosch legal discovers gap during integration audit; "
        "or conversely, Aircon employees unable to access product designs in MZ "
        "causing engineering disruption.",
        "Intellectual property dispute between JCI and Bosch post-deal; "
        "legal remediation cost EUR 500K+; product development continuity interrupted "
        "for engineering teams; potential deal impairment if IP ownership is contested.",
        "2028-01-01", "JCI Legal Counsel", "Formal Risk Review",
        "High", "30%", "threat",
        "IP segregation is a legal obligation for clean carve-out; "
        "PLM and document management must be explicitly included in scope.",
        "500000", "200000",
        "Avoid",
        "Include PLM and all document management systems in Wave 1 migration scope by QG1; "
        "conduct IP segregation audit jointly by JCI and Bosch legal teams by 2027-01-31; "
        "JCI read-access to Aircon IP systems revoked within 30 days post-GoLive; "
        "Bosch IP counsel signs off IP segregation completeness at QG4; "
        "independent IP audit report commissioned before GoLive.",
        "2026-10-01", "not started",
        "PLM scope confirmation required in Phase 1 app inventory"
    ),

    # R024 - Infosys subcontractor ethics (Ecosystems & Ethics)
    (
        "R024", "Ecosystems & Ethics",
        "Infosys may use subcontractors in low-cost regions for MZ infrastructure build "
        "and data migration components; subcontractors may have access to sensitive JCI "
        "Aircon SAP production data, 12,000 user credentials, and M&A transaction details; "
        "ethical data handling standards and background screening requirements "
        "may not match Bosch Supplier Code of Conduct; subcontractor list not disclosed "
        "at project start.",
        "Infosys subcontractor mishandles sensitive Aircon IP or SAP credentials; "
        "information security incident traced to subcontractor access; "
        "Bosch and JCI contractual data protection obligations breached; "
        "regulatory investigation triggered; deal completion risk if breach is material.",
        "Infosys primary contract liability but reputational risk for Bosch; "
        "regulatory investigation cost EUR 1-3M; "
        "potential programme pause pending security audit; "
        "Bosch board-level disclosure required.",
        "2027-12-31", "KPMG PMO Lead", "Formal Risk Review",
        "High", "10%", "threat",
        "Subcontractor supply chain risk is increasingly scrutinised by regulators; "
        "contractual and audit controls on Infosys supply chain are essential.",
        "1000000", "0",
        "Transfer",
        "Require Infosys to disclose full subcontractor list before Phase 2 start; "
        "subcontractor vetting per Bosch Supplier Code of Conduct enforced by Infosys; "
        "data processing agreements (DPA) signed with all subcontractors handling "
        "personal data or SAP credentials; "
        "KPMG audit rights over Infosys supply chain contractually secured; "
        "annual subcontractor compliance review built into governance calendar.",
        "2026-10-01", "not started",
        "Subcontractor disclosure is a pre-condition to Infosys contract signature"
    ),

    # R025 - OPPORTUNITY: Cloud modernisation during MZ migration
    (
        "R025", "Technology, R&D",
        "MZ infrastructure build provides an architectural fresh start free from JCI "
        "technical debt; Infosys has strong Azure and AWS expertise; migrating 1,800+ "
        "JCI on-premise applications and SAP to MZ creates a unique opportunity to "
        "architect for cloud-native from the outset rather than replicating the JCI "
        "on-premise model in MZ.",
        "Structured cloud migration review during Phase 1 architecture identifies "
        "300-400 JCI on-premise workloads suitable for Azure/AWS cloud-native design "
        "within MZ; cloud MZ reduces hardware procurement risk; "
        "Bosch ultimately inherits a modern cloud estate rather than on-premise MZ.",
        "Cloud-native MZ reduces long-term Bosch operational cost by EUR 3-5M/year "
        "post-integration; eliminates hardware procurement lead-time risk; "
        "accelerates Bosch cloud strategy; positions Aircon business as cloud-ready at GoLive.",
        "2027-02-01", "Infosys Cloud Lead", "Formal Risk Review",
        "High", "50%", "opportunity",
        "Cloud-first MZ design is a strategic opportunity to modernise the Aircon IT estate "
        "at zero additional effort cost during the mandatory migration programme.",
        "0", "5000000",
        "Exploit",
        "Mandate cloud suitability flag in Phase 1 application inventory; "
        "Infosys cloud architect to design MZ with Azure-first architecture from QG1; "
        "present cloud ROI business case at QG1 for Bosch steering approval; "
        "pilot cloud MZ with Wave 1 non-SAP apps; "
        "target 300+ apps on cloud-native by GoLive; "
        "decommission physical MZ hardware within 6 months post-GoLive.",
        "2026-10-01", "not started",
        "Cloud-first decision must be made at QG1 to influence MZ hardware procurement"
    ),
]


# =============================================================================
# MAIN
# =============================================================================

def main():
    print(f"[Trinity-CAM] Generating risk register from template...")
    print(f"  Template: {TEMPLATE}")

    if not TEMPLATE.exists():
        print(f"  ERROR: Template not found: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(str(TEMPLATE))

    # --- Info sheet ---
    info = wb["Info"]
    info["C4"] = "TCM-RR-001"
    info["C5"] = "Trinity-CAM_Risk_Register.xlsx"
    info["C6"] = "Trinity-CAM"
    info["C7"] = "TCM-2026"
    info["C8"] = "KPMG PMO Lead"

    # --- Risk Register sheet ---
    rr = wb["Risk Register"]
    first_data_row = 5

    for i, risk in enumerate(RISKS):
        r = first_data_row + i
        (rid, cat, cause, event, effect, evt_date,
         owner, source, impact, prob, p_or_t, qualitative,
         eur_cy, eur_3y, strategy, measure, due_date,
         status, notes) = risk

        rr.cell(r, 2).value  = rid
        rr.cell(r, 3).value  = TODAY
        rr.cell(r, 4).value  = cat
        rr.cell(r, 5).value  = cause
        rr.cell(r, 6).value  = event
        rr.cell(r, 7).value  = effect
        rr.cell(r, 8).value  = evt_date
        rr.cell(r, 9).value  = owner
        rr.cell(r, 10).value = source
        # col 11 (K) unused/spacer
        rr.cell(r, 12).value = impact      # L
        rr.cell(r, 13).value = f'=_xlfn.IFNA(VLOOKUP(L{r},$D$182:$E$186,2,FALSE),"")' # M
        rr.cell(r, 14).value = prob        # N
        rr.cell(r, 15).value = f'=_xlfn.IFNA(VLOOKUP(N{r},$D$189:$E$193,2,FALSE),"")' # O
        rr.cell(r, 16).value = p_or_t     # P
        rr.cell(r, 17).value = f'=M{r}*O{r}'  # Q
        rr.cell(r, 18).value = qualitative # R
        rr.cell(r, 19).value = int(eur_cy) # S
        rr.cell(r, 20).value = int(eur_3y) # T
        # col 21 (U) spacer
        rr.cell(r, 22).value = strategy   # V
        rr.cell(r, 23).value = measure    # W
        rr.cell(r, 24).value = due_date   # X
        # col 25 (Y) spacer
        rr.cell(r, 26).value = status     # Z
        rr.cell(r, 27).value = REPORT     # AA
        rr.cell(r, 28).value = impact     # AB (actual = initial)
        rr.cell(r, 29).value = f'=_xlfn.IFNA(VLOOKUP(AB{r},$D$182:$E$186,2,FALSE),"")' # AC
        rr.cell(r, 30).value = prob       # AD (actual = initial)
        rr.cell(r, 31).value = f'=_xlfn.IFNA(VLOOKUP(AD{r},$D$189:$E$193,2,FALSE),"")' # AE
        rr.cell(r, 32).value = f'=AC{r}'  # AF
        rr.cell(r, 33).value = f'=AE{r}'  # AG
        rr.cell(r, 34).value = f'=AF{r}*AG{r}'  # AH
        rr.cell(r, 35).value = notes      # AI

    # --- Fix Matrix sheet: black font on yellow cells ---
    if "Matrix " in wb.sheetnames:
        matrix_ws = wb["Matrix "]
        yellow_fills = {"FFFFFF00", "FFFFFFCC", "FFF2CC", "FFFF00"}
        for row in matrix_ws.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.fgColor:
                    colour = cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else ""
                    if colour in yellow_fills or colour.upper() in yellow_fills:
                        cell.font = Font(
                            name=cell.font.name if cell.font else "Calibri",
                            size=cell.font.size if cell.font else 11,
                            bold=cell.font.bold if cell.font else False,
                            color="000000"
                        )

    wb.save(str(OUTPUT))
    print(f"  Output: {OUTPUT}")
    print(f"  Risks written: {len(RISKS)}")
    print(f"\n[Trinity-CAM] Risk register complete.")
    print(f"  Top threats by score (P x I, 5-scale):")
    score_preview = [
        (r[0], r[1], r[8], r[9]) for r in RISKS if r[10] == "threat"
    ]
    prob_map = {"10%": 1, "30%": 2, "50%": 3, "70%": 4, "90%": 5}
    imp_map  = {"Very Low": 1, "Low": 2, "Moderate": 3, "High": 4, "Very High": 5}
    scored = sorted(
        [(rid, cat, imp, pr, imp_map.get(imp, 0) * prob_map.get(pr, 0))
         for rid, cat, imp, pr in score_preview],
        key=lambda x: -x[4]
    )
    for rid, cat, imp, pr, sc in scored[:10]:
        print(f"    {rid}  Score={sc}  [{pr} x {imp}]  {cat}")


if __name__ == "__main__":
    main()
