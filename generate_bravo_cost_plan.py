"""
generate_bravo_cost_plan.py
Generates Bravo_Cost_Plan.xlsx (formatted Excel workbook)

Project Bravo: BGSW AI Business Carve-Out into 50/50 JV with Tata
Seller: Bosch BGSW | Buyer: Tata
7 months: 01 Apr 2026 -> 30 Oct 2026 | 2 India sites | 70 users | 17 apps | No ERP | No TSA
PM: Riyaz Ahmed Syed Ahmed (BD/MIL-PSM4)

Run:  python generate_bravo_cost_plan.py
Output: Bravo\\Bravo_Cost_Plan.xlsx
"""

from pathlib import Path
from datetime import datetime as _dt

HERE = Path(__file__).parent
OUT_DIR = HERE / "Bravo"
OUT_DIR.mkdir(exist_ok=True)
XLSX_PATH = OUT_DIR / "Bravo_Cost_Plan.xlsx"

# ── Data ─────────────────────────────────────────────────────────────────────
# Row types: "meta", "category", "resource", "subtotal", "total", "blank",
#            "section_header", "breakdown_row", "capex_row", "note"

METADATA = [
    ("Based on:", "Bravo_Project_Schedule.xlsx"),
    ("Risk-aligned:", "Bravo_Risk_Register.xlsx"),
    ("Project:", "Bravo  |  Seller: Bosch BGSW  |  Buyer: Tata"),
    ("PM:", "Riyaz Ahmed Syed Ahmed (BD/MIL-PSM4)"),
    ("Sites / Users / Apps:", "2 (India)  |  70 users  |  17 apps  |  No ERP  |  No TSA"),
    ("Duration:", "01 Apr 2026 – 30 Oct 2026  (7 months)"),
    ("Budget Baseline:", "TBC - to be approved at QG0"),
]

# Category sections: (category_label, [(resource, days, hrs, rate, cost), ...])
CATEGORIES = [
    ("PROGRAMME MANAGEMENT", [
        ("Riyaz Ahmed (BD/MIL-PSM4)",       "Programme Management",              85,  680, 175, 119000),
        ("PMO + Riyaz Ahmed",               "PMO Coordination",                  30,  240, 100,  24000),
        ("Riyaz Ahmed",                     "Steering Committee Facilitation",   15,  120, 175,  21000),
    ]),
    ("IT PROJECT MANAGEMENT", [
        ("BGSW IT PM",                      "IT Project Management",             85,  680, 120,  81600),
        ("WS Leads",                        "Workstream Coordination",           40,  320, 100,  32000),
    ]),
    ("INFRASTRUCTURE & CLOUD", [
        ("BGSW AD Team",                    "Active Directory Build (JV Forest)", 15, 120, 110,  13200),
        ("BGSW Azure Team",                 "M365 Tenant & Azure Setup",         15,  120, 110,  13200),
        ("BGSW Infra",                      "Network Setup (2 India Sites)",     15,  120, 100,  12000),
        ("BGSW CISO",                       "Security & IAM Platform",           12,   96, 120,  11520),
        ("BGSW Cloud Team",                 "Cloud Environment Configuration",   10,   80, 110,   8800),
    ]),
    ("APPLICATION MIGRATION", [
        ("BGSW IT Architects",  "Application Deep-Dive & Classification (17 apps)",  15, 120, 110, 13200),
        ("Dev Teams",           "Application Reconfiguration for JV (17 apps)",      20, 160, 100, 16000),
        ("App Teams + Test Team","Integration & UAT Testing (17 apps)",              15, 120,  90, 10800),
        ("BGSW IT + Riyaz Ahmed","Data Migration Execution",                         12,  96, 100,  9600),
        ("BGSW Procurement + IT","License & Contract Review (17 apps)",              12,  96,  90,  8640),
    ]),
    ("CLIENT WORKPLACE", [
        ("BGSW IT + Asset Mgmt","Device Inventory & Assessment (70 devices)",         5,  40,  90,  3600),
        ("BGSW CWP",            "Device Reimaging - JV Standard Image (70 devices)", 10,  80,  90,  7200),
        ("IT Ops",              "Help Desk Activation & Training",                    5,  40,  90,  3600),
    ]),
    ("ARCHITECTURE & DESIGN", [
        ("BGSW IT Architects",      "As-Is Analysis & Architecture Design",      20, 160, 120, 19200),
        ("BGSW + Tata Architects",  "JV Architecture Workshop Facilitation",      5,  40, 120,  4800),
        ("BGSW CISO",               "Security Architecture Design",               8,  64, 120,  7680),
    ]),
    ("LEGAL & COMPLIANCE", [
        ("Legal + Finance",         "JV Legal Entity Setup (India)",             10,  80, 150, 12000),
        ("Legal + Riyaz Ahmed",     "Contract & Data Separation Review",         10,  80, 150, 12000),
    ]),
    ("HR IT", [
        ("BGSW HR IT",              "HR IT Mapping (70 Users)",                   5,  40,  90,  3600),
    ]),
    ("HYPERCARE & CLOSURE", [
        ("IT Ops + Help Desk",              "Hypercare Operations (60 working days)",     60, 480, 90, 43200),
        ("App Teams + IT Ops",              "Application Stability Monitoring (17 apps)", 30, 240, 90, 21600),
        ("Riyaz Ahmed + Bosch BD/MIL",      "Programme Closure",                          15, 120, 175, 21000),
    ]),
]

PHASE_BREAKDOWN = [
    ("Phase 0 - Initialization (01 Apr – 30 Apr 2026)",          55000),
    ("Phase 1 - Concept (01 May – 15 May 2026)",                 38000),
    ("Phase 2 - Architecture & Design (11 May – 29 May 2026)",   72000),
    ("Phase 3 - Build and Test (01 Jun – 26 Jun 2026)",         165000),
    ("Phase 4 - GoLive & Closure (29 Jun – 30 Oct 2026)",       224040),
]

CAPEX_ROWS = [
    ("M365/Azure JV Tenant Licensing (<70 users)",      "BGSW Procurement",  "TBC - to be confirmed at QG0"),
    ("India Cloud Infrastructure (2 sites)",            "BGSW Cloud Team",   "TBC - to be confirmed at QG0"),
    ("Network Circuit Setup (2 India JV Sites)",        "BGSW Infra",        "TBC - to be confirmed at QG0"),
    ("Device Hardware Replacement if required (≤70 devices)", "BGSW CWP",   "TBC - to be confirmed at QG0"),
    ("Risk Contingency (Risk #1 - schedule risk)",      "See Risk Register",  "15% of labour = EUR 83,106"),
    ("Risk Contingency (Risk #3 - cloud cost overrun)", "See Risk Register",  "EUR 15,000 reserved"),
]

NOTES = [
    "All figures are labour-only estimates; hardware, licensing, WAN, and cloud subscription costs are excluded from the labour total.",
    "CAPEX items (M365, Azure, networking) to be formally quoted and approved at QG0 SteerCo.",
    "Risk contingency lines derived from Bravo_Risk_Register.xlsx: Risk #1 (Amber; P4xI4=16) and Risk #3 (Amber; P3xI3=9).",
    "No TSA cost provision required: Bosch-led JV; BGSW services available via normal governance channels post-GoLive.",
    "Budget baseline state: TBC - to be approved at QG0 (30 Apr 2026).",
]


# ── Writer ────────────────────────────────────────────────────────────────────
def write_xlsx(path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Colours
    C_DARK_BLUE   = "003B6E"
    C_MID_BLUE    = "0066CC"
    C_LIGHT_BLUE  = "C6D4E8"
    C_ALT         = "EFF4FB"
    C_AMBER       = "FFF2CC"
    C_GREY_HEADER = "F2F2F2"
    C_WHITE       = "FFFFFF"
    C_BLACK       = "000000"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color=C_BLACK, size=9, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    thin = Side(style="thin", color="CCCCCC")
    med  = Side(style="medium", color="888888")

    def border(left=thin, right=thin, top=thin, bottom=thin):
        return Border(left=left, right=right, top=top, bottom=bottom)

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Plan"

    # Column layout: A=Category/Task, B=Resource, C=Days, D=Hrs, E=Rate(EUR), F=Cost(EUR)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.freeze_panes = "A3"

    r = 1  # current row

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="PROJECT BRAVO — IT Cost Plan (Labour Only)")
    c.font = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align(h="center")
    c.border = border()
    ws.row_dimensions[r].height = 26
    r += 1

    # ── Column headers ────────────────────────────────────────────────────────
    COLS = ["CATEGORY / TASK", "RESOURCE", "TOTAL DAYS", "TOTAL HRS", "RATE (EUR/hr)", "TOTAL COST (EUR)"]
    for ci, h in enumerate(COLS, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = font(bold=True, color=C_WHITE, size=9)
        c.fill = fill(C_MID_BLUE)
        c.alignment = align(h="center")
        c.border = border()
    ws.row_dimensions[r].height = 18
    r += 1

    # ── Metadata block ────────────────────────────────────────────────────────
    for label, value in METADATA:
        ws.cell(row=r, column=1, value=label).font = font(bold=True, size=8, color="555555")
        ws.cell(row=r, column=1).fill = fill(C_GREY_HEADER)
        ws.cell(row=r, column=1).border = border()
        ws.cell(row=r, column=1).alignment = align()
        ws.merge_cells(f"B{r}:F{r}")
        c = ws.cell(row=r, column=2, value=value)
        c.font = font(size=8)
        c.fill = fill(C_GREY_HEADER)
        c.border = border()
        c.alignment = align()
        for ci in range(3, 7):
            ws.cell(row=r, column=ci).fill = fill(C_GREY_HEADER)
            ws.cell(row=r, column=ci).border = border()
        ws.row_dimensions[r].height = 14
        r += 1

    r += 1  # blank spacer

    # ── Category sections ─────────────────────────────────────────────────────
    alt = False
    for cat_name, resources in CATEGORIES:
        # Category header
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(row=r, column=1, value=cat_name)
        c.font = font(bold=True, color=C_WHITE, size=9)
        c.fill = fill(C_MID_BLUE)
        c.alignment = align()
        c.border = border(left=Side(style="medium", color=C_DARK_BLUE),
                          right=Side(style="medium", color=C_DARK_BLUE),
                          top=Side(style="medium", color=C_DARK_BLUE),
                          bottom=thin)
        ws.row_dimensions[r].height = 16
        r += 1

        subtotal_days = 0
        subtotal_hrs  = 0
        subtotal_cost = 0

        for resource, task, days, hrs, rate, cost in resources:
            row_fill = fill(C_ALT) if alt else fill(C_WHITE)
            alt = not alt
            ws.cell(row=r, column=1, value=task).font      = font(size=9)
            ws.cell(row=r, column=1).fill                  = row_fill
            ws.cell(row=r, column=1).border                = border()
            ws.cell(row=r, column=1).alignment             = align(wrap=True)
            ws.cell(row=r, column=2, value=resource).font  = font(size=9)
            ws.cell(row=r, column=2).fill                  = row_fill
            ws.cell(row=r, column=2).border                = border()
            ws.cell(row=r, column=2).alignment             = align(wrap=True)
            ws.cell(row=r, column=3, value=days).font      = font(size=9)
            ws.cell(row=r, column=3).fill                  = row_fill
            ws.cell(row=r, column=3).border                = border()
            ws.cell(row=r, column=3).alignment             = align(h="center")
            ws.cell(row=r, column=4, value=hrs).font       = font(size=9)
            ws.cell(row=r, column=4).fill                  = row_fill
            ws.cell(row=r, column=4).border                = border()
            ws.cell(row=r, column=4).alignment             = align(h="center")
            ws.cell(row=r, column=5, value=rate).font      = font(size=9)
            ws.cell(row=r, column=5).fill                  = row_fill
            ws.cell(row=r, column=5).border                = border()
            ws.cell(row=r, column=5).alignment             = align(h="right")
            ws.cell(row=r, column=5).number_format         = '#,##0'
            ws.cell(row=r, column=6, value=cost).font      = font(size=9)
            ws.cell(row=r, column=6).fill                  = row_fill
            ws.cell(row=r, column=6).border                = border()
            ws.cell(row=r, column=6).alignment             = align(h="right")
            ws.cell(row=r, column=6).number_format         = '#,##0'
            ws.row_dimensions[r].height = 15
            subtotal_days += days
            subtotal_hrs  += hrs
            subtotal_cost += cost
            r += 1

        # Subtotal row
        ws.cell(row=r, column=1, value=f"SUBTOTAL — {cat_name}").font = font(bold=True, size=9)
        ws.cell(row=r, column=1).fill      = fill(C_LIGHT_BLUE)
        ws.cell(row=r, column=1).border    = border(bottom=Side(style="medium", color=C_MID_BLUE))
        ws.cell(row=r, column=1).alignment = align()
        ws.cell(row=r, column=2).fill      = fill(C_LIGHT_BLUE)
        ws.cell(row=r, column=2).border    = border(bottom=Side(style="medium", color=C_MID_BLUE))
        ws.cell(row=r, column=3, value=subtotal_days).font = font(bold=True, size=9)
        ws.cell(row=r, column=3).fill      = fill(C_LIGHT_BLUE)
        ws.cell(row=r, column=3).border    = border(bottom=Side(style="medium", color=C_MID_BLUE))
        ws.cell(row=r, column=3).alignment = align(h="center")
        ws.cell(row=r, column=4, value=subtotal_hrs).font = font(bold=True, size=9)
        ws.cell(row=r, column=4).fill      = fill(C_LIGHT_BLUE)
        ws.cell(row=r, column=4).border    = border(bottom=Side(style="medium", color=C_MID_BLUE))
        ws.cell(row=r, column=4).alignment = align(h="center")
        ws.cell(row=r, column=5).fill      = fill(C_LIGHT_BLUE)
        ws.cell(row=r, column=5).border    = border(bottom=Side(style="medium", color=C_MID_BLUE))
        ws.cell(row=r, column=6, value=subtotal_cost).font = font(bold=True, size=9)
        ws.cell(row=r, column=6).fill      = fill(C_LIGHT_BLUE)
        ws.cell(row=r, column=6).border    = border(bottom=Side(style="medium", color=C_MID_BLUE))
        ws.cell(row=r, column=6).alignment = align(h="right")
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.row_dimensions[r].height = 15
        r += 1
        r += 1  # blank between categories

    # ── Overall Total ─────────────────────────────────────────────────────────
    GRAND_TOTAL = sum(cost for _, resources in CATEGORIES for _, _, _, _, _, cost in resources)
    GRAND_HRS   = sum(hrs  for _, resources in CATEGORIES for _, _, _, hrs, _, _ in resources)
    GRAND_DAYS  = sum(days for _, resources in CATEGORIES for _, _, days, _, _, _ in resources)

    ws.cell(row=r, column=1, value="OVERALL PROJECT TOTAL (LABOUR ONLY)").font = font(bold=True, color=C_WHITE, size=10)
    ws.cell(row=r, column=1).fill      = fill(C_DARK_BLUE)
    ws.cell(row=r, column=1).border    = border()
    ws.cell(row=r, column=1).alignment = align()
    ws.cell(row=r, column=2).fill      = fill(C_DARK_BLUE)
    ws.cell(row=r, column=2).border    = border()
    ws.cell(row=r, column=3, value=GRAND_DAYS).font = font(bold=True, color=C_WHITE, size=10)
    ws.cell(row=r, column=3).fill      = fill(C_DARK_BLUE)
    ws.cell(row=r, column=3).border    = border()
    ws.cell(row=r, column=3).alignment = align(h="center")
    ws.cell(row=r, column=4, value=GRAND_HRS).font  = font(bold=True, color=C_WHITE, size=10)
    ws.cell(row=r, column=4).fill      = fill(C_DARK_BLUE)
    ws.cell(row=r, column=4).border    = border()
    ws.cell(row=r, column=4).alignment = align(h="center")
    ws.cell(row=r, column=5).fill      = fill(C_DARK_BLUE)
    ws.cell(row=r, column=5).border    = border()
    ws.cell(row=r, column=6, value=GRAND_TOTAL).font = font(bold=True, color=C_WHITE, size=10)
    ws.cell(row=r, column=6).fill      = fill(C_DARK_BLUE)
    ws.cell(row=r, column=6).border    = border()
    ws.cell(row=r, column=6).alignment = align(h="right")
    ws.cell(row=r, column=6).number_format = '#,##0'
    ws.row_dimensions[r].height = 20
    r += 2

    # ── Cost Breakdown by Category ────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="COST BREAKDOWN BY CATEGORY")
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_MID_BLUE)
    c.alignment = align()
    c.border = border()
    ws.row_dimensions[r].height = 16
    r += 1

    alt2 = False
    for cat_name, resources in CATEGORIES:
        cat_cost = sum(cost for _, _, _, _, _, cost in resources)
        row_fill = fill(C_ALT) if alt2 else fill(C_WHITE)
        alt2 = not alt2
        ws.merge_cells(f"A{r}:E{r}")
        ws.cell(row=r, column=1, value=cat_name).font  = font(size=9)
        ws.cell(row=r, column=1).fill      = row_fill
        ws.cell(row=r, column=1).border    = border()
        ws.cell(row=r, column=1).alignment = align()
        for ci in range(2, 6):
            ws.cell(row=r, column=ci).fill   = row_fill
            ws.cell(row=r, column=ci).border = border()
        ws.cell(row=r, column=6, value=cat_cost).font   = font(size=9)
        ws.cell(row=r, column=6).fill       = row_fill
        ws.cell(row=r, column=6).border     = border()
        ws.cell(row=r, column=6).alignment  = align(h="right")
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.row_dimensions[r].height = 14
        r += 1

    # Category total
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1, value="TOTAL LABOUR").font = font(bold=True, size=9)
    ws.cell(row=r, column=1).fill      = fill(C_LIGHT_BLUE)
    ws.cell(row=r, column=1).border    = border()
    ws.cell(row=r, column=1).alignment = align()
    for ci in range(2, 6):
        ws.cell(row=r, column=ci).fill   = fill(C_LIGHT_BLUE)
        ws.cell(row=r, column=ci).border = border()
    ws.cell(row=r, column=6, value=GRAND_TOTAL).font = font(bold=True, size=9)
    ws.cell(row=r, column=6).fill       = fill(C_LIGHT_BLUE)
    ws.cell(row=r, column=6).border     = border()
    ws.cell(row=r, column=6).alignment  = align(h="right")
    ws.cell(row=r, column=6).number_format = '#,##0'
    ws.row_dimensions[r].height = 15
    r += 2

    # ── Cost Breakdown by Phase ───────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="COST BREAKDOWN BY PHASE")
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_MID_BLUE)
    c.alignment = align()
    c.border = border()
    ws.row_dimensions[r].height = 16
    r += 1

    alt3 = False
    phase_total = 0
    for phase_name, phase_cost in PHASE_BREAKDOWN:
        row_fill = fill(C_ALT) if alt3 else fill(C_WHITE)
        alt3 = not alt3
        ws.merge_cells(f"A{r}:E{r}")
        ws.cell(row=r, column=1, value=phase_name).font  = font(size=9)
        ws.cell(row=r, column=1).fill      = row_fill
        ws.cell(row=r, column=1).border    = border()
        ws.cell(row=r, column=1).alignment = align()
        for ci in range(2, 6):
            ws.cell(row=r, column=ci).fill   = row_fill
            ws.cell(row=r, column=ci).border = border()
        ws.cell(row=r, column=6, value=phase_cost).font   = font(size=9)
        ws.cell(row=r, column=6).fill       = row_fill
        ws.cell(row=r, column=6).border     = border()
        ws.cell(row=r, column=6).alignment  = align(h="right")
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.row_dimensions[r].height = 14
        phase_total += phase_cost
        r += 1

    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1, value="TOTAL LABOUR").font = font(bold=True, size=9)
    ws.cell(row=r, column=1).fill      = fill(C_LIGHT_BLUE)
    ws.cell(row=r, column=1).border    = border()
    ws.cell(row=r, column=1).alignment = align()
    for ci in range(2, 6):
        ws.cell(row=r, column=ci).fill   = fill(C_LIGHT_BLUE)
        ws.cell(row=r, column=ci).border = border()
    ws.cell(row=r, column=6, value=phase_total).font = font(bold=True, size=9)
    ws.cell(row=r, column=6).fill       = fill(C_LIGHT_BLUE)
    ws.cell(row=r, column=6).border     = border()
    ws.cell(row=r, column=6).alignment  = align(h="right")
    ws.cell(row=r, column=6).number_format = '#,##0'
    ws.row_dimensions[r].height = 15
    r += 2

    # ── CAPEX / Additional Costs ──────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="CAPEX / ADDITIONAL COSTS (excluded from labour total)")
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_MID_BLUE)
    c.alignment = align()
    c.border = border()
    ws.row_dimensions[r].height = 16
    r += 1

    # Sub-header
    for ci, h in enumerate(["ITEM", "OWNER", "ESTIMATE / NOTE", "", "", ""], start=1):
        ws.cell(row=r, column=ci, value=h).font = font(bold=True, size=8, color="555555")
        ws.cell(row=r, column=ci).fill      = fill(C_GREY_HEADER)
        ws.cell(row=r, column=ci).border    = border()
        ws.cell(row=r, column=ci).alignment = align(h="center")
    ws.merge_cells(f"C{r}:F{r}")
    ws.row_dimensions[r].height = 14
    r += 1

    alt4 = False
    for item, owner, estimate in CAPEX_ROWS:
        row_fill = fill(C_ALT) if alt4 else fill(C_WHITE)
        alt4 = not alt4
        ws.cell(row=r, column=1, value=item).font   = font(size=9)
        ws.cell(row=r, column=1).fill     = row_fill
        ws.cell(row=r, column=1).border   = border()
        ws.cell(row=r, column=1).alignment = align(wrap=True)
        ws.cell(row=r, column=2, value=owner).font  = font(size=9)
        ws.cell(row=r, column=2).fill     = row_fill
        ws.cell(row=r, column=2).border   = border()
        ws.cell(row=r, column=2).alignment = align(wrap=True)
        ws.merge_cells(f"C{r}:F{r}")
        ws.cell(row=r, column=3, value=estimate).font = font(size=9, italic=True)
        ws.cell(row=r, column=3).fill     = row_fill
        ws.cell(row=r, column=3).border   = border()
        ws.cell(row=r, column=3).alignment = align(wrap=True)
        for ci in range(4, 7):
            ws.cell(row=r, column=ci).fill   = row_fill
            ws.cell(row=r, column=ci).border = border()
        ws.row_dimensions[r].height = 15
        r += 1

    r += 1

    # ── Notes ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="NOTES")
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align()
    c.border = border()
    ws.row_dimensions[r].height = 16
    r += 1

    for note in NOTES:
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(row=r, column=1, value=f"• {note}")
        c.font      = font(size=8, italic=True)
        c.fill      = fill(C_GREY_HEADER)
        c.border    = border()
        c.alignment = align(wrap=True)
        for ci in range(2, 7):
            ws.cell(row=r, column=ci).fill   = fill(C_GREY_HEADER)
            ws.cell(row=r, column=ci).border = border()
        ws.row_dimensions[r].height = 28
        r += 1

    wb.save(path)
    print(f"XLSX written: {path}")


if __name__ == "__main__":
    t0 = _dt.now()
    print(f"Started : {t0.strftime('%Y-%m-%d %H:%M:%S')}")
    write_xlsx(XLSX_PATH)
    t1 = _dt.now()
    print(f"Finished: {t1.strftime('%Y-%m-%d %H:%M:%S')}  ({(t1-t0).total_seconds():.1f}s elapsed)")
