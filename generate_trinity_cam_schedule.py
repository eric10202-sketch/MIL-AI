#!/usr/bin/env python3
"""
Generate Trinity-CAM IT Carve-out Project Schedule (XLSX + CSV + XML).

Project:          Trinity-CAM
Seller:           Johnson Controls International (JCI)
Buyer:            Robert Bosch GmbH
Business:         Aircondition business (JCI Aircon division)
Carve-out Model:  Integration (JCI IT -> Merger Zone -> Bosch IT)
PMO Lead:         KPMG
Sites:            48 worldwide
IT Users:         12,000 (all on JCI side at start)
Applications:     1,800+ including major SAP landscape
IT Delivery:      Infosys (Merger Zone setup, operation, IT & app migration)
TSA:              JCI provides services until 2026-06-30; carve-out starts 2026-07-01

Timeline:
  - Start:       2026-07-01
  - QG0:         2026-07-01
  - QG1:         2026-10-01
  - QG2&3:       2027-07-31
  - QG4:         2027-12-08
  - GoLive:      2028-01-01
  - QG5:         2028-04-01  (90-day hypercare complete)
  - Completion:  2028-04-01

Migration path:  JCI IT -> Merger Zone (Infosys-operated) -> Bosch IT
"""

from pathlib import Path
import sys
import subprocess
from datetime import datetime, timedelta

HERE = Path(__file__).parent
PROJECT_NAME = "Trinity-CAM"
XLSX_PATH = HERE / "active-projects" / PROJECT_NAME / f"{PROJECT_NAME}_Project_Schedule.xlsx"
XML_PATH  = HERE / "active-projects" / PROJECT_NAME / f"{PROJECT_NAME}_Project_Schedule.xml"
CSV_PATH  = HERE / "active-projects" / PROJECT_NAME / f"{PROJECT_NAME}_Project_Schedule.csv"

XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

START_DATE      = datetime(2026, 7, 1)
GOLIVE_DATE     = datetime(2028, 1, 1)
COMPLETION_DATE = datetime(2028, 4, 1)   # QG5 / end of 90-day hypercare


def date_str(d: datetime) -> str:
    return d.strftime("%Y-%m-%d")


def dp(base: datetime, days: int) -> str:
    """Add calendar days and return ISO string."""
    return date_str(base + timedelta(days=days))


S = START_DATE   # shorthand

# =============================================================================
# TRINITY-CAM TASK LIST  (118 tasks)
# Schema: (ID, OutlineLevel, Name, Duration, Start, Finish,
#          Predecessors, ResourceNames, Notes, Milestone)
#
# Key day offsets from START (2026-07-01):
#   day   0  = 2026-07-01  QG0 / project start
#   day  92  = 2026-10-01  QG1 - Concept Approval
#   day 215  = 2027-02-01  Phase 2 start (Merger Zone Build)
#   day 395  = 2027-07-31  QG2&3 - Build Complete
#   day 396  = 2027-08-01  Phase 3 start
#   day 518  = 2027-12-01  Phase 4 start
#   day 525  = 2027-12-08  QG4 - Pre-GoLive Gate
#   day 549  = 2028-01-01  GoLive
#   day 550  = 2028-01-02  Phase 5 start
#   day 640  = 2028-04-01  QG5 / Completion
# =============================================================================

TASKS = [

    # =========================================================================
    # PHASE 0: PROJECT INITIATION & GOVERNANCE  (day 0 - 91)
    # =========================================================================
    (1,  1, "Phase 0: Project Initiation & Governance",
     "92 days", dp(S, 0), dp(S, 92), "", "",
     "PMO mobilisation, governance, Infosys onboarding, legal, QG1 approval", "No"),

    # --- 0.1 Project Governance & PMO Setup ---
    (2,  2, "0.1 Project Governance & PMO Setup",
     "46 days", dp(S, 0), dp(S, 46), "", "",
     "KPMG PMO, steering, RACI, Infosys onboarding", "No"),

    (3,  3, "QG0 - Programme Kickoff & Mobilisation",
     "0 days", dp(S, 0), dp(S, 0), "",
     "KPMG PMO Lead + JCI Programme Sponsor + Bosch Programme Sponsor",
     "Programme officially started; all workstream leads mobilised; Infosys contractually engaged", "Yes"),

    (4,  3, "Establish KPMG PMO & Project Management Office",
     "14 days", dp(S, 0), dp(S, 14), "3",
     "KPMG PMO Lead + KPMG Project Manager",
     "PMO governance tools active; reporting cadence; SharePoint site live", "No"),

    (5,  3, "Appoint Project Steering Committee (JCI + Bosch + KPMG)",
     "10 days", dp(S, 0), dp(S, 10), "3",
     "KPMG PMO Lead + JCI Programme Sponsor + Bosch Programme Sponsor",
     "Steering committee chartered; meeting calendar confirmed", "No"),

    (6,  3, "Define RACI & IT Workstream Structure",
     "14 days", dp(S, 11), dp(S, 25), "5",
     "KPMG Project Manager + KPMG PMO Lead",
     "RACI approved; nine workstreams defined: Infrastructure, SAP, Applications, Identity, End-User Workplace, Data Migration, Security/GDPR, TSA Exit/HR/Legal, Programme Control", "No"),

    (7,  3, "Onboard Infosys as IT Delivery Partner (Merger Zone)",
     "45 days", dp(S, 1), dp(S, 46), "3",
     "KPMG PMO Lead + Infosys Programme Manager",
     "Infosys Statement of Work executed; MZ scope confirmed; delivery team onsite", "No"),

    (8,  3, "Stakeholder Communication Plan (JCI / Bosch / Infosys / 48 Sites)",
     "14 days", dp(S, 15), dp(S, 29), "4",
     "KPMG Project Manager + KPMG PMO Lead",
     "Comms plan approved; escalation path; change management approach", "No"),

    (9,  3, "Initial Risk & Issue Register Baseline",
     "21 days", dp(S, 22), dp(S, 43), "8",
     "KPMG Project Manager",
     "Risk register v1.0 baselined; top 25 risks scored and owned (24 threats, 1 opportunity)", "No"),

    # --- 0.2 Legal & Regulatory Compliance ---
    (10, 2, "0.2 Legal & Regulatory Compliance Framework",
     "71 days", dp(S, 10), dp(S, 81), "", "",
     "GDPR, legal entity separation, competition law", "No"),

    (11, 3, "GDPR & Data Privacy Assessment (12,000 Users - 48 Sites)",
     "46 days", dp(S, 10), dp(S, 56), "6",
     "KPMG Project Manager + JCI Legal Counsel",
     "Privacy impact assessment; data residency; cross-border transfer rules", "No"),

    (12, 3, "Legal Entity Separation & Carve-out Structure",
     "46 days", dp(S, 10), dp(S, 56), "6",
     "KPMG Project Manager + JCI Legal Counsel + Bosch Legal Counsel",
     "NewCo SPA; legal entity readiness; Integration model structures confirmed", "No"),

    (13, 3, "Regulatory & Competition Law Review (Multi-jurisdiction)",
     "46 days", dp(S, 30), dp(S, 76), "11",
     "KPMG Project Manager + JCI Legal Counsel",
     "Multi-jurisdiction clearance tracking; antitrust filings; no blocking conditions", "No"),

    (14, 3, "QG1 - Concept Approval",
     "0 days", dp(S, 92), dp(S, 92), "9,13",
     "KPMG PMO Lead + JCI Programme Sponsor + Bosch Programme Sponsor",
     "Project concept approved; Infosys scope confirmed; proceed to Discovery & Architecture", "Yes"),

    # =========================================================================
    # PHASE 1: DISCOVERY & ARCHITECTURE DESIGN  (day 92 - 214)
    # =========================================================================
    (15, 1, "Phase 1: Discovery & Architecture Design",
     "122 days", dp(S, 92), dp(S, 214), "", "",
     "App inventory 1800+, MZ architecture, TSA catalogue, data strategy", "No"),

    # --- 1.1 Application Landscape Assessment ---
    (16, 2, "1.1 Application Landscape Assessment (1,800+ Applications)",
     "111 days", dp(S, 92), dp(S, 203), "", "",
     "SAP analysis, non-SAP inventory, dependency mapping, wave planning", "No"),

    (17, 3, "SAP System Landscape Analysis (Full SAP Scope)",
     "45 days", dp(S, 92), dp(S, 137), "14",
     "KPMG SAP Architect + Infosys SAP Lead + JCI SAP Owner",
     "SAP component list; modules in scope; interface inventory; carve-out complexity", "No"),

    (18, 3, "Non-SAP Application Inventory & Classification (1,800+ Apps)",
     "60 days", dp(S, 92), dp(S, 152), "14",
     "KPMG Data Architect + Infosys Application Lead + JCI IT Manager",
     "Full app inventory; ownership; migration classification: Keep/Retire/Replace/Re-host", "No"),

    (19, 3, "Application Dependency Mapping & Interface Analysis",
     "35 days", dp(S, 137), dp(S, 172), "17",
     "KPMG Data Architect + Infosys Application Lead",
     "Cross-app and SAP dependencies; integration interface inventory; circular deps flagged", "No"),

    (20, 3, "Application Migration Wave Planning (Waves 1-3)",
     "31 days", dp(S, 172), dp(S, 203), "18,19",
     "KPMG Data Architect + KPMG PMO Lead + Infosys Application Lead",
     "Wave 1: ~400 critical apps; Wave 2: ~800 standard; Wave 3: ~600 local/specialist", "No"),

    # --- 1.2 Infrastructure & Network Assessment ---
    (21, 2, "1.2 Infrastructure & Network Assessment (48 Sites)",
     "77 days", dp(S, 92), dp(S, 169), "", "",
     "JCI network, DC/cloud, segregation design, device estate", "No"),

    (22, 3, "JCI Network Architecture Review (48 Sites)",
     "30 days", dp(S, 92), dp(S, 122), "14",
     "KPMG Infrastructure Architect + Infosys Network Lead + JCI IT Manager",
     "WAN, LAN, SD-WAN, firewall topology per site; bandwidth baseline", "No"),

    (23, 3, "Data Centre & Cloud Inventory (JCI Hosted Assets)",
     "31 days", dp(S, 92), dp(S, 123), "14",
     "KPMG Infrastructure Architect + JCI IT Manager",
     "Primary & secondary DC; cloud workloads; contracts; hosting obligations", "No"),

    (24, 3, "Network Segregation Design (48 Sites to Merger Zone)",
     "46 days", dp(S, 123), dp(S, 169), "22,23",
     "KPMG Infrastructure Architect + Infosys Network Lead",
     "Site-by-site network separation design; MZ connectivity model; SD-WAN approach", "No"),

    (25, 3, "End-User Device Estate Assessment (12,000 Users)",
     "45 days", dp(S, 92), dp(S, 137), "14",
     "KPMG Infrastructure Architect + JCI IT Manager",
     "Device inventory by site; OS; age; refresh plan; Intune/SCCM readiness", "No"),

    # --- 1.3 Merger Zone Architecture Design ---
    (26, 2, "1.3 Merger Zone Architecture Design (Infosys-led)",
     "84 days", dp(S, 92), dp(S, 176), "", "",
     "MZ blueprint, AD design, core services architecture", "No"),

    (27, 3, "Merger Zone Infrastructure Blueprint (Infosys)",
     "61 days", dp(S, 92), dp(S, 153), "22,23",
     "Infosys Infrastructure Architect + KPMG Infrastructure Architect",
     "MZ DC/cloud topology; compute; storage; network zones; security perimeter", "No"),

    (28, 3, "Identity & Active Directory Design (Merger Zone)",
     "46 days", dp(S, 92), dp(S, 138), "25",
     "KPMG Infrastructure Architect + Infosys IAM Lead",
     "MZ AD forest design; trust model with JCI and Bosch; federation approach", "No"),

    (29, 3, "Core Services Architecture (Email, M365, Collaboration, Security)",
     "61 days", dp(S, 115), dp(S, 176), "27",
     "KPMG Infrastructure Architect + Infosys Cloud Lead",
     "MZ M365 tenant architecture; Teams; security tooling; ITSM platform selection", "No"),

    # --- 1.4 TSA Service Catalogue & Exit Plan ---
    (30, 2, "1.4 TSA Service Catalogue & Exit Plan",
     "92 days", dp(S, 92), dp(S, 184), "", "",
     "JCI TSA services catalogued; SLA agreed; exit criteria defined", "No"),

    (31, 3, "TSA Service Definition & SLA Agreement per Service Line",
     "31 days", dp(S, 92), dp(S, 123), "14",
     "KPMG Project Manager + JCI IT Manager + Bosch IT Manager",
     "All JCI services to continue under TSA catalogued; SLA and KPI per service line", "No"),

    (32, 3, "TSA Governance & Transition Reporting Framework",
     "30 days", dp(S, 123), dp(S, 153), "31",
     "KPMG Project Manager + JCI IT Manager",
     "Monthly TSA service reports; escalation path; service exit trigger criteria", "No"),

    (33, 3, "TSA Exit Criteria per Service Line",
     "31 days", dp(S, 153), dp(S, 184), "32",
     "KPMG Project Manager + KPMG PMO Lead",
     "Exit criteria defined per service; acceptance testing approach; sign-off process", "No"),

    # --- 1.5 Data Segregation & Migration Strategy ---
    (34, 2, "1.5 Data Segregation & Migration Strategy",
     "101 days", dp(S, 92), dp(S, 193), "", "",
     "Data classification, migration approach JCI > MZ > Bosch", "No"),

    (35, 3, "Data Classification & Ownership (JCI Aircon Business)",
     "46 days", dp(S, 92), dp(S, 138), "14",
     "KPMG Data Architect + JCI Data Owner + Bosch Data Owner",
     "PII, IP, regulated data identified; data ownership matrix approved", "No"),

    (36, 3, "Data Migration Approach: JCI to Merger Zone to Bosch",
     "55 days", dp(S, 138), dp(S, 193), "35",
     "KPMG Data Architect + Infosys Data Migration Lead",
     "Tooling selection; ETL methodology; data validation; SAP non-selective approach", "No"),

    # --- 1.6 HR & People Transition Planning ---
    (37, 2, "1.6 HR & People Transition Planning",
     "92 days", dp(S, 92), dp(S, 184), "", "",
     "12,000-user mapping; multi-jurisdiction employment; comms plan", "No"),

    (38, 3, "Employee Mapping: JCI Aircon to Carve-out Entity (12,000 Users)",
     "45 days", dp(S, 92), dp(S, 137), "14",
     "KPMG Project Manager + JCI HR Director + Bosch HR Director",
     "User list by site; role mapping; TUPE and transfer rules per jurisdiction", "No"),

    (39, 3, "People Communication Plan (Multi-region, 48 Sites)",
     "47 days", dp(S, 137), dp(S, 184), "38",
     "KPMG Project Manager + JCI HR Director",
     "Phased comms per wave; change management; union consultation plans", "No"),

    # =========================================================================
    # PHASE 2: MERGER ZONE BUILD & CONFIGURATION  (day 215 - 395)
    # =========================================================================
    (40, 1, "Phase 2: Merger Zone Build & Configuration (Infosys)",
     "180 days", dp(S, 215), dp(S, 395), "", "",
     "MZ infrastructure build, IAM, core IT, SAP system copy, app wave 1", "No"),

    # --- 2.1 Infrastructure Build ---
    (41, 2, "2.1 Infrastructure Build - Merger Zone (Infosys)",
     "121 days", dp(S, 215), dp(S, 336), "", "",
     "DC/cloud, network, security, backup/DR", "No"),

    (42, 3, "Merger Zone DC / Cloud Environment Setup (Infosys)",
     "60 days", dp(S, 215), dp(S, 275), "29,27",
     "Infosys Infrastructure Architect + KPMG Infrastructure Architect",
     "Primary and secondary MZ hosting environments provisioned and validated", "No"),

    (43, 3, "Network Connectivity: All 48 Sites to Merger Zone (Infosys)",
     "91 days", dp(S, 245), dp(S, 336), "24",
     "Infosys Network Lead + KPMG Infrastructure Architect + JCI IT Manager",
     "Site-by-site WAN/SD-WAN activation; phased rollout; all 48 sites connected to MZ", "No"),

    (44, 3, "Security Architecture & Firewall Implementation (Merger Zone)",
     "60 days", dp(S, 230), dp(S, 290), "29",
     "Infosys Security Lead + KPMG Infrastructure Architect",
     "MZ perimeter security; DLP; endpoint protection; SOC integration; IDS/IPS", "No"),

    (45, 3, "Backup, Disaster Recovery & Business Continuity Setup",
     "45 days", dp(S, 275), dp(S, 320), "44",
     "Infosys Infrastructure Architect + KPMG Infrastructure Architect",
     "DR test plan; RPO/RTO targets defined and validated; backup cadence confirmed", "No"),

    # --- 2.2 Identity & Access Management ---
    (46, 2, "2.2 Identity & Access Management (Merger Zone)",
     "90 days", dp(S, 215), dp(S, 305), "", "",
     "AD forest, identity federation, PAM, MFA", "No"),

    (47, 3, "Active Directory Forest Setup (Merger Zone)",
     "43 days", dp(S, 215), dp(S, 258), "28",
     "Infosys IAM Lead + KPMG Infrastructure Architect",
     "MZ AD forest live; OU structure; delegation model; trust relationships", "No"),

    (48, 3, "Identity Federation: JCI to Merger Zone (SSO)",
     "46 days", dp(S, 258), dp(S, 304), "47",
     "Infosys IAM Lead + KPMG Infrastructure Architect",
     "Federation live; ADFS/Azure AD Connect; JCI users authenticate to MZ services", "No"),

    (49, 3, "Privileged Access Management (PAM) Setup",
     "30 days", dp(S, 259), dp(S, 289), "47",
     "Infosys Security Lead + KPMG Infrastructure Architect",
     "PAM tooling deployed; all admin accounts vaulted; session recording active", "No"),

    (50, 3, "MFA & Security Policy Enforcement (12,000 Users)",
     "30 days", dp(S, 275), dp(S, 305), "49",
     "Infosys Security Lead + KPMG Infrastructure Architect",
     "MFA enrolled for all 12,000 users; conditional access; security baselines enforced", "No"),

    # --- 2.3 Core IT Services Build ---
    (51, 2, "2.3 Core IT Services Build (Merger Zone)",
     "121 days", dp(S, 215), dp(S, 336), "", "",
     "Email/M365, SharePoint, Teams, ITSM service desk", "No"),

    (52, 3, "Email & M365 Tenant Setup (Merger Zone)",
     "60 days", dp(S, 215), dp(S, 275), "29",
     "Infosys Cloud Lead + KPMG Infrastructure Architect",
     "MZ M365 tenant provisioned; mail flow validated; DNS prepared for cutover", "No"),

    (53, 3, "SharePoint & OneDrive Migration Tooling Setup",
     "45 days", dp(S, 275), dp(S, 320), "52,42",
     "Infosys Cloud Lead + KPMG Infrastructure Architect",
     "Migration tooling configured and tested; user data migration pipeline validated", "No"),

    (54, 3, "Collaboration Tools: Teams & VOIP (Merger Zone)",
     "46 days", dp(S, 275), dp(S, 321), "52",
     "Infosys Cloud Lead + KPMG Infrastructure Architect",
     "Teams provisioned; VOIP platform migrated; telephony operational for MZ users", "No"),

    (55, 3, "Service Desk & ITSM Tool Configuration (Infosys)",
     "61 days", dp(S, 245), dp(S, 306), "29",
     "Infosys Service Delivery Lead + KPMG Project Manager",
     "ITSM platform live; service catalogue; SLA monitoring; L1/L2 support model", "No"),

    # --- 2.4 SAP Build & Configuration ---
    (56, 2, "2.4 SAP Build & Configuration (Infosys SAP Team)",
     "180 days", dp(S, 215), dp(S, 395), "", "",
     "SAP system copy, client sep, interface rewiring, security, data prep", "No"),

    (57, 3, "SAP System Copy: JCI to Merger Zone",
     "60 days", dp(S, 275), dp(S, 335), "20,42",
     "Infosys SAP Architect + KPMG SAP Architect + JCI SAP Owner",
     "Full SAP landscape replication into MZ environment; technical validation complete", "No"),

    (58, 3, "SAP Client Separation & Carve-out Configuration",
     "30 days", dp(S, 335), dp(S, 365), "57",
     "Infosys SAP Architect + KPMG SAP Architect + JCI SAP Owner",
     "Aircon business data separated; shared client access revoked; configuration validated", "No"),

    (59, 3, "SAP Interface Rewiring: JCI to Merger Zone Endpoints",
     "30 days", dp(S, 365), dp(S, 395), "58",
     "Infosys SAP Architect + KPMG SAP Architect",
     "All outbound/inbound SAP interfaces rewired to MZ endpoints; EDI remapped", "No"),

    (60, 3, "SAP Security Role Redesign & Authorization Model",
     "30 days", dp(S, 335), dp(S, 365), "57",
     "Infosys SAP Architect + KPMG SAP Architect",
     "Role matrix redesigned; SoD rules applied; SAP GRC configured for MZ", "No"),

    (61, 3, "SAP Data Migration Preparation (Technical Setup)",
     "75 days", dp(S, 215), dp(S, 290), "36",
     "Infosys Data Migration Lead + KPMG Data Architect",
     "Migration tooling; data mapping; extraction rules; transformation logic; test datasets", "No"),

    # --- 2.5 Non-SAP App Migration Wave 1 ---
    (62, 2, "2.5 Non-SAP Application Migration - Wave 1 (~400 Critical Apps)",
     "176 days", dp(S, 215), dp(S, 391), "", "",
     "Prioritization, setup in MZ, data migration, smoke testing", "No"),

    (63, 3, "Wave 1 Application Prioritization & Scope Finalization",
     "30 days", dp(S, 215), dp(S, 245), "20",
     "KPMG Data Architect + Infosys Application Lead",
     "Wave 1 app list frozen; business criticality confirmed; migration sequence defined", "No"),

    (64, 3, "Wave 1 Application Setup in Merger Zone",
     "46 days", dp(S, 275), dp(S, 321), "63,42",
     "Infosys Application Lead + KPMG Infrastructure Architect",
     "All Wave 1 apps installed and configured in MZ; integration points validated", "No"),

    (65, 3, "Wave 1 Data Migration Execution",
     "45 days", dp(S, 321), dp(S, 366), "64",
     "Infosys Data Migration Lead + KPMG Data Architect",
     "Data extracted from JCI; transformed; loaded into MZ; integrity checks passed", "No"),

    (66, 3, "Wave 1 Application Smoke Testing & Validation",
     "25 days", dp(S, 366), dp(S, 391), "65",
     "Infosys Application Lead + KPMG Infrastructure Architect + JCI IT Manager",
     "Functional smoke tests passed; business owners sign-off Wave 1 app readiness", "No"),

    # --- QG2&3 Milestone ---
    (67, 3, "QG2&3 - Build Complete & Test Ready",
     "0 days", dp(S, 395), dp(S, 395), "59,60,66",
     "KPMG PMO Lead + JCI Programme Sponsor + Bosch Programme Sponsor",
     "MZ infrastructure live; SAP interfaces rewired; Wave 1 apps validated; proceed to testing", "Yes"),

    # =========================================================================
    # PHASE 3: TESTING & MIGRATION WAVES  (day 396 - 518)
    # =========================================================================
    (68, 1, "Phase 3: Testing & Migration Waves",
     "122 days", dp(S, 396), dp(S, 518), "", "",
     "SIT, UAT, 12k user migrations, app waves 2-3, SAP cutover rehearsal", "No"),

    # --- 3.1 Systems Integration Testing ---
    (69, 2, "3.1 Systems Integration Testing",
     "62 days", dp(S, 396), dp(S, 458), "", "",
     "E2E, performance, security pen test, DR test", "No"),

    (70, 3, "End-to-End Integration Test: SAP & Non-SAP (Wave 1)",
     "46 days", dp(S, 396), dp(S, 442), "67",
     "Infosys SAP Architect + KPMG SAP Architect + KPMG Infrastructure Architect",
     "Full E2E flows across SAP, Wave 1 apps, MZ services; defects logged and resolved", "No"),

    (71, 3, "Performance & Load Testing (12,000 Concurrent Users)",
     "47 days", dp(S, 396), dp(S, 443), "67",
     "Infosys Infrastructure Architect + KPMG Infrastructure Architect",
     "Peak load simulation; 12,000 concurrent users; MZ performance baselines met", "No"),

    (72, 3, "Security Penetration Testing (Merger Zone & Applications)",
     "30 days", dp(S, 420), dp(S, 450), "67",
     "Infosys Security Lead + KPMG Infrastructure Architect",
     "External and internal pen test; critical findings remediated; security report signed", "No"),

    (73, 3, "Disaster Recovery Failover Test",
     "30 days", dp(S, 420), dp(S, 450), "45",
     "Infosys Infrastructure Architect + KPMG Infrastructure Architect",
     "Full DR failover executed; RTO/RPO targets validated; runbook confirmed", "No"),

    # --- 3.2 User Acceptance Testing ---
    (74, 2, "3.2 User Acceptance Testing (UAT)",
     "109 days", dp(S, 396), dp(S, 505), "", "",
     "UAT planning, execution, defect resolution, sign-off", "No"),

    (75, 3, "UAT Planning & Test Case Design (All Workstreams)",
     "31 days", dp(S, 396), dp(S, 427), "67",
     "KPMG Project Manager + JCI IT Manager + Bosch IT Manager",
     "Test scenarios per workstream; 48-site scope covered; business testers briefed", "No"),

    (76, 3, "UAT Execution: All User Groups (12,000 Users, Phased)",
     "46 days", dp(S, 427), dp(S, 473), "75,70",
     "KPMG Project Manager + Infosys Service Delivery Lead + JCI IT Manager",
     "Phased UAT across all 48 sites; SAP transactions; Wave 1 apps; collaboration tools", "No"),

    (77, 3, "UAT Defect Triage & Resolution",
     "31 days", dp(S, 457), dp(S, 488), "76",
     "Infosys Application Lead + Infosys SAP Architect + KPMG Project Manager",
     "All critical and high defects resolved; regression tested; defect log shows zero blocking", "No"),

    (78, 3, "UAT Sign-off by Business & IT Owners (All Sites)",
     "15 days", dp(S, 488), dp(S, 503), "77",
     "KPMG PMO Lead + JCI IT Manager + Bosch IT Manager",
     "Formal UAT sign-off received; go-forward confirmation from all 48 sites", "No"),

    # --- 3.3 User Migration Waves ---
    (79, 2, "3.3 User Migration Waves (12,000 Users - 48 Sites)",
     "119 days", dp(S, 396), dp(S, 515), "", "",
     "Four waves covering all 12,000 users across 48 sites - completed pre-QG4", "No"),

    (80, 3, "User Migration Wave 1: Sites 1-12 (3,000 Users)",
     "32 days", dp(S, 411), dp(S, 443), "67",
     "Infosys Service Delivery Lead + KPMG Infrastructure Architect + JCI IT Manager",
     "Wave 1 users migrated to MZ; devices enrolled; M365 mailboxes cut; service desk live", "No"),

    (81, 3, "User Migration Wave 2: Sites 13-24 (3,000 Users)",
     "31 days", dp(S, 443), dp(S, 474), "80",
     "Infosys Service Delivery Lead + KPMG Infrastructure Architect + JCI IT Manager",
     "Wave 2 sites migrated; L1/L2 support cover maintained throughout window", "No"),

    (82, 3, "User Migration Wave 3: Sites 25-36 (3,000 Users)",
     "23 days", dp(S, 474), dp(S, 497), "81",
     "Infosys Service Delivery Lead + KPMG Infrastructure Architect + JCI IT Manager",
     "Wave 3 complete; all APAC and EMEA tail sites on MZ services", "No"),

    (83, 3, "User Migration Wave 4: Sites 37-48 (3,000 Users)",
     "18 days", dp(S, 497), dp(S, 515), "82",
     "Infosys Service Delivery Lead + KPMG Infrastructure Architect + JCI IT Manager",
     "All 12,000 users migrated to Merger Zone; JCI user access fully transitioned", "No"),

    # --- 3.4 Non-SAP App Migration Wave 2 ---
    (84, 2, "3.4 Non-SAP Application Migration - Wave 2 (~800 Standard Apps)",
     "92 days", dp(S, 396), dp(S, 488), "", "",
     "Wave 2 app migration and data migration - completed pre-QG4", "No"),

    (85, 3, "Wave 2 Application Migration & Integration Testing",
     "46 days", dp(S, 396), dp(S, 442), "67",
     "Infosys Application Lead + KPMG Data Architect",
     "Wave 2 apps migrated to MZ; integration tests passed; business sign-off", "No"),

    (86, 3, "Wave 2 Data Migration Execution",
     "46 days", dp(S, 442), dp(S, 488), "85",
     "Infosys Data Migration Lead + KPMG Data Architect",
     "Wave 2 data extracted, transformed, loaded; integrity validated; reconciliation done", "No"),

    # --- 3.5 Non-SAP App Migration Wave 3 ---
    (87, 2, "3.5 Non-SAP Application Migration - Wave 3 (~600 Specialist Apps)",
     "61 days", dp(S, 457), dp(S, 518), "", "",
     "Wave 3 app and data migration - completed pre-QG4", "No"),

    (88, 3, "Wave 3 Application Migration & Integration Testing",
     "46 days", dp(S, 457), dp(S, 503), "86",
     "Infosys Application Lead + KPMG Data Architect",
     "Wave 3 specialist/regional apps migrated; integration validated; defects resolved", "No"),

    (89, 3, "Wave 3 Data Migration Execution",
     "29 days", dp(S, 488), dp(S, 517), "86",
     "Infosys Data Migration Lead + KPMG Data Architect",
     "Wave 3 data migration complete; all 1,800+ app data in MZ; no outstanding items", "No"),

    # --- 3.6 SAP Cutover Planning ---
    (90, 2, "3.6 SAP Production Cutover Planning & Rehearsal",
     "86 days", dp(S, 427), dp(S, 513), "", "",
     "SAP cutover run book, two mock cutovers - completed pre-QG4", "No"),

    (91, 3, "SAP GoLive Cutover Strategy & Run Book",
     "45 days", dp(S, 427), dp(S, 472), "61",
     "Infosys SAP Architect + KPMG SAP Architect + JCI SAP Owner",
     "Detailed run book; cutover sequence; rollback plan; timing and resources confirmed", "No"),

    (92, 3, "SAP Mock Cutover Exercise 1 (Full Dry Run)",
     "20 days", dp(S, 472), dp(S, 492), "91",
     "Infosys SAP Architect + KPMG SAP Architect + JCI SAP Owner",
     "Full dry run; timing measured; issues logged; rollback tested successfully", "No"),

    (93, 3, "SAP Mock Cutover Exercise 2 (Final Rehearsal)",
     "21 days", dp(S, 492), dp(S, 513), "92",
     "Infosys SAP Architect + KPMG SAP Architect + JCI SAP Owner",
     "Issues from Mock 1 corrected; cutover timing confirmed; go/no-go criteria met", "No"),

    # =========================================================================
    # PHASE 4: PRE-GOLIVE READINESS  (day 518 - 549)
    # =========================================================================
    (94, 1, "Phase 4: Pre-GoLive Readiness",
     "31 days", dp(S, 518), dp(S, 549), "", "",
     "QG4 preparation, final readiness checks, GoLive cutover", "No"),

    # --- 4.1 QG4 Gate Preparation ---
    (95, 2, "4.1 QG4 Gate Preparation",
     "7 days", dp(S, 518), dp(S, 525), "", "",
     "All migration verification, infrastructure and SAP readiness, rollback plan", "No"),

    (96, 3, "All Migration Waves Completed & Verified (Users + Apps + Data)",
     "6 days", dp(S, 518), dp(S, 524), "83,89,93,78",
     "KPMG PMO Lead + Infosys Programme Manager",
     "12,000 users migrated; all 1,800+ apps on MZ; all data reconciled; UAT signed off", "No"),

    (97, 3, "Infrastructure Readiness Sign-off (Infosys + KPMG)",
     "6 days", dp(S, 518), dp(S, 524), "73,50",
     "Infosys Infrastructure Architect + KPMG Infrastructure Architect",
     "MZ infrastructure fully operational; DR validated; security cleared; monitoring active", "No"),

    (98, 3, "SAP Production Readiness Confirmed",
     "6 days", dp(S, 518), dp(S, 524), "93",
     "Infosys SAP Architect + KPMG SAP Architect + JCI SAP Owner",
     "SAP production environment validated; interfaces live; authorization model approved", "No"),

    (99, 3, "Cutover Rollback Plan Approved",
     "6 days", dp(S, 518), dp(S, 524), "91",
     "KPMG PMO Lead + Infosys Programme Manager",
     "Rollback plan reviewed and approved by steering; go/no-go criteria documented", "No"),

    (100, 3, "QG4 - Pre-GoLive Gate Approval",
     "0 days", dp(S, 525), dp(S, 525), "96,97,98,99",
     "KPMG PMO Lead + JCI Programme Sponsor + Bosch Programme Sponsor",
     "All workstreams ready; QG4 approved; proceed to Final Readiness & GoLive preparation", "Yes"),

    # --- 4.2 Final Readiness & Open Item Closure ---
    (101, 2, "4.2 Final Readiness & Open Item Closure",
     "21 days", dp(S, 526), dp(S, 547), "", "",
     "Mandatory: UAT defects, infra checks, rollback test, steering sign-off", "No"),

    (102, 3, "All UAT Defects Resolved - QA Sign-off",
     "8 days", dp(S, 526), dp(S, 534), "100",
     "KPMG Project Manager + Infosys Service Delivery Lead",
     "Zero blocking defects; QA formally signs off UAT closure; test evidence archived", "No"),

    (103, 3, "Final Infrastructure & Systems Health Checks (No Critical Alerts)",
     "12 days", dp(S, 526), dp(S, 538), "100",
     "Infosys Infrastructure Architect + KPMG Infrastructure Architect",
     "All MZ components green; no open P1 tickets; capacity validated for GoLive load", "No"),

    (104, 3, "Cutover Rollback Test Executed (Production-like)",
     "10 days", dp(S, 527), dp(S, 537), "100",
     "Infosys SAP Architect + KPMG Project Manager",
     "Full rollback test completed; rollback timing confirmed within window; approved", "No"),

    (105, 3, "Business & Steering Committee Sign-off (GoLive Authorisation)",
     "7 days", dp(S, 539), dp(S, 546), "102,103,104",
     "KPMG PMO Lead + JCI Programme Sponsor + Bosch Programme Sponsor",
     "Formal steering approval to proceed to GoLive; all action items closed; date confirmed", "No"),

    # --- GoLive Milestone ---
    (106, 3, "GoLive - Merger Zone Day 1 Cutover (2028-01-01)",
     "0 days", dp(S, 549), dp(S, 549), "105",
     "KPMG PMO Lead + Infosys Programme Manager + JCI Programme Sponsor + Bosch Programme Sponsor",
     "All 12,000 users live on Merger Zone; SAP production activated; 1,800+ apps operational; 24/7 support active", "Yes"),

    # =========================================================================
    # PHASE 5: HYPERCARE & PROJECT CLOSURE  (day 550 - 640)
    # =========================================================================
    (107, 1, "Phase 5: Hypercare & Project Closure",
     "90 days", dp(S, 550), dp(S, 640), "", "",
     "90-day hypercare, Bosch integration handover, TSA exit, PMO closure", "No"),

    # --- 5.1 Hypercare Support ---
    (108, 2, "5.1 Hypercare Support (90 Days - Stabilisation Only)",
     "90 days", dp(S, 550), dp(S, 640), "", "",
     "24/7 L3 support, hotfixes, user enablement, Bosch knowledge transfer", "No"),

    (109, 3, "24/7 L3 Support & Incident Triage (Infosys)",
     "90 days", dp(S, 550), dp(S, 640), "106",
     "Infosys Service Delivery Lead + KPMG Project Manager",
     "Round-the-clock L3 support; P1/P2 hotline; daily ops standups all regions", "No"),

    (110, 3, "Hotfix Release Management (Critical Bug Fixes Only)",
     "60 days", dp(S, 550), dp(S, 610), "106",
     "Infosys Application Lead + Infosys SAP Architect",
     "Bug fixes only; no new features; change freeze enforced; release cadence weekly", "No"),

    (111, 3, "User Enablement & Adoption Programs (12,000 Users)",
     "45 days", dp(S, 550), dp(S, 595), "106",
     "KPMG Project Manager + Infosys Service Delivery Lead + JCI IT Manager",
     "Self-service guides; regional training; adoption dashboards; help desk KPIs tracked", "No"),

    (112, 3, "Knowledge Transfer to Bosch IT Operations",
     "60 days", dp(S, 564), dp(S, 624), "106",
     "KPMG Infrastructure Architect + Infosys Infrastructure Architect + Bosch IT Manager",
     "Run books; architecture docs; SAP operational guides; ITSM transition to Bosch team", "No"),

    # --- 5.2 Bosch Integration Handover & Project Closure ---
    (113, 2, "5.2 Bosch Integration Handover & Project Closure",
     "76 days", dp(S, 564), dp(S, 640), "", "",
     "TSA exit, Bosch IT handover, documentation, lessons learned", "No"),

    (114, 3, "JCI TSA Service Exit Execution",
     "46 days", dp(S, 564), dp(S, 610), "106",
     "KPMG Project Manager + JCI IT Manager + Bosch IT Manager",
     "JCI services formally exited per TSA exit criteria; acceptance tests passed; contracts closed", "No"),

    (115, 3, "Bosch IT Systems Integration Handover",
     "56 days", dp(S, 580), dp(S, 636), "112",
     "Infosys Programme Manager + KPMG Infrastructure Architect + Bosch IT Manager",
     "MZ services formally handed to Bosch IT operations; Infosys role concluded; ITSM transferred", "No"),

    (116, 3, "Documentation & Runbook Finalization (All Workstreams)",
     "44 days", dp(S, 580), dp(S, 624), "112",
     "KPMG PMO Lead + Infosys Programme Manager",
     "Architecture docs; SAP guides; operational runbooks; security playbooks; all archived", "No"),

    (117, 3, "Lessons Learned & PMO Closure Report (KPMG)",
     "31 days", dp(S, 609), dp(S, 640), "115,116",
     "KPMG PMO Lead + KPMG Project Manager",
     "Lessons learned workshop; final PMO report; project artifacts archived; team debrief", "No"),

    (118, 3, "QG5 - Project Completion",
     "0 days", dp(S, 640), dp(S, 640), "109,117",
     "KPMG PMO Lead + JCI Programme Sponsor + Bosch Programme Sponsor",
     "90-day hypercare complete; JCI TSA fully exited; Bosch IT operations live; programme closed", "Yes"),
]


# =============================================================================
# WRITE CSV  (kept permanently alongside XLSX and XML)
# =============================================================================

def _write_csv(path: Path) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write("ID,Outline Level,Name,Duration,Start,Finish,Predecessors,Resource Names,Notes,Milestone\n")
        for tid, ol, name, dur, start, finish, preds, res, notes, ms in TASKS:
            name   = name.replace('"', '""')
            notes  = notes.replace('"', '""')
            res    = res.replace('"', '""')
            f.write(f'{tid},"{ol}","{name}","{dur}","{start}","{finish}","{preds}","{res}","{notes}","{ms}"\n')


# =============================================================================
# GENERATE XLSX  (Bosch Blue theme)
# =============================================================================

def _generate_excel() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime as _dt

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    headers = [
        "ID", "Outline Level", "Name", "Duration",
        "Start", "Finish", "Predecessors", "Resource Names", "Notes", "Milestone"
    ]
    ws.append(headers)

    hdr_fill = PatternFill(start_color="002147", end_color="002147", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _to_date(v: str):
        try:
            return _dt.strptime(v, "%Y-%m-%d")
        except Exception:
            return v

    fill_l1        = PatternFill(start_color="003B6E", end_color="003B6E", fill_type="solid")
    fill_l2        = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    fill_milestone = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_even      = PatternFill(start_color="EFF4FB", end_color="EFF4FB", fill_type="solid")
    fill_odd       = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    font_white_bold = Font(bold=True, color="FFFFFF", size=10)
    font_ms         = Font(bold=True, color="000000", size=10)
    font_detail     = Font(color="000000", size=9)

    indent_map = {1: "", 2: "  ", 3: "    "}

    for row_idx, (tid, ol, name, dur, start, finish, preds, res, notes, ms) in enumerate(TASKS, start=2):
        indent = indent_map.get(ol, "")
        ws.append([
            tid, ol, indent + name, dur,
            _to_date(start), _to_date(finish),
            preds, res, notes, ms
        ])
        row = ws[row_idx]

        if ms == "Yes":
            cell_fill = fill_milestone
            cell_font = font_ms
        elif ol == 1:
            cell_fill = fill_l1
            cell_font = font_white_bold
        elif ol == 2:
            cell_fill = fill_l2
            cell_font = font_white_bold
        else:
            cell_fill = fill_even if (row_idx % 2 == 0) else fill_odd
            cell_font = font_detail

        for cell in row:
            cell.fill = cell_fill
            cell.font = cell_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Date format for Start/Finish columns
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 54
    ws.column_dimensions["I"].width = 44
    ws.column_dimensions["J"].width = 10

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J1"

    wb.save(str(XLSX_PATH))
    print(f"  XLSX: {XLSX_PATH}")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print(f"\n[{PROJECT_NAME}] Generating project schedule...")
    print(f"  {len(TASKS)} tasks | Start: {dp(S, 0)} | GoLive: {dp(S, 549)} | QG5: {dp(S, 640)}")

    _generate_excel()
    _write_csv(CSV_PATH)
    print(f"  CSV:  {CSV_PATH}")

    result = subprocess.run(
        [sys.executable, str(HERE / "generate_msp_xml.py"),
         "--csv",     str(CSV_PATH),
         "--out",     str(XML_PATH),
         "--project", PROJECT_NAME],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"  ERROR generating XML:\n{result.stderr}")
        sys.exit(1)
    print(f"  XML:  {XML_PATH}")
    print(f"\n[{PROJECT_NAME}] Schedule generation complete.")
    print("  QG0:   2026-07-01")
    print("  QG1:   2026-10-01")
    print("  QG2/3: 2027-07-31")
    print("  QG4:   2027-12-08  (GoLive minus 24 days)")
    print("  GL:    2028-01-01")
    print("  QG5:   2028-04-01  (90-day hypercare complete)")
