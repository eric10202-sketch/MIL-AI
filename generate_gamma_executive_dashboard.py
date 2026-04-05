#!/usr/bin/env python3
"""Generate the Gamma executive dashboard as self-contained HTML."""

from __future__ import annotations

import base64
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


HERE = Path(__file__).parent
PROJECT_NAME = "Gamma"
OUTPUT_FOLDER_NAME = "Gamma v1.0"
SELLER = "Robert Bosch China"
BUYER = "Alibaba"
BUSINESS = "Bosch Cloud business"
MODEL = "Combination"
PMO = "EY"
REPORT_DATE = datetime(2026, 4, 5)
SCHEDULE_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Risk_Register.xlsx"
COST_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Cost_Plan.xlsx"
OUTPUT_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Executive_Dashboard.html"
LOGO_PATH = HERE / "Bosch.png"

IMPACT_VALUES = {"Very Low": 1, "Low": 2, "Moderate": 3, "High": 4, "Very High": 5}
PROBABILITY_VALUES = {"10%": 1, "30%": 2, "50%": 3, "70%": 4, "90%": 5}


def fmt_date(value) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%d %b %Y")
    return str(value)


def days_until(date_str: str) -> int:
    target = datetime.strptime(date_str, "%Y-%m-%d")
    return (target - REPORT_DATE).days


def load_schedule_metrics():
    wb = load_workbook(SCHEDULE_PATH, data_only=False)
    ws = wb["Schedule"]
    phases = []
    milestones = []
    task_count = 0
    resource_groups = set()
    total_person_hours = 0
    current_phase = None
    phase_detail_counts = defaultdict(int)
    for row in range(2, ws.max_row + 1):
        level = ws.cell(row, 2).value
        if level is None:
            continue
        task_count += 1
        level = int(level)
        name = str(ws.cell(row, 3).value or "").strip()
        start = ws.cell(row, 5).value
        finish = ws.cell(row, 6).value
        if level == 1:
            current_phase = name
            phases.append((name, fmt_date(start), fmt_date(finish), start.strftime("%Y-%m-%d"), finish.strftime("%Y-%m-%d")))
        resources = str(ws.cell(row, 8).value or "")
        if level >= 3 and resources:
            days = int(str(ws.cell(row, 4).value).split()[0])
            tokens = [token.strip() for token in resources.split("+") if token.strip()]
            for token in tokens:
                resource_groups.add(token)
                total_person_hours += days * 8
            if current_phase:
                phase_detail_counts[current_phase] += 1
        if str(ws.cell(row, 10).value) == "Yes" and ("QG" in name or "GoLive" in name or "Closure" in name):
            iso_date = start.strftime("%Y-%m-%d")
            milestones.append((name, fmt_date(start), iso_date, days_until(iso_date)))
    return phases, milestones, task_count, len(resource_groups), total_person_hours, phase_detail_counts


def load_risk_metrics():
    wb = load_workbook(RISK_PATH, data_only=False)
    ws = wb["Risk Register"]
    high = 0
    medium = 0
    low = 0
    risks = []
    for row in range(5, 140):
        risk_id = ws.cell(row, 2).value
        if risk_id is None:
            continue
        impact = ws.cell(row, 12).value
        probability = ws.cell(row, 14).value
        risk_type = ws.cell(row, 16).value
        score = IMPACT_VALUES.get(impact, 0) * PROBABILITY_VALUES.get(probability, 0)
        if score >= 12:
            high += 1
        elif score >= 6:
            medium += 1
        else:
            low += 1
        risks.append(
            {
                "id": int(risk_id),
                "category": ws.cell(row, 4).value,
                "event": ws.cell(row, 6).value,
                "owner": ws.cell(row, 9).value,
                "score": score,
                "type": risk_type,
            }
        )
    risks.sort(key=lambda item: item["score"], reverse=True)
    return high, medium, low, risks[:6]


def load_cost_metrics():
    wb = load_workbook(COST_PATH, data_only=False)
    ws = wb["Cost Plan"]
    labour_total = 0
    categories = []
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label == "OVERALL PROJECT TOTAL - LABOUR ONLY":
            labour_total = int(ws.cell(row, 6).value or 0)
        if isinstance(label, str) and label.startswith("SUBTOTAL -"):
            categories.append((label.replace("SUBTOTAL - ", ""), int(ws.cell(row, 6).value or 0)))
    categories = [item for item in categories if item[1] > 0]
    categories.sort(key=lambda item: item[1], reverse=True)
    return labour_total, categories


def workstream_cards():
    return [
        ("WS1", "PMO & Governance", "amber", ["Restricted clean-team governance", "Shared JV decision path", "Daily legal dependency control"]),
        ("WS2", "Legal & Confidentiality", "red", ["Antitrust diligence unresolved", "Controlled stakeholder expansion", "Disclosure sequencing required"]),
        ("WS3", "Infrastructure & Hosting", "amber", ["Shared Bosch services must be isolated", "Day 1 hosting baseline", "5-site connectivity control"]),
        ("WS4", "Network & Connectivity", "amber", ["Provider lead-time sensitivity", "Certificate and DNS control", "Fallback routing required"]),
        ("WS5", "Security & Identity", "amber", ["Access recertification required", "Privileged access redesign", "Data-segregation evidence for QG4"]),
        ("WS6", "Applications & Data", "green", ["20 apps only", "No SAP complexity", "Focused transition scope"]),
        ("WS7", "Testing & Readiness", "amber", ["Restricted UAT participant group", "Rollback rehearsal mandatory", "QG4 buffer tightly managed"]),
        ("WS8", "Deployment & Change", "amber", ["250 users across 5 sites", "Phased communication approvals", "Training expansion depends on legal sign-off"]),
        ("WS9", "Hypercare & Handover", "green", ["90-day stabilization locked", "JV steady-state handover", "Transition-service exit tracked"]),
    ]


def main() -> None:
    phases, milestones, task_count, resource_group_count, total_person_hours, phase_detail_counts = load_schedule_metrics()
    high_risks, medium_risks, low_risks, top_risks = load_risk_metrics()
    labour_total, cost_categories = load_cost_metrics()
    logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode()

    kickoff_days = days_until("2026-08-01")
    golive_days = days_until("2027-02-01")
    hypercare_exit_days = days_until("2027-05-02")

    max_category = max((value for _, value in cost_categories), default=1)
    category_rows = "\n".join(
        f"<div class='budget-row'><span>{name}</span><span>{value:,.0f} EUR</span><div class='budget-bar'><div style='width:{(value / max_category) * 100:.0f}%;'></div></div></div>"
        for name, value in cost_categories
    )
    milestone_rows = "\n".join(
        f"<tr><td>{name}</td><td>{date}</td><td>{days:+d}</td><td><span class='pill {'red' if days < 0 else 'green' if days > 90 else 'amber'}'>{'past' if days < 0 else 'upcoming'}</span></td></tr>"
        for name, date, _, days in milestones[:6]
    )
    qg_rows = "\n".join(
        [
            "<div class='qg-item'><strong>QG0</strong><span>Restricted planning controls, perimeter confirmation, and inventory complete</span></div>",
            "<div class='qg-item'><strong>QG1</strong><span>Concept approved with operating model, transition principles, and risk baseline</span></div>",
            "<div class='qg-item'><strong>QG2 and QG3</strong><span>Build and testing complete, cutover rehearsal and migration packs approved</span></div>",
            "<div class='qg-item'><strong>QG4</strong><span>All pre-GoLive migrations complete and final-readiness window opened</span></div>",
            "<div class='qg-item'><strong>QG5</strong><span>Hypercare complete, transition services exited, steady-state handover accepted</span></div>",
        ]
    )
    workstream_html = "\n".join(
        f"<div class='ws-card'><div class='ws-top'><span>{code}</span><span class='pill {status}'>{status}</span></div><h4>{title}</h4><ul>{''.join(f'<li>{line}</li>' for line in bullets)}</ul></div>"
        for code, title, status, bullets in workstream_cards()
    )
    risk_cards = f"""
      <div class='risk-card red'><div class='risk-num'>{high_risks}</div><div>High</div></div>
      <div class='risk-card amber'><div class='risk-num'>{medium_risks}</div><div>Medium</div></div>
      <div class='risk-card green'><div class='risk-num'>{low_risks}</div><div>Low</div></div>
    """
    top_risk_rows = "\n".join(
        f"<tr><td>#{risk['id']}</td><td>{risk['category']}</td><td>{risk['score']}</td><td>{risk['owner']}</td></tr>"
        for risk in top_risks
    )
    wave_rows = "\n".join(
        [
            "<div class='wave'><span>Wave 1 - Foundation services</span><div class='wave-bar'><div style='width:30%;'></div></div><strong>6 apps</strong></div>",
            "<div class='wave'><span>Wave 2 - Core operations</span><div class='wave-bar'><div style='width:40%;'></div></div><strong>8 apps</strong></div>",
            "<div class='wave'><span>Wave 3 - Residual shared services</span><div class='wave-bar'><div style='width:30%;'></div></div><strong>6 apps</strong></div>",
        ]
    )
    phase_timeline = "".join(
        f"<div class='timeline-phase' style='flex:{(datetime.strptime(end, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')).days + 1};'><span>{name.replace('Phase ', '').replace(' - ', ': ')}</span></div>"
        for name, _, _, start, end in phases
    )

    html = f"""<!DOCTYPE html>
<html lang='en'>
<head>
  <meta charset='UTF-8' />
  <meta name='viewport' content='width=device-width, initial-scale=1.0' />
  <title>{PROJECT_NAME} Executive Dashboard</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; background: #f4f6f9; color: #1a1a1a; }}
    .page {{ max-width: 1024px; margin: 0 auto 20px; background: #fff; padding: 22px; box-shadow: 0 10px 30px rgba(0,0,0,0.08); }}
    .page-break {{ page-break-before: always; }}
    .header {{ background: linear-gradient(135deg, #002147 0%, #003b6e 55%, #0066CC 100%); color: #fff; border-radius: 12px; padding: 22px; display: grid; grid-template-columns: 1.6fr 1fr; gap: 16px; margin-bottom: 16px; }}
    .bosch-logo {{ display:flex; align-items:center; background:#fff; padding:4px 8px; border-radius:4px; width:fit-content; margin-bottom:12px; }}
    .title {{ font-size: 30px; font-weight: 700; margin-bottom: 4px; }}
    .subtitle {{ font-size: 14px; opacity: 0.92; line-height: 1.6; }}
    .date-panel {{ text-align: right; font-size: 12px; }}
    .days-strip {{ display:grid; grid-template-columns: repeat(3, 1fr); gap:12px; background:#e20015; color:#fff; border-radius:10px; padding:12px; margin-bottom:16px; }}
    .days-box {{ background:rgba(255,255,255,0.12); border-radius:8px; padding:10px; }}
    .days-box .n {{ font-size: 26px; font-weight: 700; }}
    .section {{ margin-bottom: 16px; border: 1px solid #d7deea; border-radius: 10px; overflow: hidden; }}
    .section-header {{ background:#005199; color:#fff; padding:10px 14px; font-weight:700; font-size:14px; }}
    .section-content {{ padding:14px; }}
    .two-col {{ display:grid; grid-template-columns: 1fr 1fr; gap:16px; }}
    .stats {{ display:grid; grid-template-columns: repeat(6, 1fr); gap:12px; margin-bottom:16px; }}
    .stat {{ background:#fff; border:1px solid #d7deea; border-top:4px solid #0066CC; border-radius:10px; padding:12px; text-align:center; }}
    .stat .v {{ font-size:20px; font-weight:700; color:#003b6e; }}
    .stat .l {{ font-size:11px; color:#5d6d83; text-transform:uppercase; letter-spacing:.6px; margin-top:4px; }}
    .timeline {{ display:flex; height:44px; overflow:hidden; border-radius:8px; margin-bottom:8px; }}
    .timeline-phase {{ display:flex; align-items:center; justify-content:center; color:#fff; font-size:11px; font-weight:700; padding:0 8px; text-align:center; }}
    .timeline-phase:nth-child(1) {{ background:#6e8bb7; }}
    .timeline-phase:nth-child(2) {{ background:#3f6db1; }}
    .timeline-phase:nth-child(3) {{ background:#0f7dd8; }}
    .timeline-phase:nth-child(4) {{ background:#0080a8; }}
    .timeline-phase:nth-child(5) {{ background:#007A33; }}
    .overview-box, .model-box {{ background:#eff4fb; border-radius:10px; padding:14px; border-left:4px solid #0066CC; }}
    .meta-list div {{ margin-bottom:6px; font-size:12px; }}
    table {{ width:100%; border-collapse:collapse; font-size:12px; }}
    th, td {{ padding:8px 10px; border-bottom:1px solid #e1e7f0; text-align:left; vertical-align:top; }}
    th {{ background:#eff4fb; color:#003b6e; }}
    .pill {{ display:inline-block; padding:2px 8px; border-radius:12px; font-size:11px; font-weight:700; color:#fff; text-transform:uppercase; }}
    .pill.green {{ background:#007A33; }}
    .pill.amber {{ background:#E8A000; color:#1a1a1a; }}
    .pill.red {{ background:#CC0000; }}
    .budget-row {{ margin-bottom:10px; font-size:12px; }}
    .budget-row span:last-child {{ float:right; font-weight:700; }}
    .budget-bar {{ clear:both; height:9px; background:#dce7f7; border-radius:999px; overflow:hidden; margin-top:6px; }}
    .budget-bar div {{ height:100%; background:#0066CC; }}
    .ws-grid {{ display:grid; grid-template-columns: repeat(3, 1fr); gap:12px; }}
    .ws-card {{ border:1px solid #d7deea; border-radius:10px; padding:12px; background:#fff; }}
    .ws-top {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:8px; font-size:12px; color:#5d6d83; }}
    .ws-card h4 {{ margin:0 0 8px; color:#003b6e; font-size:14px; }}
    .ws-card ul {{ margin:0; padding-left:18px; font-size:12px; }}
    .qg-item {{ display:grid; grid-template-columns: 110px 1fr; gap:12px; padding:8px 0; border-bottom:1px solid #e1e7f0; font-size:12px; }}
    .bars .bar {{ margin-bottom:12px; }}
    .bars .bar-label {{ font-size:12px; margin-bottom:4px; }}
    .bars .bar-track {{ height:12px; background:#e1e7f0; border-radius:999px; overflow:hidden; }}
    .bars .bar-track div {{ height:100%; background:#0066CC; }}
    .risk-grid {{ display:grid; grid-template-columns: repeat(3, 1fr); gap:10px; margin-bottom:12px; }}
    .risk-card {{ border-radius:10px; color:#fff; text-align:center; padding:12px; }}
    .risk-card.red {{ background:#CC0000; }}
    .risk-card.amber {{ background:#E8A000; color:#1a1a1a; }}
    .risk-card.green {{ background:#007A33; }}
    .risk-num {{ font-size:24px; font-weight:700; margin-bottom:4px; }}
    .wave {{ display:grid; grid-template-columns: 1.4fr 2fr .6fr; gap:12px; align-items:center; margin-bottom:12px; font-size:12px; }}
    .wave-bar {{ height:14px; background:#e1e7f0; border-radius:999px; overflow:hidden; }}
    .wave-bar div {{ height:100%; background:#0077BB; }}
    .hotspot-grid {{ display:grid; grid-template-columns: repeat(2, 1fr); gap:12px; }}
    .hotspot {{ background:#eff4fb; border-radius:10px; padding:12px; border:1px solid #d7deea; }}
    .hotspot h4 {{ margin:0 0 8px; color:#003b6e; }}
    .stats-strip {{ background:#003b6e; color:#fff; border-radius:10px; padding:14px; display:grid; grid-template-columns: repeat(4, 1fr); gap:12px; margin:16px 0; }}
    .stats-strip div {{ text-align:center; }}
    .stats-strip strong {{ display:block; font-size:22px; }}
    .cp-grid {{ display:grid; grid-template-columns: repeat(4, 1fr); gap:12px; }}
    .cp-card {{ border:1px solid #d7deea; border-radius:10px; padding:12px; background:#fff; }}
    .cp-card h4 {{ margin:0 0 8px; color:#003b6e; }}
    .footer {{ margin-top:16px; font-size:11px; color:#5d6d83; text-align:center; }}
  </style>
</head>
<body>
  <div class='page'>
    <div class='header'>
      <div>
        <div class='bosch-logo'><img src='data:image/png;base64,{logo_b64}' alt='Bosch - Invented for Life' style='height:36px;display:block;' /></div>
        <div class='title'>{PROJECT_NAME} Executive Dashboard</div>
        <div class='subtitle'>{BUSINESS} carve-out from {SELLER} into a 50/50 JV with {BUYER}. Executive view aligned to the current Gamma schedule, risk register, and cost plan baselines.</div>
      </div>
      <div class='date-panel'>
        <div><strong>Dashboard Date</strong></div>
        <div>{REPORT_DATE.strftime('%d %b %Y')}</div>
        <div style='margin-top:18px;'><strong>GoLive Countdown</strong></div>
        <div style='font-size:28px;font-weight:700;'>{golive_days}</div>
        <div>days to Day 1</div>
      </div>
    </div>

    <div class='days-strip'>
      <div class='days-box'><div>Kickoff</div><div class='n'>{kickoff_days}</div><div>days to 01 Aug 2026</div></div>
      <div class='days-box'><div>Day 1 GoLive</div><div class='n'>{golive_days}</div><div>days to 01 Feb 2027</div></div>
      <div class='days-box'><div>Hypercare Exit</div><div class='n'>{hypercare_exit_days}</div><div>days to 02 May 2027</div></div>
    </div>

    <div class='section'>
      <div class='section-header'>PROJECT OVERVIEW</div>
      <div class='section-content two-col'>
        <div class='overview-box'>
          <p>Gamma is a confidentiality-constrained infrastructure carve-out. The project separates the Bosch Cloud business from Robert Bosch China, creates the Day 1 operating baseline for a jointly managed JV with Alibaba, and controls legal, disclosure, and shared-service dependencies through a narrow clean-team model.</p>
          <p>The programme is intentionally lighter on application complexity than a typical enterprise carve-out because only 20 applications are in scope and there is no SAP. Delivery risk is concentrated in legal timing, infrastructure separation, identity, operational tooling, and final readiness discipline between QG4 and GoLive.</p>
        </div>
        <div class='model-box'>
          <div class='meta-list'>
            <div><strong>Carve-out model:</strong> {MODEL}</div>
            <div><strong>PMO / methodology lead:</strong> {PMO}</div>
            <div><strong>Budget baseline:</strong> {labour_total:,.0f} EUR labour only</div>
            <div><strong>Governance note:</strong> 50/50 JV with shared management and restricted disclosure</div>
            <div><strong>Legal note:</strong> Antitrust due diligence still active</div>
          </div>
        </div>
      </div>
    </div>

    <div class='stats'>
      <div class='stat'><div class='v'>5</div><div class='l'>Global Sites</div></div>
      <div class='stat'><div class='v'>250</div><div class='l'>IT Users</div></div>
      <div class='stat'><div class='v'>20</div><div class='l'>Applications</div></div>
      <div class='stat'><div class='v'>10</div><div class='l'>Project Months</div></div>
      <div class='stat'><div class='v'>90</div><div class='l'>Hypercare Days</div></div>
      <div class='stat'><div class='v'>{resource_group_count}</div><div class='l'>Resource Groups</div></div>
    </div>

    <div class='timeline'>{phase_timeline}</div>

    <div class='two-col'>
      <div class='section'>
        <div class='section-header'>KEY MILESTONES & QUALITY GATES</div>
        <div class='section-content'>
          <table>
            <tr><th>Milestone</th><th>Date</th><th>Days</th><th>Status</th></tr>
            {milestone_rows}
          </table>
        </div>
      </div>
      <div class='section'>
        <div class='section-header'>BUDGET DISTRIBUTION</div>
        <div class='section-content'>
          <div style='font-size:30px;font-weight:700;color:#003b6e;margin-bottom:6px;'>{labour_total:,.0f} EUR</div>
          <div style='font-size:12px;color:#5d6d83;margin-bottom:14px;'>Labour baseline only. Risk-linked CAPEX and contingency are held separately in the cost plan.</div>
          {category_rows}
        </div>
      </div>
    </div>
  </div>

  <div class='page page-break'>
    <div class='section'>
      <div class='section-header'>IT WORKSTREAM COVERAGE</div>
      <div class='section-content ws-grid'>
        {workstream_html}
      </div>
    </div>
    <div class='two-col'>
      <div class='section'>
        <div class='section-header'>QUALITY GATE TRACKER</div>
        <div class='section-content'>
          {qg_rows}
        </div>
      </div>
      <div class='section'>
        <div class='section-header'>SCOPE & SCALE INDICATORS</div>
        <div class='section-content'>
          <div class='bars'>
            <div class='bar'>
              <div class='bar-label'>Confirmed clean-team site set</div>
              <div class='bar-track'><div style='width:100%;'></div></div>
            </div>
            <div style='font-size:12px;color:#5d6d83;margin-bottom:16px;'>Detailed regional split is intentionally withheld in the dashboard because legal confidentiality still limits broad disclosure of site geography.</div>
          </div>
          <div class='risk-grid'>{risk_cards}</div>
          <table>
            <tr><th>Risk</th><th>Category</th><th>Score</th><th>Owner</th></tr>
            {top_risk_rows}
          </table>
        </div>
      </div>
    </div>
  </div>

  <div class='page page-break'>
    <div class='section'>
      <div class='section-header'>APPLICATION MIGRATION WAVES</div>
      <div class='section-content'>
        <p style='font-size:12px;color:#5d6d83;margin-top:0;'>Indicative executive grouping for the 20 in-scope applications based on the current scope baseline and wave-planning workstream.</p>
        {wave_rows}
      </div>
    </div>
    <div class='section'>
      <div class='section-header'>COMPLEXITY HOTSPOTS UNDER RESTRICTED DISCLOSURE</div>
      <div class='section-content hotspot-grid'>
        <div class='hotspot'><h4>Legal & Disclosure</h4><p>Antitrust due diligence and phased disclosure directly shape the timing of decisions, approvals, and stakeholder access.</p></div>
        <div class='hotspot'><h4>Shared Services</h4><p>Hidden Bosch China infrastructure dependencies remain the most material technical threat to build certainty.</p></div>
        <div class='hotspot'><h4>Identity Boundary</h4><p>Joint governance requires careful identity and privileged-access design to avoid retained Bosch access after cutover.</p></div>
        <div class='hotspot'><h4>Final Readiness</h4><p>The QG4 to GoLive interval is fixed and short, so residual defects or open actions must be burned down aggressively.</p></div>
      </div>
    </div>
    <div class='stats-strip'>
      <div><strong>{task_count}</strong>Total Tasks</div>
      <div><strong>{resource_group_count}</strong>Resource Groups</div>
      <div><strong>{total_person_hours:,.0f}</strong>Person-Hours</div>
      <div><strong>{len(phases)}</strong>Baseline Phases</div>
    </div>
    <div class='section'>
      <div class='section-header'>CRITICAL PATH & GUIDING PRINCIPLES</div>
      <div class='section-content cp-grid'>
        <div class='cp-card'><h4>Infrastructure Critical Path</h4><p>Shared-service discovery, landing-zone build, connectivity, and operational tooling readiness define the technical backbone of Gamma.</p></div>
        <div class='cp-card'><h4>Security Critical Path</h4><p>Identity segregation, privileged-access certification, and data-protection evidence must be complete before QG4.</p></div>
        <div class='cp-card'><h4>Deployment Critical Path</h4><p>User, device, and site transitions must be completed before QG4, with no migration activity after GoLive.</p></div>
        <div class='cp-card'><h4>Programme Principles</h4><p>Protect confidentiality, avoid copied assumptions, keep Day 1 scope lean, and escalate any gate-threatening issue within 24 hours.</p></div>
      </div>
    </div>
    <div class='footer'>{PROJECT_NAME} | Dashboard date {REPORT_DATE.strftime('%d %b %Y')} | Data sources: {SCHEDULE_PATH.name}, {RISK_PATH.name}, {COST_PATH.name} | Confidential executive material</div>
  </div>
</body>
</html>
"""

    OUTPUT_PATH.write_text(html, encoding="utf-8")
    print(f"[{PROJECT_NAME}] Executive dashboard written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()