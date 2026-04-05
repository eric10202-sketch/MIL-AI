#!/usr/bin/env python3
"""Generate Trinity-CAM (GPT) carve-out schedule as XLSX, CSV, and MS Project XML."""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
import subprocess
import sys


HERE = Path(__file__).parent
PROJECT_NAME = "Trinity-CAM (GPT)"
OUTPUT_FOLDER_NAME = "Trinity-CAM (GPT) v1.1"
DOCUMENT_VERSION = "Version 1.1 - Change Request 1"
SELLER = "Johnson Controls (JCI)"
BUYER = "Bosch"
BUSINESS = "Air conditioning business"
PMO_LEAD = "KPMG"
DELIVERY_PARTNER = "Infosys"
WORLDWIDE_SITES = 48
IT_USERS = 12000
APPLICATIONS = 1800

START_DATE = datetime(2026, 7, 1)
GOLIVE_DATE = datetime(2028, 1, 1)
COMPLETION_DATE = datetime(2028, 4, 1)
APPROVED_TSA_END_DATE = datetime(2027, 7, 31)

OUTPUT_DIR = HERE / "active-projects" / OUTPUT_FOLDER_NAME
XLSX_PATH = OUTPUT_DIR / f"{PROJECT_NAME}_Project_Schedule.xlsx"
CSV_PATH = OUTPUT_DIR / f"{PROJECT_NAME}_Project_Schedule.csv"
XML_PATH = OUTPUT_DIR / f"{PROJECT_NAME}_Project_Schedule.xml"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def iso_date(offset_days: int) -> str:
    return (START_DATE + timedelta(days=offset_days)).strftime("%Y-%m-%d")


def make_task(
    task_id: int,
    outline_level: int,
    name: str,
    start_offset: int,
    finish_offset: int,
    predecessors: str = "",
    resources: str = "",
    notes: str = "",
    milestone: bool = False,
) -> tuple[int, int, str, str, str, str, str, str, str, str]:
    duration_days = 1 if milestone else max(finish_offset - start_offset, 1)
    duration_label = f"{duration_days} day" if duration_days == 1 else f"{duration_days} days"
    milestone_flag = "Yes" if milestone else "No"
    return (
        task_id,
        outline_level,
        name,
        duration_label,
        iso_date(start_offset),
        iso_date(finish_offset),
        predecessors,
        resources,
        notes,
        milestone_flag,
    )


TASKS = [
    make_task(
        1,
        1,
        "Phase 0: Mobilize carve-out governance and TSA baseline",
        0,
        92,
        notes="Programme launch, TSA framing, approved extension control, partner mobilisation, and QG1 concept approval under change request version 1.1.",
    ),
    make_task(2, 2, "0.1 Governance and partner mobilization", 0, 60, notes="PMO setup with JCI, Bosch, KPMG, and Infosys."),
    make_task(
        3,
        3,
        "QG0 - Programme kickoff approved",
        0,
        0,
        resources="KPMG PMO Lead + JCI Sponsor + Bosch Sponsor",
        notes="Formal launch of the carve-out starting 2026-07-01. Users continue to work in the legacy JCI environment while Infosys builds the merger zone and the approved TSA extension remains available through 2027-07-31.",
        milestone=True,
    ),
    make_task(4, 3, "Stand up PMO, governance cadence, and workstream RACI", 0, 14, "3", "KPMG PMO Lead + KPMG Project Manager", "Weekly governance, steering, RAID, and reporting in place."),
    make_task(5, 3, "Onboard Infosys for merger zone delivery and managed services", 1, 30, "3", "KPMG PMO Lead + Infosys Programme Manager", "Delivery scope confirms merger zone setup, operations, and migration execution."),
    make_task(6, 3, "Baseline TSA towers, service owners, and exit principles", 14, 60, "4,5", "KPMG Project Manager + JCI IT Manager + Bosch IT Manager", "JCI TSA catalogue, approved extension through 2027-07-31, and exit logic defined for all in-scope IT towers."),
    make_task(7, 2, "0.2 Separation compliance and stakeholder alignment", 10, 92, notes="Legal, privacy, communications, and regional alignment."),
    make_task(8, 3, "Define legal perimeter, contracts, and supplier dependencies", 10, 45, "4", "KPMG Project Manager + JCI Legal Counsel + Bosch Legal Counsel", "Commercial and legal separation perimeter agreed."),
    make_task(9, 3, "Assess privacy, works council, and cross-border data constraints", 20, 70, "8", "KPMG Data Architect + JCI Legal Counsel", "Regional data transfer and employee impact constraints captured."),
    make_task(10, 3, "Launch site change network and stakeholder communications", 30, 92, "6,9", "KPMG Change Lead + JCI HR Lead + Bosch HR Lead", "Regional comms and escalation channels activated across 48 sites."),
    make_task(
        11,
        3,
        "QG1 - Concept and transition model approved",
        92,
        92,
        "6,9,10",
        "KPMG PMO Lead + JCI Sponsor + Bosch Sponsor",
        "Concept sign-off for JCI to merger zone to Bosch integration model.",
        milestone=True,
    ),
    make_task(
        12,
        1,
        "Phase 1: Discover landscape and design the merger zone",
        93,
        214,
        notes="Detailed discovery of applications, SAP, sites, data, and target merger zone design.",
    ),
    make_task(13, 2, "1.1 Application, SAP, and data discovery", 93, 180, notes="Discovery across more than 1,800 applications including major SAP scope."),
    make_task(14, 3, "Catalogue 1,800-plus applications and assign transition disposition", 93, 145, "11", "KPMG Enterprise Architect + JCI Application Owner + Infosys Application Lead", "Classify applications by retain, migrate, replace, retire, and TSA dependency."),
    make_task(15, 3, "Assess SAP landscape, interfaces, and shared services exposure", 93, 155, "11", "KPMG SAP Architect + JCI SAP Owner + Infosys SAP Lead", "Identify carve-out complexity across SAP systems, interfaces, and shared master data."),
    make_task(16, 3, "Map dependencies and define migration waves for apps and data", 145, 180, "14,15", "KPMG Enterprise Architect + Infosys Application Lead", "Wave sequencing aligns critical apps, regional apps, and local edge cases."),
    make_task(17, 2, "1.2 Merger zone and target-state architecture", 93, 205, notes="Merger zone blueprint, security, connectivity, workplace, and Bosch end-state landing."),
    make_task(18, 3, "Design merger zone hosting, network zones, and security blueprint", 93, 150, "11", "Infosys Infrastructure Architect + KPMG Infrastructure Architect", "Target blueprint for temporary merger zone with secure JCI and Bosch connectivity."),
    make_task(19, 3, "Design identity, Active Directory trust, and privileged access model", 110, 165, "18", "Infosys IAM Lead + KPMG Security Architect", "Trust path and identity controls support JCI to merger zone to Bosch transition."),
    make_task(20, 3, "Design workplace, M365, collaboration, and service management stack", 120, 185, "18", "Infosys Cloud Lead + KPMG Workplace Lead", "Email, M365, Teams, endpoint management, and service desk operating design complete."),
    make_task(21, 3, "Design Bosch landing zone integration for post-merger-zone operations", 160, 205, "18,19,20", "Bosch IT Manager + KPMG Infrastructure Architect + Infosys Programme Manager", "Bosch target interfaces, identity landing, and support integration agreed."),
    make_task(22, 2, "1.3 TSA, migration, and site sequencing design", 120, 214, notes="TSA service tower design, migration controls, and regional sequencing."),
    make_task(23, 3, "Define TSA service towers, KPIs, and operational reporting", 120, 155, "6,11", "KPMG Project Manager + JCI IT Manager + Bosch IT Manager", "Service lines and TSA metrics defined for transition governance with the approved seller-service buffer through 2027-07-31."),
    make_task(24, 3, "Define exit criteria, acceptance tests, and service cutover checkpoints", 155, 190, "23", "KPMG Project Manager + Bosch IT Manager", "Each tower receives explicit readiness and exit conditions."),
    make_task(25, 3, "Define data separation controls and reconciliation approach", 130, 175, "14,15", "KPMG Data Architect + Infosys Data Migration Lead", "Controls for business data separation, validation, and audit trail agreed."),
    make_task(26, 3, "Sequence 48 sites and 12,000 users into migration waves", 160, 200, "16,20", "KPMG Deployment Lead + Infosys Service Delivery Lead", "Regional wave plan balances business criticality and support capacity."),
    make_task(27, 3, "Design command center, cutover, and hypercare support model", 185, 214, "24,26", "KPMG PMO Lead + Infosys Service Delivery Lead", "Pre-GoLive command center and post-GoLive stabilization model defined."),
    make_task(
        28,
        1,
        "Phase 2: Build merger zone and prepare transition factory",
        215,
        395,
        notes="Build core platforms, prepare migration factory, and complete test readiness for QG2 and QG3.",
    ),
    make_task(29, 2, "2.1 Merger zone platform build", 215, 340, notes="Infrastructure, security, identity, workplace, and monitoring build."),
    make_task(30, 3, "Provision merger zone hosting, compute, storage, and base services", 215, 285, "18,21", "Infosys Infrastructure Architect + KPMG Infrastructure Architect", "Core hosting platform established for temporary operational landing zone."),
    make_task(31, 3, "Deploy network connectivity, segmentation, and core security controls", 240, 340, "18,19", "Infosys Network Lead + Infosys Security Lead", "Connectivity enables progressive site onboarding with required security segregation."),
    make_task(32, 3, "Build identity federation, AD services, and privileged access tooling", 230, 320, "19", "Infosys IAM Lead + KPMG Security Architect", "Core identity services ready for user migration into the merger zone."),
    make_task(33, 3, "Stand up M365 tenant, collaboration tools, and service desk platform", 245, 330, "20,27", "Infosys Cloud Lead + Infosys Service Delivery Lead", "Workplace and support stack prepared for all migrated users."),
    make_task(34, 3, "Implement monitoring, backup, disaster recovery, and observability", 285, 340, "30,31,33", "Infosys Infrastructure Architect + KPMG Infrastructure Architect", "Operational readiness controls active before broad migration starts."),
    make_task(35, 2, "2.2 SAP and application transition factory", 215, 395, notes="SAP carve-out build plus app migration factory and tooling."),
    make_task(36, 3, "Prepare SAP environments, copies, and carve-out configuration path", 245, 330, "15,25,30", "Infosys SAP Lead + KPMG SAP Architect", "SAP technical path into merger zone prepared and validated."),
    make_task(37, 3, "Execute SAP separation build, role redesign, and interface rewiring", 330, 370, "36", "Infosys SAP Lead + KPMG SAP Architect + JCI SAP Owner", "Shared services, roles, and interfaces aligned to carved-out business scope."),
    make_task(38, 3, "Build migration factory, tooling, and automation for application moves", 215, 300, "16,25", "Infosys Application Lead + Infosys Data Migration Lead", "Migration tooling covers data movement, validation, and evidence collection."),
    make_task(39, 3, "Build and validate wave 1 critical applications in merger zone", 300, 365, "30,38", "Infosys Application Lead + KPMG Enterprise Architect", "Critical application stack prepared first for integrated testing."),
    make_task(40, 3, "Pre-build wave 2 standard applications and package wave 3 edge apps", 315, 390, "38", "Infosys Application Lead + KPMG Enterprise Architect", "Standard and specialist applications prepared for accelerated migration execution."),
    make_task(41, 2, "2.3 Site and user migration readiness", 240, 395, notes="Endpoint readiness, local playbooks, and pilot readiness across 48 sites."),
    make_task(42, 3, "Build device images, enrollment policies, and endpoint packaging", 240, 320, "20,26", "Infosys Workplace Lead + KPMG Workplace Lead", "Endpoint build supports phased user migration into merger zone services."),
    make_task(43, 3, "Prepare regional migration playbooks, floorwalker model, and logistics", 300, 365, "26,27,42", "KPMG Deployment Lead + Infosys Service Delivery Lead", "Regional cutover packs and local support rosters ready."),
    make_task(44, 3, "Run pilot site connectivity and service desk rehearsals", 340, 395, "31,33,43", "Infosys Service Delivery Lead + JCI IT Manager", "Pilot proves merger zone onboarding process and support flows before scaled migration."),
    make_task(45, 2, "2.4 Test planning and go-no-go preparation", 300, 395, notes="Formal test planning, mock rehearsals, and readiness evidence model."),
    make_task(46, 3, "Build SIT and UAT plans, evidence templates, and sign-off workflow", 300, 340, "27,33", "KPMG Test Lead + Infosys QA Lead", "Common evidence and defect lifecycle agreed for all workstreams."),
    make_task(47, 3, "Run SAP mock cutover rehearsal 1 and capture rollback evidence", 350, 380, "37,46", "Infosys SAP Lead + KPMG SAP Architect", "Initial mock validates timing, dependencies, and rollback controls."),
    make_task(48, 3, "Complete operational readiness dossier for command center launch", 340, 395, "34,44,46,47", "KPMG PMO Lead + Infosys Programme Manager", "Evidence package prepared for test-ready milestone and steering review."),
    make_task(
        49,
        3,
        "QG2 and QG3 - Build complete and test entry approved",
        395,
        395,
        "39,40,44,47,48",
        "KPMG PMO Lead + JCI Sponsor + Bosch Sponsor",
        "Merger zone build complete and programme cleared for integrated testing and migration execution.",
        milestone=True,
    ),
    make_task(
        50,
        1,
        "Phase 3: Test, migrate to merger zone, and prepare Bosch landing",
        396,
        527,
        notes="Formal testing, migration waves, Bosch operational readiness, and QG4 entry.",
    ),
    make_task(51, 2, "3.1 Formal testing and defect closure", 396, 500, notes="Integrated testing, UAT, security, performance, DR, and regression closure."),
    make_task(52, 3, "Execute end-to-end SIT across SAP, applications, and core services", 396, 450, "49", "Infosys QA Lead + KPMG Test Lead + Infosys SAP Lead", "Full end-to-end process flows validated inside the merger zone."),
    make_task(53, 3, "Run performance, security, and disaster recovery validation", 410, 470, "49,34", "Infosys Security Lead + Infosys Infrastructure Architect", "Capacity, resilience, and control effectiveness validated for GoLive load."),
    make_task(54, 3, "Run UAT with business teams across all regions and site groups", 430, 485, "46,52", "KPMG Test Lead + JCI Business Lead + Bosch Business Lead", "Business users validate SAP and non-SAP journeys needed at cutover."),
    make_task(55, 3, "Close defects, run regression, and obtain UAT closure evidence", 460, 500, "52,53,54", "Infosys QA Lead + KPMG Test Lead", "Critical and high defects resolved before final readiness stage."),
    make_task(56, 2, "3.2 Migration waves into merger zone", 410, 520, notes="Users, applications, and data move from JCI to the merger zone after the approved TSA buffer period while overall programme progress and GoLive remain unchanged."),
    make_task(57, 3, "Migrate user waves 1 and 2 with workplace and identity cutover", 410, 470, "44,49", "Infosys Service Delivery Lead + Infosys Workplace Lead + JCI IT Manager", "First half of the user base transitions from JCI to merger zone services."),
    make_task(58, 3, "Migrate user waves 3 and 4 and complete 48-site onboarding", 470, 520, "57,55", "Infosys Service Delivery Lead + Infosys Workplace Lead + JCI IT Manager", "All 12,000 users operate from the merger zone before QG4."),
    make_task(59, 3, "Execute wave 1 application and data migration into merger zone", 410, 455, "39,52", "Infosys Application Lead + Infosys Data Migration Lead", "Critical application and data moves completed with reconciliation."),
    make_task(60, 3, "Execute wave 2 application and data migration into merger zone", 455, 495, "40,59", "Infosys Application Lead + Infosys Data Migration Lead", "Standard application estate transitioned and validated."),
    make_task(61, 3, "Execute wave 3 specialist application completion and data closure", 495, 520, "60,55", "Infosys Application Lead + Infosys Data Migration Lead", "Specialist and regional edge applications completed before QG4."),
    make_task(62, 2, "3.3 Bosch landing and operating model readiness", 430, 520, notes="Bosch acceptance of connectivity, support processes, and eventual integration path."),
    make_task(63, 3, "Enable Bosch connectivity, identity trust, and operational acceptance", 430, 495, "21,32,33", "Bosch IT Manager + Infosys IAM Lead + Infosys Network Lead", "Bosch confirms connectivity and identity patterns for downstream integration."),
    make_task(64, 3, "Align resolver groups, service desk, and runbooks with Bosch operations", 470, 515, "55,63", "Infosys Service Delivery Lead + Bosch IT Manager", "Support ownership model agreed for post-GoLive stabilization and later handover."),
    make_task(65, 3, "Run final SAP cutover rehearsal and certify rollback readiness", 485, 520, "37,55", "Infosys SAP Lead + KPMG SAP Architect", "Final rehearsal confirms cutover steps and fallback timing."),
    make_task(66, 2, "3.4 Pre-GoLive evidence pack", 518, 527, notes="Final evidence pack proving all migration and readiness activities are complete."),
    make_task(67, 3, "Reconcile all migrations and archive formal UAT sign-off", 518, 523, "58,61,55", "KPMG PMO Lead + Infosys Programme Manager", "Users, applications, and data all reconciled before gate approval."),
    make_task(68, 3, "Finalize infrastructure, SAP, and operational readiness dossier", 518, 524, "53,64,65", "Infosys Programme Manager + KPMG PMO Lead", "Operations, security, SAP, and service evidence complete for steering."),
    make_task(69, 3, "Prepare steering gate deck and cutover recommendation", 523, 527, "67,68", "KPMG PMO Lead + Bosch Sponsor + JCI Sponsor", "Formal gate recommendation prepared for QG4 decision."),
    make_task(
        70,
        3,
        "QG4 - Pre-GoLive gate approved",
        527,
        527,
        "69",
        "KPMG PMO Lead + JCI Sponsor + Bosch Sponsor",
        "Pre-GoLive approval granted; final readiness and open-item closure begins.",
        milestone=True,
    ),
    make_task(
        71,
        1,
        "Phase 4: Final readiness and GoLive decision",
        528,
        549,
        notes="Mandatory buffer between QG4 and GoLive for closure, checks, rollback rehearsal, and sign-off.",
    ),
    make_task(72, 2, "4.1 Final readiness and open item closure", 528, 548, notes="No migrations remain; only closure, checks, sign-offs, and command center activation."),
    make_task(73, 3, "Resolve residual defects, confirm health checks, and secure business sign-off", 528, 548, "70", "KPMG Test Lead + Infosys Service Delivery Lead + Bosch Business Lead", "All QG4 actions closed and no critical alerts remain before cutover."),
    make_task(74, 3, "Activate cutover command center and issue final go decision", 546, 548, "73", "KPMG PMO Lead + Infosys Programme Manager + JCI Sponsor + Bosch Sponsor", "Rollback path revalidated and final go decision issued before GoLive day 1."),
    make_task(
        75,
        3,
        "GoLive - Day 1 cutover to merger zone complete",
        549,
        549,
        "74",
        "KPMG PMO Lead + Infosys Programme Manager + JCI Sponsor + Bosch Sponsor",
        "All in-scope users and services operate from the merger zone on GoLive.",
        milestone=True,
    ),
    make_task(
        76,
        1,
        "Phase 5: Hypercare, TSA exit, and Bosch handover",
        550,
        640,
        notes="Ninety-day stabilization only; no new development or migration after GoLive.",
    ),
    make_task(77, 2, "5.1 Hypercare stabilization only", 550, 640, notes="Support, hotfixes, enablement, and operational stabilization only."),
    make_task(78, 3, "Run 24x7 incident triage and command center support", 550, 640, "75", "Infosys Service Delivery Lead + KPMG Project Manager", "Round-the-clock stabilization support across all regions."),
    make_task(79, 3, "Operate hotfix-only release cadence and eliminate recurring problems", 550, 610, "75", "Infosys Application Lead + Infosys SAP Lead", "Bug fixes only after GoLive; no new features or migrations."),
    make_task(80, 3, "Provide user enablement, local floorwalking, and adoption support", 550, 590, "75", "KPMG Change Lead + Infosys Workplace Lead", "Support users in the new merger zone operating model after cutover."),
    make_task(81, 2, "5.2 TSA exit and Bosch operations handover", 560, 640, notes="Tower-by-tower TSA exit followed by Bosch operational ownership uplift."),
    make_task(82, 3, "Execute TSA exits by service tower and confirm acceptance", 560, 620, "75", "KPMG Project Manager + JCI IT Manager + Bosch IT Manager", "JCI TSA services closed according to approved exit criteria and acceptance tests."),
    make_task(83, 3, "Transfer knowledge, runbooks, and service ownership to Bosch operations", 560, 625, "75,64", "Bosch IT Manager + Infosys Programme Manager + KPMG Infrastructure Architect", "Bosch operations absorb merger zone support knowledge and control model."),
    make_task(84, 3, "Transition merger zone operating model to Bosch steady-state teams", 590, 635, "82,83", "Bosch IT Manager + Infosys Programme Manager", "Operational ownership transitions from project mode to Bosch steady-state teams."),
    make_task(85, 2, "5.3 Governance closure and archive", 610, 640, notes="Closure reporting, lessons learned, and artefact archive."),
    make_task(86, 3, "Close financials, capture lessons learned, and archive artefacts", 610, 635, "82,83", "KPMG PMO Lead + KPMG Project Manager", "Programme closure outputs completed and repository artefacts archived."),
    make_task(
        87,
        3,
        "QG5 - Project completion approved",
        640,
        640,
        "78,84,86",
        "KPMG PMO Lead + JCI Sponsor + Bosch Sponsor",
        "Ninety-day hypercare complete, TSA exited, and handover to Bosch operations complete.",
        milestone=True,
    ),
]


def _write_temp_csv(path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        handle.write("ID,Outline Level,Name,Duration,Start,Finish,Predecessors,Resource Names,Notes,Milestone\n")
        for task_id, level, name, duration, start, finish, preds, resources, notes, milestone in TASKS:
            handle.write(
                f'{task_id},"{level}","{name.replace("\"", "\"\"")}",'
                f'"{duration}","{start}","{finish}","{preds}",'
                f'"{resources.replace("\"", "\"\"")}","{notes.replace("\"", "\"\"")}","{milestone}"\n'
            )


def _generate_excel() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"

    headers = [
        "ID",
        "Outline Level",
        "Name",
        "Duration",
        "Start",
        "Finish",
        "Predecessors",
        "Resource Names",
        "Notes",
        "Milestone",
    ]
    sheet.append(headers)

    header_fill = PatternFill(start_color="002147", end_color="002147", fill_type="solid")
    phase_fill = PatternFill(start_color="003B6E", end_color="003B6E", fill_type="solid")
    section_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    milestone_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    even_fill = PatternFill(start_color="EFF4FB", end_color="EFF4FB", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    white_bold_font = Font(bold=True, color="FFFFFF", size=10)
    milestone_font = Font(bold=True, color="000000", size=10)
    detail_font = Font(color="000000", size=9)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    indent_map = {1: "", 2: "  ", 3: "    "}

    for row_index, task in enumerate(TASKS, start=2):
        task_id, level, name, duration, start, finish, preds, resources, notes, milestone = task
        sheet.append(
            [
                task_id,
                level,
                indent_map.get(level, "") + name,
                duration,
                datetime.strptime(start, "%Y-%m-%d"),
                datetime.strptime(finish, "%Y-%m-%d"),
                preds,
                resources,
                notes,
                milestone,
            ]
        )
        row = sheet[row_index]
        if milestone == "Yes":
            row_fill = milestone_fill
            row_font = milestone_font
        elif level == 1:
            row_fill = phase_fill
            row_font = white_bold_font
        elif level == 2:
            row_fill = section_fill
            row_font = white_bold_font
        else:
            row_fill = even_fill if row_index % 2 == 0 else odd_fill
            row_font = detail_font

        for cell in row:
            cell.fill = row_fill
            cell.font = row_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"

    sheet.column_dimensions["A"].width = 6
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 62
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 14
    sheet.column_dimensions["F"].width = 14
    sheet.column_dimensions["G"].width = 16
    sheet.column_dimensions["H"].width = 46
    sheet.column_dimensions["I"].width = 60
    sheet.column_dimensions["J"].width = 10

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:J1"

    workbook.save(XLSX_PATH)


if __name__ == "__main__":
    print(f"[{PROJECT_NAME}] Generating project schedule")
    print(f"  Seller: {SELLER}")
    print(f"  Buyer: {BUYER}")
    print(f"  Business: {BUSINESS}")
    print(f"  Delivery partner: {DELIVERY_PARTNER}")
    print(f"  Scope: {WORLDWIDE_SITES} sites | {IT_USERS} users | {APPLICATIONS}+ applications")
    print(f"  Start: {START_DATE:%Y-%m-%d} | GoLive: {GOLIVE_DATE:%Y-%m-%d} | Completion: {COMPLETION_DATE:%Y-%m-%d}")
    print(f"  Total tasks: {len(TASKS)}")

    _generate_excel()
    _write_temp_csv(CSV_PATH)

    result = subprocess.run(
        [
            sys.executable,
            str(HERE / "generate_msp_xml.py"),
            "--csv",
            str(CSV_PATH),
            "--out",
            str(XML_PATH),
            "--project",
            PROJECT_NAME,
        ],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print(f"XML generation failed:\n{result.stderr}")
        sys.exit(1)

    print(f"  Output folder version: {OUTPUT_FOLDER_NAME} | {DOCUMENT_VERSION}")
    print(f"  Approved TSA end: {APPROVED_TSA_END_DATE:%Y-%m-%d}")
    print(f"  XLSX: {XLSX_PATH}")
    print(f"  CSV:  {CSV_PATH}")
    print(f"  XML:  {XML_PATH}")
    print("  Quality gates: QG0, QG1, QG2/QG3, QG4, GoLive, QG5")