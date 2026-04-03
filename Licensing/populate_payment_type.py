"""
Fill blank 'Type of Payment' cells in 'Astra Shoppping List with SWS.xlsx'
using:
  1. Existing vendor/product patterns from the same file
  2. A curated knowledge base built from research + web lookups
  3. Product keyword heuristics as fallback

AI-filled values are prefixed with: "(Populated by AI) "
Output: Astra_Shoppping_List_with_SWS_AI_Payment.xlsx
"""
import os, re, json, collections
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

BASE = r'c:\Users\kho1sgp\OneDrive - Bosch Group\My Work Documents\AI topics\BDMIL use case\Software_license_project'

AI_PREFIX    = '(Populated by AI) '
AI_FILL_HEX  = 'FFF2CC'   # warm yellow
AI_FONT_COL  = '7A4F00'   # dark amber text
COL_NAVY     = '003c64'

# ═══════════════════════════════════════════════════════════════
# CURATED KNOWLEDGE BASE
# key = lowercase(vendor) or (lowercase(vendor), product_keyword)
# value = (payment_type, confidence, source_note)
# ═══════════════════════════════════════════════════════════════

# Canonical payment types used in the file:
S   = 'subscription'
P   = 'perpetual'
F   = 'free of payment'
OSS = 'open source'
PM  = 'perpetual + maintenance'
SS  = 'subscription (commercial)'

# ── (a) Vendor-level lookups (deterministic majority) ──────────────────────
VENDOR_MAP = {
    # A
    'acatec software / revalize'        : (S,   'Revalize/Speedmaxx = SaaS subscription'),
    'allyx technology'                  : (S,   'Xmaint is SaaS/subscription asset mgmt'),
    'alm toolkit'                       : (F,   'ALM Toolkit for Power BI is open source/free'),
    'amazon'                            : (F,   'Amazon Corretto/OpenJDK = free open source JDK'),
    'ametek'                            : (P,   'AMETEK test & measurement = perpetual software'),
    'anaconda'                          : (S,   'Anaconda Business/Enterprise = subscription'),
    'andrea vacondio'                   : (F,   'PDF Split and Merge = open source/freeware'),
    'ansgar becker'                     : (F,   'HeidiSQL = open source, free'),
    'ansys'                             : (S,   'Ansys moved to subscription model (2019+)'),
    'antoine potten'                    : (F,   'Ant Renamer = freeware'),
    'apple'                             : (F,   'iTunes = free of payment'),
    'arduino de'                        : (F,   'Arduino = open source/free'),
    'at&t'                              : (F,   'Graphviz = open source (AT&T Bell Labs, EPL)'),
    'autem gmbh'                        : (P,   'SPS-ANALYZER = industrial, perpetual'),
    'autohotkey foundation llc'         : (F,   'AutoHotkey = open source/free GPL'),
    # B
    'bambu lab'                         : (F,   'Bambu Studio = open source slicer'),
    'benito van der zander'             : (F,   'TeXstudio = open source/free'),
    'beyondtrust'                       : (S,   'BeyondTrust PAM/remote access = subscription'),
    'bosch'                             : (P,   'Internal Bosch tools = perpetual/internal'),
    'bosch bd'                          : (F,   'Bosch BD internal software = internal/free'),
    'bosch dc'                          : (P,   'Bosch DC internal software = perpetual/internal'),
    'bosch-bd'                          : (P,   'Bosch-BD internal tool = perpetual/internal'),
    'broadcom'                          : (S,   'VMware/Broadcom products = subscription post-2024'),
    'bryce harrington'                  : (F,   'Inkscape = open source/free (GPL)'),
    # C
    'cadence design systems'            : (PM,  'Cadence EDA tools = perpetual + maintenance'),
    'caresoft global'                   : (S,   'CARESoft = SaaS subscription'),
    'carl  zeiss microscopy gmbh'       : (P,   'Zeiss ZEN software = perpetual with HW'),
    'cegid'                             : (S,   'Cegid ERP/Finance/HR = SaaS subscription'),
    'charles lechasseur'                : (F,   'Folder Size = freeware'),
    'checkmk'                           : (S,   'Checkmk enterprise = subscription (community=free)'),
    'christian schenk'                  : (F,   'MikTeX = open source/free LaTeX distribution'),
    'cisco'                             : (S,   'Cisco software = subscription (Smart Licensing)'),
    'citrix'                            : (S,   'Citrix Virtual Apps/Desktops = subscription'),
    'colin mackie'                      : (F,   'WinMD5Free = freeware'),
    # D
    'dax studio'                        : (F,   'DAX Studio = open source/free (Power BI tool)'),
    'dimitry polivaev'                  : (F,   'Freeplane = open source/free'),
    'dominik reichl'                    : (F,   'KeePass = open source/free'),
    'don ho'                            : (F,   'Notepad++ = open source/free GPL'),
    # E
    'eclipse foundation'                : (F,   'Eclipse IDE = open source/free EPL'),
    'elasticsearch'                     : (S,   'Elasticsearch enterprise = subscription (basic=free)'),
    'empower'                           : (S,   'Empower software = subscription'),
    'entrust'                           : (S,   'Entrust certificate/identity mgmt = subscription'),
    'epic games'                        : (F,   'Unreal Engine = free (royalty for games)'),
    'eplan software & service'          : (PM,  'EPLAN = perpetual + annual maintenance'),
    'ewsoftware'                        : (F,   'Sandcastle SHFB = open source/free'),
    'exakom'                            : (P,   'Exakom maintenance software = perpetual'),
    # F
    'famatech'                          : (F,   'Radmin Viewer = freeware'),
    'famos'                             : (P,   'imc FAMOS = perpetual measurement software'),
    'festo'                             : (F,   'Festo automation software = free with hardware'),
    'flexera'                           : (S,   'FlexNet Manager/Flexera = subscription SaaS'),
    'fortinet'                          : (S,   'Fortinet FortiGate/security = subscription'),
    # G
    'genotrance-ganesh viswanathan'     : (F,   'Small utility freeware'),
    'ghisler software'                  : (P,   'Total Commander = perpetual shareware license'),
    'gingco.net'                        : (S,   'GINGCo PLM = subscription'),
    'git-team'                          : (F,   'Git = open source/free GPL'),
    'github'                            : (S,   'GitHub Enterprise = subscription'),
    'glavsoft'                          : (F,   'Qt VNC/TigerVNC = open source/free'),
    'godaddy'                           : (S,   'GoDaddy domains/SSL = annual subscription'),
    'gordon lyon'                       : (F,   'Nmap = open source/free GPL'),
    # H
    'hbk(hottinger brueel & kjaer)'    : (S,   'HBK measurement software = subscription'),
    'head acoustics'                    : (PM,  'HEAD acoustics software = perpetual + maintenance'),
    'hexagon'                           : (S,   'Hexagon MI/SHM software = subscription'),
    'hikikomori82'                      : (F,   'Small freeware utility'),
    'hp'                                : (F,   'HP drivers/utilities = free of payment'),
    # I
    'ibm'                               : (S,   'IBM software = subscription (IBM NEXT)'),
    'igor pavlov'                       : (F,   '7-Zip = open source/free LGPL'),
    'imagewriter developers'            : (F,   'imageWriter = freeware/open source'),
    'inkscape/bryce harrington'         : (F,   'Inkscape = open source/free'),
    'ion'                               : (S,   'ION financial software = subscription'),
    'irfan skiljan'                     : (F,   'IrfanView = freeware (free for non-commercial)'),
    # J
    'jabra/gn audio'                    : (F,   'Jabra Direct = free companion software'),
    'jamie o connell'                   : (F,   'Small freeware utility'),
    'kapacity'                          : (P,   'Kapacity workforce mgmt = perpetual'),
    # K
    'keyence'                           : (F,   'Keyence software = free with hardware/device'),
    'krzysztof kowalczyk'               : (F,   'SumatraPDF = open source/free'),
    # L
    'learnpulse'                        : (P,   'Learnpulse LMS = perpetual'),
    'lingo4you'                         : (F,   'Language learning = freeware'),
    'logitech'                          : (F,   'Logitech G HUB / Options = free driver software'),
    'lrs'                               : (S,   'LRS output mgmt = subscription'),
    'lucy security'                     : (S,   'Lucy Security awareness training = subscription'),
    # M
    'magtrol'                           : (F,   'Magtrol MTEST = free with hardware'),
    'marek jasinski'                    : (F,   'FreeCommander = freeware'),
    'martin prikryl'                    : (F,   'WinSCP = open source/free GPL'),
    'meltytech'                         : (F,   'Shotcut = open source/free GPL'),
    'mercateo'                          : (F,   'Mercateo procurement platform = free access'),
    'micro focus'                       : (S,   'Micro Focus (now OpenText) = subscription'),
    'microsoft sysinternals'            : (F,   'Sysinternals = free utilities from Microsoft'),
    'midox/jamie o connell'             : (F,   'Small freeware utility'),
    'mozilla'                           : (F,   'Firefox/Thunderbird = open source/free'),
    'multimedia'                        : (P,   'Multimedia software = perpetual'),
    # N
    'national instruments'              : (S,   'NI LabVIEW/TestStand = subscription (2020+)'),
    'nexthink'                          : (S,   'Nexthink DEX platform = subscription'),
    'obs studio contributors'           : (F,   'OBS Studio = open source/free GPL'),
    'obs studio contributors-tricaster' : (F,   'OBS Studio = open source/free GPL'),
    'open text'                         : (S,   'OpenText enterprise = subscription platform'),
    'openshot studios'                  : (F,   'OpenShot = open source/free LGPL'),
    'optimala'                          : (S,   'Optimala workforce = subscription SaaS'),
    # P
    'pilz'                              : (P,   'Pilz safety software = perpetual with hardware'),
    'pkware'                            : (S,   'PKWARE/WinZip = subscription (V11+)'),
    'plantronics'                       : (F,   'Plantronics Hub = free driver software'),
    'ptc'                               : (S,   'PTC CAD/PLM = subscription'),
    'putty/simon tatham'                : (F,   'PuTTY = open source/free MIT'),
    # R
    'radacad'                           : (S,   'SQLBI/RADACAD Power BI tools = subscription'),
    'realtek'                           : (F,   'Realtek audio/network drivers = free'),
    'rocket software'                   : (S,   'Rocket Software enterprise = subscription'),
    # S
    's.a.x. software'                   : (P,   'S.A.X. document mgmt = perpetual'),
    'scooter software'                  : (P,   'Beyond Compare = perpetual license'),
    'segger microcontroller'            : (P,   'SEGGER J-Link/Embedded Studio = perpetual'),
    'sew-eurodrive'                     : (F,   'MOVITOOLS = free with SEW hardware'),
    'sick'                              : (F,   'SICK SOPAS/SICKView = free with sensor hardware'),
    'splunk'                            : (S,   'Splunk SIEM = subscription'),
    'sqlbi'                             : (F,   'SQLBI DAX tools = free/open source'),
    'stethos'                           : (P,   'Stethos industrial software = perpetual'),
    'stiegele datensysteme'             : (P,   'Stiegele document mgmt = perpetual'),
    'swift'                             : (S,   'SWIFT financial messaging = subscription'),
    'synaptics'                         : (F,   'Synaptics drivers = free'),
    # T
    'tabular editor'                    : (S,   'Tabular Editor 3 = subscription (per user/month)'),
    'tenfold securities'                : (P,   'Tenfold IAM = perpetual'),
    'tenorshare'                        : (P,   'Tenorshare data recovery = perpetual'),
    'testo'                             : (F,   'Testo ComSoft = free companion software'),
    'the gimp team'                     : (F,   'GIMP = open source/free GPL'),
    'the mathworks'                     : (S,   'MATLAB = subscription (2024+ model)'),
    'think-cell'                        : (S,   'think-cell PowerPoint add-in = subscription'),
    'topsolid'                          : (PM,  'TopSolid CAD/CAM = perpetual + maintenance'),
    'trendmicro'                        : (S,   'Trend Micro security = subscription'),
    'tricat'                            : (S,   'TriCAT e-learning LMS = subscription'),
    'trivit'                            : (P,   'TRIVIT industrial software = perpetual'),
    'ultravnc'                          : (F,   'UltraVNC = open source/free GPL'),
    # V
    'veeam'                             : (S,   'Veeam Data Platform = subscription'),
    'versiondog'                        : (S,   'VersionDog/Octoplant backup = subscription'),
    # W
    'walkme'                            : (S,   'WalkMe DAP = subscription SaaS (SAP-owned 2023)'),
    'wibu systems'                      : (PM,  'Wibu CodeMeter = perpetual + maintenance'),
    'wireshark development team'        : (F,   'Wireshark = open source/free GPL'),
    'wireshark foundation'              : (F,   'Wireshark = open source/free GPL'),
    # Y
    'yokogawa elect'                    : (S,   'Yokogawa measurement = subscription'),
    # Z
    'zebra technologies'                : (F,   'Zebra ZPL/ZDesigner = free with hardware'),
}

# ── (b) Product-level overrides (vendor+keyword combo) ────────────────────
# Format: (vendor_lower_substring, product_keyword_lower) -> payment_type
PRODUCT_OVERRIDES = {
    # adobe patch/hotfix is always free
    ('adobe',     'hotfix')     : F,
    ('adobe',     'patch')      : F,
    # autodesk free viewers
    ('autodesk',  'trueview')   : F,
    ('autodesk',  'design review') : F,
    ('autodesk',  'licensing service') : F,
    # microsoft free items
    ('microsoft', 'sysinternals'): F,
    ('microsoft', 'default')    : F,
    ('microsoft', 'redistribut'): F,
    # anaconda free community
    ('anaconda',  '2020')       : S,
    # amazon free
    ('amazon',    'corretto')   : F,
    ('amazon',    'openjdk')    : F,
    # mathworks free downloads
    ('mathworks', 'compiler runtime'): F,
    ('mathworks', 'runtime')    : F,
    # siemens TBD vs subscription
    ('siemens',   'tia')        : S,
    ('siemens',   'step 7')     : PM,
    # checkmk free community
    ('checkmk',   'community')  : F,
    # bosch internal
    ('bosch bd',  'catia')      : PM,
    ('bosch bd',  'windows')    : F,
    # GitHub enterprise vs personal
    ('github',    'enterprise') : S,
    ('github',    'actions')    : S,
    # Oracle java
    ('oracle',    'java se')    : SS,
    ('oracle',    'jdk')        : SS,
}

# ── (c) Product keyword fallbacks ────────────────────────────────────────────
# If vendor not in map, scan product name for these keywords
PROD_KEYWORD_RULES = [
    (r'\bfreeware\b',                    F),
    (r'\bopen.?source\b',                F),
    (r'\bgpl\b',                         F),
    (r'\bopen.?jdk\b',                   F),
    (r'\bcorretto\b',                    F),
    (r'\bhotfix\b',                      F),
    (r'\bpatch\b',                       F),
    (r'\bplugin\b',                      F),
    (r'\bextension\b',                   F),
    (r'\bdriver\b',                      F),
    (r'\bfirmware\b',                    F),
    (r'\bviewer\b',                      F),
    (r'\b(trueview|design.?review|licensing.?service|sysinternals)\b', F),
    (r'\bsaas\b',                        S),
    (r'\bsubscription\b',                S),
    (r'\bcloud\b',                       S),
    (r'\benterprise\b',                  S),
    (r'\baas\b',                         S),   # software-as-a-service suffix
    (r'\bperpetual\b',                   P),
    (r'\bmaintenance\b',                 PM),
]

# ── Load data ─────────────────────────────────────────────────────────────────
def clean(v):
    s = str(v or '').strip()
    return s if s not in ('None', '') else ''

wb = openpyxl.load_workbook(os.path.join(BASE, '_sws_copy2.xlsx'))
ws = wb['Shoppinglist']
rows_raw = list(ws.iter_rows(values_only=True))

# Build vendor->payment map from existing filled rows
vendor_pt = collections.defaultdict(collections.Counter)
for r in rows_raw[1:]:
    vendor = clean(r[1])
    ptype  = clean(r[5])
    if vendor and ptype and 'tbd' not in ptype.lower():
        vendor_pt[vendor.lower()][ptype.lower()] += 1

def dominant_existing(vendor_key, threshold=0.6):
    ctr = vendor_pt.get(vendor_key, collections.Counter())
    total = sum(ctr.values())
    if not total:
        return None
    top, n = ctr.most_common(1)[0]
    return top if n/total >= threshold else None

# ── Classifier ────────────────────────────────────────────────────────────────
def classify(vendor, product):
    vk = vendor.lower().strip()
    pk = product.lower().strip()
    combined = vk + ' ' + pk

    # 1. Product-level override
    for (vv, pv), pt in PRODUCT_OVERRIDES.items():
        if vv in vk and pv in pk:
            return pt, f'product override ({vv}/{pv})'

    # 2. Vendor map (curated)
    if vk in VENDOR_MAP:
        pt, note = VENDOR_MAP[vk]
        return pt, f'curated knowledge base ({note})'

    # Partial vendor match in curated map
    for kv, (pt, note) in VENDOR_MAP.items():
        if len(kv) >= 5 and (kv in vk or vk in kv):
            return pt, f'partial vendor match curated ({kv})'

    # 3. Existing file pattern for vendor
    dom = dominant_existing(vk)
    if dom:
        return dom, f'existing file pattern ({vendor})'

    # 4. Product keyword rules
    for pattern, pt in PROD_KEYWORD_RULES:
        if re.search(pattern, combined, re.IGNORECASE):
            return pt, f'product keyword match ({pattern})'

    # 5. Fallback: majority class in file
    return S, 'fallback (subscription is majority observed pattern)'

# ── Apply to sheet ────────────────────────────────────────────────────────────
ai_fill  = PatternFill('solid', fgColor=AI_FILL_HEX)
ai_font  = Font(color=AI_FONT_COL, size=9, italic=True, bold=False)
centered = Alignment(horizontal='center', vertical='center', wrap_text=True)

applied  = 0
skipped  = 0
log      = []

for row_idx in range(2, ws.max_row + 1):
    vendor  = clean(ws.cell(row_idx, 2).value)
    product = clean(ws.cell(row_idx, 3).value)
    ptype   = clean(ws.cell(row_idx, 6).value)

    if not vendor and not product:
        continue
    if ptype:               # already has a value — do NOT overwrite
        skipped += 1
        continue

    proposed, reason = classify(vendor, product)
    label = AI_PREFIX + proposed

    cell = ws.cell(row=row_idx, column=6, value=label)
    cell.fill      = ai_fill
    cell.font      = ai_font
    cell.alignment = centered

    log.append({'row': row_idx, 'vendor': vendor, 'product': product,
                'proposed': proposed, 'reason': reason})
    applied += 1

print(f'Applied AI payment type to {applied} rows  |  Skipped {skipped} rows (already filled)')

# ── Distribution summary ──────────────────────────────────────────────────────
dist = collections.Counter(x['proposed'] for x in log)
print('Distribution of AI-proposed types:')
for k, v in dist.most_common():
    print(f'  {k!r}: {v}')

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(BASE, 'Astra_Shoppping_List_with_SWS_AI_Payment.xlsx')
wb.save(out_path)
print(f'\nSaved: {out_path}')

# ── Detailed log ──────────────────────────────────────────────────────────────
print('\n=== AI PAYMENT TYPE LOG ===')
for x in log:
    print('row=%-4d  %-35s  %-40s  ->  %-28s  [%s]' % (
        x['row'], x['vendor'][:35], x['product'][:40],
        x['proposed'], x['reason'][:70]))
