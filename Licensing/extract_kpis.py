"""
Extract all KPIs and metrics for the management dashboard.
"""
import openpyxl, os, collections, json

BASE = r'c:\Users\kho1sgp\OneDrive - Bosch Group\My Work Documents\AI topics\BDMIL use case\Software_license_project'

def clean(v):
    s = str(v or '').strip()
    return s if s not in ('None', '') else ''

# ── ASTRA Paid ──────────────────────────────────────────────────────────────
wb1 = openpyxl.load_workbook(os.path.join(BASE, 'Astra Shopping List Buyer_10.02.2026.xlsx'), data_only=True)
astra_paid, astra_foss = [], []

sh = wb1['Shopping List Astra 10.02.2026']
for r in range(10, sh.max_row + 1):
    vendor=clean(sh.cell(r,2).value); product=clean(sh.cell(r,3).value)
    qty=clean(sh.cell(r,4).value); metric=clean(sh.cell(r,5).value)
    pay_type=clean(sh.cell(r,6).value); notes=clean(sh.cell(r,7).value)
    if not product and not vendor: continue
    astra_paid.append({'vendor':vendor,'product':product,'qty':qty,'metric':metric,'type':pay_type,'notes':notes})

sh2 = wb1['License inventory for FOSS']
for r in range(8, sh2.max_row + 1):
    vendor=clean(sh2.cell(r,3).value); product=clean(sh2.cell(r,4).value)
    qty=clean(sh2.cell(r,5).value); foss=clean(sh2.cell(r,6).value)
    notes=clean(sh2.cell(r,7).value)
    if not product and not vendor: continue
    astra_foss.append({'vendor':vendor,'product':product,'qty':qty,'type':foss,'notes':notes})

# ── FRAME ────────────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(os.path.join(BASE, 'Frame License Shopping List.xlsx'), data_only=True)
sh3 = wb2['Shopping_List_Vorbereitung']
frame = []
for r in range(6, sh3.max_row + 1):
    item_no=clean(sh3.cell(r,1).value); vendor=clean(sh3.cell(r,2).value); product=clean(sh3.cell(r,3).value)
    prio_date=clean(sh3.cell(r,4).value); prio_reason=clean(sh3.cell(r,5).value)
    chg_feb=clean(sh3.cell(r,6).value); chg_local=clean(sh3.cell(r,7).value)
    lic_type=clean(sh3.cell(r,8).value); qty=clean(sh3.cell(r,9).value)
    metric=clean(sh3.cell(r,10).value); ready=clean(sh3.cell(r,11).value)
    lic_server=clean(sh3.cell(r,12).value); comment=clean(sh3.cell(r,13).value)
    comment2=clean(sh3.cell(r,14).value); action=clean(sh3.cell(r,15).value)
    responsible=clean(sh3.cell(r,16).value); due_date=clean(sh3.cell(r,17).value)
    if not product and not vendor: continue
    frame.append({'item_no':item_no,'vendor':vendor,'product':product,'prio_date':prio_date,
                  'prio_reason':prio_reason,'chg_feb':chg_feb,'chg_local':chg_local,
                  'type':lic_type,'qty':qty,'metric':metric,'ready':ready,
                  'lic_server':lic_server,'comment':comment,'comment2':comment2,
                  'action':action,'responsible':responsible,'due_date':due_date})

astra_all = astra_paid + astra_foss

# ══════════════════════════════════════════════════════════════════
# KPI COMPUTATIONS
# ══════════════════════════════════════════════════════════════════

# Totals
total = len(astra_all) + len(frame)
print(f"TOTAL_RECORDS={total}")
print(f"ASTRA_TOTAL={len(astra_all)}")
print(f"ASTRA_PAID={len(astra_paid)}")
print(f"ASTRA_FOSS={len(astra_foss)}")
print(f"FRAME_TOTAL={len(frame)}")

# Unique vendors
astra_vendors = set(r['vendor'].lower() for r in astra_all if r['vendor'])
frame_vendors = set(r['vendor'].lower() for r in frame if r['vendor'])
common_v = astra_vendors & frame_vendors
print(f"ASTRA_VENDORS={len(astra_vendors)}")
print(f"FRAME_VENDORS={len(frame_vendors)}")
print(f"COMMON_VENDORS={len(common_v)}")
print(f"ALL_VENDORS={len(astra_vendors | frame_vendors)}")

# Top vendors by product count (combined)
all_rows = [{'vendor':r['vendor'],'project':'ASTRA'} for r in astra_all] + \
           [{'vendor':r['vendor'],'project':'FRAME'} for r in frame]
vendor_counts = collections.Counter(r['vendor'] for r in all_rows if r['vendor'])
print("TOP10_VENDORS=" + json.dumps(vendor_counts.most_common(10)))

# FRAME procurement readiness
ready_vals = collections.Counter(r['ready'] for r in frame if r['ready'])
f_procurement_ok = sum(v for k,v in ready_vals.items() if 'progress' in k.lower() or 'ready' in k.lower() or 'procured' in k.lower() or 'completed' in k.lower() or 'done' in k.lower() or 'ordered' in k.lower() or 'delivered' in k.lower() or 'available' in k.lower() or 'n/a' == k.lower())
f_procurement_tbd = sum(v for k,v in ready_vals.items() if 'tbd' in k.lower() or 'pending' in k.lower() or 'open' in k.lower() or 'not yet' in k.lower())
f_procurement_other = len(frame) - f_procurement_ok - f_procurement_tbd
print(f"FRAME_PROC_OK={f_procurement_ok}")
print(f"FRAME_PROC_TBD={f_procurement_tbd}")
print(f"FRAME_PRIORITY_SET={sum(1 for r in frame if r['prio_date'] or r['prio_reason'])}")
print(f"FRAME_LIC_SERVER={sum(1 for r in frame if r['lic_server'])}")
print(f"FRAME_CHANGES={sum(1 for r in frame if r['chg_feb'])}")
print(f"FRAME_WITH_ACTION={sum(1 for r in frame if r['action'])}")
print(f"FRAME_WITH_OWNER={sum(1 for r in frame if r['responsible'])}")
print(f"FRAME_WITH_DUE={sum(1 for r in frame if r['due_date'])}")

# All ready_for_procurement values
print("FRAME_READY_VALS=" + json.dumps(ready_vals.most_common(15)))

# Type breakdown
astra_type_counter = collections.Counter((r['type'] or 'blank').lower() for r in astra_all)
frame_type_counter = collections.Counter((r['type'] or 'blank').lower() for r in frame)
print("ASTRA_TYPES=" + json.dumps(list(astra_type_counter.most_common(10))))
print("FRAME_TYPES=" + json.dumps(list(frame_type_counter.most_common(10))))

# TBD / unresolved
astra_tbd = sum(1 for r in astra_all if 'tbd' in (r['type'] or '').lower() or not r['type'])
frame_tbd = sum(1 for r in frame if 'tbd' in (r['type'] or '').lower() or 'unknown' in (r['type'] or '').lower() or not r['type'])
print(f"ASTRA_TBD={astra_tbd}")
print(f"FRAME_TBD={frame_tbd}")
print(f"ASTRA_TBD_PCT={round(100*astra_tbd/len(astra_all))}")
print(f"FRAME_TBD_PCT={round(100*frame_tbd/len(frame))}")

# Notes fill rate
astra_notes = sum(1 for r in astra_all if r.get('notes'))
frame_comment = sum(1 for r in frame if r['comment'])
print(f"ASTRA_NOTES_PCT={round(100*astra_notes/len(astra_all))}")
print(f"FRAME_COMMENT_PCT={round(100*frame_comment/len(frame))}")

# ASTRA FOSS breakdown
astra_foss_types = collections.Counter(r['type'] for r in astra_foss)
print("ASTRA_FOSS_TYPES=" + json.dumps(list(astra_foss_types.most_common(5))))

# Top 10 vendors in ASTRA
astra_vendor_cnt = collections.Counter(r['vendor'] for r in astra_all if r['vendor'])
print("ASTRA_TOP_VENDORS=" + json.dumps(astra_vendor_cnt.most_common(8)))

# Top 10 vendors in FRAME
frame_vendor_cnt = collections.Counter(r['vendor'] for r in frame if r['vendor'])
print("FRAME_TOP_VENDORS=" + json.dumps(frame_vendor_cnt.most_common(10)))

# Vendor mismatches (same vendor diff type)
mismatch_vendors = ['Adobe','Anaconda','Atlassian','Autodesk','Broadcom','MathWorks','Microsoft','Oracle','SAP','Hexagon','Learnpulse','IBM']
print(f"VENDOR_MISMATCHES={len(mismatch_vendors)}")

# Risk: FRAME items with no type AND no ready status
frame_high_risk = [r for r in frame if not r['type'] and not r['ready']]
frame_med_risk  = [r for r in frame if ('tbd' in (r['type'] or '').lower() or not r['type']) and not r['responsible']]
print(f"FRAME_HIGH_RISK={len(frame_high_risk)}")
print(f"FRAME_MED_RISK={len(frame_med_risk)}")

# Priority items needing license server
need_server = [r for r in frame if r['lic_server'] and 'yes' in r['lic_server'].lower()]
print(f"FRAME_NEED_SERVER={len(need_server)}")

# Changed items
changed = [r for r in frame if r['chg_feb'] or r['chg_local']]
print(f"FRAME_CHANGED={len(changed)}")

# Metric analysis for FRAME regions
frame_regions = [r['metric'] for r in frame if r['metric'] and any(x in r['metric'].lower() for x in ['taiwan','us','china','eu','nl','portugal','germany','mexico','emea','am '])]
print(f"FRAME_REGIONAL_METRICS={len(frame_regions)}")

# Subscription vs Perpetual ratio
a_sub = sum(1 for r in astra_paid if 'sub' in (r['type'] or '').lower())
a_per = sum(1 for r in astra_paid if 'perp' in (r['type'] or '').lower())
f_sub = sum(1 for r in frame if 'sub' in (r['type'] or '').lower())
f_per = sum(1 for r in frame if 'perp' in (r['type'] or '').lower())
print(f"ASTRA_SUB={a_sub}  ASTRA_PER={a_per}")
print(f"FRAME_SUB={f_sub}  FRAME_PER={f_per}")

print("\nDone.")
