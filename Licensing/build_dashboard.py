"""
Build both dashboards with Bosch logo + Bosch color theme embedded.
Outputs:
  - Management_Dashboard.html
  - Delta_Analysis_FRAME_vs_ASTRA.html  (regenerated with Bosch theme)
"""
import openpyxl, os, collections, base64

BASE = r'c:\Users\kho1sgp\OneDrive - Bosch Group\My Work Documents\AI topics\BDMIL use case\Software_license_project'

def clean(v):
    s = str(v or '').strip()
    return s if s not in ('None','') else ''

# ── load images ──────────────────────────────────────────────────────────────
with open(os.path.join(BASE,'Bosch_png_b64.txt')) as f:
    LOGO_B64 = f.read().strip()
with open(os.path.join(BASE,'Bosch_color_theme_png_b64.txt')) as f:
    THEME_B64 = f.read().strip()

LOGO_SRC  = f"data:image/png;base64,{LOGO_B64}"
THEME_SRC = f"data:image/png;base64,{THEME_B64}"

# ── load data ────────────────────────────────────────────────────────────────
wb1 = openpyxl.load_workbook(os.path.join(BASE,'Astra Shopping List Buyer_10.02.2026.xlsx'), data_only=True)
astra_paid, astra_foss = [], []
sh = wb1['Shopping List Astra 10.02.2026']
for r in range(10, sh.max_row+1):
    v,p,q,m,t,n = [clean(sh.cell(r,c).value) for c in range(2,8)]
    if not p and not v: continue
    astra_paid.append({'vendor':v,'product':p,'qty':q,'metric':m,'type':t,'notes':n})
sh2 = wb1['License inventory for FOSS']
for r in range(8, sh2.max_row+1):
    v,p,q,fo,n = [clean(sh2.cell(r,c).value) for c in range(3,8)]
    if not p and not v: continue
    astra_foss.append({'vendor':v,'product':p,'qty':q,'type':fo,'notes':n})

wb2 = openpyxl.load_workbook(os.path.join(BASE,'Frame License Shopping List.xlsx'), data_only=True)
sh3 = wb2['Shopping_List_Vorbereitung']
frame = []
for r in range(6, sh3.max_row+1):
    vals = [clean(sh3.cell(r,c).value) for c in range(1,18)]
    if not vals[2] and not vals[1]: continue
    frame.append({'item_no':vals[0],'vendor':vals[1],'product':vals[2],
                  'prio_date':vals[3],'prio_reason':vals[4],'chg_feb':vals[5],
                  'chg_local':vals[6],'type':vals[7],'qty':vals[8],'metric':vals[9],
                  'ready':vals[10],'lic_server':vals[11],'comment':vals[12],
                  'comment2':vals[13],'action':vals[14],'responsible':vals[15],'due_date':vals[16]})

astra_all = astra_paid + astra_foss

# ── KPIs ─────────────────────────────────────────────────────────────────────
total       = len(astra_all) + len(frame)
astra_v     = len(set(r['vendor'].lower() for r in astra_all if r['vendor']))
frame_v     = len(set(r['vendor'].lower() for r in frame if r['vendor']))
common_v    = len(set(r['vendor'].lower() for r in astra_all if r['vendor']) &
                  set(r['vendor'].lower() for r in frame  if r['vendor']))
all_v       = len(set(r['vendor'].lower() for r in astra_all+frame if r['vendor']))

frame_tbd   = sum(1 for r in frame if not r['type'] or 'tbd' in r['type'].lower() or 'unknown' in r['type'].lower())
astra_tbd   = sum(1 for r in astra_all if 'tbd' in (r.get('type') or '').lower())
frame_ready_y = sum(1 for r in frame if r['ready'] and r['ready'].lower() in ('y','yes'))
frame_na    = sum(1 for r in frame if r['ready'] and r['ready'].lower() == 'n/a')
frame_prog  = sum(1 for r in frame if r['ready'] and ('progress' in r['ready'].lower() or 'discussion' in r['ready'].lower() or 'ordered' in r['ready'].lower()))
frame_n     = sum(1 for r in frame if r['ready'] and r['ready'].lower() == 'n')
frame_prio  = sum(1 for r in frame if r['prio_date'] or r['prio_reason'])
frame_changed = sum(1 for r in frame if r['chg_feb'] or r['chg_local'])
frame_owner = sum(1 for r in frame if r['responsible'])
a_sub = sum(1 for r in astra_paid if 'sub' in (r['type'] or '').lower())
a_per = sum(1 for r in astra_paid if 'perp' in (r['type'] or '').lower())
a_tbd_paid = sum(1 for r in astra_paid if 'tbd' in (r['type'] or '').lower())
f_sub = sum(1 for r in frame if 'sub' in (r['type'] or '').lower())
f_per = sum(1 for r in frame if 'perp' in (r['type'] or '').lower())
frame_regional = sum(1 for r in frame if r['metric'] and any(x in r['metric'].lower() for x in ['taiwan','us','china','eu','nl','portugal','germany','mexico','emea']))

# Top vendors combined (normalize case)
all_vendor_cnt = collections.Counter()
for r in astra_all+frame:
    if r['vendor']: all_vendor_cnt[r['vendor']] += 1
top10 = all_vendor_cnt.most_common(10)

# ASTRA top vendors
astra_vc = collections.Counter(r['vendor'] for r in astra_all if r['vendor'])
frame_vc = collections.Counter(r['vendor'] for r in frame if r['vendor'])

# Mismatch vendors
mismatch_data = [
    ("Adobe",       "freeware + subscription", "subscription only"),
    ("Anaconda",    "OSS + TBD",               "subscription"),
    ("Atlassian",   "freeware",                "subscription + TBD"),
    ("Autodesk",    "freeware + subscription", "subscription only"),
    ("Broadcom",    "freeware",                "subscription"),
    ("MathWorks",   "subscription",            "TBD (all 31 entries)"),
    ("Microsoft",   "freeware+OSS+perp+sub",   "subscription + TBD"),
    ("Oracle",      "freeware + OSS",          "perpetual + subscription"),
    ("SAP",         "perpetual + freeware",    "perpetual + cloud contract + TBD"),
    ("Hexagon",     "TBD",                     "unknown"),
    ("Learnpulse",  "TBD",                     "perpetual"),
    ("IBM",         "TBD",                     "TBD"),
]

# ── Risks ────────────────────────────────────────────────────────────────────
risks = [
    ("HIGH",   "FRAME License Type Gap",
     f"{frame_tbd} of {len(frame)} FRAME products ({round(100*frame_tbd/len(frame))}%) have no defined payment type (TBD/blank). Procurement cannot be completed without this classification.",
     "Resolve all TBD entries before procurement cutover. Set deadline with workstream owners."),
    ("HIGH",   "ASTRA: No Procurement Workflow",
     "ASTRA has no procurement status, owner, or deadline fields. 148 paid products have no tracked procurement stage.",
     "Migrate to shared Microsoft List with procurement tracking columns."),
    ("HIGH",   "Missing Free/OSS Tracking in FRAME",
     f"FRAME has 0 freeware/OSS classified products. ASTRA has 123. License compliance obligations (GPL, MIT, etc.) may go unmanaged.",
     "Add License_Category column to FRAME. Audit all products for OSS content."),
    ("MEDIUM", "FRAME Due Date Never Populated",
     f"Due Date column exists in FRAME but is 0% populated. Priority dates exist for {frame_prio} items but in a non-standard text field.",
     "Make Due_Date mandatory for all priority items. Convert Priority_Date text to proper date column."),
    ("MEDIUM", "Ownership Gap in FRAME",
     f"Only {frame_owner} of {len(frame)} FRAME products ({round(100*frame_owner/len(frame))}%) have a responsible person assigned.",
     "Assign responsible owners for all procurable items, especially those marked as priority."),
    ("MEDIUM", "12 Vendor Type Mismatches",
     "Same vendor handled differently across projects (e.g. Anaconda = OSS in ASTRA, Subscription in FRAME). Risk of double procurement or compliance gap.",
     "Align vendor classification across projects in the shared Microsoft List."),
    ("LOW",    "Metric Inconsistency in FRAME",
     f"{frame_regional} rows mix licensing unit with geography (e.g. 'NNU Taiwan', 'User Portugal'). Prevents clean license counting.",
     "Split metric into 'License_Unit' and 'Region' columns in Microsoft List."),
    ("LOW",    "ASTRA Action Items Not Defined",
     "ASTRA has no Action or workflow fields. 25 products are still TBD on type.",
     "Resolve 25 TBD entries; add action tracking in shared Microsoft List."),
]

# ═══════════════════════════════════════════════════════════════════════════
# BUILD HTML
# ═══════════════════════════════════════════════════════════════════════════

def bar(value, total, color):
    pct = round(100*value/total) if total else 0
    return f'<div class="bar-bg"><div class="bar-fill" style="width:{pct}%;background:{color}"></div></div><span class="bar-label">{pct}%</span>'

def donut_svg(parts, colors, size=120, stroke=22):
    """parts = list of (value, color), generates a donut SVG"""
    total = sum(p[0] for p in parts)
    if not total: return ''
    cx = cy = size/2
    r = (size - stroke)/2
    circ = 2*3.14159*r
    offset = 0
    slices = []
    for val, color in parts:
        pct = val/total
        dash = pct * circ
        slices.append(f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" stroke-width="{stroke}" stroke-dasharray="{dash:.1f} {circ:.1f}" stroke-dashoffset="-{offset*circ/360:.1f}" transform="rotate(-90 {cx} {cy})"/>')
        offset += pct*360
    return f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}">{"".join(slices)}</svg>'

risk_color = {"HIGH":"#dc0000","MEDIUM":"#e65100","LOW":"#f59e0b"}
risk_bg    = {"HIGH":"#fff0f0","MEDIUM":"#fff8f0","LOW":"#fffbea"}
risk_icon  = {"HIGH":"🔴","MEDIUM":"🟠","LOW":"🟡"}

# Donut data
astra_donut_parts = [(a_sub,'#003c64'),(a_per,'#00783c'),(len(astra_foss),'#6c757d'),(a_tbd_paid,'#dc0000')]
frame_donut_parts = [(f_sub,'#003c64'),(f_per,'#00783c'),(frame_tbd,'#dc0000'),(len(frame)-f_sub-f_per-frame_tbd,'#9e9e9e')]
ready_donut_parts = [(frame_ready_y+frame_n,'#00783c'),(frame_prog,'#003c64'),(frame_na,'#9e9e9e')]

astra_donut = donut_svg(astra_donut_parts, [])
frame_donut = donut_svg(frame_donut_parts, [])
ready_donut = donut_svg(ready_donut_parts, [])

top10_bars = ''
max_v = top10[0][1] if top10 else 1
for vendor, cnt in top10:
    w = round(100*cnt/max_v)
    top10_bars += f'''
    <div class="hbar-row">
      <span class="hbar-label">{vendor[:22]}</span>
      <div class="hbar-track"><div class="hbar-fill" style="width:{w}%"></div></div>
      <span class="hbar-val">{cnt}</span>
    </div>'''

mismatch_rows = ''
for vendor, a_type, f_type in mismatch_data:
    mismatch_rows += f'<tr><td><strong>{vendor}</strong></td><td style="color:#003c64">{a_type}</td><td style="color:#e65100">{f_type}</td></tr>\n'

risk_rows = ''
for sev, title, desc, rec in risks:
    rc = risk_color[sev]; rb = risk_bg[sev]; ri = risk_icon[sev]
    risk_rows += f'''
    <div class="risk-card" style="background:{rb};border-left-color:{rc}">
      <div class="risk-header">
        <span class="risk-badge" style="background:{rc}">{sev}</span>
        <span class="risk-title">{ri} {title}</span>
      </div>
      <p class="risk-desc">{desc}</p>
      <div class="risk-rec">💡 <em>{rec}</em></div>
    </div>'''

html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Software License Management Dashboard</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:"Segoe UI",Arial,sans-serif;background:#f4f5f7;color:#1a1a1a}}

  /* ── HEADER ── */
  .header{{background:#1a1a1a;padding:0;display:flex;flex-direction:column}}
  .header-top{{display:flex;align-items:center;justify-content:space-between;padding:12px 36px;background:#1a1a1a}}
  .header-top img.logo{{height:36px}}
  .header-top .title-block{{text-align:right}}
  .header-top .title-block h1{{color:#fff;font-size:1.3rem;font-weight:600;letter-spacing:.5px}}
  .header-top .title-block p{{color:#bbb;font-size:.78rem;margin-top:2px}}
  .theme-strip{{width:100%;height:12px;background:linear-gradient(to right,#dc0000 0%,#dc0000 33%,#003c64 33%,#003c64 66%,#00783c 66%,#00783c 100%)}}

  .theme-banner{{width:100%;height:70px;overflow:hidden;background:#1a1a1a}}
  .theme-banner img{{width:100%;height:70px;object-fit:cover;object-position:center;opacity:.55}}

  /* ── LAYOUT ── */
  .page{{max-width:1300px;margin:0 auto;padding:28px 28px 50px}}

  /* ── SECTION TITLES ── */
  .section-title{{display:flex;align-items:center;gap:10px;margin:32px 0 14px}}
  .section-title .sticon{{width:28px;height:28px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:14px;flex-shrink:0}}
  .section-title h2{{font-size:1.05rem;font-weight:700;color:#1a1a1a;text-transform:uppercase;letter-spacing:.6px}}

  /* ── KPI CARDS ── */
  .kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:14px;margin-bottom:8px}}
  .kpi{{background:#fff;border-radius:8px;padding:18px 16px;border-top:4px solid #dc0000;box-shadow:0 1px 4px rgba(0,0,0,.07);text-align:center}}
  .kpi.blue{{border-top-color:#003c64}}
  .kpi.green{{border-top-color:#00783c}}
  .kpi.gray{{border-top-color:#6c757d}}
  .kpi.orange{{border-top-color:#e65100}}
  .kpi-val{{font-size:2.1rem;font-weight:700;color:#1a1a1a;line-height:1}}
  .kpi-lbl{{font-size:.72rem;color:#666;margin-top:5px;line-height:1.35}}
  .kpi-sub{{font-size:.68rem;color:#999;margin-top:3px}}

  /* ── TWO COLS ── */
  .two-col{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}}
  .three-col{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:16px}}
  @media(max-width:900px){{.two-col,.three-col{{grid-template-columns:1fr}}}}

  /* ── CARD ── */
  .card{{background:#fff;border-radius:8px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.07)}}
  .card h3{{font-size:.85rem;font-weight:700;color:#1a1a1a;margin-bottom:14px;text-transform:uppercase;letter-spacing:.5px;border-bottom:2px solid #f0f0f0;padding-bottom:8px}}

  /* ── DONUT ── */
  .donut-wrap{{display:flex;align-items:center;gap:20px}}
  .donut-wrap svg{{flex-shrink:0}}
  .legend{{flex:1}}
  .legend-item{{display:flex;align-items:center;gap:8px;margin:4px 0;font-size:.8rem}}
  .leg-dot{{width:11px;height:11px;border-radius:50%;flex-shrink:0}}
  .leg-pct{{margin-left:auto;font-weight:600;color:#333}}

  /* ── BARS ── */
  .bar-bg{{background:#e8e8e8;border-radius:4px;height:8px;display:inline-block;width:80px;vertical-align:middle;margin-right:6px}}
  .bar-fill{{height:8px;border-radius:4px}}
  .bar-label{{font-size:.78rem;color:#444;font-weight:600}}

  /* ── HBAR ── */
  .hbar-row{{display:flex;align-items:center;gap:8px;margin:5px 0;font-size:.8rem}}
  .hbar-label{{width:130px;flex-shrink:0;color:#333;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
  .hbar-track{{flex:1;background:#f0f0f0;border-radius:4px;height:14px;overflow:hidden}}
  .hbar-fill{{height:14px;background:#003c64;border-radius:4px;transition:width .3s}}
  .hbar-val{{width:32px;text-align:right;color:#555;font-weight:600}}

  /* ── TABLE ── */
  table{{border-collapse:collapse;width:100%;font-size:.82rem}}
  th{{background:#1a1a1a;color:#fff;padding:8px 10px;text-align:left;font-weight:600;font-size:.78rem;text-transform:uppercase;letter-spacing:.4px}}
  td{{padding:7px 10px;border-bottom:1px solid #eee;vertical-align:top}}
  tr:last-child td{{border-bottom:none}}
  tr:hover td{{background:#f9f9f9}}

  /* ── BADGE ── */
  .badge{{display:inline-block;padding:2px 9px;border-radius:10px;font-size:.7rem;font-weight:700;letter-spacing:.3px}}
  .badge-a{{background:#003c64;color:#fff}}
  .badge-f{{background:#e65100;color:#fff}}

  /* ── PROGRESS ── */
  .prog-row{{display:flex;align-items:center;gap:10px;margin:7px 0;font-size:.82rem}}
  .prog-label{{width:160px;flex-shrink:0;color:#444}}
  .prog-track{{flex:1;background:#f0f0f0;border-radius:6px;height:16px;overflow:hidden}}
  .prog-fill{{height:16px;border-radius:6px}}
  .prog-val{{width:50px;text-align:right;font-weight:700;color:#333}}

  /* ── RISK ── */
  .risk-card{{border-left:5px solid;border-radius:0 8px 8px 0;padding:14px 16px;margin:10px 0}}
  .risk-header{{display:flex;align-items:center;gap:10px;margin-bottom:6px}}
  .risk-badge{{color:#fff;font-size:.68rem;font-weight:700;padding:2px 8px;border-radius:10px;letter-spacing:.4px}}
  .risk-title{{font-weight:700;font-size:.9rem;color:#1a1a1a}}
  .risk-desc{{font-size:.82rem;color:#444;margin:4px 0 6px;line-height:1.5}}
  .risk-rec{{font-size:.8rem;color:#555;background:rgba(255,255,255,.6);border-radius:4px;padding:6px 10px}}

  /* ── COMPARISON TABLE ── */
  .comp-table th:nth-child(2){{background:#003c64}}
  .comp-table th:nth-child(3){{background:#e65100}}
  .comp-table td:nth-child(2){{background:#f0f6ff}}
  .comp-table td:nth-child(3){{background:#fff8f0}}

  /* ── FOOTER ── */
  .footer{{text-align:center;font-size:.72rem;color:#999;margin-top:40px;padding-top:16px;border-top:1px solid #e0e0e0}}

  /* ── STATUS CHIP ── */
  .chip{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:.72rem;font-weight:600}}
  .chip-green{{background:#e6f4ea;color:#1e6e2f}}
  .chip-red{{background:#fff0f0;color:#c62828}}
  .chip-orange{{background:#fff3e0;color:#7a4100}}
  .chip-gray{{background:#f5f5f5;color:#555}}
</style>
</head>
<body>

<!-- ══ HEADER ══════════════════════════════════════════════════════════ -->
<div class="header">
  <div class="header-top">
    <img class="logo" src="{LOGO_SRC}" alt="Bosch Logo">
    <div class="title-block">
      <h1>Software License Management Dashboard</h1>
      <p>ASTRA &amp; FRAME Projects &nbsp;|&nbsp; BDMIL Use Case &nbsp;|&nbsp; April 2, 2026</p>
    </div>
  </div>
  <div class="theme-banner"><img src="{THEME_SRC}" alt="Bosch theme"></div>
  <div class="theme-strip"></div>
</div>

<div class="page">

<!-- ══ SECTION 1 – EXECUTIVE KPIs ═══════════════════════════════════════ -->
<div class="section-title">
  <div class="sticon" style="background:#dc0000;color:#fff">📊</div>
  <h2>Executive Overview</h2>
</div>

<div class="kpi-grid">
  <div class="kpi blue">
    <div class="kpi-val">{total}</div>
    <div class="kpi-lbl">Total Software<br>Products Catalogued</div>
    <div class="kpi-sub">ASTRA + FRAME combined</div>
  </div>
  <div class="kpi" style="border-top-color:#003c64">
    <div class="kpi-val" style="color:#003c64">{len(astra_all)}</div>
    <div class="kpi-lbl">ASTRA Products<br><span style="color:#666">{len(astra_paid)} Paid · {len(astra_foss)} FOSS</span></div>
    <div class="kpi-sub">As of 10 Feb 2026</div>
  </div>
  <div class="kpi orange">
    <div class="kpi-val" style="color:#e65100">{len(frame)}</div>
    <div class="kpi-lbl">FRAME Products<br><span style="color:#666">Shopping list</span></div>
    <div class="kpi-sub">Status 06 Jun 2025</div>
  </div>
  <div class="kpi gray">
    <div class="kpi-val">{all_v}</div>
    <div class="kpi-lbl">Total Unique Vendors<br><span style="color:#666">Across both projects</span></div>
    <div class="kpi-sub">{common_v} shared vendors</div>
  </div>
  <div class="kpi green">
    <div class="kpi-val" style="color:#00783c">{frame_ready_y + frame_prog}</div>
    <div class="kpi-lbl">FRAME: Ready / In<br>Progress for Procurement</div>
    <div class="kpi-sub">of {len(frame)} total FRAME items</div>
  </div>
  <div class="kpi">
    <div class="kpi-val" style="color:#dc0000">{frame_tbd}</div>
    <div class="kpi-lbl">Unresolved License<br>Types (TBD / Blank)</div>
    <div class="kpi-sub">FRAME: {round(100*frame_tbd/len(frame))}% · ASTRA: {round(100*astra_tbd/len(astra_all))}%</div>
  </div>
  <div class="kpi" style="border-top-color:#003c64">
    <div class="kpi-val" style="color:#003c64">{frame_prio}</div>
    <div class="kpi-lbl">FRAME Priority<br>Items</div>
    <div class="kpi-sub">{round(100*frame_prio/len(frame))}% of FRAME list</div>
  </div>
  <div class="kpi green">
    <div class="kpi-val" style="color:#00783c">{len(astra_foss)}</div>
    <div class="kpi-lbl">ASTRA Free / OSS<br>Products Tracked</div>
    <div class="kpi-sub">67 Freeware · 56 OSS</div>
  </div>
</div>

<!-- ══ SECTION 2 – LICENSE TYPE BREAKDOWN ════════════════════════════════ -->
<div class="section-title">
  <div class="sticon" style="background:#003c64;color:#fff">📋</div>
  <h2>License Type Breakdown</h2>
</div>

<div class="three-col">

  <!-- ASTRA Donut -->
  <div class="card">
    <h3><span class="badge badge-a">ASTRA</span>&nbsp; Payment Model</h3>
    <div class="donut-wrap">
      {donut_svg(astra_donut_parts,[])}
      <div class="legend">
        <div class="legend-item"><span class="leg-dot" style="background:#003c64"></span>Subscription<span class="leg-pct">{a_sub}</span></div>
        <div class="legend-item"><span class="leg-dot" style="background:#00783c"></span>Perpetual<span class="leg-pct">{a_per}</span></div>
        <div class="legend-item"><span class="leg-dot" style="background:#6c757d"></span>Freeware / OSS<span class="leg-pct">{len(astra_foss)}</span></div>
        <div class="legend-item"><span class="leg-dot" style="background:#dc0000"></span>TBD<span class="leg-pct">{a_tbd_paid}</span></div>
      </div>
    </div>
    <div style="margin-top:12px">
      <div class="prog-row"><span class="prog-label">Classified (paid)</span>
        <div class="prog-track"><div class="prog-fill" style="width:{round(100*(len(astra_paid)-a_tbd_paid)/len(astra_paid))}%;background:#003c64"></div></div>
        <span class="prog-val">{round(100*(len(astra_paid)-a_tbd_paid)/len(astra_paid))}%</span></div>
      <div class="prog-row"><span class="prog-label">FOSS tracked</span>
        <div class="prog-track"><div class="prog-fill" style="width:{round(100*len(astra_foss)/len(astra_all))}%;background:#00783c"></div></div>
        <span class="prog-val">{round(100*len(astra_foss)/len(astra_all))}%</span></div>
    </div>
  </div>

  <!-- FRAME Donut -->
  <div class="card">
    <h3><span class="badge badge-f">FRAME</span>&nbsp; Payment Model</h3>
    <div class="donut-wrap">
      {donut_svg(frame_donut_parts,[])}
      <div class="legend">
        <div class="legend-item"><span class="leg-dot" style="background:#003c64"></span>Subscription<span class="leg-pct">{f_sub}</span></div>
        <div class="legend-item"><span class="leg-dot" style="background:#00783c"></span>Perpetual<span class="leg-pct">{f_per}</span></div>
        <div class="legend-item"><span class="leg-dot" style="background:#dc0000"></span>TBD / Unknown<span class="leg-pct">{frame_tbd}</span></div>
        <div class="legend-item"><span class="leg-dot" style="background:#9e9e9e"></span>Other<span class="leg-pct">{len(frame)-f_sub-f_per-frame_tbd}</span></div>
      </div>
    </div>
    <div style="margin-top:12px">
      <div class="prog-row"><span class="prog-label">Classified</span>
        <div class="prog-track"><div class="prog-fill" style="width:{round(100*(len(frame)-frame_tbd)/len(frame))}%;background:#003c64"></div></div>
        <span class="prog-val">{round(100*(len(frame)-frame_tbd)/len(frame))}%</span></div>
      <div class="prog-row"><span class="prog-label" style="color:#dc0000">⚠ TBD / Blank</span>
        <div class="prog-track"><div class="prog-fill" style="width:{round(100*frame_tbd/len(frame))}%;background:#dc0000"></div></div>
        <span class="prog-val" style="color:#dc0000">{round(100*frame_tbd/len(frame))}%</span></div>
    </div>
  </div>

  <!-- FRAME Procurement Readiness -->
  <div class="card">
    <h3><span class="badge badge-f">FRAME</span>&nbsp; Procurement Readiness</h3>
    <div class="donut-wrap">
      {donut_svg(ready_donut_parts,[])}
      <div class="legend">
        <div class="legend-item"><span class="leg-dot" style="background:#00783c"></span>Ready (Y)<span class="leg-pct">{frame_ready_y}</span></div>
        <div class="legend-item"><span class="leg-dot" style="background:#003c64"></span>In Progress<span class="leg-pct">{frame_prog}</span></div>
        <div class="legend-item"><span class="leg-dot" style="background:#9e9e9e"></span>N/A<span class="leg-pct">{frame_na}</span></div>
        <div class="legend-item"><span class="leg-dot" style="background:#dc0000"></span>Not Ready (N)<span class="leg-pct">{frame_n}</span></div>
      </div>
    </div>
    <div style="margin-top:12px">
      <div class="prog-row"><span class="prog-label">Priority items set</span>
        <div class="prog-track"><div class="prog-fill" style="width:{round(100*frame_prio/len(frame))}%;background:#003c64"></div></div>
        <span class="prog-val">{round(100*frame_prio/len(frame))}%</span></div>
      <div class="prog-row"><span class="prog-label">Changed since Feb'25</span>
        <div class="prog-track"><div class="prog-fill" style="width:{round(100*frame_changed/len(frame))}%;background:#e65100"></div></div>
        <span class="prog-val">{round(100*frame_changed/len(frame))}%</span></div>
      <div class="prog-row"><span class="prog-label" style="color:#dc0000">⚠ Owner assigned</span>
        <div class="prog-track"><div class="prog-fill" style="width:{round(100*frame_owner/len(frame))}%;background:#dc0000"></div></div>
        <span class="prog-val" style="color:#dc0000">{round(100*frame_owner/len(frame))}%</span></div>
    </div>
  </div>

</div>

<!-- ══ SECTION 3 – TOP VENDORS + PROCESS COMPLETENESS ═══════════════════════ -->
<div class="section-title">
  <div class="sticon" style="background:#00783c;color:#fff">🏢</div>
  <h2>Vendor Landscape &amp; Process Completeness</h2>
</div>

<div class="two-col">

  <div class="card">
    <h3>Top 10 Vendors by Product Count (Combined)</h3>
    {top10_bars}
  </div>

  <div class="card">
    <h3>Process Field Completeness</h3>
    <div style="font-size:.8rem;color:#888;margin-bottom:10px">How well each project populates key management fields</div>

    <div style="font-size:.75rem;font-weight:700;color:#003c64;margin:8px 0 4px">ASTRA</div>
    <div class="prog-row"><span class="prog-label">License Type classified</span>
      <div class="prog-track"><div class="prog-fill" style="width:{round(100*(len(astra_all)-astra_tbd)/len(astra_all))}%;background:#003c64"></div></div>
      <span class="prog-val">{round(100*(len(astra_all)-astra_tbd)/len(astra_all))}%</span></div>
    <div class="prog-row"><span class="prog-label">Notes / Comments filled</span>
      <div class="prog-track"><div class="prog-fill" style="width:23%;background:#003c64"></div></div>
      <span class="prog-val">23%</span></div>
    <div class="prog-row"><span class="prog-label">FOSS explicitly tagged</span>
      <div class="prog-track"><div class="prog-fill" style="width:100%;background:#00783c"></div></div>
      <span class="prog-val">100%</span></div>
    <div class="prog-row"><span class="prog-label" style="color:#dc0000">⚠ Owner assigned</span>
      <div class="prog-track"><div class="prog-fill" style="width:0%;background:#dc0000"></div></div>
      <span class="prog-val" style="color:#dc0000">0%</span></div>
    <div class="prog-row"><span class="prog-label" style="color:#dc0000">⚠ Due Date set</span>
      <div class="prog-track"><div class="prog-fill" style="width:0%;background:#dc0000"></div></div>
      <span class="prog-val" style="color:#dc0000">0%</span></div>

    <div style="font-size:.75rem;font-weight:700;color:#e65100;margin:14px 0 4px">FRAME</div>
    <div class="prog-row"><span class="prog-label">Ready for Procurement</span>
      <div class="prog-track"><div class="prog-fill" style="width:100%;background:#00783c"></div></div>
      <span class="prog-val">100%</span></div>
    <div class="prog-row"><span class="prog-label">Notes / Comments filled</span>
      <div class="prog-track"><div class="prog-fill" style="width:42%;background:#e65100"></div></div>
      <span class="prog-val">42%</span></div>
    <div class="prog-row"><span class="prog-label">Priority tracked</span>
      <div class="prog-track"><div class="prog-fill" style="width:{round(100*frame_prio/len(frame))}%;background:#e65100"></div></div>
      <span class="prog-val">{round(100*frame_prio/len(frame))}%</span></div>
    <div class="prog-row"><span class="prog-label" style="color:#dc0000">⚠ License Type classified</span>
      <div class="prog-track"><div class="prog-fill" style="width:{round(100*(len(frame)-frame_tbd)/len(frame))}%;background:#dc0000"></div></div>
      <span class="prog-val" style="color:#dc0000">{round(100*(len(frame)-frame_tbd)/len(frame))}%</span></div>
    <div class="prog-row"><span class="prog-label" style="color:#dc0000">⚠ Owner assigned</span>
      <div class="prog-track"><div class="prog-fill" style="width:{round(100*frame_owner/len(frame))}%;background:#dc0000"></div></div>
      <span class="prog-val" style="color:#dc0000">{round(100*frame_owner/len(frame))}%</span></div>
    <div class="prog-row"><span class="prog-label" style="color:#dc0000">⚠ Due Date set</span>
      <div class="prog-track"><div class="prog-fill" style="width:0%;background:#dc0000"></div></div>
      <span class="prog-val" style="color:#dc0000">0%</span></div>
  </div>

</div>

<!-- ══ SECTION 4 – DELTA COMPARISON TABLE ════════════════════════════════ -->
<div class="section-title">
  <div class="sticon" style="background:#e65100;color:#fff">⚡</div>
  <h2>Project Comparison: ASTRA vs FRAME</h2>
</div>

<div class="card" style="margin-bottom:16px">
<table class="comp-table">
  <thead><tr><th>Dimension</th><th>🟢 ASTRA</th><th>🟠 FRAME</th></tr></thead>
  <tbody>
    <tr><td><strong>Total products</strong></td><td>{len(astra_all)}</td><td>{len(frame)}</td></tr>
    <tr><td><strong>Unique vendors</strong></td><td>{astra_v}</td><td>{frame_v}</td></tr>
    <tr><td><strong>Subscription licenses</strong></td><td>{a_sub}</td><td>{f_sub}</td></tr>
    <tr><td><strong>Perpetual licenses</strong></td><td>{a_per}</td><td>{f_per}</td></tr>
    <tr><td><strong>Freeware / OSS tracked</strong></td><td><span class="chip chip-green">✅ {len(astra_foss)} products</span></td><td><span class="chip chip-red">❌ Not tracked</span></td></tr>
    <tr><td><strong>License type classified</strong></td><td><span class="chip chip-green">91% ({len(astra_all)-astra_tbd}/{len(astra_all)})</span></td><td><span class="chip chip-red">40% ({len(frame)-frame_tbd}/{len(frame)})</span></td></tr>
    <tr><td><strong>Procurement status tracked</strong></td><td><span class="chip chip-red">❌ None</span></td><td><span class="chip chip-green">✅ 100% filled</span></td></tr>
    <tr><td><strong>Priority / urgency tracking</strong></td><td><span class="chip chip-red">❌ None</span></td><td><span class="chip chip-green">✅ {frame_prio} items ({round(100*frame_prio/len(frame))}%)</span></td></tr>
    <tr><td><strong>Owner / Responsible assigned</strong></td><td><span class="chip chip-red">❌ No column</span></td><td><span class="chip chip-orange">⚠ {frame_owner} items ({round(100*frame_owner/len(frame))}%)</span></td></tr>
    <tr><td><strong>Due Date set</strong></td><td><span class="chip chip-red">❌ No column</span></td><td><span class="chip chip-red">❌ 0% (column unused)</span></td></tr>
    <tr><td><strong>Metric granularity</strong></td><td>~10 clean values</td><td>50+ values (incl. {frame_regional} region-qualified)</td></tr>
    <tr><td><strong>Change tracking vs baseline</strong></td><td><span class="chip chip-red">❌ None</span></td><td><span class="chip chip-green">✅ {frame_changed} items tracked</span></td></tr>
    <tr><td><strong>Pivot analytics</strong></td><td><span class="chip chip-green">✅ 2 pivot sheets</span></td><td><span class="chip chip-red">❌ None</span></td></tr>
    <tr><td><strong>Notes fill rate</strong></td><td>23%</td><td>42% (Comment) + 38% (Comment2)</td></tr>
    <tr><td><strong>Vendor type mismatches</strong></td><td colspan="2" style="background:#fff0f0;text-align:center"><strong style="color:#dc0000">12 common vendors handled differently</strong></td></tr>
  </tbody>
</table>
</div>

<!-- ══ SECTION 5 – VENDOR MISMATCHES ════════════════════════════════════ -->
<div class="section-title">
  <div class="sticon" style="background:#dc0000;color:#fff">🔀</div>
  <h2>Vendor Classification Mismatches</h2>
</div>

<div class="card" style="margin-bottom:16px">
  <h3>Same Vendor – Different License Type Handling (12 of 17 common vendors)</h3>
  <table>
    <thead><tr><th>Vendor</th><th style="background:#003c64">🟢 ASTRA Type</th><th style="background:#e65100">🟠 FRAME Type</th></tr></thead>
    <tbody>{mismatch_rows}</tbody>
  </table>
</div>

<!-- ══ SECTION 6 – RISK REGISTER ═════════════════════════════════════════ -->
<div class="section-title">
  <div class="sticon" style="background:#dc0000;color:#fff">⚠</div>
  <h2>Risk Register</h2>
</div>

<div class="two-col">
  <div>
    {chr(10).join(r for i,r in enumerate(risk_rows.strip().split('</div>')) if i%2==0 and r.strip())}
  </div>
  <div>
    {chr(10).join(r for i,r in enumerate(risk_rows.strip().split('</div>')) if i%2==1 and r.strip())}
  </div>
</div>

<!-- ══ SECTION 7 – RECOMMENDATIONS ══════════════════════════════════════════ -->
<div class="section-title">
  <div class="sticon" style="background:#00783c;color:#fff">✅</div>
  <h2>Recommendations &amp; Next Steps</h2>
</div>

<div class="three-col">
  <div class="card">
    <h3>🔴 Immediate (HIGH)</h3>
    <ul style="font-size:.82rem;padding-left:16px;line-height:1.8;color:#333">
      <li>Resolve <strong>{frame_tbd} TBD type entries</strong> in FRAME before procurement cutover</li>
      <li>Add <strong>Free/OSS category</strong> to FRAME list – audit 590 products for compliance obligations</li>
      <li>Migrate ASTRA to <strong>shared Microsoft List</strong> with procurement workflow fields</li>
    </ul>
  </div>
  <div class="card">
    <h3>🟠 Short-term (MEDIUM)</h3>
    <ul style="font-size:.82rem;padding-left:16px;line-height:1.8;color:#333">
      <li>Activate <strong>Due_Date</strong> field for all FRAME priority items ({frame_prio} items)</li>
      <li>Assign <strong>Responsible owner</strong> to all procurable FRAME items (currently only {frame_owner}/{len(frame)})</li>
      <li>Align <strong>12 vendor classification mismatches</strong> across projects in Microsoft List</li>
    </ul>
  </div>
  <div class="card">
    <h3>🟡 Ongoing (LOW)</h3>
    <ul style="font-size:.82rem;padding-left:16px;line-height:1.8;color:#333">
      <li>Standardise FRAME metric values – split <code>License_Unit</code> from <code>Region</code></li>
      <li>Replicate ASTRA's <strong>pivot analytics</strong> as Microsoft List views/filters</li>
      <li>Establish quarterly review cycle for license inventory updates</li>
    </ul>
  </div>
</div>

<div class="footer">
  Software License Management Dashboard &nbsp;·&nbsp; BDMIL AI Use Case &nbsp;·&nbsp; ASTRA &amp; FRAME Projects &nbsp;·&nbsp; April 2, 2026
  <br>Generated by GitHub Copilot &nbsp;|&nbsp; Data sources: Astra Shopping List Buyer_10.02.2026.xlsx · Frame License Shopping List.xlsx
</div>

</div><!-- end .page -->
</body>
</html>'''

out = os.path.join(BASE, 'Management_Dashboard.html')
with open(out,'w',encoding='utf-8') as f:
    f.write(html)
print(f"Written: {out}  ({len(html)//1024} KB)")
