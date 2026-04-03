import openpyxl
import csv
import os

path = r'c:\Users\kho1sgp\OneDrive - Bosch Group\My Work Documents\AI topics\BDMIL use case\Software_license_project'

# ============================================================
# FILE 1: Astra Shopping List
# ============================================================
wb1 = openpyxl.load_workbook(os.path.join(path, 'Astra Shopping List Buyer_10.02.2026.xlsx'), data_only=True)

print("=== FILE 1 Sheets:", wb1.sheetnames)

sh = wb1['Shopping List Astra 10.02.2026']
print(f"\n--- Sheet: Shopping List Astra 10.02.2026 | rows={sh.max_row} cols={sh.max_column} ---")
for r in range(1, min(sh.max_row + 1, 20)):
    row = [str(sh.cell(r, c).value or '').strip() for c in range(1, sh.max_column + 1)]
    if any(v for v in row):
        print(f"  R{r}: {' | '.join(row)}")

sh2 = wb1['License inventory for FOSS']
print(f"\n--- Sheet: License inventory for FOSS | rows={sh2.max_row} cols={sh2.max_column} ---")
for r in range(1, min(sh2.max_row + 1, 20)):
    row = [str(sh2.cell(r, c).value or '').strip() for c in range(1, sh2.max_column + 1)]
    if any(v for v in row):
        print(f"  R{r}: {' | '.join(row)}")

# ============================================================
# FILE 2: Frame License Shopping List
# ============================================================
wb2 = openpyxl.load_workbook(os.path.join(path, 'Frame License Shopping List.xlsx'), data_only=True)
print("\n\n=== FILE 2 Sheets:", wb2.sheetnames)

sh3 = wb2['Shopping_List_Vorbereitung']
print(f"\n--- Sheet: Shopping_List_Vorbereitung | rows={sh3.max_row} cols={sh3.max_column} ---")
for r in range(1, min(sh3.max_row + 1, 20)):
    row = [str(sh3.cell(r, c).value or '').strip() for c in range(1, sh3.max_column + 1)]
    if any(v for v in row):
        print(f"  R{r}: {' | '.join(row)}")
