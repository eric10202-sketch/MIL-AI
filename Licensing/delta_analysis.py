"""
Delta analysis: FRAME vs ASTRA – behavioural differences in product handling.
Focus: process fields, not quantity.
"""
import openpyxl, csv, os, collections

BASE = r'c:\Users\kho1sgp\OneDrive - Bosch Group\My Work Documents\AI topics\BDMIL use case\Software_license_project'

def clean(v):
    s = str(v or '').strip()
    return s if s not in ('None', '') else ''

# ── ASTRA Paid ──────────────────────────────────────────────────────────────
wb1 = openpyxl.load_workbook(os.path.join(BASE, 'Astra Shopping List Buyer_10.02.2026.xlsx'), data_only=True)

astra_paid = []
sh = wb1['Shopping List Astra 10.02.2026']
for r in range(10, sh.max_row + 1):
    vendor   = clean(sh.cell(r,2).value)
    product  = clean(sh.cell(r,3).value)
    qty      = clean(sh.cell(r,4).value)
    metric   = clean(sh.cell(r,5).value)
    pay_type = clean(sh.cell(r,6).value)
    notes    = clean(sh.cell(r,7).value)
    if not product and not vendor: continue
    astra_paid.append({'vendor':vendor,'product':product,'qty':qty,
                       'metric':metric,'type':pay_type,'notes':notes,'cat':'Paid'})

astra_foss = []
sh2 = wb1['License inventory for FOSS']
for r in range(8, sh2.max_row + 1):
    vendor  = clean(sh2.cell(r,3).value)
    product = clean(sh2.cell(r,4).value)
    qty     = clean(sh2.cell(r,5).value)
    foss    = clean(sh2.cell(r,6).value)
    notes   = clean(sh2.cell(r,7).value)
    if not product and not vendor: continue
    astra_foss.append({'vendor':vendor,'product':product,'qty':qty,
                       'metric':'','type':foss,'notes':notes,'cat':foss})

# ── FRAME ────────────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(os.path.join(BASE, 'Frame License Shopping List.xlsx'), data_only=True)
sh3 = wb2['Shopping_List_Vorbereitung']

frame = []
for r in range(6, sh3.max_row + 1):
    item_no     = clean(sh3.cell(r,1).value)
    vendor      = clean(sh3.cell(r,2).value)
    product     = clean(sh3.cell(r,3).value)
    prio_date   = clean(sh3.cell(r,4).value)
    prio_reason = clean(sh3.cell(r,5).value)
    chg_feb     = clean(sh3.cell(r,6).value)
    chg_local   = clean(sh3.cell(r,7).value)
    lic_type    = clean(sh3.cell(r,8).value)
    qty         = clean(sh3.cell(r,9).value)
    metric      = clean(sh3.cell(r,10).value)
    ready       = clean(sh3.cell(r,11).value)
    lic_server  = clean(sh3.cell(r,12).value)
    comment     = clean(sh3.cell(r,13).value)
    comment2    = clean(sh3.cell(r,14).value)
    action      = clean(sh3.cell(r,15).value)
    responsible = clean(sh3.cell(r,16).value)
    due_date    = clean(sh3.cell(r,17).value)
    if not product and not vendor: continue
    frame.append({'vendor':vendor,'product':product,'qty':qty,'metric':metric,
                  'type':lic_type,'prio_date':prio_date,'prio_reason':prio_reason,
                  'chg_feb':chg_feb,'chg_local':chg_local,'ready':ready,
                  'lic_server':lic_server,'comment':comment,'comment2':comment2,
                  'action':action,'responsible':responsible,'due_date':due_date})

astra_all = astra_paid + astra_foss
print(f"ASTRA total: {len(astra_all)}  (paid={len(astra_paid)}, foss={len(astra_foss)})")
print(f"FRAME total: {len(frame)}")

# ════════════════════════════════════════════════════════════════════
# ANALYSIS 1 – Columns / Tracking Fields used by each project
# ════════════════════════════════════════════════════════════════════
print("\n\n══════════════════════════════════════════════════════")
print("ANALYSIS 1 – Tracking Fields Available per Project")
print("══════════════════════════════════════════════════════")
print("""
ASTRA fields : vendor | product | quantity | metric | type_of_payment | notes
FRAME fields : vendor | product | quantity | metric | type | priority_date |
               priority_reason | changes_since_feb_2025 | changes_from_local_list |
               ready_for_procurement | license_server_planned | comment |
               comment_2 | action | responsible | due_date
""")
astra_extra = set()  # fields ASTRA has but FRAME doesn't (none in this case)
frame_extra = ['priority_date','priority_reason','changes_since_feb_2025',
               'changes_from_local_list','ready_for_procurement',
               'license_server_planned','action','responsible','due_date','comment_2']
print("Fields ONLY in FRAME (not tracked in ASTRA):")
for f in frame_extra:
    print(f"  + {f}")

# ════════════════════════════════════════════════════════════════════
# ANALYSIS 2 – License Type / Payment Model breakdown
# ════════════════════════════════════════════════════════════════════
print("\n\n══════════════════════════════════════════════════════")
print("ANALYSIS 2 – License Type / Payment Model")
print("══════════════════════════════════════════════════════")

astra_types = collections.Counter(r['type'].lower() if r['type'] else '(blank)' for r in astra_all)
frame_types = collections.Counter(r['type'].lower() if r['type'] else '(blank)' for r in frame)
print("\nASTRA type distribution:")
for k,v in sorted(astra_types.items(), key=lambda x:-x[1]):
    print(f"  {k:<40} {v:>5}")
print("\nFRAME type distribution:")
for k,v in sorted(frame_types.items(), key=lambda x:-x[1]):
    print(f"  {k:<40} {v:>5}")

# ════════════════════════════════════════════════════════════════════
# ANALYSIS 3 – Metric breakdown
# ════════════════════════════════════════════════════════════════════
print("\n\n══════════════════════════════════════════════════════")
print("ANALYSIS 3 – Metric (Licensing Unit) Usage")
print("══════════════════════════════════════════════════════")
astra_metrics = collections.Counter(r['metric'].lower() if r['metric'] else '(blank)' for r in astra_all)
frame_metrics = collections.Counter(r['metric'].lower() if r['metric'] else '(blank)' for r in frame)
print("\nASTRA metrics:")
for k,v in sorted(astra_metrics.items(), key=lambda x:-x[1]):
    print(f"  {k:<30} {v:>5}")
print("\nFRAME metrics:")
for k,v in sorted(frame_metrics.items(), key=lambda x:-x[1]):
    print(f"  {k:<30} {v:>5}")

# ════════════════════════════════════════════════════════════════════
# ANALYSIS 4 – Notes / Comments population rate
# ════════════════════════════════════════════════════════════════════
print("\n\n══════════════════════════════════════════════════════")
print("ANALYSIS 4 – Comment / Notes Population Rate")
print("══════════════════════════════════════════════════════")
a_notes = sum(1 for r in astra_all if r['notes'])
f_comment = sum(1 for r in frame if r['comment'])
f_comment2 = sum(1 for r in frame if r['comment2'])
f_action = sum(1 for r in frame if r['action'])
f_responsible = sum(1 for r in frame if r['responsible'])
f_due = sum(1 for r in frame if r['due_date'])
f_ready = sum(1 for r in frame if r['ready'])
f_server = sum(1 for r in frame if r['lic_server'])
f_chg = sum(1 for r in frame if r['chg_feb'])
print(f"\nASTRA:")
print(f"  Notes populated          : {a_notes}/{len(astra_all)}  ({100*a_notes//len(astra_all)}%)")
print(f"\nFRAME:")
print(f"  Comment populated        : {f_comment}/{len(frame)}  ({100*f_comment//len(frame)}%)")
print(f"  Comment2 populated       : {f_comment2}/{len(frame)}  ({100*f_comment2//len(frame)}%)")
print(f"  Action populated         : {f_action}/{len(frame)}  ({100*f_action//len(frame)}%)")
print(f"  Responsible populated    : {f_responsible}/{len(frame)}  ({100*f_responsible//len(frame)}%)")
print(f"  Due Date populated       : {f_due}/{len(frame)}  ({100*f_due//len(frame)}%)")
print(f"  Ready for Procurement    : {f_ready}/{len(frame)}  ({100*f_ready//len(frame)}%)")
print(f"  License Server planned?  : {f_server}/{len(frame)}  ({100*f_server//len(frame)}%)")
print(f"  Changes since Feb 2025   : {f_chg}/{len(frame)}  ({100*f_chg//len(frame)}%)")

# ════════════════════════════════════════════════════════════════════
# ANALYSIS 5 – Common vendors - how same vendor is handled differently
# ════════════════════════════════════════════════════════════════════
print("\n\n══════════════════════════════════════════════════════")
print("ANALYSIS 5 – Common Vendors: Different Handling")
print("══════════════════════════════════════════════════════")
astra_vendors = {r['vendor'].lower():r['vendor'] for r in astra_all}
frame_vendors = {r['vendor'].lower():r['vendor'] for r in frame}
common_lc = set(astra_vendors.keys()) & set(frame_vendors.keys())
print(f"\nCommon vendors (both projects): {len(common_lc)}")
for v in sorted(common_lc):
    a_rows = [r for r in astra_all if r['vendor'].lower()==v]
    f_rows = [r for r in frame if r['vendor'].lower()==v]
    a_types = set(r['type'].lower() for r in a_rows if r['type'])
    f_types = set(r['type'].lower() for r in f_rows if r['type'])
    a_prods = set(r['product'] for r in a_rows)
    f_prods = set(r['product'] for r in f_rows)
    shared_prods = {ap for ap in a_prods for fp in f_prods if ap.lower() in fp.lower() or fp.lower() in ap.lower()}
    diff_type = a_types != f_types and a_types and f_types
    print(f"\n  Vendor: {astra_vendors[v]}")
    print(f"    ASTRA products ({len(a_prods)}): {', '.join(sorted(a_prods)[:5])}{'...' if len(a_prods)>5 else ''}")
    print(f"    FRAME products ({len(f_prods)}): {', '.join(sorted(f_prods)[:5])}{'...' if len(f_prods)>5 else ''}")
    print(f"    ASTRA types: {a_types or '(none)'}  |  FRAME types: {f_types or '(none)'}")
    if diff_type:
        print(f"    *** TYPE MISMATCH ***")

# ════════════════════════════════════════════════════════════════════
# ANALYSIS 6 – Priority / Urgency tracking (FRAME only)
# ════════════════════════════════════════════════════════════════════
print("\n\n══════════════════════════════════════════════════════")
print("ANALYSIS 6 – Priority / Urgency Tracking (FRAME only)")
print("══════════════════════════════════════════════════════")
prio = [r for r in frame if r['prio_date'] or r['prio_reason']]
print(f"\nFRAME rows with Priority Date or Reason: {len(prio)} / {len(frame)}")
for r in prio[:15]:
    print(f"  {r['vendor'][:20]:<20} | {r['product'][:35]:<35} | date={r['prio_date'][:20]} | {r['prio_reason'][:50]}")

# ════════════════════════════════════════════════════════════════════
# ANALYSIS 7 – Distinct behavior: ASTRA has FOSS categorisation, FRAME doesn't
# ════════════════════════════════════════════════════════════════════
print("\n\n══════════════════════════════════════════════════════")
print("ANALYSIS 7 – Open Source / Freeware Handling")
print("══════════════════════════════════════════════════════")
print(f"""
ASTRA: Explicitly separates paid vs freeware/OSS into TWO sheets:
  - Sheet 'Shopping List Astra 10.02.2026'  → Paid licenses ({len(astra_paid)} products)
  - Sheet 'License inventory for FOSS'      → Freeware & OSS ({len(astra_foss)} products)
  ASTRA FOSS sub-categories:""")
foss_cats = collections.Counter(r['cat'] for r in astra_foss)
for k,v in sorted(foss_cats.items(), key=lambda x:-x[1]):
    print(f"    {k:<40} {v:>4}")

print(f"""
FRAME: All licenses in ONE sheet. No explicit OSS/Freeware category column.
  Type column values related to free/OSS in FRAME:""")
frame_free = [r for r in frame if 'free' in (r['type'] or '').lower() or 'oss' in (r['type'] or '').lower() or 'open' in (r['type'] or '').lower()]
print(f"    Found {len(frame_free)} rows mentioning free/OSS in 'Type' column")

# ════════════════════════════════════════════════════════════════════
# ANALYSIS 8 – Procurement Status tracking
# ════════════════════════════════════════════════════════════════════
print("\n\n══════════════════════════════════════════════════════")
print("ANALYSIS 8 – Procurement Status Tracking")
print("══════════════════════════════════════════════════════")
print("""
ASTRA: NO dedicated procurement status column.
  → Only free-text 'Notes' field. No stage/workflow tracking.
  
FRAME: Explicit 'Ready_for_Procurement' column with status values:""")
frame_ready_vals = collections.Counter(r['ready'] for r in frame if r['ready'])
for k,v in sorted(frame_ready_vals.items(), key=lambda x:-x[1])[:20]:
    print(f"    '{k[:60]}'  →  {v} items")

print("\n\nFRAME also tracks 'Changes since Feb 2025' and 'Changes from locally managed list':")
chg_vals = collections.Counter(r['chg_feb'] for r in frame if r['chg_feb'])
for k,v in list(sorted(chg_vals.items(), key=lambda x:-x[1]))[:10]:
    print(f"    '{k[:60]}'  →  {v} items")

# ════════════════════════════════════════════════════════════════════
# ANALYSIS 9 – Responsible person / ownership
# ════════════════════════════════════════════════════════════════════
print("\n\n══════════════════════════════════════════════════════")
print("ANALYSIS 9 – Ownership / Responsible Person")
print("══════════════════════════════════════════════════════")
print("ASTRA: No 'Responsible' or 'Owner' column at all.")
frame_resp = [r for r in frame if r['responsible']]
resp_vals = collections.Counter(r['responsible'] for r in frame_resp)
print(f"\nFRAME: 'Responsible' field populated for {len(frame_resp)}/{len(frame)} rows")
print("  Top responsible persons/teams:")
for k,v in sorted(resp_vals.items(), key=lambda x:-x[1])[:10]:
    print(f"    {k:<40} {v:>4} items")

print("\n\nDone.")
