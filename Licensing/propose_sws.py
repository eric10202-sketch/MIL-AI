"""
Propose SWS group for items in 'Astra Shopping List Buyer_10.02.2026.xlsx'
which have no SWS assignment, by learning patterns from
'Astra Shoppping List with SWS.xlsx'.

Output: Astra_Shopping_List_Buyer_AI_SWS.xlsx
        (copy of the Buyer file with an extra 'SWS (AI-proposed)' column)
"""
import os, re, collections
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r'c:\Users\kho1sgp\OneDrive - Bosch Group\My Work Documents\AI topics\BDMIL use case\Software_license_project'

# ─── colours ────────────────────────────────────────────────────────────────
COL_SWS  = {'SWS1':'003c64','SWS2':'dc0000','SWS3':'00783c',
            'SWS4':'e65100','SWS1/3':'6c757d','SWS8':'795548'}
COL_AI_FILL  = 'FFF9E6'     # warm yellow background for AI cells
COL_HDR_FILL = '1A1A1A'     # dark header

# ─── 1. Load training data from SWS file ─────────────────────────────────────
def clean(v):
    s = str(v or '').strip()
    return s if s not in ('None', '') else ''

wb_sws = openpyxl.load_workbook(
    os.path.join(BASE, '_sws_copy.xlsx'))
ws_sws = wb_sws['Shoppinglist']

training = []
for r in ws_sws.iter_rows(min_row=2, values_only=True):
    vendor  = clean(r[1])
    product = clean(r[2])
    sws     = clean(r[6])
    bdbu    = clean(r[8])
    if (vendor or product) and sws:
        training.append({'vendor': vendor, 'product': product,
                         'sws': sws, 'bdbu': bdbu})

print(f'Training items: {len(training)}')

# ─── 2. Build lookup tables ───────────────────────────────────────────────────
# vendor (lower) → Counter of SWS
vendor_sws = collections.defaultdict(collections.Counter)
for t in training:
    if t['vendor']:
        vendor_sws[t['vendor'].lower()][t['sws']] += 1

# product first-word (lower) → Counter
prod_word_sws = collections.defaultdict(collections.Counter)
for t in training:
    if t['product']:
        words = re.findall(r'\b[a-zA-Z][a-zA-Z0-9+#.]{2,}\b', t['product'])
        for w in words[:4]:                 # use first 4 significant words
            prod_word_sws[w.lower()][t['sws']] += 1

# Exact vendor name → dominant SWS (if coverage ≥ 55%)
def dominant(counter, threshold=0.55):
    total = sum(counter.values())
    if not total:
        return None, 0.0
    top, top_n = counter.most_common(1)[0]
    conf = top_n / total
    return (top, conf) if conf >= threshold else (None, conf)

# Full vendor->SWS reference (show all)
print('\nVendor->SWS distribution (top 20 vendors in training):')
for vendor, ctr in sorted(vendor_sws.items(),
                          key=lambda x: -sum(x[1].values()))[:20]:
    dom, conf = dominant(ctr)
    print(f'  {vendor!r:40s}: {dict(ctr)} → dom={dom} ({conf:.0%})')

# ─── 3. Claassifier ──────────────────────────────────────────────────────────
# Engineer-type keywords that strongly suggest SWS4
ENG_KEYWORDS = {
    'catia','solidworks','nx','creo','autocad','inventor',
    'teamcenter','windchill','enovia','abaqus','nastran','ansys',
    'hypermesh','ansa','simufact','fluent','cfx','star-ccm',
    'siemens','dassault','ptc','mathworks','matlab','simulink',
    'labview','teststand','pspice','orcad','altium','zuken',
    'eplan','canape','candb','canalyzer','inca','mda','etk',
    'doors','polarion','codebeamer','mechatronic','plm',
    'cam','cnc','robotics','kuka','fanuc','pcs7','step7','tia',
    'scada','wonderware','osisoft','pi system','codesys','iec61131',
}
# IT-type keywords that suggest SWS1
IT_KEYWORDS = {
    'office','windows','azure','sharepoint','teams','outlook',
    'excel','word','access','visio','project','dynamics',
    'oracle','sap','jira','confluence','bitbucket','gitlab',
    'github','slack','zoom','webex','salesforce','servicenow',
    'adobe','acrobat','illustrator','photoshop','creative',
    'vmware','citrix','cisco','paloalto','fortinet','symantec',
    'kaspersky','mcafee','sophos','crowdstrike','sentinel',
    'backup','veeam','arcserve','commvault','sql','mysql',
    'postgresql','mongodb','redis','elasticsearch','splunk',
    'tableau','powerbi','qlik','python','anaconda','jupyter',
    'intellij','eclipse','vscode','visual studio',
}

def keyword_score(text, kw_set):
    """Return count of keywords found in text."""
    t_lower = text.lower()
    return sum(1 for kw in kw_set
               if re.search(r'\b' + re.escape(kw) + r'\b', t_lower))

def propose_sws(vendor, product, bdbu=''):
    """
    Return (proposed_sws, confidence_str, reason)
    """
    v_key = vendor.lower().strip()
    p_key = product.lower().strip()
    combined = (vendor + ' ' + product).lower()

    # --- step 1: exact vendor match with dominant SWS ----------------------
    if v_key in vendor_sws:
        dom, conf = dominant(vendor_sws[v_key], threshold=0.55)
        if dom:
            return dom, f'{conf:.0%}', f'vendor exact match ({vendor})'
        # ambiguous vendor - use product keywords to disambiguate
        ctr = vendor_sws[v_key]
        # add product keyword votes on top
        eng = keyword_score(combined, ENG_KEYWORDS)
        it  = keyword_score(combined, IT_KEYWORDS)
        if eng > it and 'SWS4' in ctr:
            return 'SWS4', '70%', f'vendor={vendor} + eng keywords ({eng}>{it})'
        if it > eng and 'SWS1' in ctr:
            return 'SWS1', '70%', f'vendor={vendor} + IT keywords ({it}>{eng})'
        # fall back to most common for this vendor
        best = ctr.most_common(1)[0][0]
        return best, f'{conf:.0%}', f'vendor ambiguous - picked most common ({vendor})'

    # --- step 2: partial vendor name match ---------------------------------
    for known_v, ctr in vendor_sws.items():
        if (known_v in v_key or v_key in known_v) and len(known_v) >= 4:
            dom, conf = dominant(ctr, threshold=0.55)
            if dom:
                return dom, f'{conf:.0%}', f'partial vendor match ({known_v})'

    # --- step 3: product keyword scoring -----------------------------------
    eng = keyword_score(combined, ENG_KEYWORDS)
    it  = keyword_score(combined, IT_KEYWORDS)
    if eng > 0 and eng >= it:
        return 'SWS4', '60%', f'product eng keywords ({eng} hits)'
    if it > 0 and it > eng:
        return 'SWS1', '60%', f'product IT keywords ({it} hits)'

    # --- step 4: product word table ----------------------------------------
    words = re.findall(r'\b[a-zA-Z][a-zA-Z0-9+#.]{2,}\b', product)
    word_votes = collections.Counter()
    for w in words[:5]:
        wl = w.lower()
        if wl in prod_word_sws:
            for sws, cnt in prod_word_sws[wl].items():
                word_votes[sws] += cnt
    if word_votes:
        best_sws, best_n = word_votes.most_common(1)[0]
        total_v = sum(word_votes.values())
        conf = best_n / total_v
        if conf >= 0.5:
            return best_sws, f'{conf:.0%}', f'product word match ({words[:3]})'

    # --- step 5: global fallback (majority class = SWS1) ------------------
    return 'SWS1', '30%', 'global fallback (SWS1 is majority class)'


# ─── 4. Load Buyer file ───────────────────────────────────────────────────────
wb_buyer = openpyxl.load_workbook(
    os.path.join(BASE, '_buyer_copy.xlsx'))

ai_fill   = PatternFill('solid', fgColor=COL_AI_FILL)
ai_font   = Font(color='1A1A1A', size=9, italic=True)
bold_font = Font(bold=True, color='FFFFFF', size=9)
centered  = Alignment(horizontal='center', vertical='center', wrap_text=True)

results_log = []

for sheet_name, header_row, data_start, v_col, p_col in [
    ('Shopping List Astra 10.02.2026', 9,  10, 2, 3),
    ('License inventory for FOSS',      7,  8,  3, 4),
]:
    ws_b = wb_buyer[sheet_name]
    max_col = ws_b.max_column
    new_col = max_col + 1
    new_col_letter = get_column_letter(new_col)

    # Write header for new column
    hdr_cell = ws_b.cell(row=header_row, column=new_col, value='SWS (AI-proposed)')
    hdr_cell.font  = Font(bold=True, color='FFFFFF', size=9)
    hdr_cell.fill  = PatternFill('solid', fgColor=COL_HDR_FILL)
    hdr_cell.alignment = centered

    applied = 0
    for row_idx in range(data_start, ws_b.max_row + 1):
        vendor  = clean(ws_b.cell(row_idx, v_col).value)
        product = clean(ws_b.cell(row_idx, p_col).value)
        if not vendor and not product:
            continue

        sws, conf, reason = propose_sws(vendor, product)
        label = f'(AI-proposed) {sws}'
        results_log.append({
            'sheet': sheet_name, 'row': row_idx,
            'vendor': vendor, 'product': product,
            'proposed': sws, 'conf': conf, 'reason': reason,
        })

        cell = ws_b.cell(row=row_idx, column=new_col, value=label)
        cell.fill      = ai_fill
        cell.font      = Font(color=COL_SWS.get(sws, '333333'), size=9,
                              italic=True, bold=True)
        cell.alignment = centered
        applied += 1

    # Auto-size
    ws_b.column_dimensions[new_col_letter].width = 22
    print(f'Sheet {sheet_name!r}: applied to {applied} rows')

# ─── 5. Save ─────────────────────────────────────────────────────────────────
out_path = os.path.join(BASE, 'Astra_Shopping_List_Buyer_AI_SWS.xlsx')
wb_buyer.save(out_path)
print(f'\nSaved: {out_path}')

# ─── 6. Print summary log ─────────────────────────────────────────────────────
print('\n--- AI SWS Proposal Log ---')
dist = collections.Counter(r['proposed'] for r in results_log)
print('Distribution:', dict(dist.most_common()))
conf_low = [r for r in results_log if int(r['conf'].rstrip('%')) < 60]
print(f'Low-confidence items (< 60%): {len(conf_low)}')
for r in conf_low[:20]:
    print(f"  row={r['row']} vendor={r['vendor']!r} product={r['product']!r} "
          f"-> {r['proposed']} ({r['conf']}) | {r['reason']}")
