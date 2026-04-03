"""
Reads both Excel files and exports a unified CSV for Microsoft List import.

Files:
  - Astra Shopping List Buyer_10.02.2026.xlsx  (Project ASTRA)
  - Frame License Shopping List.xlsx           (Project FRAME)

Output:
  - Software_License_MicrosoftList_Import.csv
"""

import openpyxl
import csv
import os

BASE = r'c:\Users\kho1sgp\OneDrive - Bosch Group\My Work Documents\AI topics\BDMIL use case\Software_license_project'

# ------------------------------------------------------------------
# Unified column names for Microsoft List
# ------------------------------------------------------------------
COLUMNS = [
    "Title",                    # Product name  (MS Lists mandatory column)
    "Vendor",
    "Source_Project",           # ASTRA | FRAME
    "License_Category",         # Paid License | Freeware | Open Source Software (OSS)
    "Item_No",                  # Row number in source file
    "Quantity",
    "Metric",
    "Type",                     # Subscription | Perpetual | Freeware | OSS | TBD
    "Ready_for_Procurement",
    "License_Server_Planned",
    "Priority_Date",
    "Priority_Reason",
    "Changes_Since_Feb_2025",
    "Changes_from_Local_List",
    "Action",
    "Responsible",
    "Due_Date",
    "Comment",
    "Comment_2",
]

rows_out = []

def clean(val):
    """Strip whitespace and None-ify empty strings."""
    s = str(val or '').strip()
    return s if s not in ('None', '') else ''


# ==================================================================
# FILE 1 – ASTRA  |  Sheet: Shopping List Astra 10.02.2026
# Columns (row 9): #(A) | Vendor(B) | Product name(C) | Quantity(D) | Metric(E) | Type of Payment*(F) | Notes(G)
# ==================================================================
wb1 = openpyxl.load_workbook(os.path.join(BASE, 'Astra Shopping List Buyer_10.02.2026.xlsx'), data_only=True)

sh = wb1['Shopping List Astra 10.02.2026']
print(f"Reading ASTRA Paid sheet ... rows={sh.max_row}")
for r in range(10, sh.max_row + 1):          # data from row 10
    item_no  = clean(sh.cell(r, 1).value)
    vendor   = clean(sh.cell(r, 2).value)
    product  = clean(sh.cell(r, 3).value)
    qty      = clean(sh.cell(r, 4).value)
    metric   = clean(sh.cell(r, 5).value)
    pay_type = clean(sh.cell(r, 6).value)
    notes    = clean(sh.cell(r, 7).value)

    if not product and not vendor:
        continue   # skip empty / footer rows

    rows_out.append({
        "Title":                  product,
        "Vendor":                 vendor,
        "Source_Project":         "ASTRA",
        "License_Category":       "Paid License",
        "Item_No":                item_no,
        "Quantity":               qty,
        "Metric":                 metric,
        "Type":                   pay_type,
        "Ready_for_Procurement":  '',
        "License_Server_Planned": '',
        "Priority_Date":          '',
        "Priority_Reason":        '',
        "Changes_Since_Feb_2025": '',
        "Changes_from_Local_List":'',
        "Action":                 '',
        "Responsible":            '',
        "Due_Date":               '',
        "Comment":                notes,
        "Comment_2":              '',
    })

print(f"  -> {sum(1 for r in rows_out if r['Source_Project']=='ASTRA' and r['License_Category']=='Paid License')} paid records")

# ==================================================================
# FILE 1 – ASTRA  |  Sheet: License inventory for FOSS
# Columns (row 7): (blank)(A) | #(B) | Vendor(C) | Product name(D) | Quantity(E) | Free or OSS(F) | Notes(G)
# ==================================================================
sh2 = wb1['License inventory for FOSS']
print(f"Reading ASTRA FOSS sheet ... rows={sh2.max_row}")
count_before = len(rows_out)
for r in range(8, sh2.max_row + 1):          # data from row 8
    item_no  = clean(sh2.cell(r, 2).value)
    vendor   = clean(sh2.cell(r, 3).value)
    product  = clean(sh2.cell(r, 4).value)
    qty      = clean(sh2.cell(r, 5).value)
    foss     = clean(sh2.cell(r, 6).value)
    notes    = clean(sh2.cell(r, 7).value)

    if not product and not vendor:
        continue

    rows_out.append({
        "Title":                  product,
        "Vendor":                 vendor,
        "Source_Project":         "ASTRA",
        "License_Category":       foss if foss else "Freeware / OSS",
        "Item_No":                item_no,
        "Quantity":               qty,
        "Metric":                 '',
        "Type":                   foss if foss else '',
        "Ready_for_Procurement":  '',
        "License_Server_Planned": '',
        "Priority_Date":          '',
        "Priority_Reason":        '',
        "Changes_Since_Feb_2025": '',
        "Changes_from_Local_List":'',
        "Action":                 '',
        "Responsible":            '',
        "Due_Date":               '',
        "Comment":                notes,
        "Comment_2":              '',
    })

print(f"  -> {len(rows_out) - count_before} FOSS records")

# ==================================================================
# FILE 2 – FRAME  |  Sheet: Shopping_List_Vorbereitung
# Columns (row 5):
#   A:#1 | B:Vendor | C:Product | D:Priority Date | E:Priority Reason
#   F:Changes since Feb 2025 | G:Changes from local list | H:Type
#   I:Quantity | J:Metric | K:Ready for Procurement | L:License Server planned?
#   M:Comment | N:Comment 2 | O:Action | P:Responsible | Q:Due Date
# ==================================================================
wb2 = openpyxl.load_workbook(os.path.join(BASE, 'Frame License Shopping List.xlsx'), data_only=True)
sh3 = wb2['Shopping_List_Vorbereitung']
print(f"Reading FRAME sheet ... rows={sh3.max_row}")
count_before = len(rows_out)
for r in range(6, sh3.max_row + 1):          # data from row 6
    item_no     = clean(sh3.cell(r, 1).value)
    vendor      = clean(sh3.cell(r, 2).value)
    product     = clean(sh3.cell(r, 3).value)
    prio_date   = clean(sh3.cell(r, 4).value)
    prio_reason = clean(sh3.cell(r, 5).value)
    chg_feb     = clean(sh3.cell(r, 6).value)
    chg_local   = clean(sh3.cell(r, 7).value)
    lic_type    = clean(sh3.cell(r, 8).value)
    qty         = clean(sh3.cell(r, 9).value)
    metric      = clean(sh3.cell(r, 10).value)
    ready       = clean(sh3.cell(r, 11).value)
    lic_server  = clean(sh3.cell(r, 12).value)
    comment     = clean(sh3.cell(r, 13).value)
    comment2    = clean(sh3.cell(r, 14).value)
    action      = clean(sh3.cell(r, 15).value)
    responsible = clean(sh3.cell(r, 16).value)
    due_date    = clean(sh3.cell(r, 17).value)

    if not product and not vendor:
        continue

    rows_out.append({
        "Title":                  product,
        "Vendor":                 vendor,
        "Source_Project":         "FRAME",
        "License_Category":       lic_type if lic_type else "Paid License",
        "Item_No":                item_no,
        "Quantity":               qty,
        "Metric":                 metric,
        "Type":                   lic_type,
        "Ready_for_Procurement":  ready,
        "License_Server_Planned": lic_server,
        "Priority_Date":          prio_date,
        "Priority_Reason":        prio_reason,
        "Changes_Since_Feb_2025": chg_feb,
        "Changes_from_Local_List":chg_local,
        "Action":                 action,
        "Responsible":            responsible,
        "Due_Date":               due_date,
        "Comment":                comment,
        "Comment_2":              comment2,
    })

print(f"  -> {len(rows_out) - count_before} FRAME records")
print(f"\nTotal records to export: {len(rows_out)}")

# ==================================================================
# Write CSV
# ==================================================================
out_path = os.path.join(BASE, 'Software_License_MicrosoftList_Import.csv')
with open(out_path, 'w', newline='', encoding='utf-8-sig') as f:   # utf-8-sig = BOM for Excel
    writer = csv.DictWriter(f, fieldnames=COLUMNS)
    writer.writeheader()
    writer.writerows(rows_out)

print(f"\nCSV written to: {out_path}")
print(f"Columns: {', '.join(COLUMNS)}")
