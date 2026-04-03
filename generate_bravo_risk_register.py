"""
generate_bravo_risk_register.py
Generates Bravo_Risk_Register.xlsx by populating Risk_analysis_template.xlsx
with Bravo project risk data.

Project Bravo: BGSW AI Business Carve-Out into 50/50 JV with Tata
Seller: Bosch BGSW | Buyer: Tata | PM: Riyaz Ahmed Syed Ahmed (BD/MIL-PSM4)
2 India sites | 70 users | 17 apps | No ERP | No TSA | Bosch JV leadership control

Usage:
    python generate_bravo_risk_register.py
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment

HERE = Path(__file__).parent
TEMPLATE = HERE / "Risk_analysis_template.xlsx"
OUT_DIR  = HERE / "Bravo"
OUT_DIR.mkdir(exist_ok=True)
OUTPUT   = OUT_DIR / "Bravo_Risk_Register.xlsx"

# Bravo Risk Data — 14 identified risks, probability x impact (1-5 scale)
# Columns: Category, Description, Effects, Root Cause,
#          Probability(W), Impact(T), Mitigation, Owner, Deadline, Status, Remarks
BRAVO_RISKS = [
    {
        "Category":    "ScR - Schedule",
        "Description": "BGSW and Tata JV management bandwidth constrained during 3-month build-to-GoLive sprint (Apr-Jun 2026)",
        "Effects":     "Phase 3 build and testing delays; GoLive of 1 Jul 2026 missed",
        "Causes":      "Carve-out running concurrently with live AI business operations; split leadership focus",
        "Probability": 4, "Impact": 4,
        "Mitigation":  "Dedicate BGSW carve-out lead with explicit protected time allocation; Deputy named; Weekly SteerCo cadence from Day 1",
        "Owner":       "Riyaz Ahmed, BGSW CIO",
        "Deadline":    "30-Apr-26",
        "Status":      "Amber",
        "Remarks":     "High priority. Mitigation plan in place. Monitor weekly.",
    },
    {
        "Category":    "SR - Scope",
        "Description": "JV legal entity registration in India delayed beyond QG0 target (30 Apr 2026)",
        "Effects":     "QG0 gate cannot be passed; entire programme timeline shifts right",
        "Causes":      "MCA India bureaucratic processing times (4-8 weeks); first-time JV registration complexity",
        "Probability": 3, "Impact": 5,
        "Mitigation":  "Engage external India legal counsel by Apr 3; initiate MCA filing within first week of project",
        "Owner":       "Riyaz Ahmed, Legal, Finance",
        "Deadline":    "10-Apr-26",
        "Status":      "Green",
        "Remarks":     "Critical path item. India legal counsel essential. No antitrust filing required.",
    },
    {
        "Category":    "BtR - Budget",
        "Description": "India cloud and M365 licensing costs exceed initial estimate for JV tenant provisioning",
        "Effects":     "Budget overrun in Phase 3; delay in cloud environment readiness",
        "Causes":      "Azure India region pricing; JV tenant licensing model complexity for 70 users",
        "Probability": 3, "Impact": 3,
        "Mitigation":  "Conduct M365/Azure RFQ by end of Apr; lock licensing model at QG0; budget contingency of 15% reserved",
        "Owner":       "Finance, BGSW Procurement",
        "Deadline":    "30-Apr-26",
        "Status":      "Amber",
        "Remarks":     "Small user base (70) limits financial exposure. Containable risk.",
    },
    {
        "Category":    "ScR - Schedule",
        "Description": "Phase 3 build (Jun 2026) delayed due to BGSW infrastructure resource contention",
        "Effects":     "GoLive date of 1 Jul 2026 missed; Phase 4 cutover cannot proceed",
        "Causes":      "BGSW infra team shared across multiple India projects; AD and network build sequential dependencies",
        "Probability": 3, "Impact": 4,
        "Mitigation":  "Reserve BGSW infra team capacity from May; confirm resource allocation at QG1/2/3 gate",
        "Owner":       "Riyaz Ahmed, BGSW Infra Lead",
        "Deadline":    "01-Jun-26",
        "Status":      "Green",
        "Remarks":     "Short build window (Jun only). Resource reservation critical.",
    },
    {
        "Category":    "RR - Resource",
        "Description": "Key BGSW IT personnel unavailable during compressed Jun 2026 build-and-test window",
        "Effects":     "Build quality impacted; integration testing incomplete by QG1/2/3 (26 Jun)",
        "Causes":      "Annual leave; BGSW parallel project demands; carve-out knowledge concentrated in few individuals",
        "Probability": 3, "Impact": 4,
        "Mitigation":  "Enforce Jun leave freeze for critical roles; identify/train backups; document knowledge by end of May",
        "Owner":       "BGSW HR, Riyaz Ahmed",
        "Deadline":    "29-May-26",
        "Status":      "Amber",
        "Remarks":     "June is the highest-risk month. Contingency plan required.",
    },
    {
        "Category":    "SR - Scope",
        "Description": "Application migration scope expands beyond 17 AI apps due to late-discovered dependencies",
        "Effects":     "Phase 3 timeline overrun; additional testing cycles required",
        "Causes":      "Incomplete initial inventory; business users requesting additional apps post-QG0",
        "Probability": 2, "Impact": 4,
        "Mitigation":  "Hard scope freeze at QG0 (30 Apr); no additions after that date without SteerCo approval and timeline impact assessment",
        "Owner":       "Riyaz Ahmed, WS Leads",
        "Deadline":    "30-Apr-26",
        "Status":      "Green",
        "Remarks":     "Small app count (17) makes scope creep manageable if freeze enforced.",
    },
    {
        "Category":    "CR - Change",
        "Description": "Active Directory forest separation for JV incomplete by GoLive (1 Jul 2026)",
        "Effects":     "Day 1 users unable to log on to JV systems; GoLive blocked",
        "Causes":      "AD tooling complexity; configuration errors in JV forest build; insufficient testing time",
        "Probability": 2, "Impact": 5,
        "Mitigation":  "Start AD build in week 1 of Jun; run integration and regression testing in parallel; conduct dress rehearsal Jun 24-26",
        "Owner":       "BGSW AD Team, Riyaz Ahmed",
        "Deadline":    "24-Jun-26",
        "Status":      "Green",
        "Remarks":     "AD is foundational for Day 1. No backout. Pilot and rehearsal mandatory.",
    },
    {
        "Category":    "BtR - Budget",
        "Description": "Licence change-of-control costs for 17 AI applications exceed estimates",
        "Effects":     "Budget overrun; delay in app migration if contracts not renegotiated in time",
        "Causes":      "ISVs invoking change-of-control clauses; unexpected vendor lock-in terms",
        "Probability": 2, "Impact": 3,
        "Mitigation":  "Complete licence review with change-of-control clause analysis by May 14; negotiate amendments ahead of Phase 3",
        "Owner":       "BGSW Procurement, Legal",
        "Deadline":    "14-May-26",
        "Status":      "Green",
        "Remarks":     "17 apps is manageable. Early review essential.",
    },
    {
        "Category":    "QR - Quality",
        "Description": "UAT failure rate high due to compressed testing window (18-24 Jun 2026)",
        "Effects":     "QG1/2/3 gate (26 Jun) delayed; GoLive 1 Jul 2026 at risk",
        "Causes":      "Business users not available for full UAT; test environment not stable",
        "Probability": 3, "Impact": 4,
        "Mitigation":  "Engage business UAT leads by May 1; establish test environment in week 1 of Jun; run smoke tests daily from Jun 15",
        "Owner":       "Riyaz Ahmed, Business Leads",
        "Deadline":    "15-Jun-26",
        "Status":      "Green",
        "Remarks":     "UAT window is tight (5 working days). Early prep critical.",
    },
    {
        "Category":    "SR - Scope",
        "Description": "Tata JV counterpart team (architecture, IT) not staffed in time for May concept workshops",
        "Effects":     "JV IT architecture design delayed; Phase 2 deliverables incomplete",
        "Causes":      "Tata internal hiring/assignment process for JV roles not yet started",
        "Probability": 3, "Impact": 3,
        "Mitigation":  "Confirm Tata JV IT team nominees at Steering Committee kickoff (Apr 1); escalate to executive sponsor if not confirmed by Apr 10",
        "Owner":       "Riyaz Ahmed, Tata Executive Sponsor",
        "Deadline":    "10-Apr-26",
        "Status":      "Amber",
        "Remarks":     "Bosch leadership control mitigates: BGSW can proceed if Tata team is delayed.",
    },
    {
        "Category":    "LR - Legal",
        "Description": "JV shareholder agreement final terms delay IT separation decisions",
        "Effects":     "Scope boundary between BGSW and JV IT undefined; Phase 2 architecture blocked",
        "Causes":      "Legal negotiations between Bosch and Tata on IP ownership and data rights extending beyond Apr",
        "Probability": 2, "Impact": 4,
        "Mitigation":  "Identify IT decisions independent of full SHA; escalate blocking items to SteerCo; resolve core IP scope by QG0",
        "Owner":       "Legal, Riyaz Ahmed",
        "Deadline":    "30-Apr-26",
        "Status":      "Green",
        "Remarks":     "Bosch leadership control reduces legal complexity significantly.",
    },
    {
        "Category":    "CR - Change",
        "Description": "Data separation between BGSW corporate and JV AI business data is more complex than anticipated",
        "Effects":     "Data migration execution delayed past Jun 22; Day 1 data completeness at risk",
        "Causes":      "AI training datasets and model artefacts stored in shared BGSW repositories; unclear data lineage",
        "Probability": 3, "Impact": 3,
        "Mitigation":  "Complete data ownership mapping by May 19; agree separation rules with legal and data owners before Phase 3 starts",
        "Owner":       "Riyaz Ahmed, Legal, BGSW IT",
        "Deadline":    "29-May-26",
        "Status":      "Green",
        "Remarks":     "Bosch-led JV reduces risk: BGSW data stewardship retained post-GoLive.",
    },
    {
        "Category":    "ScR - Schedule",
        "Description": "Hypercare issues in Jul-Sep 2026 require BGSW resources beyond planned capacity",
        "Effects":     "BGSW IT cost overrun; closure milestone (Oct 30) delayed",
        "Causes":      "Application instability post-GoLive; 17 apps not fully stabilised before hypercare close",
        "Probability": 2, "Impact": 3,
        "Mitigation":  "Ensure all 17 apps fully tested before GoLive; define clear P1/P2 escalation matrix; 60-day hypercare window is adequate buffer",
        "Owner":       "Riyaz Ahmed, IT Ops",
        "Deadline":    "01-Jul-26",
        "Status":      "Green",
        "Remarks":     "No TSA risk as JV is Bosch-led; escalation to BGSW via governance channel available.",
    },
    {
        "Category":    "BR - Business",
        "Description": "AI business continuity disrupted during Jun 2026 build and cutover activities",
        "Effects":     "Tata / BGSW AI business projects delayed; executive sponsor dissatisfaction",
        "Causes":      "Carve-out build activities running concurrently with live AI project deliveries",
        "Probability": 2, "Impact": 3,
        "Mitigation":  "Define AI business activity blackout window for Week of Jun 29-Jul 1; agree cutover communications with business leads",
        "Owner":       "BGSW AI Business Lead, Riyaz Ahmed",
        "Deadline":    "29-Jun-26",
        "Status":      "Green",
        "Remarks":     "Business continuity plan to be approved at QG1/2/3 gate.",
    },
]


def populate_risk_register():
    wb = load_workbook(TEMPLATE)

    # --- Cover Sheet ---
    ws_cover = wb["Cover sheet"]
    ws_cover["D1"] = "BGSW CIO"
    ws_cover["D2"] = "Riyaz Ahmed Syed Ahmed (BD/MIL-PSM4)"
    ws_cover["D3"] = "Riyaz Ahmed Syed Ahmed (BD/MIL-PSM4)"

    # --- Analysis Sheet ---
    ws = wb["Analysis of project risks"]
    FIRST_ROW = 9

    for i, risk in enumerate(BRAVO_RISKS):
        row_num = FIRST_ROW + i
        ws.cell(row_num, 1).value  = i + 1                          # No.
        ws.cell(row_num, 2).value  = "Bravo"                        # Sub-project
        ws.cell(row_num, 3).value  = "03-Apr-26"                    # Entry Date
        ws.cell(row_num, 4).value  = risk["Category"]               # Risk Category
        ws.cell(row_num, 5).value  = risk["Description"]            # Risk Description
        ws.cell(row_num, 6).value  = risk["Effects"]                # Effects (Column F)
        ws.cell(row_num, 7).value  = risk["Causes"]                 # Causes
        ws.cell(row_num, 8).value  = risk["Probability"]            # W - Probability
        ws.cell(row_num, 9).value  = risk["Impact"]                 # T - Impact
        # Column J: RZ formula — set explicitly for rows beyond template pre-built range
        ws.cell(row_num, 10).value = f"=$H{row_num}*$I{row_num}"
        ws.cell(row_num, 11).value = risk["Mitigation"]             # Actions
        ws.cell(row_num, 12).value = risk["Owner"]                  # Responsible
        ws.cell(row_num, 13).value = risk["Deadline"]               # Deadline
        ws.cell(row_num, 14).value = risk["Status"]                 # Status
        ws.cell(row_num, 15).value = risk["Remarks"]                # Remarks

        # Wrap text in description columns
        for col in [5, 6, 7, 11, 15]:
            ws.cell(row_num, col).alignment = Alignment(wrap_text=True)

    wb.save(OUTPUT)
    print(f"Risk register written: {OUTPUT}  ({len(BRAVO_RISKS)} risks)")


if __name__ == "__main__":
    from datetime import datetime as _dt
    _t0 = _dt.now()
    print(f"Started : {_t0.strftime('%Y-%m-%d %H:%M:%S')}")
    if not TEMPLATE.exists():
        print(f"ERROR: Template not found: {TEMPLATE}")
        sys.exit(1)
    populate_risk_register()
    _t1 = _dt.now()
    print(f"Finished: {_t1.strftime('%Y-%m-%d %H:%M:%S')}  ({(_t1-_t0).total_seconds():.1f}s elapsed)")
