#!/usr/bin/env python3
"""Generate Gamma cost plan from the schedule and risk register."""

from __future__ import annotations

import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HERE = Path(__file__).parent
PROJECT_NAME = "Gamma"
OUTPUT_FOLDER_NAME = "Gamma v1.0"
DOCUMENT_VERSION = "Version 1.0 - Initial Baseline"
SCHEDULE_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Risk_Register.xlsx"
OUTPUT_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Cost_Plan.xlsx"

HOURLY_RATES = {
    "Alibaba Cloud": 125,
    "Alibaba IT": 110,
    "Alibaba ITSM": 90,
    "Alibaba Leadership": 180,
    "Alibaba PMO": 100,
    "Alibaba Security": 125,
    "Application Owners": 100,
    "Business Key Users": 80,
    "Business Leads": 100,
    "Communications": 80,
    "Data Leads": 105,
    "Deployment Lead": 100,
    "EY": 135,
    "EY Test Lead": 125,
    "Finance": 95,
    "HR": 80,
    "IT Operations": 95,
    "ITSM Leads": 95,
    "Infrastructure Leads": 120,
    "Infrastructure Operations": 90,
    "Legal": 175,
    "PMO": 95,
    "Procurement": 95,
    "Robert Bosch China IT": 110,
    "Robert Bosch China IT Security": 130,
    "Robert Bosch China Infrastructure": 115,
    "Robert Bosch China Leadership": 180,
    "Robert Bosch China PMO": 100,
    "Robert Bosch China Security": 130,
    "Security Leads": 120,
    "Service Desk": 65,
    "Steering Committee": 210,
    "Test Lead": 110,
    "Training Lead": 80,
    "Workstream Leads": 110,
}

CATEGORY_MAP = {
    "Alibaba Cloud": "INFRASTRUCTURE & CLOUD",
    "Alibaba IT": "INFRASTRUCTURE & CLOUD",
    "Alibaba ITSM": "OPERATIONS & HYPERCARE",
    "Alibaba Leadership": "EXECUTIVE SPONSORSHIP",
    "Alibaba PMO": "PROGRAMME MANAGEMENT & GOVERNANCE",
    "Alibaba Security": "SECURITY & IDENTITY",
    "Application Owners": "APPLICATION & DATA TRANSITION",
    "Business Key Users": "TESTING & BUSINESS READINESS",
    "Business Leads": "TESTING & BUSINESS READINESS",
    "Communications": "CHANGE, DEPLOYMENT & TRAINING",
    "Data Leads": "APPLICATION & DATA TRANSITION",
    "Deployment Lead": "CHANGE, DEPLOYMENT & TRAINING",
    "EY": "PROGRAMME MANAGEMENT & GOVERNANCE",
    "EY Test Lead": "TESTING & BUSINESS READINESS",
    "Finance": "PROGRAMME MANAGEMENT & GOVERNANCE",
    "HR": "CHANGE, DEPLOYMENT & TRAINING",
    "IT Operations": "OPERATIONS & HYPERCARE",
    "ITSM Leads": "OPERATIONS & HYPERCARE",
    "Infrastructure Leads": "INFRASTRUCTURE & CLOUD",
    "Infrastructure Operations": "OPERATIONS & HYPERCARE",
    "Legal": "LEGAL, PROCUREMENT & COMPLIANCE",
    "PMO": "PROGRAMME MANAGEMENT & GOVERNANCE",
    "Procurement": "LEGAL, PROCUREMENT & COMPLIANCE",
    "Robert Bosch China IT": "INFRASTRUCTURE & CLOUD",
    "Robert Bosch China IT Security": "SECURITY & IDENTITY",
    "Robert Bosch China Infrastructure": "INFRASTRUCTURE & CLOUD",
    "Robert Bosch China Leadership": "EXECUTIVE SPONSORSHIP",
    "Robert Bosch China PMO": "PROGRAMME MANAGEMENT & GOVERNANCE",
    "Robert Bosch China Security": "SECURITY & IDENTITY",
    "Security Leads": "SECURITY & IDENTITY",
    "Service Desk": "OPERATIONS & HYPERCARE",
    "Steering Committee": "EXECUTIVE SPONSORSHIP",
    "Test Lead": "TESTING & BUSINESS READINESS",
    "Training Lead": "CHANGE, DEPLOYMENT & TRAINING",
    "Workstream Leads": "PROGRAMME MANAGEMENT & GOVERNANCE",
}

CATEGORY_ORDER = [
    "PROGRAMME MANAGEMENT & GOVERNANCE",
    "LEGAL, PROCUREMENT & COMPLIANCE",
    "INFRASTRUCTURE & CLOUD",
    "SECURITY & IDENTITY",
    "APPLICATION & DATA TRANSITION",
    "TESTING & BUSINESS READINESS",
    "CHANGE, DEPLOYMENT & TRAINING",
    "OPERATIONS & HYPERCARE",
    "EXECUTIVE SPONSORSHIP",
]

IMPACT_VALUES = {"Very Low": 1, "Low": 2, "Moderate": 3, "High": 4, "Very High": 5}
PROBABILITY_VALUES = {"10%": 1, "30%": 2, "50%": 3, "70%": 4, "90%": 5}

CONTINGENCY_MAP = {
    1: ("Risk Register #1 - External antitrust and JV legal reserve", 45000, "External legal support for antitrust and governance scenario rework"),
    5: ("Risk Register #5 - Shared infrastructure dependency assessment reserve", 60000, "External specialist support for hidden shared-service separation issues"),
    10: ("Risk Register #10 - Network and certificate acceleration reserve", 30000, "Provider escalation, temporary routing, and certificate contingency"),
    11: ("Risk Register #11 - Dual-running and transition-service reserve", 55000, "Temporary duplicate environments and service continuity uplift"),
    13: ("Risk Register #13 - Contract novation and licence transfer reserve", 40000, "External legal and vendor charges for change-of-control approvals"),
    16: ("Risk Register #16 - Final readiness and rollback contingency", 35000, "Additional command-center, testing, and rollback support around QG4"),
}


def parse_duration_days(value: str) -> int:
    return int(str(value).split()[0])


def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


def split_resources(value: str) -> list[str]:
    return [token.strip() for token in str(value).split("+") if token and token.strip()]


def load_schedule_inputs() -> tuple[list[tuple[str, str, str]], dict[str, int], dict[str, int]]:
    if not SCHEDULE_PATH.exists():
        raise FileNotFoundError(f"Missing schedule input: {SCHEDULE_PATH}")

    wb = load_workbook(SCHEDULE_PATH, data_only=False)
    ws = wb["Schedule"]

    phases: list[tuple[str, str, str]] = []
    resource_days: dict[str, int] = defaultdict(int)
    phase_costs: dict[str, int] = defaultdict(int)
    current_phase = None

    for row in range(2, ws.max_row + 1):
        level = ws.cell(row, 2).value
        if level is None:
            continue
        level = int(level)
        name = str(ws.cell(row, 3).value or "").strip()
        start = format_date(ws.cell(row, 5).value)
        finish = format_date(ws.cell(row, 6).value)
        if level == 1:
            current_phase = name
            phases.append((name, start, finish))
            continue

        resources = split_resources(ws.cell(row, 8).value or "")
        if level < 3 or not resources or current_phase is None:
            continue

        days = parse_duration_days(ws.cell(row, 4).value)
        for resource in resources:
            resource_days[resource] += days
            phase_costs[current_phase] += days * 8 * HOURLY_RATES[resource]

    return phases, dict(resource_days), dict(phase_costs)


def load_high_risk_contingencies() -> tuple[list[tuple[str, int, str]], list[int]]:
    if not RISK_PATH.exists():
        raise FileNotFoundError(f"Missing risk input: {RISK_PATH}")

    wb = load_workbook(RISK_PATH, data_only=False)
    ws = wb["Risk Register"]
    high_risk_ids: list[int] = []

    for row in range(5, 140):
        risk_id = ws.cell(row, 2).value
        if risk_id is None:
            continue
        impact = ws.cell(row, 12).value
        probability = ws.cell(row, 14).value
        risk_type = ws.cell(row, 16).value
        score = IMPACT_VALUES.get(impact, 0) * PROBABILITY_VALUES.get(probability, 0)
        if risk_type == "threat" and score >= 12:
            high_risk_ids.append(int(risk_id))

    contingencies = [CONTINGENCY_MAP[risk_id] for risk_id in high_risk_ids if risk_id in CONTINGENCY_MAP]
    return contingencies, high_risk_ids


def validate_resource_mappings(resources: dict[str, int]) -> None:
    missing_rates = sorted(resource for resource in resources if resource not in HOURLY_RATES)
    missing_categories = sorted(resource for resource in resources if resource not in CATEGORY_MAP)
    if missing_rates or missing_categories:
        problems = []
        if missing_rates:
            problems.append(f"missing hourly rates for: {', '.join(missing_rates)}")
        if missing_categories:
            problems.append(f"missing categories for: {', '.join(missing_categories)}")
        raise ValueError("; ".join(problems))


def build_category_rows(resource_days: dict[str, int]) -> tuple[dict[str, list[tuple[str, int, int, int, int]]], dict[str, int], int]:
    grouped: dict[str, list[tuple[str, int, int, int, int]]] = {category: [] for category in CATEGORY_ORDER}
    category_totals: dict[str, int] = {category: 0 for category in CATEGORY_ORDER}
    total_cost = 0

    for resource in sorted(resource_days):
        days = resource_days[resource]
        hours = days * 8
        rate = HOURLY_RATES[resource]
        cost = hours * rate
        category = CATEGORY_MAP[resource]
        grouped[category].append((resource, days, hours, rate, cost))
        category_totals[category] += cost
        total_cost += cost

    return grouped, category_totals, total_cost


def write_xlsx(
    phases: list[tuple[str, str, str]],
    grouped_rows: dict[str, list[tuple[str, int, int, int, int]]],
    category_totals: dict[str, int],
    phase_costs: dict[str, int],
    labor_total: int,
    contingencies: list[tuple[str, int, str]],
    high_risk_ids: list[int],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Plan"

    dark_blue = PatternFill(start_color="003B6E", end_color="003B6E", fill_type="solid")
    mid_blue = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    light_blue = PatternFill(start_color="EFF4FB", end_color="EFF4FB", fill_type="solid")
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    subtotal_fill = PatternFill(start_color="C6D4E8", end_color="C6D4E8", fill_type="solid")

    title_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    regular_font = Font(name="Calibri", size=9, color="000000")
    bold_font = Font(name="Calibri", size=9, bold=True, color="000000")
    total_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    meta_font = Font(name="Calibri", size=8, color="000000")
    note_font = Font(name="Calibri", size=8, italic=True, color="000000")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "GAMMA PROJECT COST PLAN"
    ws[f"A{row}"].font = title_font
    ws[f"A{row}"].fill = dark_blue
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    metadata_rows = [
        ("Version:", DOCUMENT_VERSION),
        ("Project:", "Gamma | Seller: Robert Bosch China | Buyer: Alibaba"),
        ("Scope:", "5 sites | 250 users | 20 applications | Pure infrastructure services | No SAP"),
        ("Duration:", "2026-08-01 to 2027-05-31 | GoLive 2027-02-01 | Combination model"),
        ("Based on:", SCHEDULE_PATH.name),
        ("Risk-aligned:", RISK_PATH.name),
        ("Budget baseline:", "Draft labour baseline - to be approved at QG1"),
        ("High-risk cross-check:", f"Threat risks with score >= 12 reviewed: {', '.join(str(risk_id) for risk_id in high_risk_ids)}"),
    ]
    for label, value in metadata_rows:
        ws[f"A{row}"].value = label
        ws[f"B{row}"].value = value
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.fill = grey_fill
            cell.font = meta_font
        row += 1

    row += 1
    header_row = row
    headers = ["CATEGORY", "RESOURCE", "TOTAL DAYS", "TOTAL HRS", "HOURLY RATE (EUR)", "TOTAL COST (EUR)"]
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row, index)
        cell.value = header
        cell.font = header_font
        cell.fill = mid_blue
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    row += 1

    for category in CATEGORY_ORDER:
        resources = grouped_rows[category]
        if not resources:
            continue
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = category
        cell.font = header_font
        cell.fill = mid_blue
        cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        for index, (resource, days, hours, rate, cost) in enumerate(resources, start=1):
            row_fill = light_blue if index % 2 == 0 else None
            values = [category, resource, days, hours, rate, cost]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row, col)
                cell.value = value if col != 1 else ""
                cell.font = regular_font
                if row_fill:
                    cell.fill = row_fill
                cell.border = border
                if col >= 3:
                    cell.number_format = "#,##0"
            row += 1

        ws.merge_cells(f"A{row}:E{row}")
        subtotal_cell = ws[f"A{row}"]
        subtotal_cell.value = f"SUBTOTAL - {category}"
        subtotal_cell.font = bold_font
        subtotal_cell.fill = subtotal_fill
        subtotal_cell.border = border
        total_cell = ws.cell(row, 6)
        total_cell.value = category_totals[category]
        total_cell.font = bold_font
        total_cell.fill = subtotal_fill
        total_cell.border = border
        total_cell.number_format = "#,##0"
        row += 1

    ws.merge_cells(f"A{row}:E{row}")
    total_label = ws[f"A{row}"]
    total_label.value = "OVERALL PROJECT TOTAL - LABOUR ONLY"
    total_label.font = total_font
    total_label.fill = dark_blue
    total_label.border = border
    total_value = ws.cell(row, 6)
    total_value.value = labor_total
    total_value.font = total_font
    total_value.fill = dark_blue
    total_value.border = border
    total_value.number_format = "#,##0"
    row += 2

    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "COST BREAKDOWN BY CATEGORY"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = mid_blue
    row += 1
    for index, category in enumerate(CATEGORY_ORDER, start=1):
        if category_totals[category] == 0:
            continue
        row_fill = light_blue if index % 2 == 0 else None
        ws.cell(row, 1).value = category
        ws.cell(row, 6).value = category_totals[category]
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.font = regular_font
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if col == 6:
                cell.number_format = "#,##0"
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "COST BREAKDOWN BY PHASE"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = mid_blue
    row += 1
    for index, (phase_name, start, finish) in enumerate(phases, start=1):
        row_fill = light_blue if index % 2 == 0 else None
        ws.cell(row, 1).value = phase_name
        ws.cell(row, 2).value = f"{start} to {finish}"
        ws.cell(row, 6).value = phase_costs.get(phase_name, 0)
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.font = regular_font
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if col == 6:
                cell.number_format = "#,##0"
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "CAPEX / ADDITIONAL COSTS - EXCLUDED FROM LABOUR TOTAL"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = mid_blue
    row += 1
    capex_total = 0
    for index, (description, amount, note) in enumerate(contingencies, start=1):
        row_fill = light_blue if index % 2 == 0 else None
        ws.cell(row, 1).value = description
        ws.cell(row, 2).value = note
        ws.cell(row, 6).value = amount
        capex_total += amount
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.font = regular_font
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if col == 6:
                cell.number_format = "#,##0"
        row += 1

    ws.merge_cells(f"A{row}:E{row}")
    capex_label = ws[f"A{row}"]
    capex_label.value = "TOTAL CAPEX / ADDITIONAL COSTS"
    capex_label.font = bold_font
    capex_label.fill = subtotal_fill
    capex_label.border = border
    capex_value = ws.cell(row, 6)
    capex_value.value = capex_total
    capex_value.font = bold_font
    capex_value.fill = subtotal_fill
    capex_value.border = border
    capex_value.number_format = "#,##0"
    row += 2

    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "NOTES"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = mid_blue
    row += 1
    notes = [
        "Labour totals are derived from Gamma schedule resource assignments using 8 hours per day and exact resource tokens from the schedule.",
        "CAPEX and additional cost lines are excluded from the labour total and cross-reference Gamma risk register items with score >= 12 where external spend may be required.",
        "Combination model note: interim service continuity is assumed only where required until the jointly managed JV steady-state model is accepted.",
        "No SAP is in scope; the cost profile is intentionally infrastructure-heavy and sized for 5 sites, 250 users, and 20 applications.",
    ]
    for note in notes:
        ws.merge_cells(f"A{row}:F{row}")
        note_cell = ws[f"A{row}"]
        note_cell.value = note
        note_cell.font = note_font
        note_cell.fill = grey_fill
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    wb.save(OUTPUT_PATH)


def main() -> None:
    print(f"[{PROJECT_NAME}] Generating cost plan")
    phases, resource_days, phase_costs = load_schedule_inputs()
    validate_resource_mappings(resource_days)
    contingencies, high_risk_ids = load_high_risk_contingencies()
    grouped_rows, category_totals, labor_total = build_category_rows(resource_days)
    write_xlsx(phases, grouped_rows, category_totals, phase_costs, labor_total, contingencies, high_risk_ids)
    print(f"  Output: {OUTPUT_PATH}")
    print(f"  Labour total: {labor_total:,.0f} EUR")
    print(f"  CAPEX / additional cost lines: {len(contingencies)}")


if __name__ == "__main__":
    main()