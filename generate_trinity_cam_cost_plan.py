#!/usr/bin/env python3
"""
Generate Trinity-CAM IT Cost Plan (XLSX).

Derived from:  active-projects/Trinity-CAM/Trinity-CAM_Project_Schedule.xlsx
Risk-aligned:  active-projects/Trinity-CAM/Trinity-CAM_Risk_Register.xlsx

Labour-only plan covering KPMG (PMO + advisory) and Infosys (IT delivery).
JCI and Bosch internal resource effort is tracked but not costed here.
CAPEX / infrastructure / licence costs listed separately and excluded from labour total.

Integration model:  JCI IT -> Merger Zone (Infosys) -> Bosch IT
Timeline:           2026-07-01 to 2028-04-01  (18 months delivery + 3 months hypercare)
"""

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE    = Path(__file__).parent
OUTPUT  = HERE / "active-projects" / "Trinity-CAM" / "Trinity-CAM_Cost_Plan.xlsx"
OUTPUT.parent.mkdir(parents=True, exist_ok=True)

# =============================================================================
# COST DATA — Labour resources derived from Trinity-CAM schedule assignments
#
# Each entry: (category_header, resource_name, total_days, hourly_rate_eur)
# Total hours  = total_days * 8
# Total cost   = total_hours * hourly_rate
#
# Phases referenced in layout:
#   Phase 0 Initiation         2026-07-01 – 2026-10-01    91 days
#   Phase 1 Discovery          2026-10-02 – 2027-01-31   122 days
#   Phase 2 MZ Build           2027-02-01 – 2027-07-31   180 days
#   Phase 3 Testing/Migration  2027-08-01 – 2027-11-30   122 days
#   Phase 4 Pre-GoLive         2027-12-01 – 2028-01-01    31 days
#   Phase 5 Hypercare/Closure  2028-01-02 – 2028-04-01    90 days
# =============================================================================

CATEGORIES = [
    {
        "header": "KPMG — PMO & Programme Management",
        "rows": [
            ("KPMG PMO Lead",        400, 250),
            ("KPMG Project Manager", 380, 190),
        ],
    },
    {
        "header": "KPMG — Architecture & Technical Advisory",
        "rows": [
            ("KPMG Infrastructure Architect", 290, 200),
            ("KPMG Data Architect",           180, 200),
        ],
    },
    {
        "header": "KPMG — SAP Advisory",
        "rows": [
            ("KPMG SAP Architect", 240, 225),
        ],
    },
    {
        "header": "Infosys — Programme Management",
        "rows": [
            ("Infosys Programme Manager", 560, 185),
        ],
    },
    {
        "header": "Infosys — Infrastructure, Network & Security",
        "rows": [
            ("Infosys Infrastructure Architect", 520, 150),
            ("Infosys Network Lead",             320, 140),
            ("Infosys Cloud Lead",               300, 145),
            ("Infosys Security Lead",            260, 150),
        ],
    },
    {
        "header": "Infosys — Identity & Access Management",
        "rows": [
            ("Infosys IAM Lead", 280, 140),
        ],
    },
    {
        "header": "Infosys — SAP Build & Configuration",
        "rows": [
            ("Infosys SAP Architect", 480, 165),
            ("Infosys SAP Lead",      420, 150),
        ],
    },
    {
        "header": "Infosys — Application Migration",
        "rows": [
            ("Infosys Application Lead", 480, 135),
        ],
    },
    {
        "header": "Infosys — Data Migration",
        "rows": [
            ("Infosys Data Migration Lead", 360, 140),
        ],
    },
    {
        "header": "Infosys — Service Delivery & Hypercare Support",
        "rows": [
            ("Infosys Service Delivery Lead", 450, 130),
        ],
    },
]

# Phase breakdown (% of total labour cost)
PHASE_BREAKDOWN = [
    ("Phase 0: Project Initiation",        "2026-07-01", "2026-10-01",  0.05),
    ("Phase 1: Discovery & Architecture",  "2026-10-02", "2027-01-31",  0.15),
    ("Phase 2: Merger Zone Build",         "2027-02-01", "2027-07-31",  0.35),
    ("Phase 3: Testing & Migration Waves", "2027-08-01", "2027-11-30",  0.30),
    ("Phase 4: Pre-GoLive Readiness",      "2027-12-01", "2028-01-01",  0.05),
    ("Phase 5: Hypercare & Closure",       "2028-01-02", "2028-04-01",  0.10),
]

# CAPEX / additional costs — excluded from labour total
# Risk register references align to Trinity-CAM_Risk_Register.xlsx
CAPEX_ROWS = [
    ("Merger Zone DC / Cloud Infrastructure (Infosys-managed)",
     "TBC - to be approved at QG1",
     "Risk Register R002; subject to cloud-vs-hardware decision at QG1"),
    ("Software Licences — M365, ITSM, IAM, Security tooling (MZ)",
     "TBC - to be approved at QG1",
     "Per-user licence cost for 12,000 users; to be validated at architecture sign-off"),
    ("WAN/SD-WAN Network Connectivity — 48 Sites to MZ",
     "TBC - to be approved at QG1",
     "Risk Register R012; ISP circuit costs per region; include APAC/LATAM premium"),
    ("SAP SE Expert Services — System Copy & Client Separation",
     "TBC - to be approved at QG1",
     "Risk Register R001; SAP SE engagement for complex 1,800+ app SAP copy"),
    ("Risk Contingency Reserve — SAP Delay & MZ Build",
     "TBC - to be approved at QG1",
     "Risk Register R001, R002, R003; recommended 15% of total labour as contingency"),
    ("Cyber Insurance — Merger Zone & Data Migration Coverage",
     "TBC - to be approved at QG1",
     "Risk Register R006, R016; Infosys cyber liability insurance for MZ operations"),
    ("Independent Security Audit (GDPR / ISO 27001)",
     "TBC - to be approved at QG1",
     "Risk Register R006; pre-Phase 3 independent security assessment"),
    ("JCI IT Staff Retention Bonus Pool",
     "TBC - to be agreed in SPA",
     "Risk Register R017; retention incentives for key JCI SAP and IT roles through GoLive"),
    ("Travel & Expenses — 48 Sites, Multi-region (CAPEX)",
     "TBC - to be approved at QG1",
     "Site visits across EMEA, APAC, LATAM for migration waves; excluded from labour total"),
]

NOTES = [
    "1. Labour-only plan; CAPEX costs (infrastructure, licences, network, travel) listed separately above.",
    "2. JCI and Bosch internal resource effort is tracked in the schedule but not costed here.",
    "3. All EUR figures are indicative; subject to contract negotiation and QG1 budget board approval.",
    "4. Infosys rates are blended by role; actual contract rates to be validated at SOW signature.",
    "5. Standard billing day = 8 hours; standard working year = 220 days.",
    "6. Risk contingency of 15% of total labour is recommended for Risk Register items R001, R002, R003.",
    "7. Integration model (JCI > Merger Zone > Bosch) implies no Merger Zone tear-down post-GoLive; "
       "Infosys demobilisation schedule from GoLive +90 days.",
    "8. Based on: Trinity-CAM_Project_Schedule.xlsx (118 tasks, 2026-07-01 to 2028-04-01).",
    "9. Risk-aligned: Trinity-CAM_Risk_Register.xlsx (25 risks; top threat R001 SAP Complexity score=20).",
]


# =============================================================================
# XLSX WRITER
# =============================================================================

def write_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Plan"

    # --- Styles ---
    BLUE_DARK  = "003B6E"
    BLUE_MID   = "0066CC"
    BLUE_LIGHT = "EFF4FB"
    GREY       = "F2F2F2"
    SUBTOTAL   = "C6D4E8"
    WHITE      = "FFFFFF"

    def fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def font(color=None, bold=False, size=9, italic=False):
        return Font(
            color=color or "000000",
            bold=bold,
            size=size,
            italic=italic,
            name="Calibri"
        )

    def align(wrap=True, h="left", v="top"):
        return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin)

    # Column widths
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 17
    ws.column_dimensions["F"].width = 17

    row = 1

    # --- Title banner ---
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "TRINITY-CAM  |  IT CARVE-OUT COST PLAN  |  Labour Resource Plan"
    ws[f"A{row}"].fill   = fill(BLUE_DARK)
    ws[f"A{row}"].font   = font(WHITE, bold=True, size=13)
    ws[f"A{row}"].alignment = align(h="center", v="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # --- Metadata ---
    meta = [
        ("Project",       "Trinity-CAM — JCI Aircon IT Carve-out"),
        ("Seller",        "Johnson Controls International (JCI)"),
        ("Buyer",         "Robert Bosch GmbH"),
        ("Carve-out Model", "Integration (JCI IT → Merger Zone → Bosch IT)"),
        ("PMO Lead",      "KPMG"),
        ("IT Delivery",   "Infosys (Merger Zone Build, Operation & Migration)"),
        ("GoLive",        "2028-01-01"),
        ("Completion",    "2028-04-01  (90-day hypercare complete)"),
        ("Based on",      "Trinity-CAM_Project_Schedule.xlsx"),
        ("Risk-aligned",  "Trinity-CAM_Risk_Register.xlsx  (25 risks)"),
        ("Status",        "DRAFT — to be approved at QG1"),
    ]
    for label, value in meta:
        ws.cell(row, 1).value = label
        ws.cell(row, 2).value = value
        ws.merge_cells(f"B{row}:F{row}")
        for c in range(1, 7):
            ws.cell(row, c).fill = fill(GREY)
            ws.cell(row, c).font = font(size=8)
            ws.cell(row, c).alignment = align()
        row += 1

    row += 1  # spacer

    # --- Column headers ---
    headers = ["CATEGORY / RESOURCE", "DESCRIPTION", "TOTAL DAYS",
               "TOTAL HRS", "RATE (EUR/hr)", "TOTAL COST (EUR)"]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row, ci).value = h
        ws.cell(row, ci).fill  = fill(BLUE_MID)
        ws.cell(row, ci).font  = font(WHITE, bold=True, size=9)
        ws.cell(row, ci).alignment = align(h="center")
    ws.freeze_panes = f"A{row + 1}"
    header_row = row
    row += 1

    # --- Category blocks ---
    grand_total = 0
    all_cat_subtotals = []

    for cat_idx, cat in enumerate(CATEGORIES):
        # Category header row
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = cat["header"]
        ws[f"A{row}"].fill      = fill(BLUE_MID)
        ws[f"A{row}"].font      = font(WHITE, bold=True, size=9)
        ws[f"A{row}"].alignment = align()
        row += 1

        cat_total = 0
        cat_days  = 0
        first_data_row = row

        for ri, (resource, days, rate) in enumerate(cat["rows"]):
            hours = days * 8
            cost  = hours * rate
            cat_total += cost
            cat_days  += days
            bg = WHITE if ri % 2 == 0 else BLUE_LIGHT

            ws.cell(row, 1).value = f"    {resource}"
            ws.cell(row, 2).value = ""
            ws.cell(row, 3).value = days
            ws.cell(row, 4).value = hours
            ws.cell(row, 5).value = rate
            ws.cell(row, 6).value = cost

            for ci in range(1, 7):
                ws.cell(row, ci).fill = fill(bg)
                ws.cell(row, ci).font = font()
                ws.cell(row, ci).alignment = align(h="right" if ci >= 3 else "left")
                if ci in (3, 4, 5, 6):
                    ws.cell(row, ci).number_format = "#,##0"
            row += 1

        # Subtotal
        ws.cell(row, 1).value = f"Subtotal — {cat['header']}"
        ws.cell(row, 3).value = cat_days
        ws.cell(row, 6).value = cat_total
        for ci in range(1, 7):
            ws.cell(row, ci).fill = fill(SUBTOTAL)
            ws.cell(row, ci).font = font(bold=True, size=9)
            ws.cell(row, ci).alignment = align(h="right" if ci >= 3 else "left")
            if ci in (3, 6):
                ws.cell(row, ci).number_format = "#,##0"
        grand_total += cat_total
        all_cat_subtotals.append((cat["header"], cat_days, cat_total))
        row += 1

    row += 1  # spacer

    # --- Overall Project Total ---
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "OVERALL PROJECT LABOUR TOTAL"
    ws[f"A{row}"].fill = fill(BLUE_DARK)
    ws[f"A{row}"].font = font(WHITE, bold=True, size=10)
    ws[f"A{row}"].alignment = align(h="left")
    ws.cell(row, 6).value  = grand_total
    ws.cell(row, 6).fill   = fill(BLUE_DARK)
    ws.cell(row, 6).font   = font(WHITE, bold=True, size=10)
    ws.cell(row, 6).alignment = align(h="right")
    ws.cell(row, 6).number_format = "#,##0"
    row += 2

    # --- Cost breakdown by category ---
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COST BREAKDOWN BY CATEGORY"
    ws[f"A{row}"].fill = fill(BLUE_MID)
    ws[f"A{row}"].font = font(WHITE, bold=True, size=9)
    ws[f"A{row}"].alignment = align()
    row += 1

    for bi, (cat_name, cat_days, cat_cost) in enumerate(all_cat_subtotals):
        pct = (cat_cost / grand_total * 100) if grand_total else 0
        bg  = WHITE if bi % 2 == 0 else BLUE_LIGHT
        ws.cell(row, 1).value = cat_name
        ws.cell(row, 2).value = f"{pct:.1f}% of total"
        ws.cell(row, 3).value = cat_days
        ws.cell(row, 6).value = cat_cost
        for ci in range(1, 7):
            ws.cell(row, ci).fill      = fill(bg)
            ws.cell(row, ci).font      = font()
            ws.cell(row, ci).alignment = align(h="right" if ci >= 3 else "left")
            if ci in (3, 6):
                ws.cell(row, ci).number_format = "#,##0"
        row += 1

    row += 1

    # --- Cost breakdown by phase ---
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COST BREAKDOWN BY PHASE"
    ws[f"A{row}"].fill = fill(BLUE_MID)
    ws[f"A{row}"].font = font(WHITE, bold=True, size=9)
    ws[f"A{row}"].alignment = align()
    row += 1

    for pi, (phase, start, end, pct) in enumerate(PHASE_BREAKDOWN):
        phase_cost = round(grand_total * pct)
        bg = WHITE if pi % 2 == 0 else BLUE_LIGHT
        ws.cell(row, 1).value = phase
        ws.cell(row, 2).value = f"{start}  to  {end}"
        ws.cell(row, 3).value = f"{int(pct * 100)}%"
        ws.cell(row, 6).value = phase_cost
        for ci in range(1, 7):
            ws.cell(row, ci).fill      = fill(bg)
            ws.cell(row, ci).font      = font()
            ws.cell(row, ci).alignment = align(h="right" if ci >= 3 else "left")
            if ci == 6:
                ws.cell(row, ci).number_format = "#,##0"
        row += 1

    row += 1

    # --- CAPEX / additional costs ---
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CAPEX & ADDITIONAL COSTS  (excluded from labour total)"
    ws[f"A{row}"].fill = fill(BLUE_MID)
    ws[f"A{row}"].font = font(WHITE, bold=True, size=9)
    ws[f"A{row}"].alignment = align()
    row += 1

    # CAPEX column headers
    for ci, h in enumerate(["ITEM", "ESTIMATED COST (EUR)", "NOTES / RISK REFERENCE"], start=1):
        ws.cell(row, ci).value     = h
        ws.cell(row, ci).fill      = fill(GREY)
        ws.cell(row, ci).font      = font(bold=True, size=8)
        ws.cell(row, ci).alignment = align(h="center")
    ws.merge_cells(f"C{row}:F{row}")
    row += 1

    for ci_idx, (item, cost, notes_text) in enumerate(CAPEX_ROWS):
        bg = WHITE if ci_idx % 2 == 0 else BLUE_LIGHT
        ws.cell(row, 1).value = item
        ws.cell(row, 2).value = cost
        ws.merge_cells(f"C{row}:F{row}")
        ws.cell(row, 3).value = notes_text
        for c in range(1, 7):
            ws.cell(row, c).fill      = fill(bg)
            ws.cell(row, c).font      = font(size=8)
            ws.cell(row, c).alignment = align(h="left")
        row += 1

    row += 1

    # --- Notes ---
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "NOTES"
    ws[f"A{row}"].fill = fill(GREY)
    ws[f"A{row}"].font = font(bold=True, size=8)
    row += 1

    for note in NOTES:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = note
        ws[f"A{row}"].fill      = fill(GREY)
        ws[f"A{row}"].font      = font(size=8, italic=True)
        ws[f"A{row}"].alignment = align(h="left")
        row += 1

    wb.save(str(path))
    return grand_total


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print(f"[Trinity-CAM] Generating cost plan...")
    total = write_xlsx(OUTPUT)
    print(f"  Output: {OUTPUT}")
    print(f"\n  Labour Total:    EUR {total:,.0f}")
    print(f"  Contingency 15%: EUR {total * 0.15:,.0f}")
    print(f"  Total incl. res: EUR {total * 1.15:,.0f}")
    print(f"\n  Phase breakdown:")
    for phase, start, end, pct in PHASE_BREAKDOWN:
        print(f"    {phase:45s} {int(pct*100):3d}%   EUR {round(total*pct):>10,.0f}")
    print(f"\n[Trinity-CAM] Cost plan complete.")
