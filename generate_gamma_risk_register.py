#!/usr/bin/env python3
"""Generate Gamma risk register using the Bosch template."""

from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))

from openpyxl import load_workbook
from openpyxl.styles import Font


HERE = Path(__file__).parent
PROJECT_NAME = "Gamma"
OUTPUT_FOLDER_NAME = "Gamma v1.0"
DOCUMENT_VERSION = "Version 1.0 - Initial Baseline"
REPORT_DATE = "05/04/2026"
DOCUMENT_OWNER = "EY"
TEMPLATE_PATH = HERE / "Reference" / "BD_Risk-Register_template_en_V1.0_Dec2023.xlsx"
OUTPUT_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Risk_Register.xlsx"

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)


# Evaluated but intentionally excluded from active risk rows because they are not
# materially applicable to this IT-only cloud carve-out: Engineering,
# Manufacturing, Market & Competitors, Raw Materials.

RISKS = [
    {
        "id": 1,
        "category": "Strategy & Portfolio",
        "cause": "Antitrust due diligence for the 50/50 Gamma JV between Robert Bosch China and Alibaba is still open, and legal has only allowed a restricted clean-team planning model during Phase 0 and Phase 1.",
        "event": "Legal confirmation on JV perimeter and approval mechanics arrives after the 29/10/2026 QG1 concept gate, forcing redesign of Day 1 scope, transition governance, and the interim service model.",
        "effect": "Concept rework delays Phase 2 build entry by 2 - 4 weeks, pushes vendor commitments, and increases programme management cost.",
        "event_date": "20/10/2026",
        "owner": "Bosch China Legal Lead",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "70%",
        "type": "threat",
        "qualitative_impact": "QG1 and downstream build mobilization are directly exposed to legal timing.",
        "eur_current": 180000,
        "eur_future": 420000,
        "strategy": "Mitigate",
        "measure": "Run a weekly legal dependency review, maintain an assumptions log approved by legal, prepare a fallback Day 1 perimeter by 30/09/2026, and pre-draft alternate governance approvals before QG1.",
        "due_date": "15/10/2026",
        "status": "in progress",
        "notes": "Top strategic risk because legal uncertainty is a true schedule dependency in Gamma.",
    },
    {
        "id": 2,
        "category": "Legal & Compliance",
        "cause": "High confidentiality and the very limited aware stakeholder group require restricted distribution of designs, inventories, and migration plans across Bosch China, Alibaba, and vendors.",
        "event": "Sensitive JV information is shared beyond the approved clean-team perimeter during discovery or design workshops, triggering legal escalation and an immediate freeze on collaboration sessions.",
        "effect": "Planning pauses, trust between parties deteriorates, and the project faces reputational and compliance consequences.",
        "event_date": "30/09/2026",
        "owner": "EY PMO Lead",
        "source": "Formal Risk Review",
        "impact": "Very High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "A confidentiality incident would halt planning at the most sensitive stage of the programme.",
        "eur_current": 150000,
        "eur_future": 300000,
        "strategy": "Mitigate",
        "measure": "Use named-access workspaces only, maintain an approved audience register, watermark pack distributions, and require legal review before any stakeholder expansion.",
        "due_date": "01/09/2026",
        "status": "in progress",
        "notes": "This risk is unique to Gamma because awareness is intentionally constrained.",
    },
    {
        "id": 3,
        "category": "Schedule",
        "cause": "Because only a small clean team can validate assumptions, discovery across 5 sites, 250 users, and 20 applications depends on serial rather than parallel interviews and reviews.",
        "event": "Inventory confirmation and dependency mapping remain incomplete by 03/09/2026, causing QG0 output quality gaps and compressing the concept phase.",
        "effect": "Schedule float is consumed before build starts, and downstream test preparation is forced into a shorter window.",
        "event_date": "03/09/2026",
        "owner": "EY PMO Lead",
        "source": "Status Meeting",
        "impact": "High",
        "probability": "70%",
        "type": "threat",
        "qualitative_impact": "Early timeline compression creates persistent pressure on QG1, QG2 and QG3, and QG4 readiness.",
        "eur_current": 90000,
        "eur_future": 180000,
        "strategy": "Mitigate",
        "measure": "Lock a discovery calendar for the clean team, prioritize critical infrastructure towers first, and use staged inventory validation with weekly variance closure.",
        "due_date": "28/08/2026",
        "status": "in progress",
        "notes": "Gamma schedule risk is driven more by restricted access than by raw scale.",
    },
    {
        "id": 4,
        "category": "Resources",
        "cause": "The same limited group of legal, infrastructure, and leadership staff are serving as clean-team reviewers, approvers, and subject matter experts during the first two phases.",
        "event": "Key clean-team members become unavailable during concept sign-off or build mobilization, leaving no approved back-up resources for decisions and validations.",
        "effect": "Decision latency increases, rework grows, and issue resolution extends into the critical path.",
        "event_date": "29/10/2026",
        "owner": "EY PMO Lead",
        "source": "Status Meeting",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "A small approved resource pool creates single-person dependency risk across multiple gates.",
        "eur_current": 80000,
        "eur_future": 160000,
        "strategy": "Mitigate",
        "measure": "Nominate alternates who can be pre-cleared by legal, reserve capacity for clean-team SMEs, and escalate conflicting assignments in weekly steerco.",
        "due_date": "15/09/2026",
        "status": "not started",
        "notes": "Resource concentration is high despite the smaller programme size.",
    },
    {
        "id": 5,
        "category": "Technology, R&D",
        "cause": "The Bosch Cloud business relies on shared Bosch China infrastructure services, and 20 in-scope applications may still have hidden dependencies on shared network, identity, monitoring, backup, or certificates.",
        "event": "Previously undocumented shared service dependencies are discovered during build or SIT, requiring late redesign of the target hosting and connectivity model.",
        "effect": "Phase 2 build expands, integration testing slips, and GoLive readiness is weakened.",
        "event_date": "15/12/2026",
        "owner": "Bosch China Infrastructure Lead",
        "source": "Formal Risk Review",
        "impact": "Very High",
        "probability": "70%",
        "type": "threat",
        "qualitative_impact": "Shared infrastructure blind spots are the largest technical threat to the Gamma build plan.",
        "eur_current": 200000,
        "eur_future": 350000,
        "strategy": "Mitigate",
        "measure": "Run dependency walk-throughs by tower, require evidence for each of the 20 applications, and validate shared-service assumptions before 30/10/2026 build start.",
        "due_date": "23/10/2026",
        "status": "not started",
        "notes": "The pure infrastructure scope concentrates risk in shared platform dependencies rather than SAP complexity.",
    },
    {
        "id": 6,
        "category": "Security & Data Protection",
        "cause": "Identity and privileged access are being redesigned while the project is still operating with restricted stakeholder access and partially shared Bosch China administration patterns.",
        "event": "JV users or administrators retain inherited Bosch China privileges after cutover, leaving over-entitled access to shared directories, service consoles, or operational tooling.",
        "effect": "Unauthorized access, audit findings, and manual remediation during hypercare undermine Day 1 stability.",
        "event_date": "01/02/2027",
        "owner": "Bosch China Security Lead",
        "source": "Formal Risk Review",
        "impact": "Very High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "Identity leakage across the shared-to-JV boundary would create both security and operational disruption.",
        "eur_current": 160000,
        "eur_future": 320000,
        "strategy": "Mitigate",
        "measure": "Define role matrices early, certify privileged access before QG2 and QG3, and execute an independent access recertification before QG4.",
        "due_date": "15/01/2027",
        "status": "not started",
        "notes": "Security design is central to Gamma because the future JV is jointly managed.",
    },
    {
        "id": 7,
        "category": "Security & Data Protection",
        "cause": "Monitoring, logging, backups, and shared file repositories may still contain mixed Bosch China and Gamma JV operational data after separation design is approved.",
        "event": "Data segregation gaps are discovered in logs, backups, or archive repositories during testing or after GoLive, exposing non-Gamma records to the JV support team.",
        "effect": "Compliance breach, emergency cleanup activity, and delayed acceptance of the target operational model.",
        "event_date": "20/01/2027",
        "owner": "Data Protection Officer",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "Operational datasets are often missed because they sit outside core application inventories.",
        "eur_current": 120000,
        "eur_future": 250000,
        "strategy": "Mitigate",
        "measure": "Include logs, backups, and archives in the inventory, define explicit segregation checks, and add data-protection sign-off to the QG4 evidence pack.",
        "due_date": "12/01/2027",
        "status": "not started",
        "notes": "Gamma has lower application volume but still meaningful operational data-protection exposure.",
    },
    {
        "id": 8,
        "category": "Quality",
        "cause": "UAT is intentionally limited to a narrow set of approved representatives because broader business awareness is still constrained by legal confidentiality.",
        "event": "The reduced UAT group fails to cover critical business and operational scenarios across the 20 applications and the core infrastructure workflows.",
        "effect": "Defects surface during final readiness or early hypercare, increasing incident volume and reducing user confidence in the cutover.",
        "event_date": "30/12/2026",
        "owner": "EY Test Lead",
        "source": "Status Meeting",
        "impact": "High",
        "probability": "70%",
        "type": "threat",
        "qualitative_impact": "Testing depth is constrained by confidentiality rather than by lack of planning discipline.",
        "eur_current": 95000,
        "eur_future": 170000,
        "strategy": "Mitigate",
        "measure": "Expand scenario coverage through scripted walkthroughs, supplement business UAT with SME dry runs, and require defect trend review before QG2 and QG3 approval.",
        "due_date": "18/12/2026",
        "status": "not started",
        "notes": "This is one of the strongest links between confidentiality and delivery quality.",
    },
    {
        "id": 9,
        "category": "Customers",
        "cause": "The Bosch Cloud business provides infrastructure services, so the 250 users and dependent service consumers have low tolerance for outages during identity, network, and application cutover.",
        "event": "Service continuity fails on 01/02/2027 because network, authentication, or shared tooling do not transition cleanly for the first working day in the new environment.",
        "effect": "Customer-facing service degradation, urgent rollback pressure, and reputational damage for both JV shareholders.",
        "event_date": "01/02/2027",
        "owner": "Bosch China Operations Lead",
        "source": "Stakeholder",
        "impact": "Very High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "Day 1 stability is a visible business outcome and a joint reputational event for Bosch and Alibaba.",
        "eur_current": 220000,
        "eur_future": 400000,
        "strategy": "Mitigate",
        "measure": "Rehearse the cutover, verify critical service chains, pre-stage command center runbooks, and maintain validated rollback criteria through the final readiness window.",
        "due_date": "29/01/2027",
        "status": "not started",
        "notes": "Customer/service continuity risk is elevated because Gamma is infrastructure service heavy.",
    },
    {
        "id": 10,
        "category": "Supply Chain",
        "cause": "Network circuits, DNS changes, certificates, and third-party enablement for 5 sites and 20 applications still depend on external providers and approval lead times.",
        "event": "Provider lead times exceed plan, leaving one or more sites or shared services without the required certificates, routing, or external connectivity for integrated testing and cutover.",
        "effect": "SIT defects remain open, pre-GoLive completion slips, and manual workarounds increase operational risk.",
        "event_date": "11/01/2027",
        "owner": "Bosch China Infrastructure Lead",
        "source": "Status Meeting",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "Third-party technical lead times directly influence QG2 and QG3 and QG4 readiness.",
        "eur_current": 70000,
        "eur_future": 110000,
        "strategy": "Mitigate",
        "measure": "Submit provider requests early, track certificates and circuit milestones weekly, and define temporary fallback routing for priority services.",
        "due_date": "30/11/2026",
        "status": "in progress",
        "notes": "A smaller site count lowers volume but not dependency sensitivity.",
    },
    {
        "id": 11,
        "category": "Budget",
        "cause": "If legal decisions or third-party approvals arrive late, Gamma may need extended dual running, extra transition services, and more PMO effort beyond the initial v1.0 baseline.",
        "event": "Temporary seller-hosted services and duplicated environments stay active beyond the planned build and readiness windows.",
        "effect": "The cost plan requires an uplift, and management approval is needed for unplanned run costs during late build or hypercare.",
        "event_date": "25/01/2027",
        "owner": "EY Finance Lead",
        "source": "Formal Risk Review",
        "impact": "Moderate",
        "probability": "70%",
        "type": "threat",
        "qualitative_impact": "Budget pressure is more likely than severe because the project is smaller, but dual running can still add material cost.",
        "eur_current": 110000,
        "eur_future": 140000,
        "strategy": "Mitigate",
        "measure": "Track transition-service demand monthly, define a clear exit baseline in the cost plan, and escalate any extension request through steerco within five working days.",
        "due_date": "15/01/2027",
        "status": "not started",
        "notes": "This risk will directly inform the Gamma cost-plan contingency.",
    },
    {
        "id": 12,
        "category": "Ecosystems & Ethics",
        "cause": "The future operating model is a 50/50 jointly managed JV, which can create split accountability and slower decisions if responsibilities are not explicit by QG1 and QG4.",
        "event": "Bosch China and Alibaba disagree on service ownership, priority defects, or acceptance criteria during build or final readiness.",
        "effect": "Critical decisions stall, support ownership remains ambiguous, and the post-GoLive run model is weaker than planned.",
        "event_date": "25/01/2027",
        "owner": "Gamma Steering Committee Chair",
        "source": "Stakeholder",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "Joint governance ambiguity can turn minor operational issues into gate-blocking issues.",
        "eur_current": 85000,
        "eur_future": 160000,
        "strategy": "Mitigate",
        "measure": "Document service ownership in the operating model, pre-approve decision rights, and require both parties to sign the final readiness acceptance matrix.",
        "due_date": "20/01/2027",
        "status": "not started",
        "notes": "This is a Gamma-specific governance risk driven by the 50/50 JV structure.",
    },
    {
        "id": 13,
        "category": "Legal & Compliance",
        "cause": "Several of the 20 applications and supporting services may have contracts, licences, or processing terms that require change-of-control approval for the new JV structure.",
        "event": "One or more providers do not approve the transfer or require terms that are not acceptable before cutover.",
        "effect": "Scope exceptions, delayed cutover readiness, or unplanned replacement work are required for affected services.",
        "event_date": "12/12/2026",
        "owner": "Procurement Lead",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "A small number of blocked providers could still create a large operational gap because the application landscape is compact.",
        "eur_current": 100000,
        "eur_future": 180000,
        "strategy": "Mitigate",
        "measure": "Complete contract review during concept phase, contact affected vendors by 31/10/2026, and define alternatives for any provider without timely consent.",
        "due_date": "30/11/2026",
        "status": "not started",
        "notes": "The smaller portfolio makes early vendor outreach feasible and essential.",
    },
    {
        "id": 14,
        "category": "Stakeholder Relations & Public Affairs",
        "cause": "Because only limited associates are aware of the carve-out and JV plan, rumors or partial disclosures can spread without an approved narrative or response plan.",
        "event": "Uncontrolled internal discussion reaches broader teams, suppliers, or customers before formal communication windows are approved.",
        "effect": "Stakeholder anxiety rises, cooperation drops, and management spends time on damage control instead of execution.",
        "event_date": "30/11/2026",
        "owner": "Communications Lead",
        "source": "Stakeholder",
        "impact": "Moderate",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "Narrative gaps can create avoidable friction even when technical work remains on plan.",
        "eur_current": 40000,
        "eur_future": 70000,
        "strategy": "Mitigate",
        "measure": "Prepare confidential Q and A packs, align response owners, and release controlled stakeholder messages in phases approved by legal.",
        "due_date": "15/10/2026",
        "status": "not started",
        "notes": "Communication control is a management discipline risk rather than a pure technical issue.",
    },
    {
        "id": 15,
        "category": "Technology, R&D",
        "cause": "The new JV operating model depends on monitoring, incident routing, service desk procedures, and support tooling being ready before the 01/02/2027 cutover.",
        "event": "Operational tooling is only partially configured by the time SIT closes, leaving alert routing, ticket triage, or backup reporting incomplete at GoLive.",
        "effect": "Incident response slows during hypercare and early production issues take longer to diagnose and resolve.",
        "event_date": "25/01/2027",
        "owner": "Operations Lead",
        "source": "Status Meeting",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "Service-management immaturity would turn manageable technical issues into prolonged outages.",
        "eur_current": 90000,
        "eur_future": 150000,
        "strategy": "Mitigate",
        "measure": "Treat operational tooling as a gateable deliverable, run command-center rehearsals, and include service desk proof points in the QG4 evidence pack.",
        "due_date": "19/01/2027",
        "status": "not started",
        "notes": "This risk aligns directly with schedule task 31 and the final readiness phase.",
    },
    {
        "id": 16,
        "category": "Schedule",
        "cause": "Gamma has a firm GoLive on 01/02/2027 and a five-day final-readiness window after QG4, leaving limited tolerance for residual defects, unclosed actions, or incomplete migrations.",
        "event": "One or more QG4 prerequisites remain open on 25/01/2027, leaving insufficient time to close defects, re-run checks, and secure business sign-off before GoLive.",
        "effect": "Management must choose between delaying GoLive or accepting elevated operational and governance risk.",
        "event_date": "25/01/2027",
        "owner": "EY PMO Lead",
        "source": "Formal Risk Review",
        "impact": "Very High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "This is the principal gate risk tied to the QG4 to GoLive buffer.",
        "eur_current": 140000,
        "eur_future": 220000,
        "strategy": "Mitigate",
        "measure": "Run a daily readiness burn-down from 12/01/2027, freeze non-essential changes, and escalate any red item within 24 hours to steerco.",
        "due_date": "22/01/2027",
        "status": "not started",
        "notes": "Directly linked to the mandatory QG4 and final-readiness rules in the schedule skill.",
    },
    {
        "id": 17,
        "category": "Technology, R&D",
        "cause": "Gamma has no SAP and only about 20 applications in scope, which reduces interface complexity and allows deeper engineering attention on the infrastructure and identity stack.",
        "event": "The smaller application footprint enables earlier completion of SIT and a lower open-defect count than planned before QG2 and QG3.",
        "effect": "Build stabilizes sooner, QG4 readiness improves, and management gains schedule buffer for legal or supplier-driven issues.",
        "event_date": "11/01/2027",
        "owner": "Bosch China Infrastructure Lead",
        "source": "Other",
        "impact": "Moderate",
        "probability": "70%",
        "type": "opportunity",
        "qualitative_impact": "Gamma can benefit from a relatively compact application estate if execution stays disciplined.",
        "eur_current": 60000,
        "eur_future": 120000,
        "strategy": "Exploit",
        "measure": "Prioritize early application classification, keep scope discipline, and redirect saved effort into security hardening and operational readiness.",
        "due_date": "18/12/2026",
        "status": "not started",
        "notes": "First positive risk - smaller scope can create usable contingency.",
    },
    {
        "id": 18,
        "category": "Budget",
        "cause": "The carve-out is pure infrastructure services, which creates an opportunity to standardize cloud operations, remove legacy duplications, and consolidate vendors in the new JV run model.",
        "event": "The project identifies vendor and platform simplifications during build and hypercare that can be locked into the Day 2 operating model without additional delivery risk.",
        "effect": "The JV reduces recurring run cost and simplifies the support footprint in the first three years after cutover.",
        "event_date": "07/05/2027",
        "owner": "EY Finance Lead",
        "source": "Other",
        "impact": "Moderate",
        "probability": "50%",
        "type": "opportunity",
        "qualitative_impact": "A focused infrastructure scope can unlock meaningful steady-state savings if captured intentionally.",
        "eur_current": 50000,
        "eur_future": 250000,
        "strategy": "Enhance",
        "measure": "Track standardization candidates during build, validate savings in hypercare, and include approved simplifications in the post-GoLive handover plan.",
        "due_date": "30/04/2027",
        "status": "not started",
        "notes": "Second positive risk - operating-model simplification can offset part of the carve-out cost.",
    },
    {
        "id": 19,
        "category": "Intellectual Property",
        "cause": "Bosch Cloud automation scripts, runbooks, and operational patterns may need to be shared selectively with the new JV team while confidentiality and ownership boundaries are still being defined.",
        "event": "Proprietary Bosch operational content is transferred to the JV without the right usage restrictions, inventory, or approval trail.",
        "effect": "Intellectual-property exposure creates legal cleanup work, governance friction, and re-documentation effort.",
        "event_date": "23/04/2027",
        "owner": "Bosch China Legal Lead",
        "source": "Formal Risk Review",
        "impact": "Moderate",
        "probability": "30%",
        "type": "threat",
        "qualitative_impact": "The risk is narrower than in product carve-outs but still relevant because Gamma is infrastructure-led.",
        "eur_current": 45000,
        "eur_future": 90000,
        "strategy": "Mitigate",
        "measure": "Classify operational artefacts, define approved sharing boundaries, and require legal approval for any proprietary automation transferred into the JV runbook set.",
        "due_date": "31/03/2027",
        "status": "not started",
        "notes": "Gamma IP exposure is concentrated in operational know-how rather than product IP.",
    },
    {
        "id": 20,
        "category": "Resources",
        "cause": "Because broader user exposure is delayed, training and support preparation may remain lighter than the actual demand created at GoLive across 250 users and 5 sites.",
        "event": "Incident and support demand in the first four weeks of hypercare exceeds plan, and the command center does not have enough skilled staff to sustain response times.",
        "effect": "Ticket backlogs grow, user confidence drops, and the hypercare period risks extension beyond the planned 90 days.",
        "event_date": "20/02/2027",
        "owner": "Operations Lead",
        "source": "Status Meeting",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative_impact": "Support under-capacity is a realistic early-life risk when awareness and training are intentionally constrained.",
        "eur_current": 75000,
        "eur_future": 100000,
        "strategy": "Mitigate",
        "measure": "Model hypercare demand from pilot issues, reserve surge support capacity, and refresh training packs before 31/01/2027.",
        "due_date": "25/01/2027",
        "status": "not started",
        "notes": "Links directly to the 90-day stabilization-only phase in the schedule baseline.",
    },
]


def populate_risk_register(ws, risks: list[dict]) -> None:
    start_row = 5
    for index, risk in enumerate(risks):
        row = start_row + index
        ws.cell(row, 2).value = risk["id"]
        ws.cell(row, 3).value = REPORT_DATE
        ws.cell(row, 4).value = risk["category"]
        ws.cell(row, 5).value = risk["cause"]
        ws.cell(row, 6).value = risk["event"]
        ws.cell(row, 7).value = risk["effect"]
        ws.cell(row, 8).value = risk["event_date"]
        ws.cell(row, 9).value = risk["owner"]
        ws.cell(row, 10).value = risk["source"]
        ws.cell(row, 12).value = risk["impact"]
        ws.cell(row, 14).value = risk["probability"]
        ws.cell(row, 16).value = risk["type"]
        ws.cell(row, 18).value = risk["qualitative_impact"]
        ws.cell(row, 19).value = risk["eur_current"]
        ws.cell(row, 20).value = risk["eur_future"]
        ws.cell(row, 22).value = risk["strategy"]
        ws.cell(row, 23).value = risk["measure"]
        ws.cell(row, 24).value = risk["due_date"]
        ws.cell(row, 26).value = risk["status"]
        ws.cell(row, 27).value = REPORT_DATE
        ws.cell(row, 28).value = risk["impact"]
        ws.cell(row, 30).value = risk["probability"]
        ws.cell(row, 35).value = risk["notes"]

        ws.cell(row, 13).value = f'=_xlfn.IFNA(VLOOKUP(L{row},$D$182:$E$186,2,FALSE),"")'
        ws.cell(row, 15).value = f'=_xlfn.IFNA(VLOOKUP(N{row},$D$189:$E$193,2,FALSE),"")'
        ws.cell(row, 17).value = f'=M{row}*O{row}'
        ws.cell(row, 29).value = f'=_xlfn.IFNA(VLOOKUP(AB{row},$D$182:$E$186,2,FALSE),"")'
        ws.cell(row, 31).value = f'=_xlfn.IFNA(VLOOKUP(AD{row},$D$189:$E$193,2,FALSE),"")'
        ws.cell(row, 32).value = f'=AC{row}'
        ws.cell(row, 33).value = f'=AE{row}'
        ws.cell(row, 34).value = f'=AF{row}*AG{row}'


def clear_risk_register_area(ws) -> None:
    for row in range(5, 140):
        for column in range(2, 36):
            ws.cell(row, column).value = None


def fix_matrix_font(matrix_sheet) -> None:
    for row in matrix_sheet.iter_rows():
        for cell in row:
            if not cell.fill or not cell.fill.start_color:
                continue
            fill_color = str(cell.fill.start_color.rgb)
            if fill_color in ("00FFFFFF00", "00FFFFCC", "FFFFFF00", "FFFFFFCC"):
                cell.font = Font(color="000000", bold=cell.font.bold if cell.font else False)


def main() -> None:
    print(f"[{PROJECT_NAME}] Generating risk register")
    print(f"  Version: {DOCUMENT_VERSION}")
    print(f"  Output folder: {OUTPUT_FOLDER_NAME}")
    print(f"  Risks: {len(RISKS)}")

    if not TEMPLATE_PATH.exists():
        print(f"ERROR: Template not found at {TEMPLATE_PATH}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE_PATH)
    info = wb["Info"]
    info["C4"].value = "GAM-RR-v1.0"
    info["C5"].value = OUTPUT_PATH.name
    info["C6"].value = "Gamma - Bosch Cloud carve-out to 50/50 JV"
    info["C7"].value = "Gamma v1.0"
    info["C8"].value = DOCUMENT_OWNER

    rr_sheet = wb["Risk Register"]
    clear_risk_register_area(rr_sheet)
    populate_risk_register(rr_sheet, RISKS)
    fix_matrix_font(wb["Matrix "])
    wb.save(OUTPUT_PATH)

    print(f"  Output: {OUTPUT_PATH}")
    print("  Threats and opportunities aligned to Gamma schedule baseline")


if __name__ == "__main__":
    main()