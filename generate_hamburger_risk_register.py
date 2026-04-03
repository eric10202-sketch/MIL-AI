"""
generate_hamburger_risk_register.py
Generates Hamburger_Risk_Register.xlsx by populating Risk_analysis_template.xlsx
with Hamburger project risk data.

Project Hamburger: Solar Energy Business Carve-Out (Stand Alone model)
Seller: Robert Bosch GmbH | Buyer: Undisclosed Buyer
17 worldwide sites | 2600 IT users
Start: 01 Apr 2026 | GoLive: 01 Dec 2026 | Completion: 30 May 2027
Lead: Erik Ho (BD/MIL-ICC)

Usage:
    "C:/Program Files/px/python.exe" generate_hamburger_risk_register.py

Requirements: pip install openpyxl
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))

import argparse
import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


HAMBURGER_RISKS = [
    {
        'Category': 'ScR - Schedule',
        'Description': 'GoLive (01 Dec 2026) at risk due to complexity of 17-site global cutover with 2600 users',
        'Root Cause': 'Parallel workstreams across EMEA, Americas and APAC; time-zone coordination overhead; limited cutover window',
        'Probability': 4, 'Impact': 5,
        'Mitigation': 'Phase cutover by region; EMEA first wave, then Americas and APAC; dedicate regional cutover leads; daily SteerCo from QG3',
        'Owner': 'Erik Ho (BD/MIL-ICC), IT PM',
        'Target Date': '30-Nov-26',
        'Status': 'Red',
        'Notes': 'Hard legal separation date tied to 01 Dec. No slip possible without board approval.'
    },
    {
        'Category': 'SR - Scope',
        'Description': 'SAP / ERP separation significantly more complex than initial inventory suggests',
        'Root Cause': 'Multiple shared SAP mandants; global Solar data entangled with Bosch corporate ERP; extensive custom ABAP',
        'Probability': 4, 'Impact': 5,
        'Mitigation': 'Start ERP separation assessment immediately (Jul 2026); engage dedicated SAP migration partner; greenfield decision by QG2',
        'Owner': 'Bosch ERP / SAP Team, Erik Ho (BD/MIL-ICC)',
        'Target Date': '31-Aug-26',
        'Status': 'Red',
        'Notes': 'ERP is on the critical path. Any delay directly impacts GoLive. Escalation to SteerCo required.'
    },
    {
        'Category': 'SR - Scope',
        'Description': 'Buyer identity unknown – final IT handover scope and integration requirements unclear',
        'Root Cause': 'Buyer undisclosed for legal reasons; buyer IT readiness and architecture unknown',
        'Probability': 3, 'Impact': 5,
        'Mitigation': 'Design Solar IT as fully self-sufficient Stand Alone; no buyer-side dependencies in baseline plan; update scope at buyer disclosure',
        'Owner': 'Erik Ho (BD/MIL-ICC), Legal, Steering Committee',
        'Target Date': '30-Jun-26',
        'Status': 'Amber',
        'Notes': 'Stand Alone model mitigates buyer-dependency risk. Inherent uncertainty accepted.'
    },
    {
        'Category': 'ScR - Schedule',
        'Description': 'Active Directory migration for 2600 users across 17 sites delayed due to tooling or forest complexity',
        'Root Cause': 'AD migration at scale requires extensive testing; DNS, GPO, and trust configuration errors at scale',
        'Probability': 3, 'Impact': 4,
        'Mitigation': 'Start AD build Sep 2026; pilot 2 sites first; 2x dress rehearsals; AD locked by QG3; dedicated AD team',
        'Owner': 'Bosch AD Team, IT PM',
        'Target Date': '30-Oct-26',
        'Status': 'Amber',
        'Notes': 'AD is foundational. Pilot critical. No backout option once cutover begins.'
    },
    {
        'Category': 'RR - Resource',
        'Description': 'Insufficient Bosch IT resources to manage 17-site global programme simultaneously',
        'Root Cause': 'Bosch IT teams have competing priorities; carved-out Solar team capacity limited during transition',
        'Probability': 3, 'Impact': 4,
        'Mitigation': 'Resource plan confirmed by QG1; third-party SI engagement for site execution; outsource device rollout',
        'Owner': 'Erik Ho (BD/MIL-ICC), Bosch CIO',
        'Target Date': '30-Jun-26',
        'Status': 'Amber',
        'Notes': 'Resource gap most acute in Phases 3-4. Early contracting required.'
    },
    {
        'Category': 'BtR - Budget',
        'Description': 'Programme cost exceeds initial estimate due to global WAN/network provisioning complexity',
        'Root Cause': '17 sites requiring new WAN circuits; lead times 8-16 weeks per region; APAC/Americas pricing higher than EMEA',
        'Probability': 3, 'Impact': 4,
        'Mitigation': 'Issue global WAN RFQ by QG1; obtain firm quotes with lead times; plan parallel ordering',
        'Owner': 'Finance, Bosch Procurement, WAN Provider',
        'Target Date': '30-Jun-26',
        'Status': 'Amber',
        'Notes': 'Budget contingency TBC – to be approved at QG1. Conservative estimate +25%.'
    },
    {
        'Category': 'LR - Legality',
        'Description': 'Data residency and privacy regulation compliance gaps across multiple jurisdictions (GDPR, US, APAC)',
        'Root Cause': '17 sites span multiple legal jurisdictions; Solar data flows must comply with GDPR, US data law, and APAC regulations',
        'Probability': 3, 'Impact': 4,
        'Mitigation': 'Engage regional legal counsel per jurisdiction by QG2; data residency architecture confirmed before build',
        'Owner': 'Legal, Bosch CISO, Erik Ho (BD/MIL-ICC)',
        'Target Date': '31-Aug-26',
        'Status': 'Amber',
        'Notes': 'Multi-jurisdictional compliance is inherently complex. Early legal engagement critical.'
    },
    {
        'Category': 'ScR - Schedule',
        'Description': 'WAN circuit provisioning delayed at one or more of the 17 sites, blocking IT infrastructure build',
        'Root Cause': 'Carrier lead times 8-16 weeks; site access permissions; regulatory approvals in some regions',
        'Probability': 3, 'Impact': 4,
        'Mitigation': 'Order WAN circuits by QG1; use 4G/5G backup connectivity as interim bridge; track weekly per site',
        'Owner': 'Bosch Infra, WAN Provider',
        'Target Date': '30-Jun-26',
        'Status': 'Amber',
        'Notes': 'WAN is on critical path for site readiness. Early ordering is mandatory.'
    },
    {
        'Category': 'SR - Scope',
        'Description': 'M365 / Azure tenant provisioning for 2600 mailboxes underestimated in complexity',
        'Root Cause': 'Large tenant migration; mailbox migration sequencing across time zones; SharePoint and Teams data volume',
        'Probability': 2, 'Impact': 4,
        'Mitigation': 'Begin M365 pre-ordering and test tenant in May 2026; use tiered migration waves by region',
        'Owner': 'Bosch Azure Team, Erik Ho (BD/MIL-ICC)',
        'Target Date': '30-Jun-26',
        'Status': 'Green',
        'Notes': 'M365 migration at 2600 users is large but well-understood. Phased approach required.'
    },
    {
        'Category': 'CR - Change',
        'Description': 'Change management resistance from 2600 Solar users impacting Day 1 adoption',
        'Root Cause': 'Global user base across 17 sites; cultural differences; limited communication in local languages',
        'Probability': 2, 'Impact': 4,
        'Mitigation': 'Global change and communications plan by QG2; local language change materials; regional change champions',
        'Owner': 'Comms + Regional IT, Business Leads',
        'Target Date': '31-Aug-26',
        'Status': 'Green',
        'Notes': 'Change adoption risk is manageable with structured plan and champions network.'
    },
    {
        'Category': 'BtR - Budget',
        'Description': 'Licence change-of-control costs significantly exceed budget across Solar application portfolio',
        'Root Cause': 'ISVs enforce new licence terms on change of control; some vendors use carve-out as renegotiation leverage',
        'Probability': 2, 'Impact': 4,
        'Mitigation': 'Complete licence audit with CoC clause review by QG2; negotiate amendments early',
        'Owner': 'Bosch Procurement, Legal',
        'Target Date': '31-Aug-26',
        'Status': 'Green',
        'Notes': 'Estimate 30-50% licence cost uplift for key applications. Build into budget.'
    },
    {
        'Category': 'QR - Quality',
        'Description': 'UAT failures due to incomplete ERP / SAP regression testing across Solar business processes',
        'Root Cause': 'ERP test cases insufficient; business users unavailable for global UAT coordination',
        'Probability': 2, 'Impact': 4,
        'Mitigation': 'ERP test plan locked at QG3; business sign-off on UAT cases by 1-Nov-2026; regional UAT coordinators',
        'Owner': 'Test Lead, Business Leads, Bosch ERP / SAP Team',
        'Target Date': '20-Nov-26',
        'Status': 'Green',
        'Notes': 'ERP UAT is the highest-risk test stream. Dedicated ERP test lead required.'
    },
    {
        'Category': 'RR - Resource',
        'Description': 'Key Solar IT staff retained by Bosch after separation, creating knowledge gap for Buyer',
        'Root Cause': 'TUPE/transfer rules vary by country; some IT staff may not transfer with Solar business',
        'Probability': 2, 'Impact': 4,
        'Mitigation': 'HR IT mapping by QG1; retention agreements for key knowledge holders; knowledge transfer plan from Phase 5',
        'Owner': 'Bosch HR, Erik Ho (BD/MIL-ICC), Legal',
        'Target Date': '30-Jun-26',
        'Status': 'Green',
        'Notes': '17 sites across multiple employment law jurisdictions. Country-specific HR review required.'
    },
    {
        'Category': 'CR - Change',
        'Description': 'Data migration incomplete or inconsistent for Solar business data by Day 1 across 17 sites',
        'Root Cause': 'Data classification incomplete; high data volume across global sites; migration tooling untested at scale',
        'Probability': 2, 'Impact': 4,
        'Mitigation': 'Data ownership rules locked by QG2; phased data migration rehearsals per region; dress rehearsal mandatory',
        'Owner': 'Bosch IT, Data Owner, IT PM',
        'Target Date': '30-Nov-26',
        'Status': 'Green',
        'Notes': 'Solar data is commercially sensitive. Zero tolerance for data loss.'
    },
    {
        'Category': 'ScR - Schedule',
        'Description': 'TSA exit extends beyond planned May 2027 completion date',
        'Root Cause': 'TSA service definitions unclear; Buyer IT readiness delayed; some services harder to exit than planned',
        'Probability': 2, 'Impact': 3,
        'Mitigation': 'TSA catalogue locked at QG2; weekly TSA governance reviews from Day 1; maximum TSA duration contractually limited',
        'Owner': 'Bosch IT, Buyer IT, Erik Ho (BD/MIL-ICC)',
        'Target Date': '31-May-27',
        'Status': 'Green',
        'Notes': 'TSA overrun extends Bosch carver support costs. Maximum 6-month TSA contractually enforced.'
    },
    {
        'Category': 'QR - Quality',
        'Description': 'Help Desk / IT Ops globally understaffed or untrained at GoLive across 17 sites',
        'Root Cause': 'Global recruitment timeline long; training lag for new IT staff across multiple regions',
        'Probability': 2, 'Impact': 3,
        'Mitigation': 'Global IT Ops hiring by Aug 2026; training Oct 2026; hypercare team embedded globally from Day 1',
        'Owner': 'Bosch HR, IT Ops, Regional IT Leads',
        'Target Date': '01-Nov-26',
        'Status': 'Amber',
        'Notes': '17-site support model requires regional IT Ops leads. Contractor options as fallback.'
    },
    {
        'Category': 'BR - Business',
        'Description': 'Solar business continuity disruption during cutover weekend impacts customer commitments',
        'Root Cause': 'Solar energy business is operationally critical; 24/7 monitoring systems must not be interrupted',
        'Probability': 2, 'Impact': 3,
        'Mitigation': 'Identify OT / SCADA systems in scope; plan phased cutover avoiding operational peaks; 24/7 war room during cutover',
        'Owner': 'Solar Business Operations, IT PM',
        'Target Date': '30-Oct-26',
        'Status': 'Green',
        'Notes': 'OT/SCADA scope must be confirmed by QG2. Critical operational systems require special handling.'
    },
    {
        'Category': 'LR - Legality',
        'Description': 'Regulatory approvals for IT asset transfer delayed in one or more jurisdictions',
        'Root Cause': 'IT asset transfers in certain countries require regulatory notifications or approvals (e.g. competition authorities)',
        'Probability': 1, 'Impact': 4,
        'Mitigation': 'Map regulatory requirements per country by QG1; filings submitted by QG2; buffer time in Phase 5',
        'Owner': 'Legal, Finance, Erik Ho (BD/MIL-ICC)',
        'Target Date': '30-Jun-26',
        'Status': 'Green',
        'Notes': 'Low probability but high impact if materialises. Flag to Group Legal immediately.'
    },
]


def main():
    parser = argparse.ArgumentParser(
        description="Generate Hamburger_Risk_Register.xlsx from Risk_analysis_template.xlsx"
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=Path(__file__).parent / "Risk_analysis_template.xlsx",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).parent / "Hamburger" / "Hamburger_Risk_Register.xlsx",
    )
    args = parser.parse_args()

    print(f"Loading template : {args.template}")
    wb = load_workbook(args.template)

    # ── Update Cover Sheet ────────────────────────────────────────────────────
    wc = wb["Cover sheet"]
    wc["D1"] = "Bosch IT (Seller IT Manager)"
    wc["D2"] = "Erik Ho (BD/MIL-ICC)"
    wc["D3"] = "Hamburger PMO"

    # ── Populate Analysis Sheet ───────────────────────────────────────────────
    ws = wb["Analysis of project risks"]

    DATA_START_ROW = 9
    TEMPLATE_ROWS = 43

    wrap = Alignment(wrap_text=True, vertical="top")

    for idx, risk in enumerate(HAMBURGER_RISKS):
        row_num = DATA_START_ROW + idx

        category    = risk.get('Category', '')
        description = risk.get('Description', '')
        root_cause  = risk.get('Root Cause', '')
        probability = risk.get('Probability', 0)
        impact      = risk.get('Impact', 0)
        mitigation  = risk.get('Mitigation', '')
        owner       = risk.get('Owner', '')
        target_date = risk.get('Target Date', '')
        status      = risk.get('Status', '')
        notes       = risk.get('Notes', '')

        # Set RZ formula for rows beyond template pre-built range
        if idx >= TEMPLATE_ROWS:
            ws.cell(row_num, 10).value = f"=$H{row_num}*$I{row_num}"

        ws.cell(row_num,  1).value = idx + 1       # A: No.
        ws.cell(row_num,  2).value = None           # B: Sub-project
        ws.cell(row_num,  3).value = None           # C: Entry Date
        ws.cell(row_num,  4).value = category       # D: Risk Category
        ws.cell(row_num,  5).value = description    # E: Risk Description
        ws.cell(row_num,  6).value = notes          # F: Effects
        ws.cell(row_num,  7).value = root_cause     # G: Causes
        ws.cell(row_num,  8).value = probability    # H: W (Probability)
        ws.cell(row_num,  9).value = impact         # I: T (Impact)
        # J (col 10) already has formula from template
        ws.cell(row_num, 11).value = mitigation     # K: Actions
        ws.cell(row_num, 12).value = owner          # L: Responsible
        ws.cell(row_num, 13).value = target_date    # M: Deadline
        ws.cell(row_num, 14).value = status         # N: Status
        ws.cell(row_num, 15).value = ''             # O: Remarks

        for col in range(1, 16):
            ws.cell(row_num, col).alignment = wrap

    # ── Adjust row heights ────────────────────────────────────────────────────
    for i in range(DATA_START_ROW, DATA_START_ROW + len(HAMBURGER_RISKS)):
        ws.row_dimensions[i].height = 75

    # ── Save output ───────────────────────────────────────────────────────────
    wb.save(args.out)

    high_risks = [(i+1, r['Probability'] * r['Impact'])
                  for i, r in enumerate(HAMBURGER_RISKS)
                  if r['Probability'] * r['Impact'] >= 12]

    print(f"Generated        : {args.out}")
    print(f"Total risks      : {len(HAMBURGER_RISKS)}")
    print(f"High-priority risks (P×I >= 12): {len(high_risks)}")
    for risk_id, rating in high_risks:
        print(f"  Risk #{risk_id}: rating {rating}")


if __name__ == "__main__":
    main()
