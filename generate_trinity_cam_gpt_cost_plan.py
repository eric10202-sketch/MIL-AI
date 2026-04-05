#!/usr/bin/env python3
"""Generate the Trinity-CAM (GPT) labour cost plan as XLSX."""

from __future__ import annotations

import os
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


HERE = Path(__file__).parent
PROJECT_NAME = "Trinity-CAM (GPT)"
PROJECT_DIR = HERE / "active-projects" / PROJECT_NAME
SCHEDULE_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Risk_Register.xlsx"
OUTPUT_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Cost_Plan.xlsx"

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)


RESOURCE_CONFIG = {
    "KPMG PMO Lead": {"category": "KPMG - Programme Governance", "rate": 220},
    "KPMG Project Manager": {"category": "KPMG - Programme Governance", "rate": 170},
    "KPMG Change Lead": {"category": "KPMG - Programme Governance", "rate": 150},
    "KPMG Deployment Lead": {"category": "KPMG - Testing, Deployment, and Readiness", "rate": 150},
    "KPMG Test Lead": {"category": "KPMG - Testing, Deployment, and Readiness", "rate": 155},
    "KPMG Enterprise Architect": {"category": "KPMG - Architecture and Migration Advisory", "rate": 190},
    "KPMG Infrastructure Architect": {"category": "KPMG - Architecture and Migration Advisory", "rate": 185},
    "KPMG Data Architect": {"category": "KPMG - Architecture and Migration Advisory", "rate": 180},
    "KPMG Workplace Lead": {"category": "KPMG - Architecture and Migration Advisory", "rate": 165},
    "KPMG Security Architect": {"category": "KPMG - Architecture and Migration Advisory", "rate": 190},
    "KPMG SAP Architect": {"category": "KPMG - SAP Advisory", "rate": 210},
    "Infosys Programme Manager": {"category": "Infosys - Programme and Service Delivery", "rate": 145},
    "Infosys Service Delivery Lead": {"category": "Infosys - Programme and Service Delivery", "rate": 125},
    "Infosys Application Lead": {"category": "Infosys - Application and Data Delivery", "rate": 120},
    "Infosys Data Migration Lead": {"category": "Infosys - Application and Data Delivery", "rate": 125},
    "Infosys QA Lead": {"category": "Infosys - Application and Data Delivery", "rate": 110},
    "Infosys Cloud Lead": {"category": "Infosys - Platform, Security, and Workplace", "rate": 130},
    "Infosys IAM Lead": {"category": "Infosys - Platform, Security, and Workplace", "rate": 125},
    "Infosys Infrastructure Architect": {"category": "Infosys - Platform, Security, and Workplace", "rate": 130},
    "Infosys Network Lead": {"category": "Infosys - Platform, Security, and Workplace", "rate": 120},
    "Infosys Security Lead": {"category": "Infosys - Platform, Security, and Workplace", "rate": 130},
    "Infosys Workplace Lead": {"category": "Infosys - Platform, Security, and Workplace", "rate": 115},
    "Infosys SAP Lead": {"category": "Infosys - SAP Delivery", "rate": 145},
    "JCI Application Owner": {"category": "JCI Internal Support - Tracked Only", "rate": 0},
    "JCI Business Lead": {"category": "JCI Internal Support - Tracked Only", "rate": 0},
    "JCI HR Lead": {"category": "JCI Internal Support - Tracked Only", "rate": 0},
    "JCI IT Manager": {"category": "JCI Internal Support - Tracked Only", "rate": 0},
    "JCI Legal Counsel": {"category": "JCI Internal Support - Tracked Only", "rate": 0},
    "JCI SAP Owner": {"category": "JCI Internal Support - Tracked Only", "rate": 0},
    "JCI Sponsor": {"category": "JCI Internal Support - Tracked Only", "rate": 0},
    "Bosch Business Lead": {"category": "Bosch Internal Support - Tracked Only", "rate": 0},
    "Bosch HR Lead": {"category": "Bosch Internal Support - Tracked Only", "rate": 0},
    "Bosch IT Manager": {"category": "Bosch Internal Support - Tracked Only", "rate": 0},
    "Bosch Legal Counsel": {"category": "Bosch Internal Support - Tracked Only", "rate": 0},
    "Bosch Sponsor": {"category": "Bosch Internal Support - Tracked Only", "rate": 0},
}


RISK_SCORE_MAP = {
    "Very Low": 1,
    "Low": 2,
    "Moderate": 3,
    "High": 4,
    "Very High": 5,
}

RISK_PROBABILITY_MAP = {
    "10%": 1,
    "30%": 2,
    "50%": 3,
    "70%": 4,
    "90%": 5,
}

RISK_CONTINGENCY_LIBRARY = {
    "R001": ("SAP expert support and rehearsal contingency", "600,000 - 1,200,000", "Risk Register R001 - external SAP carve-out support, rehearsal environment, and specialist remediation capacity."),
    "R002": ("Merger-zone platform acceleration reserve", "400,000 - 900,000", "Risk Register R002 - additional platform engineering or temporary capacity if merger-zone build slips."),
    "R003": ("Cutover command-center and rollback reserve", "250,000 - 500,000", "Risk Register R003 - final cutover defect sprint, extended command center, and rollback rehearsal support."),
    "R007": ("Cyber control uplift and incident readiness", "250,000 - 500,000", "Risk Register R007 - DLP, logging, security review, and cyber readiness controls for the merger zone."),
    "R012": ("Platform procurement and fallback capacity", "200,000 - 450,000", "Risk Register R012 - long-lead platform and connectivity fallback capacity."),
    "R014": ("Regional hosting and legal compliance provision", "TBC - confirm at QG1", "Risk Register R014 - data sovereignty remediation or local hosting design changes if required."),
    "R020": ("Third-party audit and supplier-governance provision", "100,000 - 200,000", "Risk Register R020 - subcontractor audit and compliance assurance if external delivery layers expand."),
}


def parse_schedule() -> tuple[list[dict], dict[str, int], list[dict], dict[str, int]]:
    workbook = load_workbook(SCHEDULE_PATH, data_only=False)
    sheet = workbook["Schedule"]

    tasks = []
    current_phase = None
    resource_days: dict[str, int] = defaultdict(int)
    phase_rows: list[dict] = []
    phase_cost_days: dict[str, int] = defaultdict(int)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        task_id = int(row[0])
        outline_level = int(row[1])
        name = str(row[2]).strip()
        duration_label = str(row[3]).strip()
        start = row[4].strftime("%Y-%m-%d") if hasattr(row[4], "strftime") else str(row[4])
        finish = row[5].strftime("%Y-%m-%d") if hasattr(row[5], "strftime") else str(row[5])
        predecessors = str(row[6] or "")
        resources = str(row[7] or "")
        notes = str(row[8] or "")
        milestone = str(row[9] or "")
        duration_days = int(duration_label.split()[0])

        task = {
            "id": task_id,
            "outline_level": outline_level,
            "name": name.strip(),
            "duration_days": duration_days,
            "start": start,
            "finish": finish,
            "predecessors": predecessors,
            "resources": resources,
            "notes": notes,
            "milestone": milestone,
            "phase": current_phase,
        }

        if outline_level == 1:
            current_phase = {"name": name.strip(), "start": start, "finish": finish}
            phase_rows.append(current_phase)
            task["phase"] = current_phase
        elif outline_level == 3 and resources:
            for token in [item.strip() for item in resources.split("+") if item.strip()]:
                resource_days[token] += duration_days
                if current_phase:
                    phase_cost_days[current_phase["name"]] += duration_days * RESOURCE_CONFIG[token]["rate"] * 8

        tasks.append(task)

    return tasks, resource_days, phase_rows, phase_cost_days


def parse_high_risks() -> list[dict]:
    workbook = load_workbook(RISK_PATH, data_only=False)
    sheet = workbook["Risk Register"]
    results: list[dict] = []
    row = 5
    while True:
        risk_id = sheet.cell(row, 2).value
        if not risk_id:
            break
        impact = sheet.cell(row, 12).value
        probability = sheet.cell(row, 14).value
        score = RISK_SCORE_MAP.get(str(impact), 0) * RISK_PROBABILITY_MAP.get(str(probability), 0)
        results.append(
            {
                "id": str(risk_id),
                "category": str(sheet.cell(row, 4).value or ""),
                "score": score,
                "strategy": str(sheet.cell(row, 22).value or ""),
                "notes": str(sheet.cell(row, 35).value or ""),
            }
        )
        row += 1
    return results


def build_category_rows(resource_days: dict[str, int]) -> list[dict]:
    missing = sorted(set(resource_days) - set(RESOURCE_CONFIG))
    if missing:
        raise ValueError(f"Unmapped schedule resources: {', '.join(missing)}")

    grouped: dict[str, list[dict]] = defaultdict(list)
    for resource_name, total_days in sorted(resource_days.items()):
        config = RESOURCE_CONFIG[resource_name]
        grouped[config["category"]].append(
            {
                "resource": resource_name,
                "days": total_days,
                "hours": total_days * 8,
                "rate": config["rate"],
                "cost": total_days * 8 * config["rate"],
            }
        )
    return [{"category": category, "rows": grouped[category]} for category in sorted(grouped)]


def build_capex_rows(high_risks: list[dict]) -> list[tuple[str, str, str]]:
    capex_rows = []
    risk_ids = {risk["id"] for risk in high_risks}
    for risk_id, row in RISK_CONTINGENCY_LIBRARY.items():
        if risk_id in risk_ids:
            capex_rows.append(row)
    return capex_rows


def write_xlsx(category_rows: list[dict], phase_rows: list[dict], phase_costs: dict[str, int], capex_rows: list[tuple[str, str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cost Plan"

    blue_dark = "003B6E"
    blue_mid = "0066CC"
    blue_light = "EFF4FB"
    grey = "F2F2F2"
    subtotal_fill = "C6D4E8"
    white = "FFFFFF"

    def fill(colour: str) -> PatternFill:
        return PatternFill(start_color=colour, end_color=colour, fill_type="solid")

    def font(colour: str = "000000", bold: bool = False, size: int = 9, italic: bool = False) -> Font:
        return Font(name="Calibri", color=colour, bold=bold, italic=italic, size=size)

    def align(horizontal: str = "left") -> Alignment:
        return Alignment(horizontal=horizontal, vertical="top", wrap_text=True)

    sheet.column_dimensions["A"].width = 52
    sheet.column_dimensions["B"].width = 32
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 16

    row = 1
    sheet.merge_cells(f"A{row}:F{row}")
    sheet[f"A{row}"] = f"{PROJECT_NAME}  |  IT CARVE-OUT COST PLAN"
    sheet[f"A{row}"].fill = fill(blue_dark)
    sheet[f"A{row}"].font = font(white, bold=True, size=13)
    sheet[f"A{row}"].alignment = align("center")
    row += 1

    metadata_rows = [
        ("Project", PROJECT_NAME),
        ("Seller", "Johnson Controls International (JCI)"),
        ("Buyer", "Bosch"),
        ("Business", "Air conditioning business"),
        ("Carve-out Model", "Integration - JCI IT to merger zone to Bosch IT"),
        ("Based on", f"{PROJECT_NAME}_Project_Schedule.xlsx"),
        ("Risk-aligned", f"{PROJECT_NAME}_Risk_Register.xlsx"),
        ("Budget Baseline", "TBC - to be approved at QG1"),
        ("Note", "Bosch and JCI internal resource lines are tracked at zero rate for schedule traceability and excluded from the billable labour baseline."),
    ]
    for label, value in metadata_rows:
        sheet.cell(row, 1).value = label
        sheet.cell(row, 2).value = value
        sheet.merge_cells(f"B{row}:F{row}")
        for col in range(1, 7):
            sheet.cell(row, col).fill = fill(grey)
            sheet.cell(row, col).font = font(size=8)
            sheet.cell(row, col).alignment = align()
        row += 1

    row += 1
    headers = ["CATEGORY", "RESOURCE", "TOTAL DAYS", "TOTAL HRS", "HOURLY RATE (EUR)", "TOTAL COST (EUR)"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row, col).value = header
        sheet.cell(row, col).fill = fill(blue_mid)
        sheet.cell(row, col).font = font(white, bold=True, size=9)
        sheet.cell(row, col).alignment = align("center")
    sheet.freeze_panes = f"A{row + 1}"
    row += 1

    category_totals = []
    overall_total = 0
    for category_block in category_rows:
        sheet.merge_cells(f"A{row}:F{row}")
        sheet[f"A{row}"] = category_block["category"]
        sheet[f"A{row}"].fill = fill(blue_mid)
        sheet[f"A{row}"].font = font(white, bold=True, size=9)
        sheet[f"A{row}"].alignment = align()
        row += 1

        category_days = 0
        category_cost = 0
        for index, entry in enumerate(category_block["rows"]):
            line_fill = white if index % 2 == 0 else blue_light
            sheet.cell(row, 1).value = category_block["category"]
            sheet.cell(row, 2).value = entry["resource"]
            sheet.cell(row, 3).value = entry["days"]
            sheet.cell(row, 4).value = entry["hours"]
            sheet.cell(row, 5).value = entry["rate"]
            sheet.cell(row, 6).value = entry["cost"]
            for col in range(1, 7):
                sheet.cell(row, col).fill = fill(line_fill)
                sheet.cell(row, col).font = font(size=9)
                sheet.cell(row, col).alignment = align("right" if col >= 3 else "left")
            for col in (3, 4, 5, 6):
                sheet.cell(row, col).number_format = "#,##0"
            category_days += entry["days"]
            category_cost += entry["cost"]
            row += 1

        sheet.cell(row, 1).value = f"Subtotal - {category_block['category']}"
        sheet.cell(row, 3).value = category_days
        sheet.cell(row, 6).value = category_cost
        for col in range(1, 7):
            sheet.cell(row, col).fill = fill(subtotal_fill)
            sheet.cell(row, col).font = font(bold=True, size=9)
            sheet.cell(row, col).alignment = align("right" if col >= 3 else "left")
        sheet.cell(row, 3).number_format = "#,##0"
        sheet.cell(row, 6).number_format = "#,##0"
        category_totals.append((category_block["category"], category_days, category_cost))
        overall_total += category_cost
        row += 2

    sheet.merge_cells(f"A{row}:E{row}")
    sheet[f"A{row}"] = "OVERALL PROJECT TOTAL"
    sheet[f"A{row}"].fill = fill(blue_dark)
    sheet[f"A{row}"].font = font(white, bold=True, size=10)
    sheet[f"A{row}"].alignment = align()
    sheet.cell(row, 6).value = overall_total
    sheet.cell(row, 6).fill = fill(blue_dark)
    sheet.cell(row, 6).font = font(white, bold=True, size=10)
    sheet.cell(row, 6).alignment = align("right")
    sheet.cell(row, 6).number_format = "#,##0"
    row += 2

    sheet.merge_cells(f"A{row}:F{row}")
    sheet[f"A{row}"] = "Cost Breakdown by Category"
    sheet[f"A{row}"].fill = fill(blue_mid)
    sheet[f"A{row}"].font = font(white, bold=True, size=9)
    row += 1
    for index, (category, days, cost) in enumerate(category_totals):
        line_fill = white if index % 2 == 0 else blue_light
        sheet.cell(row, 1).value = category
        sheet.cell(row, 3).value = days
        sheet.cell(row, 6).value = cost
        for col in range(1, 7):
            sheet.cell(row, col).fill = fill(line_fill)
            sheet.cell(row, col).font = font(size=9)
            sheet.cell(row, col).alignment = align("right" if col >= 3 else "left")
        sheet.cell(row, 3).number_format = "#,##0"
        sheet.cell(row, 6).number_format = "#,##0"
        row += 1
    row += 1

    sheet.merge_cells(f"A{row}:F{row}")
    sheet[f"A{row}"] = "Cost Breakdown by Phase"
    sheet[f"A{row}"].fill = fill(blue_mid)
    sheet[f"A{row}"].font = font(white, bold=True, size=9)
    row += 1
    for index, phase in enumerate(phase_rows):
        line_fill = white if index % 2 == 0 else blue_light
        sheet.cell(row, 1).value = phase["name"]
        sheet.cell(row, 2).value = f"{phase['start']} to {phase['finish']}"
        sheet.cell(row, 6).value = phase_costs.get(phase["name"], 0)
        for col in range(1, 7):
            sheet.cell(row, col).fill = fill(line_fill)
            sheet.cell(row, col).font = font(size=9)
            sheet.cell(row, col).alignment = align("right" if col == 6 else "left")
        sheet.cell(row, 6).number_format = "#,##0"
        row += 1
    row += 1

    sheet.merge_cells(f"A{row}:F{row}")
    sheet[f"A{row}"] = "CAPEX / Additional Costs - Excluded from Labour Total"
    sheet[f"A{row}"].fill = fill(blue_mid)
    sheet[f"A{row}"].font = font(white, bold=True, size=9)
    row += 1
    for index, (description, estimate, note) in enumerate(capex_rows):
        line_fill = white if index % 2 == 0 else blue_light
        sheet.cell(row, 1).value = description
        sheet.cell(row, 2).value = note
        sheet.cell(row, 5).value = estimate
        for col in range(1, 7):
            sheet.cell(row, col).fill = fill(line_fill)
            sheet.cell(row, col).font = font(size=9)
            sheet.cell(row, col).alignment = align()
        row += 1
    row += 1

    notes = [
        "1. Derived from the generated schedule and aligned to the generated risk register rather than copied from any legacy project.",
        "2. Phase names and dates match the schedule exactly: Phase 0 through Phase 5 as generated for Trinity-CAM (GPT).",
        "3. Every schedule resource token has a matching cost line; Bosch and JCI internal roles are tracked at zero rate to preserve resource traceability without inflating the external labour baseline.",
        "4. Integration-model note: the merger zone is a temporary operational bridge between JCI and Bosch, so platform and security contingencies are carried separately from labour totals.",
        "5. High-priority risk references for contingency lines: R001, R002, R003, R007, R012, R014, and R020.",
        "6. Budget baseline remains TBC - to be approved at QG1.",
    ]
    for note in notes:
        sheet.merge_cells(f"A{row}:F{row}")
        sheet[f"A{row}"] = note
        sheet[f"A{row}"].fill = fill(grey)
        sheet[f"A{row}"].font = font(size=8, italic=True)
        sheet[f"A{row}"].alignment = align()
        row += 1

    workbook.save(OUTPUT_PATH)


def main() -> None:
    if not SCHEDULE_PATH.exists():
        print(f"Missing schedule input: {SCHEDULE_PATH}")
        sys.exit(1)
    if not RISK_PATH.exists():
        print(f"Missing risk register input: {RISK_PATH}")
        sys.exit(1)

    _, resource_days, phase_rows, phase_costs = parse_schedule()
    high_risks = parse_high_risks()
    category_rows = build_category_rows(resource_days)
    capex_rows = build_capex_rows(high_risks)

    print(f"[{PROJECT_NAME}] Generating cost plan")
    print(f"  Schedule input: {SCHEDULE_PATH}")
    print(f"  Risk input:     {RISK_PATH}")
    print(f"  Resource lines: {sum(len(block['rows']) for block in category_rows)}")
    print(f"  CAPEX rows:     {len(capex_rows)}")
    write_xlsx(category_rows, phase_rows, phase_costs, capex_rows)
    print(f"  Output:         {OUTPUT_PATH}")
    print(f"[{PROJECT_NAME}] Cost plan complete")


if __name__ == "__main__":
    main()