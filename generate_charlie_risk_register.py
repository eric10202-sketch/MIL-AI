"""
generate_bravo_risk_register.py
Generates Charlie_Risk_Register.xlsx using BD_Risk-Register_template_en_V1.0_Dec2023.xlsx

Project Charlie: Robert Bosch GmbH AI Business Carve-Out into 50/50 JV with Undisclosed
Seller: Robert Bosch GmbH | Buyer: Undisclosed | PM: Gill Amandeep Singh (BD/MIL-PSM1)
37 worldwide sites | 3500+ users | TBD apps | No ERP | No TSA | Bosch JV leadership control

Usage:
    python generate_bravo_risk_register.py
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

HERE = Path(__file__).parent
TEMPLATE = HERE / "BD_Risk-Register_template_en_V1.0_Dec2023.xlsx"
OUT_DIR  = HERE / "Charlie"
OUT_DIR.mkdir(exist_ok=True)
OUTPUT   = OUT_DIR / "Charlie_Risk_Register.xlsx"

# New template (BD V1.0 Dec 2023) column layout in "Risk Register" sheet
# Data starts at row 5. All columns are 1-based.
# B=2  risk ID
# C=3  creation date
# D=4  risk category
# E=5  cause(s)
# F=6  event (risk description)
# G=7  effect(s)
# H=8  risk event date
# I=9  owner
# J=10 source
# L=12 impact  ('Very low' | 'Low' | 'Medium' | 'High' | 'Very high')
# M=13 VLOOKUP: impact numeric  -- formula written explicitly for all rows
# N=14 probability  ('10%' | '30%' | '50%' | '70%' | '90%')
# O=15 VLOOKUP: probability numeric -- formula written explicitly for all rows
# P=16 threat / opportunity
# Q=17 matrix score = M*O
# R=18 qualitative impact description
# S=19 monetary impact EUR current year
# T=20 monetary impact EUR 3 subsequent years
# V=22 risk response strategy
# W=23 measure (mitigation actions)
# X=24 due date
# Z=26 status
# AA=27 reporting date
# AB=28 impact actual  (same as L initially)
# AC=29 VLOOKUP: actual impact numeric
# AD=30 probability actual  (same as N initially)
# AE=31 VLOOKUP: actual probability numeric
# AF=32 = AC (actual impact numeric)
# AG=33 = AE (actual probability numeric)
# AH=34 = AF*AG (actual matrix score)
# AI=35 notes

_IMP_MAP = {1: "Very Low", 2: "Low", 3: "Moderate", 4: "High", 5: "Very High"}
_PRB_MAP = {1: "10%",      2: "30%", 3: "50%",    4: "70%",  5: "90%"}
_CAT_MAP = {
    "ScR - Schedule": "Schedule",
    "SR - Scope":     "Engineering",
    "BtR - Budget":   "Budget",
    "RR - Resource":  "Resources",
    "CR - Change":    "Engineering",
    "QR - Quality":   "Quality",
    "LR - Legal":     "Legal & Compliance",
    "BR - Business":  "Strategy & Portfolio",
}
_STA_MAP = {"Amber": "in progress", "Green": "not started", "Red": "in progress"}

# â”€â”€ Risk data â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
CHARLIE_RISKS = [
    {
        "Category":    "ScR - Schedule",
        "Description": "Robert Bosch GmbH and Undisclosed JV management bandwidth constrained during 3-month build-to-GoLive sprint (Apr-Jun 2026)",
        "Effects":     "Phase 3 build and testing delays; GoLive of 1 Jul 2026 missed",
        "Causes":      "Carve-out running concurrently with live AI business operations; split leadership focus",
        "Probability": 4, "Impact": 4,
        "Mitigation":  "Dedicate Robert Bosch GmbH carve-out lead with explicit protected time allocation; Deputy named; Weekly SteerCo cadence from Day 1",
        "Owner":       "Riyaz Ahmed, Robert Bosch GmbH CIO",
        "Deadline":    "30-Apr-26",
        "Status":      "Amber",
        "Notes":       "High priority. Mitigation plan in place. Monitor weekly.",
    },
    {
        "Category":    "SR - Scope",
        "Description": "JV legal entity registration in India delayed beyond QG0 target (30 Apr 2026)",
        "Effects":     "QG0 gate cannot be passed; entire programme timeline shifts right",
        "Causes":      "MCA India bureaucratic processing times (4-8 weeks); first-time JV registration complexity",
        "Probability": 3, "Impact": 5,
        "Mitigation":  "Engage external India legal counsel by Apr 3; initiate MCA filing within first week of project",
        "Owner":       "Riyaz Ahmed, Legal, Finance",
        "Deadline":    "10-Apr-26",
        "Status":      "Green",
        "Notes":       "Critical path item. India legal counsel essential. No antitrust filing required.",
    },
    {
        "Category":    "BtR - Budget",
        "Description": "India cloud and M365 licensing costs exceed initial estimate for JV tenant provisioning",
        "Effects":     "Budget overrun in Phase 3; delay in cloud environment readiness",
        "Causes":      "Azure India region pricing; JV tenant licensing model complexity for 3500+ users",
        "Probability": 3, "Impact": 3,
        "Mitigation":  "Conduct M365/Azure RFQ by end of Apr; lock licensing model at QG0; budget contingency of 15% reserved",
        "Owner":       "Finance, Robert Bosch GmbH Procurement",
        "Deadline":    "30-Apr-26",
        "Status":      "Amber",
        "Notes":       "Small user base (70) limits financial exposure. Containable risk.",
    },
    {
        "Category":    "ScR - Schedule",
        "Description": "Phase 3 build (Jun 2026) delayed due to Robert Bosch GmbH infrastructure resource contention",
        "Effects":     "GoLive date of 1 Jul 2026 missed; Phase 4 cutover cannot proceed",
        "Causes":      "Robert Bosch GmbH infra team shared across multiple India projects; AD and network build sequential dependencies",
        "Probability": 3, "Impact": 4,
        "Mitigation":  "Reserve Robert Bosch GmbH infra team capacity from May; confirm resource allocation at QG1/2/3 gate",
        "Owner":       "Riyaz Ahmed, Robert Bosch GmbH Infra Lead",
        "Deadline":    "01-Jun-26",
        "Status":      "Green",
        "Notes":       "Short build window (Jun only). Resource reservation critical.",
    },
    {
        "Category":    "RR - Resource",
        "Description": "Key Robert Bosch GmbH IT personnel unavailable during compressed Jun 2026 build-and-test window",
        "Effects":     "Build quality impacted; integration testing incomplete by QG1/2/3 (26 Jun)",
        "Causes":      "Annual leave; Robert Bosch GmbH parallel project demands; carve-out knowledge concentrated in few individuals",
        "Probability": 3, "Impact": 4,
        "Mitigation":  "Enforce Jun leave freeze for critical roles; identify/train backups; document knowledge by end of May",
        "Owner":       "Robert Bosch GmbH HR, Riyaz Ahmed",
        "Deadline":    "29-May-26",
        "Status":      "Amber",
        "Notes":       "June is the highest-risk month. Contingency plan required.",
    },
    {
        "Category":    "SR - Scope",
        "Description": "Application migration scope expands beyond 17 AI apps due to late-discovered dependencies",
        "Effects":     "Phase 3 timeline overrun; additional testing cycles required",
        "Causes":      "Incomplete initial inventory; business users requesting additional apps post-QG0",
        "Probability": 2, "Impact": 4,
        "Mitigation":  "Hard scope freeze at QG0 (30 Apr); no additions after that date without SteerCo approval and timeline impact assessment",
        "Owner":       "Riyaz Ahmed, WS Leads",
        "Deadline":    "30-Apr-26",
        "Status":      "Green",
        "Notes":       "Small app count (17) makes scope creep manageable if freeze enforced.",
    },
    {
        "Category":    "CR - Change",
        "Description": "Active Directory forest separation for JV incomplete by GoLive (1 Jul 2026)",
        "Effects":     "Day 1 users unable to log on to JV systems; GoLive blocked",
        "Causes":      "AD tooling complexity; configuration errors in JV forest build; insufficient testing time",
        "Probability": 2, "Impact": 5,
        "Mitigation":  "Start AD build in week 1 of Jun; run integration and regression testing in parallel; conduct dress rehearsal Jun 24-26",
        "Owner":       "Robert Bosch GmbH AD Team, Riyaz Ahmed",
        "Deadline":    "24-Jun-26",
        "Status":      "Green",
        "Notes":       "AD is foundational for Day 1. No backout. Pilot and rehearsal mandatory.",
    },
    {
        "Category":    "BtR - Budget",
        "Description": "Licence change-of-control costs for 17 AI applications exceed estimates",
        "Effects":     "Budget overrun; delay in app migration if contracts not renegotiated in time",
        "Causes":      "ISVs invoking change-of-control clauses; unexpected vendor lock-in terms",
        "Probability": 2, "Impact": 3,
        "Mitigation":  "Complete licence review with change-of-control clause analysis by May 14; negotiate amendments ahead of Phase 3",
        "Owner":       "Robert Bosch GmbH Procurement, Legal",
        "Deadline":    "14-May-26",
        "Status":      "Green",
        "Notes":       "TBD apps is manageable. Early review essential.",
    },
    {
        "Category":    "QR - Quality",
        "Description": "UAT failure rate high due to compressed testing window (18-24 Jun 2026)",
        "Effects":     "QG1/2/3 gate (26 Jun) delayed; GoLive 1 Jul 2026 at risk",
        "Causes":      "Business users not available for full UAT; test environment not stable",
        "Probability": 3, "Impact": 4,
        "Mitigation":  "Engage business UAT leads by May 1; establish test environment in week 1 of Jun; run smoke tests daily from Jun 15",
        "Owner":       "Riyaz Ahmed, Business Leads",
        "Deadline":    "15-Jun-26",
        "Status":      "Green",
        "Notes":       "UAT window is tight (5 working days). Early prep critical.",
    },
    {
        "Category":    "SR - Scope",
        "Description": "Undisclosed JV counterpart team (architecture, IT) not staffed in time for May concept workshops",
        "Effects":     "JV IT architecture design delayed; Phase 2 deliverables incomplete",
        "Causes":      "Undisclosed internal hiring/assignment process for JV roles not yet started",
        "Probability": 3, "Impact": 3,
        "Mitigation":  "Confirm Undisclosed JV IT team nominees at Steering Committee kickoff (Apr 1); escalate to executive sponsor if not confirmed by Apr 10",
        "Owner":       "Riyaz Ahmed, Undisclosed Executive Sponsor",
        "Deadline":    "10-Apr-26",
        "Status":      "Amber",
        "Notes":       "stand-alone separation mitigates: Robert Bosch GmbH can proceed if Undisclosed team is delayed.",
    },
    {
        "Category":    "LR - Legal",
        "Description": "JV shareholder agreement final terms delay IT separation decisions",
        "Effects":     "Scope boundary between Robert Bosch GmbH and JV IT undefined; Phase 2 architecture blocked",
        "Causes":      "Legal negotiations between Bosch and Undisclosed on IP ownership and data rights extending beyond Apr",
        "Probability": 2, "Impact": 4,
        "Mitigation":  "Identify IT decisions independent of full SHA; escalate blocking items to SteerCo; resolve core IP scope by QG0",
        "Owner":       "Legal, Riyaz Ahmed",
        "Deadline":    "30-Apr-26",
        "Status":      "Green",
        "Notes":       "stand-alone separation reduces legal complexity significantly.",
    },
    {
        "Category":    "CR - Change",
        "Description": "Data separation between Robert Bosch GmbH corporate and JV AI business data is more complex than anticipated",
        "Effects":     "Data migration execution delayed past Jun 22; Day 1 data completeness at risk",
        "Causes":      "AI training datasets and model artefacts stored in shared Robert Bosch GmbH repositories; unclear data lineage",
        "Probability": 3, "Impact": 3,
        "Mitigation":  "Complete data ownership mapping by May 19; agree separation rules with legal and data owners before Phase 3 starts",
        "Owner":       "Riyaz Ahmed, Legal, Robert Bosch GmbH IT",
        "Deadline":    "29-May-26",
        "Status":      "Green",
        "Notes":       "Bosch-led JV reduces risk: Robert Bosch GmbH data stewardship retained post-GoLive.",
    },
    {
        "Category":    "ScR - Schedule",
        "Description": "Hypercare issues in Jul-Sep 2026 require Robert Bosch GmbH resources beyond planned capacity",
        "Effects":     "Robert Bosch GmbH IT cost overrun; closure milestone (Oct 30) delayed",
        "Causes":      "Application instability post-GoLive; TBD apps not fully stabilised before hypercare close",
        "Probability": 2, "Impact": 3,
        "Mitigation":  "Ensure all TBD apps fully tested before GoLive; define clear P1/P2 escalation matrix; 60-day hypercare window is adequate buffer",
        "Owner":       "Riyaz Ahmed, IT Ops",
        "Deadline":    "01-Jul-26",
        "Status":      "Green",
        "Notes":       "No TSA risk as JV is Bosch-led; escalation to Robert Bosch GmbH via governance channel available.",
    },
    {
        "Category":    "BR - Business",
        "Description": "AI business continuity disrupted during Jun 2026 build and cutover activities",
        "Effects":     "Undisclosed / Robert Bosch GmbH AI business projects delayed; executive sponsor dissatisfaction",
        "Causes":      "Carve-out build activities running concurrently with live AI project deliveries",
        "Probability": 2, "Impact": 3,
        "Mitigation":  "Define AI business activity blackout window for week of Jun 29-Jul 1; agree cutover communications with business leads",
        "Owner":       "Robert Bosch GmbH AI Business Lead, Riyaz Ahmed",
        "Deadline":    "29-Jun-26",
        "Status":      "Green",
        "Notes":       "Business continuity plan to be approved at QG1/2/3 gate.",
    },
]

FIRST_ROW = 5  # data starts at row 5 in new template


def populate_risk_register():
    wb = load_workbook(TEMPLATE)

    # â”€â”€ Info sheet â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws_info = wb["Info"]
    ws_info["C4"] = "Charlie-RR-001"
    ws_info["C5"] = "Charlie_Risk_Register.xlsx"
    ws_info["C6"] = "Project Charlie - Robert Bosch GmbH AI Business Carve-Out into 50/50 JV with Undisclosed"
    ws_info["C7"] = "Charlie"
    ws_info["C8"] = "Gill Amandeep Singh (BD/MIL-PSM1)"

    # â”€â”€ Risk Register sheet â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws = wb["Risk Register"]

    for i, risk in enumerate(CHARLIE_RISKS):
        r = FIRST_ROW + i
        imp_label = _IMP_MAP[risk["Impact"]]
        prb_label = _PRB_MAP[risk["Probability"]]
        cat_label = _CAT_MAP.get(risk["Category"], risk["Category"])
        sta_label = _STA_MAP.get(risk["Status"], "not started")

        ws.cell(r,  2).value = i + 1                          # risk ID
        ws.cell(r,  3).value = "03-Apr-26"                    # creation date
        ws.cell(r,  4).value = cat_label                      # risk category
        ws.cell(r,  5).value = risk["Causes"]                 # cause(s)
        ws.cell(r,  6).value = risk["Description"]            # event
        ws.cell(r,  7).value = risk["Effects"]                # effect(s)
        ws.cell(r,  8).value = "n.a."                         # risk event date
        ws.cell(r,  9).value = risk["Owner"]                  # owner
        ws.cell(r, 10).value = "Project Assessment"           # source
        ws.cell(r, 12).value = imp_label                      # impact (text)
        ws.cell(r, 13).value = f'=_xlfn.IFNA(VLOOKUP(L{r},$D$182:$E$186,2,FALSE),"")'  # M
        ws.cell(r, 14).value = prb_label                      # probability (text %)
        ws.cell(r, 15).value = f'=_xlfn.IFNA(VLOOKUP(N{r},$D$189:$E$193,2,FALSE),"")'  # O
        ws.cell(r, 16).value = "threat"                       # threat/opportunity
        ws.cell(r, 17).value = f"=M{r}*O{r}"                 # matrix score
        ws.cell(r, 18).value = (
            f"{imp_label} impact - {risk['Effects']}"
        )                                                      # qualitative impact
        ws.cell(r, 19).value = 0                              # monetary EUR cur yr
        ws.cell(r, 20).value = 0                              # monetary EUR 3 yrs
        ws.cell(r, 22).value = "Mitigate"                     # response strategy
        ws.cell(r, 23).value = risk["Mitigation"]             # measure
        ws.cell(r, 24).value = risk["Deadline"]               # due date
        ws.cell(r, 26).value = sta_label                      # status
        ws.cell(r, 27).value = "03-Apr-26"                    # reporting date
        ws.cell(r, 28).value = imp_label                      # impact actual (=initial)
        ws.cell(r, 29).value = f'=_xlfn.IFNA(VLOOKUP(AB{r},$D$182:$E$186,2,FALSE),"")'  # AC
        ws.cell(r, 30).value = prb_label                      # prob actual (=initial)
        ws.cell(r, 31).value = f'=_xlfn.IFNA(VLOOKUP(AD{r},$D$189:$E$193,2,FALSE),"")'  # AE
        ws.cell(r, 32).value = f"=AC{r}"                     # AF
        ws.cell(r, 33).value = f"=AE{r}"                     # AG
        ws.cell(r, 34).value = f"=AF{r}*AG{r}"               # AH matrix score actual
        ws.cell(r, 35).value = risk["Notes"]                  # notes

        # Wrap text in description/mitigation columns
        for col in [5, 6, 7, 18, 23, 35]:
            ws.cell(r, col).alignment = Alignment(wrap_text=True, vertical="top")

    # Fix Matrix tab: openpyxl does not preserve theme-inherited black font on
    # coloured cells when re-saving.  Explicitly set black font on all yellow-
    # filled cells so they remain readable (yellow fill + white/theme font = invisible).
    _YELLOW_FILLS = {"FFFFFF00", "FFFFFFCC"}
    ws_matrix = wb["Matrix "]
    for row in ws_matrix.iter_rows():
        for cell in row:
            try:
                if cell.fill.fgColor.type == "rgb" and cell.fill.fgColor.rgb in _YELLOW_FILLS:
                    f = cell.font
                    cell.font = Font(
                        name=f.name, size=f.size, bold=f.bold,
                        italic=f.italic, underline=f.underline,
                        strike=f.strike, color="FF000000",
                    )
            except Exception:
                pass

    wb.save(OUTPUT)
    print(f"Risk register written: {OUTPUT}  ({len(CHARLIE_RISKS)} risks)")


if __name__ == "__main__":
    from datetime import datetime as _dt
    _t0 = _dt.now()
    print(f"Started : {_t0.strftime('%Y-%m-%d %H:%M:%S')}")
    if not TEMPLATE.exists():
        print(f"ERROR: Template not found: {TEMPLATE}")
        sys.exit(1)
    populate_risk_register()
    _t1 = _dt.now()
    print(f"Finished: {_t1.strftime('%Y-%m-%d %H:%M:%S')}  ({(_t1-_t0).total_seconds():.1f}s elapsed)")



