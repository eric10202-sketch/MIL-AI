#!/usr/bin/env python3

from __future__ import annotations

import base64
from datetime import date, datetime, timedelta
from pathlib import Path
import re

from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent
PROJECT_NAME = "Trinity-CAM (GPT)"
OUTPUT_FOLDER_NAME = "Trinity-CAM (GPT) v1.1"
SELLER = "Johnson Controls International (JCI)"
BUYER = "Robert Bosch GmbH"
MODEL = "Integration"
SITE_COUNT = 48
USER_COUNT = 12000
APPLICATION_COUNT = 1800
PROJECT_DIR = BASE_DIR / "active-projects" / OUTPUT_FOLDER_NAME
SCHEDULE_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Risk_Register.xlsx"
COST_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Cost_Plan.xlsx"
OUTPUT_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Management_KPI_Dashboard.html"
LOGO_PATH = BASE_DIR / "Bosch.png"

IMPACT_SCORES = {"Very Low": 1, "Low": 2, "Moderate": 3, "High": 4, "Very High": 5}
PROBABILITY_SCORES = {"10%": 1, "30%": 2, "50%": 3, "70%": 4, "90%": 5}
WORKSTREAM_CONFIDENCE = [
    ("PMO and governance", 84, "good"),
    ("Merger-zone platform", 78, "good"),
    ("SAP carve-out", 56, "warn"),
    ("Application migration", 61, "warn"),
    ("Identity and workplace", 76, "good"),
    ("Security and compliance", 63, "warn"),
    ("Data migration", 74, "good"),
    ("Regional deployment", 60, "warn"),
    ("Hypercare and handover", 81, "good"),
]


def html_escape(text: object) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def as_date(value: datetime | date) -> date:
    if isinstance(value, datetime):
        return value.date()
    return value


def fmt_date(value: datetime | date) -> str:
    return as_date(value).strftime("%d %b %Y")


def fmt_eur(value: int | float) -> str:
    return f"EUR {value:,.0f}"


def days_to(value: datetime | date) -> int:
    return (as_date(value) - date.today()).days


def countdown(value: datetime | date) -> str:
    days = days_to(value)
    if days > 0:
        return f"{days} days"
    if days < 0:
        return f"{abs(days)}d ago"
    return "TODAY"


def load_schedule() -> dict[str, object]:
    ws = load_workbook(SCHEDULE_PATH, data_only=True)["Schedule"]
    tasks: list[dict[str, object]] = []
    milestones: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        task_id, outline_level, name, duration, start, finish, predecessors, resources, notes, milestone = row
        if not name:
            continue
        item = {
            "id": task_id,
            "outline_level": outline_level,
            "name": str(name).strip(),
            "start": start,
            "finish": finish,
            "resources": str(resources or ""),
            "notes": str(notes or ""),
        }
        tasks.append(item)
        if str(milestone).strip().lower() == "yes":
            milestones.append(item)
    milestones_by_name = {item["name"]: item for item in milestones}
    start_date = as_date(tasks[0]["start"])
    end_date = as_date(tasks[-1]["finish"])
    total_days = (end_date - start_date).days or 1
    elapsed_days = min(max((date.today() - start_date).days, 0), total_days)
    time_elapsed_pct = round((elapsed_days / total_days) * 100)
    forecast_cutoff = date.today() + timedelta(days=90)
    next_actions = []
    for task in tasks:
        if task["outline_level"] != 3:
            continue
        task_start = as_date(task["start"])
        if date.today() <= task_start <= forecast_cutoff:
            next_actions.append(task)
    return {
        "tasks": tasks,
        "milestones": milestones,
        "milestones_by_name": milestones_by_name,
        "start_date": start_date,
        "end_date": end_date,
        "total_days": total_days,
        "elapsed_days": elapsed_days,
        "time_elapsed_pct": time_elapsed_pct,
        "next_actions": next_actions[:6],
    }


def load_costs() -> dict[str, object]:
    ws = load_workbook(COST_PATH, data_only=True)["Cost Plan"]
    total_labour = 0.0
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == "OVERALL PROJECT TOTAL":
            total_labour = float(ws.cell(row, 6).value or 0)
            break
    return {"total_labour": total_labour}


def load_risks() -> dict[str, object]:
    ws = load_workbook(RISK_PATH, data_only=True)["Risk Register"]
    threats: list[dict[str, object]] = []
    for row_num in range(5, ws.max_row + 1):
        risk_id = ws[f"B{row_num}"].value
        if not risk_id or not re.fullmatch(r"R\d{3}", str(risk_id).strip()):
            continue
        if str(ws[f"P{row_num}"].value or "").strip().lower() == "opportunity":
            continue
        impact = str(ws[f"L{row_num}"].value or "")
        probability = str(ws[f"N{row_num}"].value or "")
        score = IMPACT_SCORES.get(impact, 0) * PROBABILITY_SCORES.get(probability, 0)
        threats.append(
            {
                "id": str(risk_id).strip(),
                "category": str(ws[f"D{row_num}"].value or ""),
                "event": str(ws[f"F{row_num}"].value or ""),
                "owner": str(ws[f"I{row_num}"].value or ""),
                "impact": impact,
                "probability": probability,
                "score": score,
            }
        )
    threats.sort(key=lambda item: (-int(item["score"]), item["id"]))
    return {"top": threats[:5], "top_score": int(threats[0]["score"]) if threats else 0}


def build_html() -> str:
    schedule = load_schedule()
    costs = load_costs()
    risks = load_risks()

    report_date = date.today().strftime("%d %B %Y")
    logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii") if LOGO_PATH.exists() else ""
    logo_tag = (
        f'<img src="data:image/png;base64,{logo_b64}" alt="Bosch" style="height:36px;display:block;" />'
        if logo_b64
        else ""
    )
    qg0 = schedule["milestones_by_name"]["QG0 - Programme kickoff approved"]["start"]
    qg1 = schedule["milestones_by_name"]["QG1 - Concept and transition model approved"]["start"]
    qg23 = schedule["milestones_by_name"]["QG2 and QG3 - Build complete and test entry approved"]["start"]
    qg4 = schedule["milestones_by_name"]["QG4 - Pre-GoLive gate approved"]["start"]
    golive = schedule["milestones_by_name"]["GoLive - Day 1 cutover to merger zone complete"]["start"]
    qg5 = schedule["milestones_by_name"]["QG5 - Project completion approved"]["start"]

    spi = 1.00
    cpi = 1.00
    pv = 0
    ev = 0
    ac = 0
    readiness_pct = 0
    overall_rag = "AMBER"
    tsa_confidence = "AMBER"

    workstream_rows = "".join(
        f'<div class="conf-bar-wrap"><div class="conf-bar-label"><span>{html_escape(name)}</span><span style="color:{"#007A33" if colour == "good" else "#E8A000"};font-weight:600;">{score}%</span></div><div class="conf-bar-outer"><div class="conf-bar-inner" style="width:{score}%;background:{"#007A33" if colour == "good" else "#E8A000"};"></div></div></div>'
        for name, score, colour in WORKSTREAM_CONFIDENCE
    )
    top_risk_rows = "".join(
        f'<tr><td>{html_escape(risk["id"])}</td><td>{html_escape(risk["event"])}</td><td>{html_escape(risk["category"])}</td><td><span class="score-badge {"sb-r" if risk["score"] >= 15 else "sb-a"}">{risk["score"]}</span></td><td>{html_escape(risk["owner"])}</td></tr>'
        for risk in risks["top"]
    )
    action_rows = "".join(
        f'<div class="action-row"><div class="action-cat">{html_escape(fmt_date(task["start"]))}</div><div class="action-text"><strong>{html_escape(task["name"])}</strong><br>{html_escape(task["notes"])}</div><div class="action-due">{html_escape(countdown(task["start"]))}</div></div>'
        for task in schedule["next_actions"]
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{html_escape(PROJECT_NAME)} Management KPI Dashboard</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Segoe UI',Arial,sans-serif;font-size:12px;color:#1a1a1a;background:#f4f6f9;}}
.hdr{{background:linear-gradient(135deg,#003b6e 0%,#005199 100%);color:#fff;padding:14px 28px;display:flex;align-items:center;gap:20px;}}
.bosch-logo{{display:flex;align-items:center;background:#fff;padding:4px 8px;border-radius:4px;}}
.hdr-center{{flex:1;}}
.hdr-center h1{{font-size:17px;font-weight:700;letter-spacing:.4px;}}
.hdr-center h2{{font-size:11px;font-weight:400;opacity:.85;margin-top:3px;}}
.hdr-right{{text-align:right;font-size:11px;opacity:.85;}}
.hdr-right strong{{font-size:13px;display:block;}}
.rag-strip{{background:#003b6e;display:flex;gap:0;border-top:2px solid rgba(255,255,255,.1);}}
.rag-box{{flex:1;border-right:1px solid rgba(255,255,255,.12);padding:7px 14px;}}
.rag-box:last-child{{border-right:none;}}
.rag-label{{font-size:9px;color:#fff;opacity:.7;text-transform:uppercase;letter-spacing:.4px;}}
.rag-val{{font-size:13px;font-weight:700;}}
.rag-sub{{font-size:9px;color:#fff;opacity:.6;}}
.rv-a{{color:#f1c40f;}} .rv-r{{color:#e74c3c;}}
.page{{max-width:1200px;margin:0 auto;padding:14px 18px 30px;}}
.grid{{display:grid;grid-template-columns:repeat(12,1fr);gap:12px;margin-bottom:12px;}}
.col-3{{grid-column:span 3;}} .col-4{{grid-column:span 4;}} .col-5{{grid-column:span 5;}} .col-6{{grid-column:span 6;}} .col-7{{grid-column:span 7;}} .col-12{{grid-column:span 12;}}
.kpi-card{{background:#fff;border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,.08);overflow:hidden;}}
.kpi-title{{background:#005199;color:#fff;font-size:10px;font-weight:700;padding:6px 12px;letter-spacing:.4px;text-transform:uppercase;}}
.kpi-body{{padding:10px 12px;}}
.gauge-wrap{{text-align:center;padding:4px 0 8px;}}
.conf-bar-wrap{{margin:4px 0;}}
.conf-bar-label{{display:flex;justify-content:space-between;font-size:10px;margin-bottom:2px;}}
.conf-bar-outer{{height:10px;background:#e8ecf2;border-radius:5px;overflow:hidden;}}
.conf-bar-inner{{height:100%;border-radius:5px;}}
.ms-timeline{{position:relative;padding:8px 0;}}
.ms-line{{position:absolute;top:22px;left:12px;right:12px;height:3px;background:#dee3ed;border-radius:2px;}}
.ms-nodes{{display:flex;justify-content:space-between;position:relative;padding:0 6px;}}
.ms-node{{text-align:center;width:84px;}}
.ms-dot{{width:20px;height:20px;border-radius:50%;margin:0 auto 4px;border:3px solid #fff;box-shadow:0 0 0 2px #005199;background:#005199;}}
.ms-dot.active{{background:#E8A000;box-shadow:0 0 0 2px #E8A000;}}
.ms-node-name{{font-size:9px;font-weight:700;color:#003b6e;}}
.ms-node-date{{font-size:8px;color:#888;}}
.ms-node-cd{{font-size:8px;color:#0066CC;font-weight:600;}}
table.kpi-tbl{{width:100%;border-collapse:collapse;font-size:10px;}}
.kpi-tbl th{{background:#003b6e;color:#fff;padding:5px 7px;text-align:left;font-size:9px;font-weight:600;}}
.kpi-tbl td{{padding:5px 7px;border-bottom:1px solid #e8ecf2;vertical-align:top;}}
.kpi-tbl tr:nth-child(even){{background:#EFF4FB;}}
.pill{{display:inline-block;padding:2px 6px;border-radius:10px;font-size:8px;font-weight:700;}}
.p-a{{background:#fef9e7;color:#b7770d;}} .p-blue{{background:#EFF4FB;color:#005199;}}
.score-badge{{display:inline-block;width:22px;height:22px;border-radius:50%;font-size:10px;font-weight:700;text-align:center;line-height:22px;color:#fff;}}
.sb-r{{background:#e74c3c;}} .sb-a{{background:#f39c12;}}
.action-row{{display:flex;gap:8px;padding:5px 0;border-bottom:1px solid #e8ecf2;align-items:flex-start;font-size:10px;}}
.action-row:last-child{{border-bottom:none;}}
.action-cat{{min-width:76px;font-weight:600;color:#005199;}}
.action-text{{flex:1;color:#333;line-height:1.5;}}
.action-due{{min-width:70px;text-align:right;font-size:9px;color:#888;}}
.footer{{text-align:center;font-size:9px;color:#aaa;padding:8px 0 16px;}}
</style>
</head>
<body>
<div class="hdr">
  <div class="bosch-logo">{logo_tag}</div>
  <div class="hdr-center">
    <h1>{html_escape(PROJECT_NAME)} - Management KPI Dashboard</h1>
    <h2>{html_escape(SELLER)} -> merger zone ({html_escape(MODEL)}) -> {html_escape(BUYER)} | {SITE_COUNT} sites | {USER_COUNT:,} users</h2>
  </div>
  <div class="hdr-right">
    <strong>{html_escape(report_date)}</strong>
    GoLive: {html_escape(fmt_date(golive))} ({html_escape(countdown(golive))})
  </div>
</div>

<div class="rag-strip">
  <div class="rag-box"><div class="rag-label">Schedule (SPI)</div><div class="rag-val rv-a">SPI {spi:.2f}</div><div class="rag-sub">Pre-kickoff baseline view</div></div>
  <div class="rag-box"><div class="rag-label">Cost (CPI)</div><div class="rag-val rv-a">CPI {cpi:.2f}</div><div class="rag-sub">No actual cost booked yet</div></div>
  <div class="rag-box"><div class="rag-label">Day 1 Readiness</div><div class="rag-val rv-a">{readiness_pct}%</div><div class="rag-sub">No milestones passed yet</div></div>
    <div class="rag-box"><div class="rag-label">TSA Confidence</div><div class="rag-val rv-a">{tsa_confidence}</div><div class="rag-sub">Pressure reduced by approved seller-side buffer</div></div>
  <div class="rag-box"><div class="rag-label">Top Risk</div><div class="rag-val rv-r">Score {risks['top_score']}</div><div class="rag-sub">SAP carve-out critical path</div></div>
  <div class="rag-box"><div class="rag-label">Overall RAG</div><div class="rag-val rv-a">{overall_rag}</div><div class="rag-sub">Planning baseline before QG0</div></div>
</div>

<div class="page">
<div class="grid">
  <div class="kpi-card col-3"><div class="kpi-title">Schedule Performance (SPI)</div><div class="kpi-body"><div class="gauge-wrap"><div style="font-size:34px;font-weight:700;color:#1a1a1a;">{spi:.2f}</div><div style="font-size:10px;color:#666;">EV / PV baseline before kick-off</div></div><div style="margin-top:8px;font-size:10px;"><div style="display:flex;justify-content:space-between;"><span>Planned Value (PV)</span><span>{fmt_eur(pv)}</span></div><div style="display:flex;justify-content:space-between;"><span>Earned Value (EV)</span><span>{fmt_eur(ev)}</span></div><div style="display:flex;justify-content:space-between;font-weight:600;margin-top:3px;"><span>SPI Status</span><span class="pill p-a">BASELINE</span></div></div></div></div>
  <div class="kpi-card col-3"><div class="kpi-title">Cost Performance (CPI)</div><div class="kpi-body"><div class="gauge-wrap"><div style="font-size:34px;font-weight:700;color:#1a1a1a;">{cpi:.2f}</div><div style="font-size:10px;color:#666;">No actuals booked before programme start</div></div><div style="margin-top:8px;font-size:10px;"><div style="display:flex;justify-content:space-between;"><span>Budget at Completion</span><span>{fmt_eur(costs['total_labour'])}</span></div><div style="display:flex;justify-content:space-between;"><span>Actual Cost (AC)</span><span>{fmt_eur(ac)}</span></div><div style="display:flex;justify-content:space-between;font-weight:600;margin-top:3px;"><span>CPI Status</span><span class="pill p-a">BASELINE</span></div></div></div></div>
  <div class="kpi-card col-3"><div class="kpi-title">Day 1 Readiness</div><div class="kpi-body"><div class="gauge-wrap"><div style="font-size:34px;font-weight:700;color:#0066CC;">{readiness_pct}%</div><div style="font-size:10px;color:#666;">Milestone completion based</div></div><div style="margin-top:8px;font-size:10px;"><div style="display:flex;justify-content:space-between;"><span>Gates Passed</span><span>0 / 6</span></div><div style="display:flex;justify-content:space-between;"><span>Time Elapsed</span><span>{schedule['time_elapsed_pct']}%</span></div><div style="display:flex;justify-content:space-between;font-weight:600;margin-top:3px;"><span>Target GoLive</span><span>{fmt_date(golive)}</span></div></div></div></div>
    <div class="kpi-card col-3"><div class="kpi-title">TSA and Integration Confidence</div><div class="kpi-body"><div style="text-align:center;padding:6px 0 8px;"><div style="font-size:28px;font-weight:700;color:#E8A000;">{tsa_confidence}</div><div style="font-size:10px;color:#666;margin-top:3px;">Seller to merger-zone to buyer transition</div></div><div style="font-size:10px;margin-top:6px;"><div style="display:flex;justify-content:space-between;margin:3px 0;"><span>Approved JCI TSA end</span><span class="pill p-a">31 Jul 2027</span></div><div style="display:flex;justify-content:space-between;margin:3px 0;"><span>Infosys role</span><span class="pill p-blue">ACTIVE</span></div><div style="display:flex;justify-content:space-between;margin:3px 0;"><span>User position</span><span class="pill p-blue">LEGACY JCI</span></div><div style="display:flex;justify-content:space-between;margin:3px 0;"><span>Model</span><span class="pill p-blue">{html_escape(MODEL)}</span></div></div></div></div>
</div>

<div class="grid">
  <div class="kpi-card col-5"><div class="kpi-title">Workstream Confidence Scores</div><div class="kpi-body">{workstream_rows}</div></div>
  <div class="kpi-card col-7"><div class="kpi-title">Milestone Gate Control Timeline</div><div class="kpi-body"><div class="ms-timeline"><div class="ms-line"></div><div class="ms-nodes"><div class="ms-node"><div class="ms-dot active"></div><div class="ms-node-name">QG0</div><div class="ms-node-date">{fmt_date(qg0)}</div><div class="ms-node-cd">{countdown(qg0)}</div></div><div class="ms-node"><div class="ms-dot"></div><div class="ms-node-name">QG1</div><div class="ms-node-date">{fmt_date(qg1)}</div><div class="ms-node-cd">{countdown(qg1)}</div></div><div class="ms-node"><div class="ms-dot"></div><div class="ms-node-name">QG2&3</div><div class="ms-node-date">{fmt_date(qg23)}</div><div class="ms-node-cd">{countdown(qg23)}</div></div><div class="ms-node"><div class="ms-dot"></div><div class="ms-node-name">QG4</div><div class="ms-node-date">{fmt_date(qg4)}</div><div class="ms-node-cd">{countdown(qg4)}</div></div><div class="ms-node"><div class="ms-dot"></div><div class="ms-node-name">GoLive</div><div class="ms-node-date">{fmt_date(golive)}</div><div class="ms-node-cd">{countdown(golive)}</div></div><div class="ms-node"><div class="ms-dot"></div><div class="ms-node-name">QG5</div><div class="ms-node-date">{fmt_date(qg5)}</div><div class="ms-node-cd">{countdown(qg5)}</div></div></div></div><div style="margin-top:10px;font-size:10px;color:#666;">Programme has not reached QG0 yet; this dashboard is a steering baseline for mobilisation and first-wave control.</div></div></div>
</div>

<div class="grid">
  <div class="kpi-card col-7"><div class="kpi-title">Top Risk Table</div><div class="kpi-body"><table class="kpi-tbl"><tr><th>ID</th><th>Risk</th><th>Category</th><th>Score</th><th>Owner</th></tr>{top_risk_rows}</table></div></div>
    <div class="kpi-card col-5"><div class="kpi-title">Model Key Differences</div><div class="kpi-body"><table class="kpi-tbl"><tr><th>Area</th><th>Current Programme Position</th></tr><tr><td>Transition path</td><td>JCI services move first into a temporary Infosys-operated merger zone before Bosch handover.</td></tr><tr><td>TSA dependency</td><td>Approved seller-service coverage through 31 Jul 2027 lets users remain on legacy JCI while Infosys completes the merger-zone build, reducing pressure without changing overall programme progress.</td></tr><tr><td>GoLive definition</td><td>Day 1 means all in-scope services operate from the merger zone with Bosch acceptance path prepared.</td></tr><tr><td>Post-GoLive scope</td><td>Only stabilisation, TSA exit, and Bosch handover remain after GoLive; no new migration scope.</td></tr></table></div></div>
</div>

<div class="grid">
  <div class="kpi-card col-12"><div class="kpi-title">Next 90 Days Action Forecast</div><div class="kpi-body">{action_rows if action_rows else '<div style="font-size:10px;color:#666;">No task starts fall within the next 90 days from the current reporting date.</div>'}</div></div>
</div>

<div class="footer">{html_escape(PROJECT_NAME)} Management KPI Dashboard | {html_escape(report_date)} | Data sources: {html_escape(SCHEDULE_PATH.name)}, {html_escape(RISK_PATH.name)}, {html_escape(COST_PATH.name)} | CONFIDENTIAL</div>
</div>
</body>
</html>
"""


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(build_html(), encoding="utf-8")
    print(f"[{PROJECT_NAME}] KPI Dashboard: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()