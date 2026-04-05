#!/usr/bin/env python3
"""
Eagle Project - Schedule Generator
Stand Alone carve-out: Security camera business
3 sites (Singapore), 300 IT users
Start: 2026-06-01, GoLive: 2026-11-01, Completion: 2027-03-31
"""

import subprocess
import sys
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
PROJECT_NAME = "Eagle"
OUTPUT_DIR = HERE / "active-projects" / PROJECT_NAME
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

XLSX_PATH = OUTPUT_DIR / f"{PROJECT_NAME}_Project_Schedule.xlsx"
CSV_PATH = OUTPUT_DIR / f"{PROJECT_NAME}_Project_Schedule.csv"
XML_PATH = OUTPUT_DIR / f"{PROJECT_NAME}_Project_Schedule.xml"

# Project Dates
START_DATE = datetime(2026, 6, 1)
GOLIVE_DATE = datetime(2026, 11, 7)  # Shifted: QG4 (Nov 1) + 5 days Final Readiness (Nov 2-6) + 1 day GoLive milestone
COMPLETION_DATE = datetime(2027, 4, 7)  # Shifted +6 days to keep 90-day hypercare + closure

def date_plus_days(base, days):
    """Add calendar days to a date"""
    return base + timedelta(days=days)

# Task list: (ID, OutlineLevel, Name, Duration, Start, Finish, Predecessors, ResourceNames, Notes, Milestone)
# STRICT QUALITY GATE SEQUENCE:
# QG0 (Intake) -> QG1 (Concept) -> QG2&3 (Build/Test) -> QG4 (Pre-GoLive) -> GoLive -> 90-day Hypercare -> QG5 (Completion) -> Closing
TASKS = [
    # Project root
    (0, 0, "Eagle - Security Camera Business Carve-out", 
     (COMPLETION_DATE - START_DATE).days, START_DATE, COMPLETION_DATE, "", "", "", "No"),
    
    # PHASE 0: Initiation (Jun 1-25, 25 days) → QG0
    (1, 1, "0. INITIATION & QG0", 25, 
     date_plus_days(START_DATE, 0), date_plus_days(START_DATE, 25), "", "", "", "No"),
    (2, 2, "0.1 Project Kickoff & Governance", 5, 
     date_plus_days(START_DATE, 0), date_plus_days(START_DATE, 5), "", "Ho Keng Hua + PMO", "", "No"),
    (3, 3, "0.1.1 Establish PMO & workstreams", 3, 
     date_plus_days(START_DATE, 0), date_plus_days(START_DATE, 3), "", "Ho Keng Hua", "", "No"),
    (4, 3, "0.1.2 Define governance & forums", 2, 
     date_plus_days(START_DATE, 3), date_plus_days(START_DATE, 5), "3", "PMO + Legal", "", "No"),
    (5, 2, "0.2 Business & IT Scope Definition", 12, 
     date_plus_days(START_DATE, 5), date_plus_days(START_DATE, 17), "", "Business Analyst + IT Lead", "", "No"),
    (6, 3, "0.2.1 Document business processes", 4, 
     date_plus_days(START_DATE, 5), date_plus_days(START_DATE, 9), "", "Business Analyst", "", "No"),
    (7, 3, "0.2.2 Inventory IT systems & data", 4, 
     date_plus_days(START_DATE, 5), date_plus_days(START_DATE, 9), "", "IT Lead", "", "No"),
    (8, 3, "0.2.3 Define separation scope", 4, 
     date_plus_days(START_DATE, 9), date_plus_days(START_DATE, 13), "6,7", "Legal + IT Lead", "", "No"),
    (9, 2, "0.3 Stakeholder & Change Readiness", 8, 
     date_plus_days(START_DATE, 5), date_plus_days(START_DATE, 13), "4", "Change Lead", "", "No"),
    (10, 3, "0.3.1 User readiness assessment", 4, 
     date_plus_days(START_DATE, 5), date_plus_days(START_DATE, 9), "4", "Change Lead", "", "No"),
    (11, 3, "0.3.2 Site consultation", 4, 
     date_plus_days(START_DATE, 9), date_plus_days(START_DATE, 13), "10", "Change Lead", "", "No"),
    (12, 2, "0.4 QG0 - Intake Approval", 1, 
     date_plus_days(START_DATE, 25), date_plus_days(START_DATE, 26), "8,11", "PMO + Steering", "Project formally approved", "Yes"),
    
    # PHASE 1: Concept & Scope (Jun 26 - Aug 20, 56 days) → QG1
    (13, 1, "1. CONCEPT & SCOPE PHASE", 56, 
     date_plus_days(START_DATE, 25), date_plus_days(START_DATE, 81), "12", "", "", "No"),
    (14, 2, "1.1 IT Architecture Baseline", 20, 
     date_plus_days(START_DATE, 26), date_plus_days(START_DATE, 46), "13", "Infrastructure Architect", "", "No"),
    (15, 3, "1.1.1 Current state infrastructure mapping", 5, 
     date_plus_days(START_DATE, 26), date_plus_days(START_DATE, 31), "13", "Infrastructure Architect", "", "No"),
    (16, 3, "1.1.2 Design target architecture", 10, 
     date_plus_days(START_DATE, 31), date_plus_days(START_DATE, 41), "15", "Infrastructure Architect", "", "No"),
    (17, 3, "1.1.3 Network & security design", 5, 
     date_plus_days(START_DATE, 41), date_plus_days(START_DATE, 46), "16", "Security Architect", "", "No"),
    (18, 2, "1.2 Application & Data Strategy", 22, 
     date_plus_days(START_DATE, 26), date_plus_days(START_DATE, 48), "13", "App Lead + Data Architect", "", "No"),
    (19, 3, "1.2.1 Application inventory & assessment", 6, 
     date_plus_days(START_DATE, 26), date_plus_days(START_DATE, 32), "13", "App Lead", "", "No"),
    (20, 3, "1.2.2 Data dependency mapping", 8, 
     date_plus_days(START_DATE, 32), date_plus_days(START_DATE, 40), "19", "Data Architect", "", "No"),
    (21, 3, "1.2.3 Carve-out strategy document", 8, 
     date_plus_days(START_DATE, 40), date_plus_days(START_DATE, 48), "20", "App Lead + Data Architect", "", "No"),
    (22, 2, "1.3 TSA & Transition Planning", 15, 
     date_plus_days(START_DATE, 45), date_plus_days(START_DATE, 60), "14,18", "IT Lead + Finance", "", "No"),
    (23, 3, "1.3.1 Define TSA scope (Seller->Buyer)", 8, 
     date_plus_days(START_DATE, 45), date_plus_days(START_DATE, 53), "14", "IT Lead", "", "No"),
    (24, 3, "1.3.2 Plan transition & handover", 7, 
     date_plus_days(START_DATE, 53), date_plus_days(START_DATE, 60), "23", "Finance + IT Lead", "", "No"),
    (25, 2, "1.4 QG1 - Concept Gate", 1, 
     date_plus_days(START_DATE, 81), date_plus_days(START_DATE, 82), "17,21,24", "PMO + Steering", "Concept phase approved", "Yes"),
    
    # PHASE 2: Design (Aug 21 - Oct 04, 45 days)
    (26, 1, "2. DESIGN PHASE", 45, 
     date_plus_days(START_DATE, 82), date_plus_days(START_DATE, 127), "25", "", "", "No"),
    (27, 2, "2.1 Infrastructure Design Details", 15, 
     date_plus_days(START_DATE, 83), date_plus_days(START_DATE, 98), "25", "Infrastructure Architect", "", "No"),
    (28, 3, "2.1.1 Network design (3 Singapore sites)", 5, 
     date_plus_days(START_DATE, 83), date_plus_days(START_DATE, 88), "25", "Infrastructure Architect", "", "No"),
    (29, 3, "2.1.2 Server & storage detail design", 5, 
     date_plus_days(START_DATE, 88), date_plus_days(START_DATE, 93), "28", "Infrastructure Architect", "", "No"),
    (30, 3, "2.1.3 Security & compliance hardening", 5, 
     date_plus_days(START_DATE, 93), date_plus_days(START_DATE, 98), "29", "Security Architect", "", "No"),
    (31, 2, "2.2 ERP & Database Design", 18, 
     date_plus_days(START_DATE, 83), date_plus_days(START_DATE, 101), "25", "ERP Lead + DBA", "", "No"),
    (32, 3, "2.2.1 ERP separation strategy", 6, 
     date_plus_days(START_DATE, 83), date_plus_days(START_DATE, 89), "25", "ERP Lead", "", "No"),
    (33, 3, "2.2.2 Data migration & replication design", 6, 
     date_plus_days(START_DATE, 89), date_plus_days(START_DATE, 95), "32", "Data Architect", "", "No"),
    (34, 3, "2.2.3 Database backup & recovery plan", 6, 
     date_plus_days(START_DATE, 95), date_plus_days(START_DATE, 101), "33", "DBA", "", "No"),
    (35, 2, "2.3 Application Carve-Out Design", 16, 
     date_plus_days(START_DATE, 89), date_plus_days(START_DATE, 105), "25", "App Lead", "", "No"),
    (36, 3, "2.3.1 Application separation design", 6, 
     date_plus_days(START_DATE, 89), date_plus_days(START_DATE, 95), "25", "App Lead", "", "No"),
    (37, 3, "2.3.2 API & integration points", 5, 
     date_plus_days(START_DATE, 95), date_plus_days(START_DATE, 100), "36", "Integration Architect", "", "No"),
    (38, 3, "2.3.3 Data protection & PII handling", 5, 
     date_plus_days(START_DATE, 100), date_plus_days(START_DATE, 105), "37", "Legal + Data Privacy", "", "No"),
    (39, 2, "2.4 Client & Endpoint Strategy", 14, 
     date_plus_days(START_DATE, 93), date_plus_days(START_DATE, 107), "25", "Client Architect", "", "No"),
    (40, 3, "2.4.1 Endpoint OS & software assessment", 4, 
     date_plus_days(START_DATE, 93), date_plus_days(START_DATE, 97), "25", "Client Architect", "", "No"),
    (41, 3, "2.4.2 Device imaging & deployment design", 5, 
     date_plus_days(START_DATE, 97), date_plus_days(START_DATE, 102), "40", "Imaging Lead", "", "No"),
    (42, 3, "2.4.3 User migration plan (300 users)", 5, 
     date_plus_days(START_DATE, 102), date_plus_days(START_DATE, 107), "41", "Change Lead", "", "No"),
    (43, 2, "2.5 Cutover & Testing Strategy", 12, 
     date_plus_days(START_DATE, 104), date_plus_days(START_DATE, 116), "34,38", "Cutover Lead", "", "No"),
    (44, 3, "2.5.1 Cutover window definition (3 sites)", 3, 
     date_plus_days(START_DATE, 104), date_plus_days(START_DATE, 107), "34", "Cutover Lead", "", "No"),
    (45, 3, "2.5.2 UAT & testing strategy", 5, 
     date_plus_days(START_DATE, 107), date_plus_days(START_DATE, 112), "44", "QA Lead", "", "No"),
    (46, 3, "2.5.3 Rollback & contingency plan", 4, 
     date_plus_days(START_DATE, 112), date_plus_days(START_DATE, 116), "45", "Cutover Lead", "", "No"),
    
    # PHASE 3: Build & Test (Oct 05 - Oct 28, 24 days) → QG2&3
    (47, 1, "3. BUILD & TEST PHASE", 24, 
     date_plus_days(START_DATE, 126), date_plus_days(START_DATE, 150), "43,46", "", "", "No"),
    (48, 2, "3.1 Infrastructure Build-Out", 12, 
     date_plus_days(START_DATE, 127), date_plus_days(START_DATE, 139), "47", "Infrastructure Team", "", "No"),
    (49, 3, "3.1.1 Procure & deploy hardware (3 sites)", 6, 
     date_plus_days(START_DATE, 127), date_plus_days(START_DATE, 133), "47", "Procurement + Infrastructure", "", "No"),
    (50, 3, "3.1.2 Network deployment & security", 4, 
     date_plus_days(START_DATE, 133), date_plus_days(START_DATE, 137), "49", "Infrastructure Team", "", "No"),
    (51, 3, "3.1.3 Infrastructure testing & validation", 2, 
     date_plus_days(START_DATE, 137), date_plus_days(START_DATE, 139), "50", "Infrastructure Team", "", "No"),
    (52, 2, "3.2 ERP Build & System Migration", 14, 
     date_plus_days(START_DATE, 127), date_plus_days(START_DATE, 141), "47", "ERP Team + DBA", "", "No"),
    (53, 3, "3.2.1 ERP system clone/build", 5, 
     date_plus_days(START_DATE, 127), date_plus_days(START_DATE, 132), "47", "ERP Team", "", "No"),
    (54, 3, "3.2.2 Data extraction & transformation", 4, 
     date_plus_days(START_DATE, 132), date_plus_days(START_DATE, 136), "53", "Data Architect", "", "No"),
    (55, 3, "3.2.3 ERP configuration & testing", 3, 
     date_plus_days(START_DATE, 136), date_plus_days(START_DATE, 139), "54", "ERP Consultant", "", "No"),
    (56, 3, "3.2.4 Data validation & reconciliation", 2, 
     date_plus_days(START_DATE, 139), date_plus_days(START_DATE, 141), "55", "ERP Lead + DBA", "", "No"),
    (57, 2, "3.3 Application Build & Config", 12, 
     date_plus_days(START_DATE, 132), date_plus_days(START_DATE, 144), "47", "App Dev Team", "", "No"),
    (58, 3, "3.3.1 Clone/provision applications", 5, 
     date_plus_days(START_DATE, 132), date_plus_days(START_DATE, 137), "47", "App Lead", "", "No"),
    (59, 3, "3.3.2 Configure integrations & APIs", 4, 
     date_plus_days(START_DATE, 137), date_plus_days(START_DATE, 141), "58", "Integration Architect", "", "No"),
    (60, 3, "3.3.3 Application testing", 3, 
     date_plus_days(START_DATE, 141), date_plus_days(START_DATE, 144), "59", "QA Team", "", "No"),
    (61, 2, "3.4 Client Imaging & Pilot", 10, 
     date_plus_days(START_DATE, 135), date_plus_days(START_DATE, 145), "47", "Imaging & Deployment", "", "No"),
    (62, 3, "3.4.1 Prepare images & deployment", 3, 
     date_plus_days(START_DATE, 135), date_plus_days(START_DATE, 138), "47", "Imaging Lead", "", "No"),
    (63, 3, "3.4.2 Pilot deployment (Site 1, ~30 users)", 3, 
     date_plus_days(START_DATE, 138), date_plus_days(START_DATE, 141), "62", "Deployment Team", "", "No"),
    (64, 3, "3.4.3 Pilot feedback & adjustments", 4, 
     date_plus_days(START_DATE, 141), date_plus_days(START_DATE, 145), "63", "Change Lead + Users", "", "No"),
    (65, 2, "3.5 QG2&3 - Build & Test Gate", 1, 
     date_plus_days(START_DATE, 150), date_plus_days(START_DATE, 151), "51,56,60,64", "PMO + QA", "Build phase complete, ready for UAT", "Yes"),
    
    # PHASE 4: Pre-GoLive UAT & Cutover (Oct 28 - Oct 31, 4 days) → QG4
    (66, 1, "4. FINAL UAT & CUTOVER PREP", 4, 
     date_plus_days(START_DATE, 150), date_plus_days(START_DATE, 154), "65", "", "", "No"),
    (67, 2, "4.1 Full UAT Execution", 2, 
     date_plus_days(START_DATE, 151), date_plus_days(START_DATE, 152), "65", "QA Lead + Business", "", "No"),
    (68, 3, "4.1.1 Run UAT test scripts (all systems)", 1, 
     date_plus_days(START_DATE, 151), date_plus_days(START_DATE, 151), "65", "QA + Business Testers", "", "No"),
    (69, 3, "4.1.2 Critical defect resolution", 1, 
     date_plus_days(START_DATE, 151), date_plus_days(START_DATE, 152), "68", "Dev + QA", "", "No"),
    (70, 2, "4.2 Pre-GoLive Readiness Check", 1, 
     date_plus_days(START_DATE, 152), date_plus_days(START_DATE, 153), "69", "PMO + Tech Leads", "", "No"),
    (71, 3, "4.2.1 Final infrastructure & system checks", 0.5, 
     date_plus_days(START_DATE, 152), date_plus_days(START_DATE, 152.5), "69", "Infrastructure Lead", "", "No"),
    (72, 3, "4.2.2 Cutover rollback test", 0.5, 
     date_plus_days(START_DATE, 152.5), date_plus_days(START_DATE, 153), "71", "Cutover Lead", "", "No"),
    (73, 2, "4.3 QG4 - Pre-GoLive Approval", 1, 
     date_plus_days(START_DATE, 153), date_plus_days(START_DATE, 154), "70", "PMO + Steering", "QG4 gate passed; ready for final cutover checks", "Yes"),
    
    # PHASE 4b: Final Readiness & Open Item Closure (Nov 1-2, 2 days)
    (74, 1, "4b. FINAL READINESS & CUTOVER VERIFICATION", 2,
     date_plus_days(START_DATE, 154), date_plus_days(START_DATE, 156), "73", "", "", "No"),
    (75, 2, "4b.1 UAT Closure & Sign-Off", 1,
     date_plus_days(START_DATE, 154), date_plus_days(START_DATE, 155), "73", "QA Lead + Business", "", "No"),
    (76, 3, "4b.1.1 Verify all UAT defects resolved", 0.5,
     date_plus_days(START_DATE, 154), date_plus_days(START_DATE, 154.5), "73", "QA Lead", "", "No"),
    (77, 3, "4b.1.2 UAT sign-off document", 0.5,
     date_plus_days(START_DATE, 154.5), date_plus_days(START_DATE, 155), "76", "Business Owner", "", "No"),
    (78, 2, "4b.2 Final Open Items Verification", 1,
     date_plus_days(START_DATE, 155), date_plus_days(START_DATE, 156), "77", "PMO + Steering", "", "No"),
    (79, 3, "4b.2.1 Confirm all action items closed", 0.5,
     date_plus_days(START_DATE, 155), date_plus_days(START_DATE, 155.5), "77", "PMO", "", "No"),
    (80, 3, "4b.2.2 Final infrastructure readiness", 0.5,
     date_plus_days(START_DATE, 155.5), date_plus_days(START_DATE, 156), "79", "Infrastructure Lead", "", "No"),
    
    # GOLIVE MILESTONE (Nov 03 = day 156)
    (80, 1, "5. GOLIVE & HYPERCARE", 91, 
     GOLIVE_DATE, date_plus_days(GOLIVE_DATE, 90), "80", "", "", "No"),
    (82, 2, "5.1 GoLive - Day 1 Cutover", 1, 
     GOLIVE_DATE, date_plus_days(GOLIVE_DATE, 1), "81", "PMO + Ops Team", "Production go-live", "Yes"),
    (83, 2, "5.2 Hypercare & Stabilization (90 days)", 90, 
     date_plus_days(GOLIVE_DATE, 1), date_plus_days(GOLIVE_DATE, 90), "82", "Support Team + SMEs", "90-day post-GoLive support", "No"),
    (84, 3, "5.2.1 24/7 L3 support & issue triage", 60, 
     date_plus_days(GOLIVE_DATE, 1), date_plus_days(GOLIVE_DATE, 60), "82", "L3 Support + Engineers", "", "No"),
    (85, 3, "5.2.2 Performance tuning & hotfixes", 45, 
     date_plus_days(GOLIVE_DATE, 1), date_plus_days(GOLIVE_DATE, 45), "82", "Performance Engineer", "", "No"),
    (86, 3, "5.2.3 User training & enablement", 60, 
     date_plus_days(GOLIVE_DATE, 1), date_plus_days(GOLIVE_DATE, 60), "82", "Training Team", "", "No"),
    (87, 3, "5.2.4 Knowledge transfer & documentation", 30, 
     date_plus_days(GOLIVE_DATE, 60), date_plus_days(GOLIVE_DATE, 90), "84", "Support Lead", "", "No"),
    
    # PHASE 6: Closure (Feb 02 - Apr 02, 60 days) → QG5
    (88, 1, "6. OPERATIONS TRANSITION & CLOSURE", 60, 
     date_plus_days(GOLIVE_DATE, 91), COMPLETION_DATE, "83", "", "", "No"),
    (89, 2, "6.1 Transition to Steady-State", 20, 
     date_plus_days(GOLIVE_DATE, 91), date_plus_days(GOLIVE_DATE, 110), "83", "Ops Lead + IT Manager", "", "No"),
    (90, 3, "6.1.1 Operations runbook finalization", 5, 
     date_plus_days(GOLIVE_DATE, 91), date_plus_days(GOLIVE_DATE, 95), "83", "Ops Lead", "", "No"),
    (91, 3, "6.1.2 Incident & escalation procedures", 5, 
     date_plus_days(GOLIVE_DATE, 95), date_plus_days(GOLIVE_DATE, 100), "90", "Ops Lead", "", "No"),
    (92, 3, "6.1.3 Capacity planning & baseline", 5, 
     date_plus_days(GOLIVE_DATE, 100), date_plus_days(GOLIVE_DATE, 105), "91", "Infrastructure", "", "No"),
    (93, 3, "6.1.4 Transition BAU support model", 5, 
     date_plus_days(GOLIVE_DATE, 105), date_plus_days(GOLIVE_DATE, 110), "92", "IT Manager", "", "No"),
    (94, 2, "6.2 Finance & Benefit Realization", 20, 
     date_plus_days(GOLIVE_DATE, 100), date_plus_days(GOLIVE_DATE, 120), "93", "Finance + PMO", "", "No"),
    (95, 3, "6.2.1 Financial separation & P&L", 10, 
     date_plus_days(GOLIVE_DATE, 100), date_plus_days(GOLIVE_DATE, 110), "93", "Finance Lead", "", "No"),
    (96, 3, "6.2.2 Cost allocation & chargeback", 10, 
     date_plus_days(GOLIVE_DATE, 110), date_plus_days(GOLIVE_DATE, 120), "95", "Finance Lead", "", "No"),
    (97, 2, "6.3 Lessons Learned & Closure", 15, 
     date_plus_days(GOLIVE_DATE, 110), date_plus_days(GOLIVE_DATE, 125), "96", "PMO + Team", "", "No"),
    (98, 3, "6.3.1 Lessons learned workshop", 3, 
     date_plus_days(GOLIVE_DATE, 110), date_plus_days(GOLIVE_DATE, 113), "96", "PMO", "", "No"),
    (99, 3, "6.3.2 Final project report & sign-off", 5, 
     date_plus_days(GOLIVE_DATE, 113), date_plus_days(GOLIVE_DATE, 118), "98", "Program Manager", "", "No"),
    (100, 3, "6.3.3 Project artifacts archival", 5, 
     date_plus_days(GOLIVE_DATE, 118), date_plus_days(GOLIVE_DATE, 123), "99", "PMO", "", "No"),
    (101, 3, "6.3.4 Final vendor settlement", 2, 
     date_plus_days(GOLIVE_DATE, 123), date_plus_days(GOLIVE_DATE, 125), "100", "Procurement + Finance", "", "No"),
    (102, 2, "6.4 QG5 - Project Completion", 1, 
     COMPLETION_DATE, date_plus_days(COMPLETION_DATE, 1), "101", "PMO + Steering", "Project formally closed", "Yes"),
]

def _generate_excel():
    """Generate formatted XLSX schedule"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Column headers
    headers = ["ID", "Level", "Task Name", "Duration (days)", "Start Date", "Finish Date", 
               "Predecessors", "Resources", "Notes", "Milestone"]
    ws.append(headers)
    
    # Header styling
    header_fill = PatternFill(start_color="002147", end_color="002147", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Task data
    for row_idx, task in enumerate(TASKS, start=2):
        task_id, level, name, duration, start, finish, preds, resources, notes, milestone = task
        
        # Format dates
        start_str = start.strftime("%Y-%m-%d")
        finish_str = finish.strftime("%Y-%m-%d")
        
        ws.append([task_id, level, name, duration, start_str, finish_str, preds, resources, notes, milestone])
        
        row = ws[row_idx]
        
        # Styling by level
        if milestone == "Yes":
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            font = Font(bold=True, color="000000", size=10)
        elif level == 1:
            fill = PatternFill(start_color="003B6E", end_color="003B6E", fill_type="solid")
            font = Font(bold=True, color="FFFFFF", size=10)
        elif level == 2:
            fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            font = Font(bold=True, color="FFFFFF", size=10)
        else:
            if row_idx % 2 == 0:
                fill = PatternFill(start_color="EFF4FB", end_color="EFF4FB", fill_type="solid")
            else:
                fill = None
            font = Font(color="000000", size=10)
        
        # Apply styles to all cells in row
        for cell in row:
            if fill:
                cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Center align numeric/date columns
        row[1].alignment = Alignment(horizontal="center", vertical="center")
        row[4].alignment = Alignment(horizontal="center", vertical="center")
        row[5].alignment = Alignment(horizontal="center", vertical="center")
        row[7].alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 12
    
    wb.save(XLSX_PATH)
    print(f"✓ XLSX generated: {XLSX_PATH}")

def _write_temp_csv(csv_path):
    """Write CSV for MS Project XML conversion"""
    with open(csv_path, 'w') as f:
        # Header
        f.write("ID,Outline Level,Name,Duration,Start,Finish,Predecessors,Resource Names,Notes,Milestone\n")
        
        # Task rows
        for task in TASKS:
            task_id, level, name, duration, start, finish, preds, resources, notes, milestone = task
            start_str = start.strftime("%Y-%m-%d")
            finish_str = finish.strftime("%Y-%m-%d")
            
            # Escape quotes in name
            name_escaped = name.replace('"', '""')
            resources_escaped = resources.replace('"', '""')
            notes_escaped = notes.replace('"', '""')
            
            f.write(f'{task_id},"{level}","{name_escaped}","{duration}","{start_str}","{finish_str}","{preds}","{resources_escaped}","{notes_escaped}","{milestone}"\n')
    
    print(f"✓ CSV generated: {csv_path}")

if __name__ == "__main__":
    print(f"\n{'='*70}")
    print(f"Eagle Project - Schedule Generator")
    print(f"{'='*70}\n")
    
    # Generate Excel
    _generate_excel()
    
    # Write CSV
    _write_temp_csv(CSV_PATH)
    
    # Generate MS Project XML
    print(f"|> Generating MS Project XML...")
    result = subprocess.run(
        [sys.executable,
         str(HERE / "generate_msp_xml.py"),
         "--csv", str(CSV_PATH),
         "--out", str(XML_PATH),
         "--project", PROJECT_NAME],
        capture_output=True,
        text=True
    )
    
    if result.returncode == 0:
        print(f"✓ XML generated: {XML_PATH}")
    else:
        print(f"✗ XML generation failed:")
        print(result.stderr)
        sys.exit(1)
    
    print(f"\n{'='*70}")
    print(f"Schedule generation complete!")
    print(f"  XLSX: {XLSX_PATH}")
    print(f"  CSV:  {CSV_PATH}")
    print(f"  XML:  {XML_PATH}")
    print(f"{'='*70}\n")
