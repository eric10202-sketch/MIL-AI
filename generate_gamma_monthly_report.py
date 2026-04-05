#!/usr/bin/env python3
"""Generate the Gamma monthly status report as a PDF."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


HERE = Path(__file__).parent
PROJECT_NAME = "Gamma"
OUTPUT_FOLDER_NAME = "Gamma v1.0"
SELLER = "Robert Bosch China"
BUYER = "Alibaba"
BUSINESS = "Bosch Cloud business"
MODEL = "Combination"
PMO = "EY"
REPORT_DATE = datetime(2026, 4, 5)
MONTH_TAG = REPORT_DATE.strftime("%b_%Y")

SCHEDULE_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Risk_Register.xlsx"
COST_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Cost_Plan.xlsx"
OUTPUT_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Monthly_Status_Report_{MONTH_TAG}.pdf"
LOGO_PATH = HERE / "Bosch.png"

IMPACT_VALUES = {"Very Low": 1, "Low": 2, "Moderate": 3, "High": 4, "Very High": 5}
PROBABILITY_VALUES = {"10%": 1, "30%": 2, "50%": 3, "70%": 4, "90%": 5}


def fmt_date(value) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%d %b %Y")
    return str(value)


def days_until(value) -> int:
    if hasattr(value, "strftime"):
        target = value
    else:
        target = datetime.strptime(str(value), "%Y-%m-%d")
    return (target - REPORT_DATE).days


def load_schedule_summary():
    wb = load_workbook(SCHEDULE_PATH, data_only=False)
    ws = wb["Schedule"]
    phases = []
    milestones = []
    start_date = None
    for row in range(2, ws.max_row + 1):
        level = ws.cell(row, 2).value
        if level is None:
            continue
        level = int(level)
        name = str(ws.cell(row, 3).value or "").strip()
        start = ws.cell(row, 5).value
        finish = ws.cell(row, 6).value
        if start_date is None and level == 1:
            start_date = start
        if level == 1:
            phases.append((name, start, finish))
        if str(ws.cell(row, 10).value) == "Yes" and ("QG" in name or "GoLive" in name or "Closure" in name):
            milestones.append((name, start, days_until(start)))
    return phases, milestones, start_date


def load_risks():
    wb = load_workbook(RISK_PATH, data_only=False)
    ws = wb["Risk Register"]
    risks = []
    high = 0
    for row in range(5, 140):
        risk_id = ws.cell(row, 2).value
        if risk_id is None:
            continue
        impact = ws.cell(row, 12).value
        probability = ws.cell(row, 14).value
        score = IMPACT_VALUES.get(impact, 0) * PROBABILITY_VALUES.get(probability, 0)
        if score >= 12 and ws.cell(row, 16).value == "threat":
            high += 1
        risks.append((int(risk_id), score, ws.cell(row, 4).value, ws.cell(row, 6).value, ws.cell(row, 9).value))
    risks.sort(key=lambda item: item[1], reverse=True)
    return high, risks[:3]


def load_budget():
    wb = load_workbook(COST_PATH, data_only=False)
    ws = wb["Cost Plan"]
    labour_total = 0
    capex_total = 0
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label == "OVERALL PROJECT TOTAL - LABOUR ONLY":
            labour_total = int(ws.cell(row, 6).value or 0)
        if label == "TOTAL CAPEX / ADDITIONAL COSTS":
            capex_total = int(ws.cell(row, 6).value or 0)
    return labour_total, capex_total


def build_section_title(text: str, styles):
    return Table(
        [[Paragraph(f"<b>{text}</b>", styles["Section"])]],
        colWidths=[170 * mm],
        style=TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#003b6e")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        ),
    )


def main() -> None:
    phases, milestones, start_date = load_schedule_summary()
    high_risks, top_risks = load_risks()
    labour_total, capex_total = load_budget()

    overall_status = "Amber"
    days_to_kickoff = days_until(start_date)
    days_to_golive = next(days for name, _, days in milestones if "GoLive" in name)
    qg0_days = next(days for name, _, days in milestones if name.startswith("QG0"))

    doc = SimpleDocTemplate(
        str(OUTPUT_PATH),
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleBlue", fontName="Helvetica-Bold", fontSize=18, textColor=colors.HexColor("#003b6e"), leading=22))
    styles.add(ParagraphStyle(name="Small", fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor("#1a1a1a")))
    styles.add(ParagraphStyle(name="Section", fontName="Helvetica-Bold", fontSize=9, leading=11, textColor=colors.white))
    styles.add(ParagraphStyle(name="BodyTight", fontName="Helvetica", fontSize=8.5, leading=10.5, textColor=colors.HexColor("#1a1a1a")))

    story = []
    story.append(Image(str(LOGO_PATH), width=32 * mm, height=9 * mm))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(f"{PROJECT_NAME} Monthly Status Report - {REPORT_DATE.strftime('%B %Y')}", styles["TitleBlue"]))
    story.append(Paragraph(f"{BUSINESS} carve-out | Seller: {SELLER} | Buyer: {BUYER} | Model: {MODEL}", styles["Small"]))
    story.append(Spacer(1, 3 * mm))

    metadata = Table(
        [
            ["Report date", REPORT_DATE.strftime("%d %b %Y"), "Overall status", overall_status, "Days to GoLive", str(days_to_golive)],
            ["Report month", REPORT_DATE.strftime("%B %Y"), "Budget status", "Draft for QG1", "Days to QG0", str(qg0_days)],
            ["PMO lead", PMO, "High threat risks", str(high_risks), "Days to kickoff", str(days_to_kickoff)],
        ],
        colWidths=[24 * mm, 28 * mm, 24 * mm, 24 * mm, 24 * mm, 22 * mm],
        style=TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f0f3f8")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d7deea")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        ),
    )
    story.append(metadata)
    story.append(Spacer(1, 3 * mm))

    summary_text = (
        "Gamma is currently in a pre-kickoff baseline state as of April 2026. The schedule, risk register, cost plan, charter, and dashboard artefacts are established, "
        "but delivery remains exposed to unresolved antitrust due diligence and a tightly restricted clean-team model. The project remains on its committed start date of 01 Aug 2026, "
        "with QG0 due on 04 Sep 2026 and GoLive fixed at 01 Feb 2027. Executive attention is required on legal timing, confidentiality control, and shared-service dependency evidence so that concept work can begin without immediate rework."
    )
    story.append(build_section_title("EXECUTIVE SUMMARY", styles))
    story.append(Spacer(1, 1.5 * mm))
    story.append(Paragraph(summary_text, styles["BodyTight"]))
    story.append(Spacer(1, 2.5 * mm))

    phase_data = [["Phase / Milestone", "Window / Date", "Status", "Days"]]
    for phase_name, start, finish in phases[:3]:
        phase_data.append([phase_name, f"{fmt_date(start)} - {fmt_date(finish)}", "Planned", str(days_until(start))])
    for name, start, days in milestones[:4]:
        phase_data.append([name, fmt_date(start), "Upcoming", str(days)])
    phase_table = Table(
        phase_data,
        colWidths=[62 * mm, 38 * mm, 28 * mm, 20 * mm],
        style=TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e4edf9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#003b6e")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d7deea")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        ),
    )
    story.append(build_section_title("PHASE STATUS AND GATE OUTLOOK", styles))
    story.append(Spacer(1, 1.5 * mm))
    story.append(phase_table)
    story.append(Spacer(1, 2.5 * mm))

    risk_data = [["Risk", "Score", "Owner", "Executive highlight"]]
    for risk_id, score, category, event, owner in top_risks:
        risk_data.append([f"#{risk_id} - {category}", str(score), owner, event])
    risk_table = Table(
        risk_data,
        colWidths=[40 * mm, 12 * mm, 34 * mm, 74 * mm],
        style=TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e4edf9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#003b6e")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d7deea")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        ),
    )
    story.append(build_section_title("RISK HIGHLIGHTS", styles))
    story.append(Spacer(1, 1.5 * mm))
    story.append(risk_table)
    story.append(Spacer(1, 2.5 * mm))

    budget_text = (
        f"Current Gamma labour baseline stands at EUR {labour_total:,.0f}. Risk-linked contingency and CAPEX lines total EUR {capex_total:,.0f} and are held outside the labour total. "
        "Budget status remains draft for QG1 approval because delivery has not yet started and the legal dependency path may still affect the final transition-service envelope."
    )
    story.append(build_section_title("BUDGET STATUS", styles))
    story.append(Spacer(1, 1.5 * mm))
    story.append(Paragraph(budget_text, styles["BodyTight"]))
    story.append(Spacer(1, 2.5 * mm))

    next_steps = [
        "Freeze the antitrust and disclosure decision path before July 2026 so mobilisation starts on stable assumptions.",
        "Complete clean-team roster confirmation and secure named infrastructure, security, and legal resource reservations.",
        "Pre-stage shared-service dependency evidence and third-party approval pre-reads for concept phase acceleration.",
        "Prepare phased communication and stakeholder-expansion controls for use once legal authorizes broader participation.",
    ]
    next_step_rows = [[Paragraph(f"<bullet>&bull;</bullet> {item}", styles["BodyTight"])] for item in next_steps]
    next_steps_table = Table(
        next_step_rows,
        colWidths=[170 * mm],
        style=TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 0.3, colors.HexColor("#d7deea")),
                ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#e6edf6")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        ),
    )
    story.append(build_section_title("NEXT 90 DAYS", styles))
    story.append(Spacer(1, 1.5 * mm))
    story.append(next_steps_table)
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(f"Data sources: {SCHEDULE_PATH.name}, {RISK_PATH.name}, {COST_PATH.name} | Confidential executive report", styles["Small"]))

    doc.build(story)
    print(f"[{PROJECT_NAME}] Monthly report written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()