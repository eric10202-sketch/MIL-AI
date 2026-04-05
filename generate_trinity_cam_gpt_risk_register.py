#!/usr/bin/env python3
"""Generate the Trinity-CAM (GPT) risk register from the BD template."""

from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))

from openpyxl import load_workbook
from openpyxl.styles import Font


HERE = Path(__file__).parent
PROJECT_NAME = "Trinity-CAM (GPT)"
PROJECT_CODE = "TCMGPT-2026"
DOCUMENT_ID = "TCMGPT-RR-001"
OWNER = "KPMG PMO Lead"
REPORT_DATE = "05.04.2026"
CREATION_DATE = "05.04.2026"

TEMPLATE_PATH = HERE / "Reference" / "BD_Risk-Register_template_en_V1.0_Dec2023.xlsx"
OUTPUT_PATH = HERE / "active-projects" / PROJECT_NAME / f"{PROJECT_NAME}_Risk_Register.xlsx"
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)


RISKS = [
    {
        "id": "R001",
        "category": "Technology, R&D",
        "cause": "More than 1,800 applications are in scope, including a large shared SAP estate with carve-out relevant finance, supply chain, service, and HR integrations. The source environment was built for JCI enterprise use rather than for legal separation, and the first clean technical dependency map will only be completed during Phase 1.",
        "event": "SAP carve-out build and system copy into the merger zone reveal shared master data, hard-coded integrations, or sequence dependencies that exceed the planned Phase 2 window and delay QG2 and QG3 on 2027-07-31.",
        "effect": "SAP build slips into Phase 3, integrated testing starts late, and the path to QG4 and GoLive 2028-01-01 compresses materially. Extended merger-zone run cost and steering escalation follow immediately.",
        "event_date": "2027-07-31",
        "owner": "KPMG SAP Architect",
        "source": "Formal Risk Review",
        "impact": "Very High",
        "probability": "70%",
        "type": "threat",
        "qualitative": "SAP is on the critical path and any delay directly affects testing, cutover rehearsal, and QG4 readiness.",
        "eur_cy": 3500000,
        "eur_3y": 1500000,
        "strategy": "Mitigate",
        "measure": "Complete a detailed SAP object and interface inventory by 2026-11-15; run rehearsal extraction and copy exercises before Phase 2 build; reserve dedicated Infosys SAP carve-out resources; maintain weekly critical-path tracking from QG1 onward.",
        "due_date": "2026-11-15",
        "status": "not started",
        "notes": "Highest schedule and cost exposure in the register.",
    },
    {
        "id": "R002",
        "category": "Technology, R&D",
        "cause": "Infosys is the primary delivery partner for merger-zone setup, operation, IT service migration, and application migration. The environment needs hosting, network, identity, service desk, collaboration services, monitoring, and security controls before scaled migration can begin.",
        "event": "Merger-zone core infrastructure is not operational by the planned Phase 2 entry date or remains unstable through build, preventing application onboarding and delaying user migration readiness.",
        "effect": "Wave planning becomes theoretical rather than executable, Phase 3 start shifts, and the programme loses the buffer needed between QG2 and QG4. Bosch is then forced to fund extra dual-run cost.",
        "event_date": "2027-07-31",
        "owner": "Infosys Programme Manager",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative": "The merger zone is the mandatory landing point between JCI and Bosch, so delays there cascade across all workstreams.",
        "eur_cy": 2200000,
        "eur_3y": 600000,
        "strategy": "Mitigate",
        "measure": "Approve the target architecture at QG1, establish milestone-based acceptance for hosting, identity, and connectivity, and run a weekly joint build board with KPMG, Bosch, and Infosys to clear blockers within five business days.",
        "due_date": "2026-10-01",
        "status": "not started",
        "notes": "Single-vendor dependency risk.",
    },
    {
        "id": "R003",
        "category": "Schedule",
        "cause": "QG4 is fixed ahead of GoLive with only a limited final-readiness window between 2027-12-10 and 2028-01-01. SAP, application waves, network readiness, UAT closure, and rollback proof all have to complete before the gate can be passed.",
        "event": "The programme reaches late November 2027 with unresolved critical defects, incomplete migration evidence, or missing steering approvals, causing QG4 to fail or move materially.",
        "effect": "GoLive no longer remains viable on 2028-01-01, JCI service dependency extends by exception, and cost and governance pressure escalate sharply across all parties.",
        "event_date": "2027-12-10",
        "owner": "KPMG PMO Lead",
        "source": "Formal Risk Review",
        "impact": "Very High",
        "probability": "50%",
        "type": "threat",
        "qualitative": "QG4 is the formal point where all streams converge, so late slippage cannot be absorbed locally.",
        "eur_cy": 3000000,
        "eur_3y": 800000,
        "strategy": "Mitigate",
        "measure": "Define QG4 entrance criteria by 2027-09-01, run bi-weekly readiness reviews from September onward, maintain a single defect burn-down across SAP, apps, and workplace, and escalate any unresolved P1 issue within 24 hours.",
        "due_date": "2027-09-01",
        "status": "not started",
        "notes": "Direct linkage to GoLive commitment.",
    },
    {
        "id": "R004",
        "category": "Budget",
        "cause": "The programme combines 48 sites, 12,000 users, 1,800 plus applications, a temporary merger zone, and a major SAP track. Change requests can originate from Bosch integration standards, JCI source-data issues, local site exceptions, or underestimated application remediation effort.",
        "event": "Approved scope expands during Phase 2 or Phase 3 without matching cost governance, producing cumulative Infosys and third-party cost changes above the planned labour envelope.",
        "effect": "Bosch budget approval cycles delay execution, the delivery team slows lower-priority work while funding is clarified, and total programme spend exceeds the acquisition planning basis.",
        "event_date": "2027-08-31",
        "owner": "Bosch Programme Sponsor",
        "source": "Status Meeting",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative": "Budget pressure is most likely once real application and migration complexity is visible in build and test.",
        "eur_cy": 2500000,
        "eur_3y": 0,
        "strategy": "Mitigate",
        "measure": "Stand up a formal change control board at QG1, pre-approve contingency handling thresholds, require quantified impact statements for all change requests, and review spend-versus-plan monthly at steering level.",
        "due_date": "2026-10-01",
        "status": "not started",
        "notes": "Contingency management risk.",
    },
    {
        "id": "R005",
        "category": "Resources",
        "cause": "The programme depends on a limited number of specialist roles across KPMG, Infosys, JCI, and Bosch, including SAP leads, merger-zone architects, data migration leads, and local business owners. The project duration of 21 months increases attrition and reassignment risk.",
        "event": "One or more key technical or business resources leave the programme during build or testing, and replacements require several weeks to gain knowledge of the JCI source landscape and the merger-zone design.",
        "effect": "Decision latency rises, defect resolution slows, and critical deliverables such as SAP design, data mapping, or site playbooks slip beyond planned dates.",
        "event_date": "2027-06-30",
        "owner": "KPMG PMO Lead",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative": "Resource continuity is essential because knowledge is highly concentrated in a few integration specialists.",
        "eur_cy": 700000,
        "eur_3y": 0,
        "strategy": "Mitigate",
        "measure": "Create named-role continuity plans, require shadowing for all critical leads, maintain an active knowledge repository, and review key-role retention monthly with partner management.",
        "due_date": "2026-11-01",
        "status": "not started",
        "notes": "Cross-party knowledge retention required.",
    },
    {
        "id": "R006",
        "category": "Legal & Compliance",
        "cause": "All 12,000 users start on JCI systems and are distributed across 48 sites and multiple legal jurisdictions. Employee consultation, labour law, and transfer obligations differ by country and can directly influence cutover timing and access changes.",
        "event": "Required employee, union, or works-council approvals are not secured in time for planned user migration waves, especially for larger EMEA sites with formal consultation windows.",
        "effect": "User migration waves are sequenced later than planned, local sites remain on JCI longer than allowed, and the global readiness picture at QG4 becomes inconsistent.",
        "event_date": "2027-08-15",
        "owner": "JCI Legal Counsel",
        "source": "Formal Risk Review",
        "impact": "Very High",
        "probability": "30%",
        "type": "threat",
        "qualitative": "A single large-country delay can materially affect the overall migration plan because the user estate is concentrated in major hubs.",
        "eur_cy": 1500000,
        "eur_3y": 400000,
        "strategy": "Avoid",
        "measure": "Map legal consultation requirements by country during Phase 1, integrate approval gates into the wave plan, and require local legal clearance before any site is marked migration-ready.",
        "due_date": "2026-12-01",
        "status": "not started",
        "notes": "High dependency for large EMEA populations.",
    },
    {
        "id": "R007",
        "category": "Security & Data Protection",
        "cause": "The programme moves personal data, service records, financial data, and application data from JCI into an Infosys-operated merger zone before Bosch landing. Cross-border data movement, elevated admin access, and transitional security controls increase exposure during the migration period.",
        "event": "A security incident, unauthorised access event, or control failure occurs during migration or while the merger zone is operating as the intermediary platform.",
        "effect": "Regulatory notification is triggered, migration activity pauses while the incident is investigated, and confidence in the programme design falls with seller, buyer, and regulators.",
        "event_date": "2027-12-31",
        "owner": "Infosys Security Lead",
        "source": "Formal Risk Review",
        "impact": "Very High",
        "probability": "30%",
        "type": "threat",
        "qualitative": "The temporary merger zone concentrates sensitive data and privileged access at the point of highest delivery pressure.",
        "eur_cy": 5000000,
        "eur_3y": 2000000,
        "strategy": "Transfer",
        "measure": "Require zero-trust access patterns, DLP and logging before production data migration, independent security review before Phase 3, cyber insurance confirmation, and rehearsed incident-response processes with a 72-hour reporting model.",
        "due_date": "2027-02-01",
        "status": "not started",
        "notes": "Board-level exposure if triggered.",
    },
    {
        "id": "R008",
        "category": "Technology, R&D",
        "cause": "The non-SAP application estate exceeds 1,800 systems and includes legacy tools, regional applications, local utilities, and custom integrations. A large portion of the inventory will only be fully understood once technical owners validate connectivity and authentication dependencies.",
        "event": "Application compatibility with merger-zone hosting, identity, or connectivity is lower than forecast, creating a remediation backlog in Waves 2 and 3 that cannot be closed before QG4.",
        "effect": "The programme reaches final readiness with missing applications, business teams request exceptions, and either GoLive scope is reduced or hypercare begins with avoidable service gaps.",
        "event_date": "2027-11-30",
        "owner": "Infosys Application Lead",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative": "Application volume and uneven ownership make late incompatibility findings likely unless triage is aggressive in Phase 1.",
        "eur_cy": 1800000,
        "eur_3y": 500000,
        "strategy": "Mitigate",
        "measure": "Classify the top 300 applications by technical risk before QG1, create remediation tiers, prove the merger-zone pattern on Wave 1, and reserve specialist remediation capacity for Wave 2 and Wave 3 peaks.",
        "due_date": "2026-10-01",
        "status": "not started",
        "notes": "Critical for Wave 2 and Wave 3 stability.",
    },
    {
        "id": "R009",
        "category": "Schedule",
        "cause": "All 12,000 users must move from JCI to the merger zone before Bosch steady-state handover can begin. The migration plan spans four user waves across 48 sites with different local readiness levels, support models, and time-zone constraints.",
        "event": "One or more user waves takes longer than planned because local site readiness, floorwalking support, or device enrollment throughput is below the assumed rate.",
        "effect": "Wave completion shifts into the narrow window before QG4, business sign-off is delayed, and the programme risks arriving at GoLive with incomplete workplace migration.",
        "event_date": "2027-11-28",
        "owner": "Infosys Service Delivery Lead",
        "source": "Status Meeting",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative": "High-volume user moves often slow once the programme reaches more complex tail sites.",
        "eur_cy": 800000,
        "eur_3y": 0,
        "strategy": "Mitigate",
        "measure": "Validate site readiness six weeks before each wave, contract local support capacity early, maintain daily volume tracking during wave execution, and hold a formal go or hold checkpoint before every regional cutover.",
        "due_date": "2027-06-01",
        "status": "not started",
        "notes": "User-volume throughput is a key operational KPI.",
    },
    {
        "id": "R010",
        "category": "Quality",
        "cause": "JCI source data for finance, materials, service, and customer processes has accumulated over many years and was not designed for rapid legal-entity separation. Data cleansing capacity depends on business users who also need to support day-to-day operations.",
        "event": "Migration rehearsals expose material data-quality defects in SAP or non-SAP datasets that require manual correction and repeated extraction cycles.",
        "effect": "Testing quality declines, reconciliation runs over schedule, and business confidence in the final cutover package falls before QG4.",
        "event_date": "2027-10-31",
        "owner": "JCI Data Owner",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "50%",
        "type": "threat",
        "qualitative": "Data quality is a core acceptance criterion for both application readiness and business sign-off.",
        "eur_cy": 1200000,
        "eur_3y": 500000,
        "strategy": "Mitigate",
        "measure": "Run data-quality assessment by workstream during Phase 1, assign business data stewards, define acceptance thresholds per object type, and execute repeated reconciliation rehearsals before production cutover planning is frozen.",
        "due_date": "2027-01-31",
        "status": "not started",
        "notes": "Finance and material masters are highest priority.",
    },
    {
        "id": "R011",
        "category": "Stakeholder Relations & Public Affairs",
        "cause": "JCI is the seller and Bosch is the buyer, while KPMG leads PMO and Infosys runs the merger zone. These four stakeholder groups do not share the same incentives, and delays or scope debates can quickly become escalation topics rather than working-level issues.",
        "event": "Cross-party decisions on scope, access, sign-off criteria, or cost allocation stall because stakeholders escalate instead of resolving issues within the agreed governance cadence.",
        "effect": "Programme decisions arrive late, workstreams operate on assumptions, and delivery speed slows precisely when the schedule needs rapid cross-functional resolution.",
        "event_date": "2027-06-30",
        "owner": "KPMG PMO Lead",
        "source": "Stakeholder",
        "impact": "High",
        "probability": "30%",
        "type": "threat",
        "qualitative": "This is a multi-party programme where weak decision governance quickly converts into schedule loss.",
        "eur_cy": 900000,
        "eur_3y": 0,
        "strategy": "Mitigate",
        "measure": "Publish a decision-rights matrix at QG1, enforce escalation response SLAs, maintain open RAID ownership by party, and review unresolved steering actions weekly until closure.",
        "due_date": "2026-10-15",
        "status": "not started",
        "notes": "Governance discipline is a control rather than an administrative task.",
    },
    {
        "id": "R012",
        "category": "Supply Chain",
        "cause": "The merger zone requires hosting capacity, connectivity, security tooling, endpoint tooling, and potentially hardware or reserved cloud capacity. Procurement timing directly affects infrastructure availability and later migration work.",
        "event": "Long-lead procurement items or commercial approvals for merger-zone capacity and connectivity are not completed early enough to support build and testing dates.",
        "effect": "Infrastructure workstreams idle waiting for equipment, capacity, or contracts, and downstream application, identity, and user migration tasks slip.",
        "event_date": "2027-04-30",
        "owner": "Infosys Infrastructure Architect",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "30%",
        "type": "threat",
        "qualitative": "Even a short procurement delay at the platform layer creates a much larger loss across dependent tasks.",
        "eur_cy": 500000,
        "eur_3y": 0,
        "strategy": "Mitigate",
        "measure": "Lock platform commercial decisions at QG1, track long-lead items in a dedicated supply register, and approve fallback cloud capacity options before physical delivery windows become critical.",
        "due_date": "2026-10-01",
        "status": "not started",
        "notes": "Platform procurement is a non-negotiable precursor.",
    },
    {
        "id": "R013",
        "category": "Customers",
        "cause": "The air conditioning business relies on customer-facing order, service, and support processes backed by the in-scope application estate. Wave 1 and Wave 2 include systems whose disruption would be visible to external customers and field service teams.",
        "event": "Migration of customer-critical applications causes service interruption, performance degradation, or interface failures that affect order capture or service execution during the transition period.",
        "effect": "Customer experience drops, revenue and service levels are impacted, and the programme attracts commercial scrutiny beyond IT governance.",
        "event_date": "2027-09-30",
        "owner": "JCI IT Manager",
        "source": "Stakeholder",
        "impact": "High",
        "probability": "30%",
        "type": "threat",
        "qualitative": "Customer-visible disruption would turn an internal carve-out issue into a market-facing business incident.",
        "eur_cy": 1400000,
        "eur_3y": 500000,
        "strategy": "Mitigate",
        "measure": "Tag customer-critical applications in Phase 1, require zero- or low-downtime migration designs for the top set, test peak usage scenarios before cutover, and obtain business-owner approval before each cut window.",
        "due_date": "2026-11-15",
        "status": "not started",
        "notes": "Commercial impact risk.",
    },
    {
        "id": "R014",
        "category": "Legal & Compliance",
        "cause": "The target operating model uses a temporary merger zone to bridge JCI and Bosch, which means data and identity flows may cross jurisdictions before final Bosch landing. Several countries in scope have data localisation or transfer restrictions.",
        "event": "Data sovereignty analysis shows that planned merger-zone hosting or data-transfer mechanisms are not compliant for one or more major countries, forcing redesign or local exceptions.",
        "effect": "Architecture decisions reopen, build work is paused or reworked, and site or user migrations for affected countries move later in the plan.",
        "event_date": "2027-02-01",
        "owner": "Bosch Legal Counsel",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "30%",
        "type": "threat",
        "qualitative": "This is a design-time legal blocker, not a late-stage operational issue, so it has to be resolved before build is locked.",
        "eur_cy": 900000,
        "eur_3y": 300000,
        "strategy": "Avoid",
        "measure": "Run country-by-country privacy and localisation review in Phase 1, include required regional hosting or control patterns in the approved merger-zone blueprint, and do not approve QG1 architecture without legal sign-off.",
        "due_date": "2026-11-01",
        "status": "not started",
        "notes": "Architecture-governance dependency.",
    },
    {
        "id": "R015",
        "category": "Engineering",
        "cause": "Air conditioning product engineering data, document management, and design-tool integrations may sit in shared JCI platforms that are not purely office IT systems. These tools can have unique latency, storage, and permission requirements compared with standard business apps.",
        "event": "Engineering and product-data systems do not transition cleanly into the merger zone because shared repositories, CAD integrations, or file-performance needs were underestimated.",
        "effect": "Engineering teams cannot access the right versions of design data or collaboration tooling, delaying product and service work while technical remediation is performed.",
        "event_date": "2027-11-15",
        "owner": "Bosch Engineering IT Lead",
        "source": "Status Meeting",
        "impact": "Moderate",
        "probability": "30%",
        "type": "threat",
        "qualitative": "Engineering toolchains are often overlooked because they sit outside the core ERP discussion but still carry operational criticality.",
        "eur_cy": 600000,
        "eur_3y": 250000,
        "strategy": "Mitigate",
        "measure": "Identify all engineering repositories and integrations in Phase 1, run targeted technical proof-of-concept tests in Phase 2, and ensure engineering sign-off is a distinct checkpoint before Wave 3 closure.",
        "due_date": "2027-01-31",
        "status": "not started",
        "notes": "Engineering tools need explicit treatment, not generic app migration handling.",
    },
    {
        "id": "R016",
        "category": "Manufacturing",
        "cause": "The business being carved out includes manufacturing operations that rely on ERP, planning, quality, and potentially OT-adjacent systems. Factory sites may require tighter cutover windows and stricter tolerance for downtime than office sites.",
        "event": "Manufacturing-critical applications or interfaces are not migrated, validated, or supported in time for planned operational cutovers at plant locations.",
        "effect": "Production planning, shop-floor coordination, or fulfilment continuity is disrupted, leading to operational loss and urgent leadership intervention.",
        "event_date": "2027-11-30",
        "owner": "JCI Manufacturing IT Lead",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "30%",
        "type": "threat",
        "qualitative": "Manufacturing sites convert standard IT migration risk into direct operational risk very quickly.",
        "eur_cy": 1600000,
        "eur_3y": 400000,
        "strategy": "Mitigate",
        "measure": "Separate manufacturing-critical systems into a dedicated readiness track, align plant cutovers to production calendars, and require plant-level business continuity approval before migration execution.",
        "due_date": "2027-03-01",
        "status": "not started",
        "notes": "Factory schedule alignment required.",
    },
    {
        "id": "R017",
        "category": "Quality",
        "cause": "GoLive performance assumptions for 12,000 users, 48 sites, and a large SAP plus application estate depend on design estimates until real workload tests begin. Peak transaction patterns may exceed assumed baseline throughput.",
        "event": "Performance and load testing in Phase 3 shows that merger-zone infrastructure or SAP sizing does not support expected concurrent usage or response-time targets.",
        "effect": "Infrastructure has to be resized or retuned, retesting consumes schedule, and final readiness enters December without enough margin.",
        "event_date": "2027-09-30",
        "owner": "Infosys Infrastructure Architect",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "30%",
        "type": "threat",
        "qualitative": "Late discovery of performance limits is expensive because it affects almost every service already loaded onto the platform.",
        "eur_cy": 700000,
        "eur_3y": 0,
        "strategy": "Mitigate",
        "measure": "Define formal performance SLAs before build, run stepped load tests well ahead of final readiness, and maintain capacity expansion options that can be activated without a full redesign.",
        "due_date": "2027-02-01",
        "status": "not started",
        "notes": "Tied to infrastructure and SAP sizing discipline.",
    },
    {
        "id": "R018",
        "category": "Strategy & Portfolio",
        "cause": "This is an integration-model carve-out, which means the merger zone is transitional by design and Bosch will eventually absorb services into its own operating model. If the handover boundary is vague, project teams may continue to add post-GoLive integration work into the carve-out programme.",
        "event": "Bosch requests additional steady-state integration scope during late Phase 3 or hypercare that was not part of the original carve-out baseline.",
        "effect": "Closure dates move, project resources remain engaged past QG5, and programme governance becomes blurred between carve-out completion and Bosch BAU integration.",
        "event_date": "2028-03-15",
        "owner": "Bosch Programme Sponsor",
        "source": "Stakeholder",
        "impact": "Moderate",
        "probability": "30%",
        "type": "threat",
        "qualitative": "Without a clear handover boundary, any integration-model project tends to accumulate additional downstream scope.",
        "eur_cy": 800000,
        "eur_3y": 0,
        "strategy": "Avoid",
        "measure": "Define the carve-out versus Bosch integration boundary at QG1, freeze hypercare scope to stabilisation only, and require any new Bosch steady-state work to be funded and governed separately.",
        "due_date": "2026-10-01",
        "status": "not started",
        "notes": "Scope boundary protection.",
    },
    {
        "id": "R019",
        "category": "Intellectual Property",
        "cause": "The air conditioning business carries design, service, and product information that may still reside in shared JCI repositories or shared access models. A carve-out requires clean separation of who can view, export, and retain that information after transition.",
        "event": "IP-related repositories, document stores, or PLM-connected datasets are not fully segregated by GoLive, leaving Bosch-owned information accessible to the seller or unavailable to the carved-out teams.",
        "effect": "Legal exposure rises, engineering continuity is impaired, and post-GoLive remediation becomes urgent and politically sensitive.",
        "event_date": "2028-01-01",
        "owner": "Bosch Legal Counsel",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "30%",
        "type": "threat",
        "qualitative": "IP segregation is a deal-protection issue, not just an application-migration issue.",
        "eur_cy": 1000000,
        "eur_3y": 500000,
        "strategy": "Avoid",
        "measure": "Include product-data repositories in Phase 1 scope validation, perform access-right audits before QG4, and require legal sign-off on IP segregation as part of final readiness evidence.",
        "due_date": "2027-10-31",
        "status": "not started",
        "notes": "Legal and engineering dependency combined.",
    },
    {
        "id": "R020",
        "category": "Ecosystems & Ethics",
        "cause": "Infosys may rely on subcontractors or offshore specialists to deliver parts of the merger-zone build, migration factory, or support model. That creates an indirect supply chain with access to sensitive systems, data, and programme information.",
        "event": "A subcontractor, external specialist, or unmanaged third party breaches data-handling, background screening, or conduct expectations while supporting the programme.",
        "effect": "Security and compliance review is triggered, Bosch and JCI challenge the vendor model, and trust in the delivery ecosystem weakens during a critical phase of the programme.",
        "event_date": "2027-12-31",
        "owner": "KPMG PMO Lead",
        "source": "Formal Risk Review",
        "impact": "Moderate",
        "probability": "10%",
        "type": "threat",
        "qualitative": "The impact is driven less by technical failure and more by governance, trust, and regulatory consequences.",
        "eur_cy": 700000,
        "eur_3y": 0,
        "strategy": "Transfer",
        "measure": "Require full subcontractor disclosure before build begins, enforce Bosch-aligned screening and conduct clauses, and preserve audit rights over all delivery layers handling sensitive systems or data.",
        "due_date": "2026-10-01",
        "status": "not started",
        "notes": "Third-party governance exposure.",
    },
    {
        "id": "R021",
        "category": "Raw Materials",
        "cause": "Material planning and inventory data in the carved-out business is likely managed through SAP and connected planning tools. If material master, procurement references, or planning parameters migrate incorrectly, operational supply decisions can become unreliable.",
        "event": "Material and planning data required for air-conditioning production or fulfilment is migrated inaccurately, leaving procurement and planning teams with incomplete or incorrect information after cutover.",
        "effect": "Production and fulfilment decisions degrade, supply continuity is threatened, and the business experiences avoidable disruption even if the technical cutover itself succeeds.",
        "event_date": "2027-11-15",
        "owner": "JCI Supply Chain Data Lead",
        "source": "Formal Risk Review",
        "impact": "Moderate",
        "probability": "30%",
        "type": "threat",
        "qualitative": "This is a business-data risk that emerges through IT migration and therefore belongs in the carve-out register.",
        "eur_cy": 650000,
        "eur_3y": 250000,
        "strategy": "Mitigate",
        "measure": "Add material-master validation to the SAP rehearsal scope, assign supply-chain business owners to reconciliation checkpoints, and block readiness sign-off if planning-critical objects fail agreed quality thresholds.",
        "due_date": "2027-08-31",
        "status": "not started",
        "notes": "Supply continuity via data integrity.",
    },
    {
        "id": "R022",
        "category": "Market & Competitors",
        "cause": "The carve-out occurs while the business continues to compete in the HVAC market. Any visible instability in ordering, service delivery, or commercial responsiveness during transition can be used by competitors to target key accounts.",
        "event": "Competitors exploit customer-facing disruption or perceived instability during migration windows to win business from major air-conditioning customers or service contracts.",
        "effect": "Commercial pressure increases on leadership, customer escalations intensify, and the programme faces external scrutiny rather than only internal delivery pressure.",
        "event_date": "2027-10-31",
        "owner": "Bosch Commercial Transition Lead",
        "source": "Stakeholder",
        "impact": "Moderate",
        "probability": "30%",
        "type": "threat",
        "qualitative": "This becomes material only if the IT transition creates externally visible service instability.",
        "eur_cy": 900000,
        "eur_3y": 400000,
        "strategy": "Mitigate",
        "measure": "Identify the top customer-facing cutover windows early, create commercial communication packs, and ensure no high-risk customer systems move without tested rollback and business-owner sign-off.",
        "due_date": "2027-06-30",
        "status": "not started",
        "notes": "Business perception risk driven by IT execution.",
    },
    {
        "id": "R023",
        "category": "Technology, R&D",
        "cause": "The merger zone provides a rare opportunity to simplify and modernise the estate instead of reproducing JCI technical debt exactly as-is. A cloud-first or service-rationalised landing pattern could reduce long-term Bosch operating cost if agreed early enough.",
        "event": "Phase 1 architecture and application triage identify a meaningful subset of applications and infrastructure services that can be landed in a more modern, lower-cost pattern inside the merger zone without delaying GoLive.",
        "effect": "Bosch inherits a cleaner estate after handover, avoids part of the future remediation backlog, and realises measurable operating-cost savings after programme closure.",
        "event_date": "2027-02-01",
        "owner": "Infosys Cloud Lead",
        "source": "Formal Risk Review",
        "impact": "High",
        "probability": "50%",
        "type": "opportunity",
        "qualitative": "The carve-out creates a once-only architecture reset point that can deliver value beyond pure separation.",
        "eur_cy": 0,
        "eur_3y": 4500000,
        "strategy": "Exploit",
        "measure": "Add a cloud and rationalisation suitability field to the Phase 1 inventory, prepare a QG1 value case for Bosch, and pilot modern landing patterns on low-risk Wave 1 services before wider adoption.",
        "due_date": "2026-10-01",
        "status": "not started",
        "notes": "Opportunity linked to architectural reset.",
    },
    {
        "id": "R024",
        "category": "Budget",
        "cause": "The transition from JCI to Bosch via a merger zone exposes software licences, support contracts, and infrastructure services that may be duplicated or fragmented across the estate. Early consolidation could create leverage before Bosch steady-state integration completes.",
        "event": "The programme identifies opportunities to consolidate licences, retire duplicate tools, or renegotiate support contracts while the estate is being re-established under Bosch control.",
        "effect": "Bosch reduces post-GoLive operating spend and simplifies the future integration backlog without additional carve-out delay.",
        "event_date": "2028-03-31",
        "owner": "Bosch Programme Sponsor",
        "source": "Formal Risk Review",
        "impact": "Moderate",
        "probability": "50%",
        "type": "opportunity",
        "qualitative": "Commercial clean-up is often easier during separation than after the new environment hardens into BAU.",
        "eur_cy": 0,
        "eur_3y": 2500000,
        "strategy": "Enhance",
        "measure": "Create a licence and contract-consolidation worklist during Phase 1, prioritise quick-win retirements in Phase 2, and hand a quantified savings roadmap to Bosch operations before QG5.",
        "due_date": "2027-01-31",
        "status": "not started",
        "notes": "Second positive risk to balance the register.",
    },
]


def fix_matrix_sheet(workbook) -> None:
    if "Matrix " not in workbook.sheetnames:
        return
    matrix_sheet = workbook["Matrix "]
    yellow_fills = {"FFFFFF00", "FFFFFFCC", "FFF2CC", "FFFF00"}
    for row in matrix_sheet.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fgColor:
                colour = cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else ""
                if colour and colour.upper() in yellow_fills:
                    cell.font = Font(
                        name=cell.font.name if cell.font else "Calibri",
                        size=cell.font.size if cell.font else 11,
                        bold=cell.font.bold if cell.font else False,
                        italic=cell.font.italic if cell.font else False,
                        color="000000",
                    )


def write_risk_register() -> None:
    if not TEMPLATE_PATH.exists():
        print(f"Template not found: {TEMPLATE_PATH}")
        sys.exit(1)

    workbook = load_workbook(TEMPLATE_PATH)

    info_sheet = workbook["Info"]
    info_sheet["C4"] = DOCUMENT_ID
    info_sheet["C5"] = f"{PROJECT_NAME}_Risk_Register.xlsx"
    info_sheet["C6"] = PROJECT_NAME
    info_sheet["C7"] = PROJECT_CODE
    info_sheet["C8"] = OWNER

    risk_sheet = workbook["Risk Register"]
    first_data_row = 5

    for index, risk in enumerate(RISKS):
        row = first_data_row + index
        risk_sheet.cell(row, 2).value = risk["id"]
        risk_sheet.cell(row, 3).value = CREATION_DATE
        risk_sheet.cell(row, 4).value = risk["category"]
        risk_sheet.cell(row, 5).value = risk["cause"]
        risk_sheet.cell(row, 6).value = risk["event"]
        risk_sheet.cell(row, 7).value = risk["effect"]
        risk_sheet.cell(row, 8).value = risk["event_date"]
        risk_sheet.cell(row, 9).value = risk["owner"]
        risk_sheet.cell(row, 10).value = risk["source"]
        risk_sheet.cell(row, 12).value = risk["impact"]
        risk_sheet.cell(row, 13).value = f'=_xlfn.IFNA(VLOOKUP(L{row},$D$182:$E$186,2,FALSE),"")'
        risk_sheet.cell(row, 14).value = risk["probability"]
        risk_sheet.cell(row, 15).value = f'=_xlfn.IFNA(VLOOKUP(N{row},$D$189:$E$193,2,FALSE),"")'
        risk_sheet.cell(row, 16).value = risk["type"]
        risk_sheet.cell(row, 17).value = f"=M{row}*O{row}"
        risk_sheet.cell(row, 18).value = risk["qualitative"]
        risk_sheet.cell(row, 19).value = risk["eur_cy"]
        risk_sheet.cell(row, 20).value = risk["eur_3y"]
        risk_sheet.cell(row, 22).value = risk["strategy"]
        risk_sheet.cell(row, 23).value = risk["measure"]
        risk_sheet.cell(row, 24).value = risk["due_date"]
        risk_sheet.cell(row, 26).value = risk["status"]
        risk_sheet.cell(row, 27).value = REPORT_DATE
        risk_sheet.cell(row, 28).value = risk["impact"]
        risk_sheet.cell(row, 29).value = f'=_xlfn.IFNA(VLOOKUP(AB{row},$D$182:$E$186,2,FALSE),"")'
        risk_sheet.cell(row, 30).value = risk["probability"]
        risk_sheet.cell(row, 31).value = f'=_xlfn.IFNA(VLOOKUP(AD{row},$D$189:$E$193,2,FALSE),"")'
        risk_sheet.cell(row, 32).value = f"=AC{row}"
        risk_sheet.cell(row, 33).value = f"=AE{row}"
        risk_sheet.cell(row, 34).value = f"=AF{row}*AG{row}"
        risk_sheet.cell(row, 35).value = risk["notes"]

    fix_matrix_sheet(workbook)
    workbook.save(OUTPUT_PATH)


def main() -> None:
    print(f"[{PROJECT_NAME}] Generating risk register")
    print(f"  Template: {TEMPLATE_PATH}")
    print(f"  Output:   {OUTPUT_PATH}")
    print(f"  Risks:    {len(RISKS)}")
    write_risk_register()
    print(f"[{PROJECT_NAME}] Risk register complete")


if __name__ == "__main__":
    main()