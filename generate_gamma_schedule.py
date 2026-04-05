#!/usr/bin/env python3
"""Generate Gamma carve-out schedule as XLSX and MS Project XML."""

from __future__ import annotations

import csv
import subprocess
import sys
from datetime import datetime
from pathlib import Path


HERE = Path(__file__).parent
PROJECT_NAME = "Gamma"
OUTPUT_FOLDER_NAME = "Gamma v1.0"
DOCUMENT_VERSION = "Version 1.0 - Initial Baseline"
SELLER = "Robert Bosch China"
BUYER = "Alibaba"
BUSINESS = "Bosch Cloud business"
PMO_LEAD = "EY"
CARVEOUT_MODEL = "Combination"
WORLDWIDE_SITES = 5
IT_USERS = 250
APPLICATIONS = 20

START_DATE = datetime(2026, 8, 1)
GOLIVE_DATE = datetime(2027, 2, 1)
COMPLETION_DATE = datetime(2027, 5, 31)

OUTPUT_DIR = HERE / "active-projects" / OUTPUT_FOLDER_NAME
XLSX_PATH = OUTPUT_DIR / f"{PROJECT_NAME}_Project_Schedule.xlsx"
XML_PATH = OUTPUT_DIR / f"{PROJECT_NAME}_Project_Schedule.xml"
TEMP_CSV_PATH = OUTPUT_DIR / f"{PROJECT_NAME}_Project_Schedule.tmp.csv"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def parse_date(value: str) -> datetime:
    return datetime.strptime(value, "%Y-%m-%d")


def duration_label(start: str, finish: str, milestone: bool) -> str:
    if milestone:
        return "1 day"
    days = (parse_date(finish) - parse_date(start)).days + 1
    return f"{days} day" if days == 1 else f"{days} days"


def make_task(
    task_id: int,
    outline_level: int,
    name: str,
    start: str,
    finish: str,
    predecessors: str = "",
    resources: str = "",
    notes: str = "",
    milestone: bool = False,
) -> tuple[int, int, str, str, str, str, str, str, str, str]:
    return (
        task_id,
        outline_level,
        name,
        duration_label(start, finish, milestone),
        start,
        finish,
        predecessors,
        resources,
        notes,
        "Yes" if milestone else "No",
    )


HEADERS = [
    "ID",
    "Outline Level",
    "Name",
    "Duration",
    "Start",
    "Finish",
    "Predecessors",
    "Resource Names",
    "Notes",
    "Milestone",
]


TASKS = [
    make_task(
        1,
        1,
        "Phase 0 - Mobilization and confidentiality control",
        "2026-08-01",
        "2026-09-04",
        notes="Programme launch, restricted planning model, legal perimeter, and initial inventory for the Bosch Cloud carve-out into a 50/50 JV with Alibaba.",
    ),
    make_task(2, 2, "0.1 Governance and restricted planning", "2026-08-01", "2026-08-19"),
    make_task(
        3,
        3,
        "Launch core programme team and initiate Gamma confidentiality controls",
        "2026-08-01",
        "2026-08-03",
        resources="EY + Robert Bosch China IT + Alibaba IT",
        notes="Only the core clean team is aware during startup because the JV and antitrust due diligence remain highly confidential.",
    ),
    make_task(
        4,
        3,
        "Establish confidentiality protocol, clean-team access, and restricted stakeholder list",
        "2026-08-01",
        "2026-08-10",
        predecessors="3",
        resources="EY + Legal + Robert Bosch China IT Security",
        notes="Planning access is tightly controlled until legal allows a broader discovery group.",
    ),
    make_task(
        5,
        3,
        "Set governance cadence, RACI, steerco path, and PMO controls",
        "2026-08-04",
        "2026-08-19",
        predecessors="3",
        resources="EY + Robert Bosch China PMO + Alibaba PMO",
        notes="Governance reflects shared management of the future 50/50 JV.",
    ),
    make_task(6, 2, "0.2 Legal perimeter and scoped discovery", "2026-08-04", "2026-09-03"),
    make_task(
        7,
        3,
        "Confirm carve-out perimeter, 50/50 JV assumptions, and decision rights",
        "2026-08-04",
        "2026-08-15",
        predecessors="3",
        resources="Legal + EY + Robert Bosch China Leadership + Alibaba Leadership",
        notes="Combination model remains the formal carve-out model; the 50/50 JV is the post-close governance construct.",
    ),
    make_task(
        8,
        3,
        "Launch antitrust due diligence tracker, legal workplan, and planning assumptions log",
        "2026-08-04",
        "2026-08-21",
        predecessors="3",
        resources="Legal + EY",
        notes="Legal uncertainty is treated as a hard planning dependency, not only a risk note.",
    ),
    make_task(
        9,
        3,
        "Inventory 5 sites, 250 users, 20 applications, and infrastructure service towers",
        "2026-08-11",
        "2026-08-28",
        predecessors="4,7",
        resources="EY + Robert Bosch China Infrastructure + Alibaba IT",
        notes="Scope is pure infrastructure services with no SAP in scope.",
    ),
    make_task(
        10,
        3,
        "Define discovery guardrails and phased stakeholder expansion path",
        "2026-08-18",
        "2026-09-03",
        predecessors="4,8,9",
        resources="EY + Legal + HR + Communications",
        notes="Broader validation remains constrained until legal confirms what can be disclosed.",
    ),
    make_task(
        11,
        3,
        "QG0 - Mobilization approved",
        "2026-09-04",
        "2026-09-04",
        predecessors="5,8,9,10",
        resources="Steering Committee",
        notes="Restricted-planning controls active, perimeter defined, and initial inventory complete.",
        milestone=True,
    ),
    make_task(
        12,
        1,
        "Phase 1 - Concept and target-state design",
        "2026-09-05",
        "2026-10-29",
        notes="Operating model, infrastructure target state, migration design, and formal readiness criteria.",
    ),
    make_task(13, 2, "1.1 Operating model and transition principles", "2026-09-05", "2026-09-30"),
    make_task(
        14,
        3,
        "Define Day 1 and Day 2 scope boundaries for the JV operating model",
        "2026-09-05",
        "2026-09-16",
        predecessors="11",
        resources="EY + Robert Bosch China IT + Alibaba IT",
        notes="Scope keeps Day 1 focused on infrastructure continuity and only the 20 in-scope applications.",
    ),
    make_task(
        15,
        3,
        "Define interim service continuity, TSA needs, and exit conditions",
        "2026-09-09",
        "2026-09-23",
        predecessors="11",
        resources="EY + Robert Bosch China IT + Alibaba IT",
        notes="Seller support is defined only where required until the new JV services are stable.",
    ),
    make_task(
        16,
        3,
        "Confirm sign-off route and approval process under legal confidentiality constraints",
        "2026-09-17",
        "2026-09-30",
        predecessors="14,15",
        resources="EY + Legal + Steering Committee",
        notes="Approvals use a narrow escalation route because only limited associates can participate.",
    ),
    make_task(17, 2, "1.2 Infrastructure and security target design", "2026-09-05", "2026-10-07"),
    make_task(
        18,
        3,
        "Design target hosting, network, connectivity, and environment segregation",
        "2026-09-05",
        "2026-09-25",
        predecessors="11",
        resources="Robert Bosch China Infrastructure + Alibaba Cloud + EY",
        notes="Target state is infrastructure-led and excludes SAP-related architecture.",
    ),
    make_task(
        19,
        3,
        "Design identity, access, endpoint, and security operating model",
        "2026-09-12",
        "2026-10-02",
        predecessors="18",
        resources="Robert Bosch China Security + Alibaba Security + EY",
        notes="Identity and access design supports joint governance without broadening access prematurely.",
    ),
    make_task(
        20,
        3,
        "Design collaboration, service management, monitoring, and support stack",
        "2026-09-19",
        "2026-10-07",
        predecessors="18",
        resources="Robert Bosch China Infrastructure + Alibaba ITSM + EY",
        notes="Service desk, monitoring, backup, and collaboration capabilities are defined for Day 1 operations.",
    ),
    make_task(21, 2, "1.3 Application and migration design", "2026-09-12", "2026-10-28"),
    make_task(
        22,
        3,
        "Classify 20 applications and decide separation pattern for each",
        "2026-09-12",
        "2026-09-30",
        predecessors="9,11",
        resources="Application Owners + EY",
        notes="Each application is classified for retain, reconfigure, replace, or retire within the JV context.",
    ),
    make_task(
        23,
        3,
        "Define migration waves for applications, users, devices, and sites",
        "2026-10-01",
        "2026-10-16",
        predecessors="19,20,22",
        resources="EY + Deployment Lead + Infrastructure Leads",
        notes="Wave design is modest because the scope is 5 sites, 250 users, and 20 applications.",
    ),
    make_task(
        24,
        3,
        "Define test strategy, cutover principles, and rollback approach",
        "2026-10-05",
        "2026-10-21",
        predecessors="19,20,22",
        resources="EY Test Lead + Infrastructure Leads",
        notes="Rollback and sign-off mechanics are formalized early due to legal and confidentiality sensitivity.",
    ),
    make_task(
        25,
        3,
        "Baseline risk register, dependency log, and entry criteria for build",
        "2026-10-12",
        "2026-10-28",
        predecessors="16,23,24",
        resources="EY + Workstream Leads",
        notes="Gamma-specific risks and dependencies are captured without borrowing reference-project content.",
    ),
    make_task(
        26,
        3,
        "QG1 - Concept approved",
        "2026-10-29",
        "2026-10-29",
        predecessors="16,19,20,23,24,25",
        resources="Steering Committee",
        notes="Target-state concept, migration design, and restricted approval path are confirmed.",
        milestone=True,
    ),
    make_task(
        27,
        1,
        "Phase 2 - Build and integrated testing",
        "2026-10-30",
        "2027-01-11",
        notes="Build core services, prepare the 20-application transition set, and complete integrated testing before final readiness.",
    ),
    make_task(28, 2, "2.1 Core infrastructure build", "2026-10-30", "2026-12-04"),
    make_task(
        29,
        3,
        "Build landing zone, hosting, and network segmentation",
        "2026-10-30",
        "2026-11-25",
        predecessors="26",
        resources="Robert Bosch China Infrastructure + Alibaba Cloud",
        notes="Core environment for the JV is established first to unblock downstream migration tasks.",
    ),
    make_task(
        30,
        3,
        "Build identity, IAM, endpoint, and access services",
        "2026-11-02",
        "2026-11-30",
        predecessors="26,19",
        resources="Robert Bosch China Security + Alibaba Security",
        notes="Identity controls are implemented with least-privilege access due to the restricted stakeholder model.",
    ),
    make_task(
        31,
        3,
        "Build collaboration, service desk, monitoring, and backup services",
        "2026-11-09",
        "2026-12-04",
        predecessors="20,29",
        resources="Infrastructure Operations + ITSM Leads",
        notes="Day 1 support tooling and operational monitoring are completed before system integration testing.",
    ),
    make_task(32, 2, "2.2 Application and data separation build", "2026-11-09", "2026-12-18"),
    make_task(
        33,
        3,
        "Reconfigure 20 applications for JV target services",
        "2026-11-09",
        "2026-12-18",
        predecessors="22,29,30",
        resources="Application Owners + Infrastructure Leads",
        notes="Application work remains intentionally lean because the scope is limited and non-SAP.",
    ),
    make_task(
        34,
        3,
        "Prepare shared data and file separation plus access remediation",
        "2026-11-16",
        "2026-12-11",
        predecessors="29,30",
        resources="Data Leads + Security Leads",
        notes="Shared repositories and access rights are remediated before formal testing starts.",
    ),
    make_task(
        35,
        3,
        "Complete third-party contract, DNS, certificate, and supplier changes",
        "2026-11-16",
        "2026-12-11",
        predecessors="15,22",
        resources="Procurement + Legal + Infrastructure Leads",
        notes="Third-party dependencies are aligned with the future JV operating model and confidentiality constraints.",
    ),
    make_task(36, 2, "2.3 Testing and migration rehearsal", "2026-12-07", "2027-01-08"),
    make_task(
        37,
        3,
        "Execute system integration testing across infrastructure services and 20 applications",
        "2026-12-07",
        "2026-12-23",
        predecessors="31,33,34,35",
        resources="Test Lead + Infrastructure Leads + Application Owners",
        notes="Integrated validation confirms that the infrastructure-led design supports the limited application set.",
    ),
    make_task(
        38,
        3,
        "Execute UAT with restricted business representatives and clean-team approvers",
        "2026-12-14",
        "2026-12-30",
        predecessors="37",
        resources="Business Key Users + EY",
        notes="UAT uses the limited approved representative group because broad awareness is still constrained.",
    ),
    make_task(
        39,
        3,
        "Run cutover rehearsal and rollback validation",
        "2026-12-21",
        "2027-01-06",
        predecessors="37,38",
        resources="Infrastructure Leads + EY",
        notes="Rollback evidence is required before QG2 and QG3 can be passed.",
    ),
    make_task(
        40,
        3,
        "Complete site, device, user, and communications migration packages",
        "2026-12-21",
        "2027-01-08",
        predecessors="31,34,35",
        resources="Deployment Lead + HR + Communications",
        notes="All migration packages are completed before QG2 and QG3 and before the QG4 readiness window.",
    ),
    make_task(
        41,
        3,
        "QG2 and QG3 - Build and test approved",
        "2027-01-11",
        "2027-01-11",
        predecessors="38,39,40",
        resources="Steering Committee",
        notes="Build complete, integrated testing passed, and final readiness activities may begin.",
        milestone=True,
    ),
    make_task(
        42,
        1,
        "Phase 3 - Final readiness and GoLive",
        "2027-01-12",
        "2027-02-01",
        notes="Complete residual pre-GoLive work, hold QG4, run the mandatory final readiness window, and execute Day 1 cutover.",
    ),
    make_task(43, 2, "3.1 Pre-GoLive completion", "2027-01-12", "2027-01-25"),
    make_task(
        44,
        3,
        "Close UAT defects and QG2 and QG3 action items",
        "2027-01-12",
        "2027-01-20",
        predecessors="41",
        resources="Test Lead + Application Owners + Infrastructure Leads",
        notes="All high-priority defects must be closed before the pre-GoLive gate.",
    ),
    make_task(
        45,
        3,
        "Complete pre-GoLive user, device, and site transitions",
        "2027-01-12",
        "2027-01-21",
        predecessors="40,41",
        resources="Deployment Lead + Infrastructure Operations",
        notes="All migrations are completed before QG4 in line with the schedule rule set.",
    ),
    make_task(
        46,
        3,
        "Complete final security, operational readiness, and support staffing check",
        "2027-01-15",
        "2027-01-22",
        predecessors="31,41,44,45",
        resources="Security Leads + IT Operations + EY",
        notes="No critical alerts, unresolved access gaps, or support staffing issues may remain at gate review.",
    ),
    make_task(
        47,
        3,
        "QG4 - Pre-GoLive approved",
        "2027-01-25",
        "2027-01-25",
        predecessors="44,45,46",
        resources="Steering Committee",
        notes="GoLive is not on the same date as QG4; a mandatory final readiness window follows.",
        milestone=True,
    ),
    make_task(48, 2, "3.2 Final readiness and open item closure", "2027-01-26", "2027-01-30"),
    make_task(
        49,
        3,
        "Verify all UAT defects and QG4 action items are closed",
        "2027-01-26",
        "2027-01-27",
        predecessors="47",
        resources="Test Lead + PMO",
        notes="This is the first mandatory final-readiness activity after QG4.",
    ),
    make_task(
        50,
        3,
        "Run final infrastructure and service readiness checks",
        "2027-01-26",
        "2027-01-28",
        predecessors="47",
        resources="Infrastructure Operations + Security Leads",
        notes="Readiness checks confirm stability, no critical alerts, and support handoff readiness.",
    ),
    make_task(
        51,
        3,
        "Execute final rollback confirmation and secure business sign-off",
        "2027-01-29",
        "2027-01-30",
        predecessors="49,50",
        resources="EY + Business Leads + Infrastructure Leads",
        notes="Final approval closes the five-day readiness period between QG4 and GoLive.",
    ),
    make_task(
        52,
        3,
        "GoLive - Day 1 cutover complete",
        "2027-02-01",
        "2027-02-01",
        predecessors="51",
        resources="Steering Committee + Infrastructure Leads",
        notes="The Gamma JV environment goes live for 5 sites, 250 users, and 20 applications on the committed date.",
        milestone=True,
    ),
    make_task(
        53,
        1,
        "Phase 4 - Hypercare and closure",
        "2027-02-02",
        "2027-05-31",
        notes="Ninety-day stabilization only after GoLive, followed by formal handover and closure.",
    ),
    make_task(54, 2, "4.1 Hypercare stabilization only", "2027-02-02", "2027-05-02"),
    make_task(
        55,
        3,
        "Run 24x7 incident triage and L3 support",
        "2027-02-02",
        "2027-05-02",
        predecessors="52",
        resources="IT Operations + Service Desk",
        notes="Hypercare is stabilization only and lasts 90 calendar days.",
    ),
    make_task(
        56,
        3,
        "Operate hotfix-only stabilization for infrastructure and applications",
        "2027-02-02",
        "2027-04-18",
        predecessors="52",
        resources="Infrastructure Leads + Application Owners",
        notes="Bug fixes only after GoLive; no new development, migration, or architecture change is allowed.",
    ),
    make_task(
        57,
        3,
        "Deliver user enablement, knowledge transfer, and adoption support",
        "2027-02-02",
        "2027-03-31",
        predecessors="52",
        resources="Training Lead + Service Desk + PMO",
        notes="Support focuses on stabilizing the new operating model for the restricted but growing user community.",
    ),
    make_task(58, 2, "4.2 Transition service exit and steady-state handover", "2027-02-15", "2027-05-07"),
    make_task(
        59,
        3,
        "Retire residual seller-hosted services and confirm exit criteria",
        "2027-02-15",
        "2027-04-09",
        predecessors="52",
        resources="Robert Bosch China IT + Alibaba IT + EY",
        notes="Only residual transition services remain after GoLive and are retired in a controlled manner.",
    ),
    make_task(
        60,
        3,
        "Finalize JV runbooks, supplier handoffs, and service ownership",
        "2027-02-15",
        "2027-04-23",
        predecessors="52",
        resources="IT Operations + Procurement + PMO",
        notes="Operational ownership is formalized for the jointly managed JV steady state.",
    ),
    make_task(
        61,
        3,
        "Confirm operational KPIs and steady-state governance handover",
        "2027-04-12",
        "2027-05-07",
        predecessors="55,56,59,60",
        resources="Steering Committee + IT Operations + EY",
        notes="Readiness for steady-state operations is confirmed after hypercare evidence and service exits are complete.",
    ),
    make_task(
        62,
        3,
        "QG5 - Project completion approved",
        "2027-05-10",
        "2027-05-10",
        predecessors="55,59,60,61",
        resources="Steering Committee",
        notes="Ninety-day hypercare is complete, transition services are exited, and operational handover is accepted.",
        milestone=True,
    ),
    make_task(63, 2, "4.3 Project closure", "2027-05-11", "2027-05-31"),
    make_task(
        64,
        3,
        "Capture lessons learned and archive project artefacts",
        "2027-05-11",
        "2027-05-21",
        predecessors="62",
        resources="PMO + Workstream Leads",
        notes="Repository artefacts and handover evidence are archived after QG5 approval.",
    ),
    make_task(
        65,
        3,
        "Close financials, confirm documentation, and issue closure report",
        "2027-05-11",
        "2027-05-28",
        predecessors="62",
        resources="Finance + PMO",
        notes="Closure report finalizes the initial v1.0 Gamma package baseline.",
    ),
    make_task(
        66,
        3,
        "Closure - Gamma programme archived",
        "2027-05-31",
        "2027-05-31",
        predecessors="64,65",
        resources="Steering Committee",
        notes="Formal closure milestone aligned with the committed completion date.",
        milestone=True,
    ),
]


def _write_temp_csv(path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(TASKS)


def _generate_excel() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"
    sheet.append(HEADERS)

    header_fill = PatternFill(start_color="002147", end_color="002147", fill_type="solid")
    phase_fill = PatternFill(start_color="003B6E", end_color="003B6E", fill_type="solid")
    section_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    milestone_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    even_fill = PatternFill(start_color="EFF4FB", end_color="EFF4FB", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    white_bold_font = Font(bold=True, color="FFFFFF", size=10)
    milestone_font = Font(bold=True, color="000000", size=10)
    detail_font = Font(color="000000", size=9)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    indent_map = {1: "", 2: "  ", 3: "    "}

    for row_index, task in enumerate(TASKS, start=2):
        task_id, level, name, duration, start, finish, preds, resources, notes, milestone = task
        sheet.append(
            [
                task_id,
                level,
                indent_map.get(level, "") + name,
                duration,
                parse_date(start),
                parse_date(finish),
                preds,
                resources,
                notes,
                milestone,
            ]
        )
        row = sheet[row_index]
        if milestone == "Yes":
            row_fill = milestone_fill
            row_font = milestone_font
        elif level == 1:
            row_fill = phase_fill
            row_font = white_bold_font
        elif level == 2:
            row_fill = section_fill
            row_font = white_bold_font
        else:
            row_fill = even_fill if row_index % 2 == 0 else odd_fill
            row_font = detail_font

        for cell in row:
            cell.fill = row_fill
            cell.font = row_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"

    sheet.column_dimensions["A"].width = 6
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 62
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 14
    sheet.column_dimensions["F"].width = 14
    sheet.column_dimensions["G"].width = 16
    sheet.column_dimensions["H"].width = 44
    sheet.column_dimensions["I"].width = 68
    sheet.column_dimensions["J"].width = 10

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:J1"

    workbook.save(XLSX_PATH)


if __name__ == "__main__":
    print(f"[{PROJECT_NAME}] Generating project schedule")
    print(f"  Version: {DOCUMENT_VERSION}")
    print(f"  Output folder: {OUTPUT_FOLDER_NAME}")
    print(f"  Seller: {SELLER}")
    print(f"  Buyer: {BUYER}")
    print(f"  PMO lead: {PMO_LEAD}")
    print(f"  Carve-out model: {CARVEOUT_MODEL}")
    print(f"  Scope: {WORLDWIDE_SITES} sites | {IT_USERS} users | {APPLICATIONS} applications | No SAP")
    print(f"  Start: {START_DATE:%Y-%m-%d} | GoLive: {GOLIVE_DATE:%Y-%m-%d} | Completion: {COMPLETION_DATE:%Y-%m-%d}")
    print(f"  Total tasks: {len(TASKS)}")

    _generate_excel()

    try:
        _write_temp_csv(TEMP_CSV_PATH)
        result = subprocess.run(
            [
                sys.executable,
                str(HERE / "generate_msp_xml.py"),
                "--csv",
                str(TEMP_CSV_PATH),
                "--out",
                str(XML_PATH),
                "--project",
                PROJECT_NAME,
            ],
            capture_output=True,
            text=True,
        )
        if result.returncode != 0:
            print(f"XML generation failed:\n{result.stderr}")
            sys.exit(1)
    finally:
        TEMP_CSV_PATH.unlink(missing_ok=True)

    print(f"  XLSX: {XLSX_PATH}")
    print(f"  XML:  {XML_PATH}")
    print("  Quality gates: QG0, QG1, QG2 and QG3, QG4, GoLive, QG5, Closure")