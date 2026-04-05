#!/usr/bin/env python3
"""
Generate Zebra carve-out project schedule (XLSX + XML) - EXPANDED WORKSTREAM VERSION.

Project: Zebra (Packaging business IT separation)
Seller: Robert Bosch GmbH
Buyer: Undisclosed (legal hold)
Business: Packaging
Model: Stand Alone
Sites: 37 worldwide (EMEA ~15 / APAC ~12 / Americas ~10)
Users: 3500+
Applications: 208 (SAP + applications)
TSA: 6 months post-GoLive

Timeline:
- Start: 1 April 2026
- GoLive: 1 June 2027 (14 months)
- Completion: 31 October 2027 (31 months total)

5 Phases with detailed workstreams:
- Phase 0: Initialization (governance, assessment, legal)
- Phase 1: Concept (requirements, landscape, architecture strategy)
- Phase 2: Design (architecture design, migration strategy, detailed planning)
- Phase 3: Build & Test (infrastructure, ERP, applications, testing)
- Phase 4: GoLive & Closure (cutover, hypercare, closure)

Quality Gates: QG0, QG1, QG2/3, QG4, QG5
Total Tasks: ~165 (comparable to AlphaX structure)
"""

from pathlib import Path
import sys
import os
import subprocess
from datetime import datetime, timedelta

HERE = Path(__file__).parent
PROJECT_NAME = "Zebra"
XLSX_PATH = HERE / "active-projects" / PROJECT_NAME / f"{PROJECT_NAME}_Project_Schedule.xlsx"
XML_PATH  = HERE / "active-projects" / PROJECT_NAME / f"{PROJECT_NAME}_Project_Schedule.xml"
CSV_PATH  = HERE / "active-projects" / PROJECT_NAME / f"{PROJECT_NAME}_Project_Schedule.csv"

# Ensure output directory exists
XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

# Project timeline
START_DATE = datetime(2026, 4, 1)
GOLIVE_DATE = datetime(2027, 6, 1)
COMPLETION_DATE = datetime(2027, 9, 15)  # 105 days post-GoLive (90-day hypercare + 15-day closure/TSA exit overlap)

def date_str(d: datetime) -> str:
    """Convert datetime to ISO YYYY-MM-DD string (used for CSV and XML generation)."""
    return d.strftime("%Y-%m-%d")

def date_plus_days(base: datetime, days: int) -> datetime:
    """Add calendar days to a date."""
    return base + timedelta(days=days)

# ============================================================================
# COMPREHENSIVE ZEBRA TASK LIST - 165 tasks across 5 phases
# Tuple: (ID, OutlineLevel, Name, Duration, Start, Finish, Predecessors, ResourceNames, Notes, Milestone)
# ============================================================================

TASKS = [
    # === PHASE 0: INITIALIZATION (60 days, ~3% of labour) ===
    (1, 1, "Phase 0: Initialization", "60 days", date_str(START_DATE), date_str(date_plus_days(START_DATE, 60)), "", "", "Governance, legal, assessment baseline", "No"),
    
    # 0.1 Project Setup & Governance (16 days)
    (2, 2, "0.1 Project Setup & Governance", "16 days", date_str(START_DATE), date_str(date_plus_days(START_DATE, 16)), "", "", "PMO, steering, RACI", "No"),
    (3, 3, "Appoint Zebra Programme Leads", "5 days", date_str(START_DATE), date_str(date_plus_days(START_DATE, 5)), "", "KPMG PMO Lead + KPMG Project Manager", "Seller and Buyer sponsors appointed", "No"),
    (4, 3, "Establish Steering Committee", "3 days", date_str(date_plus_days(START_DATE, 3)), date_str(date_plus_days(START_DATE, 6)), "3", "KPMG PMO Lead", "First steering committee session", "No"),
    (5, 3, "Define RACI and Governance Model", "5 days", date_str(date_plus_days(START_DATE, 3)), date_str(date_plus_days(START_DATE, 8)), "3", "KPMG Project Manager + KPMG PMO Lead", "Approval of governance structure", "No"),
    (6, 3, "Set Up PMO Tools and Collaboration Site", "8 days", date_str(date_plus_days(START_DATE, 8)), date_str(date_plus_days(START_DATE, 16)), "3", "KPMG PMO Lead", "SharePoint site, project tracking active", "No"),
    (7, 3, "Onboard KPMG and RoboGmbH Teams", "7 days", date_str(date_plus_days(START_DATE, 8)), date_str(date_plus_days(START_DATE, 15)), "3", "KPMG PMO Lead + RoboGmbH IT Manager", "All teams mobilized", "No"),
    
    # 0.2 IT Landscape Assessment (30 days)
    (8, 2, "0.2 IT Landscape Assessment - 37 Sites", "30 days", date_str(date_plus_days(START_DATE, 10)), date_str(date_plus_days(START_DATE, 40)), "3", "", "Inventory, apps, infra, scope", "No"),
    (9, 3, "Application Portfolio Inventory (~208 apps)", "12 days", date_str(date_plus_days(START_DATE, 10)), date_str(date_plus_days(START_DATE, 22)), "5", "KPMG Data Architect + RoboGmbH IT Manager", "Packaging-specific apps; flag shared Bosch dependencies", "No"),
    (10, 3, "IT Infrastructure As-Is Mapping (37 sites)", "15 days", date_str(date_plus_days(START_DATE, 10)), date_str(date_plus_days(START_DATE, 25)), "5", "RoboGmbH IT Manager + KPMG Infrastructure Architect", "Site surveys top 15 locations; EMEA/APAC prioritized", "No"),
    (11, 3, "ERP Landscape Analysis (SAP scope)", "15 days", date_str(date_plus_days(START_DATE, 10)), date_str(date_plus_days(START_DATE, 25)), "5", "KPMG SAP Architect + RoboGmbH ERP Specialist", "Identify packaging-specific vs shared SAP; separation complexity", "No"),
    (12, 3, "Data Ownership & Separation Rules (Draft)", "10 days", date_str(date_plus_days(START_DATE, 20)), date_str(date_plus_days(START_DATE, 30)), "9,11", "KPMG Data Architect", "Data classification baseline", "No"),
    (13, 3, "Contract & Licence Inventory (~208 apps)", "12 days", date_str(date_plus_days(START_DATE, 10)), date_str(date_plus_days(START_DATE, 22)), "5", "RoboGmbH IT Manager", "Change of control review; FOSS compliance; ~3,500 device licences", "No"),
    (14, 3, "HR IT Systems & Payroll Separation (37 sites)", "10 days", date_str(date_plus_days(START_DATE, 15)), date_str(date_plus_days(START_DATE, 25)), "5", "KPMG Data Architect", "Country-by-country HR rules (Germany, China, India, Mexico, etc.)", "No"),
    (15, 3, "OT & Production IT Assessment", "10 days", date_str(date_plus_days(START_DATE, 18)), date_str(date_plus_days(START_DATE, 28)), "5", "RoboGmbH IT Manager + KPMG Infrastructure Architect", "Production systems; safety certifications; ICS/SCADA scope", "No"),
    (16, 3, "Country-Specific Legal Requirements (Multi-region)", "8 days", date_str(date_plus_days(START_DATE, 22)), date_str(date_plus_days(START_DATE, 30)), "5", "KPMG Project Manager", "Germany works council; China PIPL; India; Mexico SAT delays", "No"),
    
    # 0.3 Project Charter & Scope (20 days)
    (17, 2, "0.3 Project Charter & Scope Definition", "20 days", date_str(date_plus_days(START_DATE, 30)), date_str(date_plus_days(START_DATE, 50)), "8", "", "Charter, approval, scope frozen", "No"),
    (18, 3, "Draft Zebra Project Charter", "8 days", date_str(date_plus_days(START_DATE, 30)), date_str(date_plus_days(START_DATE, 38)), "8", "KPMG PMO Lead + KPMG Project Manager", "Executive charter document", "No"),
    (19, 3, "Define Scope Boundaries (37 sites, 3,500+ users, 208 apps)", "8 days", date_str(date_plus_days(START_DATE, 35)), date_str(date_plus_days(START_DATE, 43)), "10,11", "KPMG Project Manager + KPMG SAP Architect", "Stand Alone model confirmed; no Merger Zone", "No"),
    (20, 3, "Guiding Principles Workshop", "3 days", date_str(date_plus_days(START_DATE, 40)), date_str(date_plus_days(START_DATE, 43)), "18,19", "KPMG PMO Lead + All Workstream Leads", "First Make It Work; priorities set", "No"),
    (21, 3, "TSA Framework Preliminary Scope", "8 days", date_str(date_plus_days(START_DATE, 40)), date_str(date_plus_days(START_DATE, 48)), "18,19", "KPMG Project Manager + RoboGmbH IT Manager", "Bosch operates all IT until Day 1; TSA scope defined", "No"),
    (22, 3, "Project Charter Approval & Gate QG0", "5 days", date_str(date_plus_days(START_DATE, 48)), date_str(date_plus_days(START_DATE, 53)), "18,19,20,21", "KPMG PMO Lead", "Charter signed; Phase 0 gate approval", "No"),
    (23, 3, "QG0 - Initialization Quality Gate", "0 days", date_str(date_plus_days(START_DATE, 53)), date_str(date_plus_days(START_DATE, 53)), "22", "KPMG PMO Lead + Buyer Sponsor", "Charter approved; inventory confirmed; governance ready", "Yes"),
    
    # === PHASE 1: CONCEPT (65 days, ~12% of labour) ===
    (24, 1, "Phase 1: Concept", "65 days", date_str(date_plus_days(START_DATE, 54)), date_str(date_plus_days(START_DATE, 119)), "23", "", "Requirements, design strategy, baseline planning", "No"),
    
    # 1.1 Detailed As-Is Analysis (40 days)
    (25, 2, "1.1 Detailed As-Is Analysis", "40 days", date_str(date_plus_days(START_DATE, 54)), date_str(date_plus_days(START_DATE, 94)), "23", "", "Landscape detail, dependencies", "No"),
    (26, 3, "Application Portfolio Deep Dive (LeanIX, 208 apps)", "20 days", date_str(date_plus_days(START_DATE, 54)), date_str(date_plus_days(START_DATE, 74)), "9", "KPMG Data Architect + KPMG Infrastructure Architect", "Flag Bosch shared dependencies; categorize critical vs standard", "No"),
    (27, 3, "SAP Landscape Detailed Analysis", "20 days", date_str(date_plus_days(START_DATE, 54)), date_str(date_plus_days(START_DATE, 74)), "11", "KPMG SAP Architect + RoboGmbH ERP Specialist", "Identify separation complexity; module scope", "No"),
    (28, 3, "Infrastructure Detailed Mapping (37 sites, top 15 surveyed)", "20 days", date_str(date_plus_days(START_DATE, 54)), date_str(date_plus_days(START_DATE, 74)), "10", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "WAN, LAN, co-lo, security topology", "No"),
    (29, 3, "Contract & Licence Final Inventory (208 apps)", "15 days", date_str(date_plus_days(START_DATE, 68)), date_str(date_plus_days(START_DATE, 83)), "13", "RoboGmbH IT Manager", "Change of control clause review; vendor dependencies", "No"),
    (30, 3, "HR & Payroll System Separation (Multi-region)", "15 days", date_str(date_plus_days(START_DATE, 70)), date_str(date_plus_days(START_DATE, 85)), "14", "KPMG Data Architect", "Region-specific payroll rules finalized", "No"),
    
    # 1.2 Architecture Design Strategy (25 days)
    (31, 2, "1.2 Stand-Alone Architecture Design Strategy", "25 days", date_str(date_plus_days(START_DATE, 60)), date_str(date_plus_days(START_DATE, 85)), "26,27,28", "", "Architecture concept (no Merger Zone)", "No"),
    (32, 3, "Architecture Concept Workshop (Stand Alone approach)", "5 days", date_str(date_plus_days(START_DATE, 60)), date_str(date_plus_days(START_DATE, 65)), "26", "KPMG SAP Architect + KPMG Infrastructure Architect", "Confirm Stand Alone; no Merger Zone; target architecture", "No"),
    (33, 3, "WAN & Network Architecture (37 sites)", "12 days", date_str(date_plus_days(START_DATE, 75)), date_str(date_plus_days(START_DATE, 87)), "32,28", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "Hub-and-spoke model; 1-2 regional co-lo; 4-6 month lead time", "No"),
    (34, 3, "Active Directory Design (New Forest)", "10 days", date_str(date_plus_days(START_DATE, 75)), date_str(date_plus_days(START_DATE, 85)), "32", "KPMG Infrastructure Architect", "New AD forest for Zebra; fully independent from Bosch", "No"),
    (35, 3, "M365 & Azure Tenant Design (3,500+ mailboxes)", "10 days", date_str(date_plus_days(START_DATE, 75)), date_str(date_plus_days(START_DATE, 85)), "32", "KPMG Infrastructure Architect", "New M365 tenant; OneDrive, Teams in scope", "No"),
    (36, 3, "Co-Location & Data Centre Selection (1-2 regional hubs)", "10 days", date_str(date_plus_days(START_DATE, 60)), date_str(date_plus_days(START_DATE, 70)), "28", "RoboGmbH IT Manager", "Primary hub selection; decision by QG1", "No"),
    (37, 3, "Security Architecture (NewCo Zebra independent)", "12 days", date_str(date_plus_days(START_DATE, 80)), date_str(date_plus_days(START_DATE, 92)), "33,34", "KPMG Infrastructure Architect", "Independent SOC; security baseline", "No"),
    (38, 3, "IAM & Identity Design (Decouple from Bosch)", "10 days", date_str(date_plus_days(START_DATE, 80)), date_str(date_plus_days(START_DATE, 90)), "34", "KPMG Infrastructure Architect", "New IAM platform; decouple from Bosch Saviynt", "No"),
    
    # 1.3 Migration Strategy (25 days)
    (39, 2, "1.3 Zebra IT Migration Strategy", "25 days", date_str(date_plus_days(START_DATE, 75)), date_str(date_plus_days(START_DATE, 100)), "27,29", "", "ERP, apps, clients, data", "No"),
    (40, 3, "ERP Migration Strategy Decision (Shell Copy vs Greenfield)", "12 days", date_str(date_plus_days(START_DATE, 75)), date_str(date_plus_days(START_DATE, 87)), "27", "KPMG SAP Architect + RoboGmbH ERP Specialist", "Shell copy benchmarked 9-12 months; decision by QG1", "No"),
    (41, 3, "Application Categorization & Wave Planning (208 apps)", "12 days", date_str(date_plus_days(START_DATE, 85)), date_str(date_plus_days(START_DATE, 97)), "26,40", "KPMG Data Architect", "Wave 1 ~60 critical; Wave 2 ~80 standard; Wave 3 ~68 local/OT", "No"),
    (42, 3, "Data Migration & Separation Strategy", "10 days", date_str(date_plus_days(START_DATE, 90)), date_str(date_plus_days(START_DATE, 100)), "12,27", "KPMG Data Architect + KPMG Data Engineer", "Non-selective approach; residual data in Bosch until TSA exit", "No"),
    (43, 3, "TSA Service Catalogue Definition", "15 days", date_str(date_plus_days(START_DATE, 85)), date_str(date_plus_days(START_DATE, 100)), "21,32", "KPMG Project Manager + RoboGmbH IT Manager", "Catalogue all Bosch services Zebra requires; exit criteria per service", "No"),
    (44, 3, "Zebra Stand-Alone Operating Model (post-TSA)", "12 days", date_str(date_plus_days(START_DATE, 95)), date_str(date_plus_days(START_DATE, 107)), "40,41", "KPMG Infrastructure Architect", "All services exit directly post-TSA; define operating model", "No"),
    
    # 1.4 Client Device & Workplace Strategy (12 days)
    (45, 2, "1.4 Client Device & Workplace Strategy", "12 days", date_str(date_plus_days(START_DATE, 75)), date_str(date_plus_days(START_DATE, 87)), "26", "", "3,500+ device migration approach", "No"),
    (46, 3, "Client Device Inventory & Assessment (~3,500 devices)", "8 days", date_str(date_plus_days(START_DATE, 75)), date_str(date_plus_days(START_DATE, 83)), "26", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "Device age, type, replacement plan for out-of-support", "No"),
    (47, 3, "Client Migration Approach (Wave-based reimaging)", "8 days", date_str(date_plus_days(START_DATE, 80)), date_str(date_plus_days(START_DATE, 88)), "46", "KPMG Infrastructure Architect", "3,500 devices across 37 sites; EMEA/APAC/Americas stagger", "No"),
    
    # 1.5 Baseline Planning (15 days)
    (48, 2, "1.5 Planning & Baseline", "15 days", date_str(date_plus_days(START_DATE, 100)), date_str(date_plus_days(START_DATE, 115)), "31,39,45", "", "Schedule, budget, risk baseline", "No"),
    (49, 3, "Detailed Project Plan Development", "10 days", date_str(date_plus_days(START_DATE, 100)), date_str(date_plus_days(START_DATE, 110)), "31,39", "KPMG PMO Lead + KPMG Project Manager", "Full phase schedule with dependencies", "No"),
    (50, 3, "Risk Register Baseline (Initial risks)", "8 days", date_str(date_plus_days(START_DATE, 100)), date_str(date_plus_days(START_DATE, 108)), "39", "KPMG Project Manager", "Risk identification and scoring", "No"),
    (51, 3, "OPL Establishment & Tracking", "5 days", date_str(date_plus_days(START_DATE, 105)), date_str(date_plus_days(START_DATE, 110)), "49", "KPMG PMO Lead", "Open Points List system active", "No"),
    (52, 3, "Resource Plan & Budget Baseline", "8 days", date_str(date_plus_days(START_DATE, 105)), date_str(date_plus_days(START_DATE, 113)), "49", "KPMG PMO Lead + KPMG Project Manager", "Labour and CAPEX baseline", "No"),
    (53, 3, "QG1 - Concept Quality Gate", "0 days", date_str(date_plus_days(START_DATE, 115)), date_str(date_plus_days(START_DATE, 115)), "49,44,43", "KPMG PMO Lead + Buyer Sponsor", "Architecture approved; ERP strategy decided; TSA catalogue signed; proceed to Design", "Yes"),
    
    # === PHASE 2: ARCHITECTURE & DESIGN (60 days, ~25% of labour) ===
    (54, 1, "Phase 2: Architecture & Design", "60 days", date_str(date_plus_days(START_DATE, 116)), date_str(date_plus_days(START_DATE, 176)), "53", "", "Detailed design, vendor selection, architecture finalization", "No"),
    
    # 2.1 Infrastructure Design & Procurement (40 days)
    (55, 2, "2.1 Infrastructure Design & Procurement", "40 days", date_str(date_plus_days(START_DATE, 116)), date_str(date_plus_days(START_DATE, 156)), "53", "", "WAN, co-lo, AD, M365, security design", "No"),
    (56, 3, "WAN Provider Selection & Ordering (37 sites)", "15 days", date_str(date_plus_days(START_DATE, 116)), date_str(date_plus_days(START_DATE, 131)), "33", "RoboGmbH IT Manager + KPMG Infrastructure Architect", "Order immediately after design; 4-6 month lead time", "No"),
    (57, 3, "Co-Location Data Centre Procurement (1-2 regional hubs)", "20 days", date_str(date_plus_days(START_DATE, 116)), date_str(date_plus_days(START_DATE, 136)), "36", "RoboGmbH IT Manager", "Primary EMEA co-locator for Zebra DC", "No"),
    (58, 3, "Active Directory Detailed Configuration", "20 days", date_str(date_plus_days(START_DATE, 130)), date_str(date_plus_days(START_DATE, 150)), "56,34", "KPMG Infrastructure Architect", "New Zebra AD forest specifications; delegation model", "No"),
    (59, 3, "M365 Tenant & Azure Configuration (3,500+ mailboxes)", "18 days", date_str(date_plus_days(START_DATE, 135)), date_str(date_plus_days(START_DATE, 153)), "35,56", "KPMG Infrastructure Architect", "Tenant isolation; licensing; governance", "No"),
    (60, 3, "Security & Firewall Architecture Finalization", "18 days", date_str(date_plus_days(START_DATE, 140)), date_str(date_plus_days(START_DATE, 158)), "37,56", "KPMG Infrastructure Architect", "Network security, endpoint security, DLP policies", "No"),
    (61, 3, "IAM Platform & SSO Design", "15 days", date_str(date_plus_days(START_DATE, 140)), date_str(date_plus_days(START_DATE, 155)), "38", "KPMG Infrastructure Architect", "Identity governance, provisioning automation", "No"),
    (62, 3, "CMDB & ITSM Platform Design", "12 days", date_str(date_plus_days(START_DATE, 145)), date_str(date_plus_days(START_DATE, 157)), "56,58", "KPMG Infrastructure Architect", "ServiceNow or alternative ITSM platform configuration", "No"),
    
    # 2.2 ERP & SAP Design (30 days)
    (63, 2, "2.2 SAP & ERP Design", "30 days", date_str(date_plus_days(START_DATE, 116)), date_str(date_plus_days(START_DATE, 146)), "40", "", "Shell copy prep, interfaces, auth design", "No"),
    (64, 3, "SAP System Landscape Detail & Separation", "15 days", date_str(date_plus_days(START_DATE, 116)), date_str(date_plus_days(START_DATE, 131)), "27", "KPMG SAP Architect + RoboGmbH ERP Specialist", "Module scope for Zebra packaging", "No"),
    (65, 3, "SAP Shell Copy Detailed Approach & Timeline", "18 days", date_str(date_plus_days(START_DATE, 125)), date_str(date_plus_days(START_DATE, 143)), "40,64", "KPMG SAP Architect + KPMG SAP Build Lead", "Extraction approach; timeline; re-configuration items", "No"),
    (66, 3, "SAP Interfaces & Integration Design (APIs)", "15 days", date_str(date_plus_days(START_DATE, 130)), date_str(date_plus_days(START_DATE, 145)), "64", "KPMG SAP Build Lead + KPMG Data Engineer", "Identify all inbound/outbound interfaces; EDI approach", "No"),
    (67, 3, "SAP Authorization & Security Design", "12 days", date_str(date_plus_days(START_DATE, 135)), date_str(date_plus_days(START_DATE, 147)), "38,64", "KPMG Infrastructure Architect", "User roles; segregation of duties; authorization matrix", "No"),
    (68, 3, "SAP Test Environment & Tools Selection", "10 days", date_str(date_plus_days(START_DATE, 135)), date_str(date_plus_days(START_DATE, 145)), "65", "KPMG SAP Architect + RoboGmbH ERP Specialist", "SIT, UAT environment specifications; performance testing tools", "No"),
    
    # 2.3 Application Migration Design (25 days)
    (69, 2, "2.3 Application Migration Design (208 apps)", "25 days", date_str(date_plus_days(START_DATE, 120)), date_str(date_plus_days(START_DATE, 145)), "41", "", "Wave definition, packaging, integration strategy", "No"),
    (70, 3, "Application Wave Refinement & Dependencies", "12 days", date_str(date_plus_days(START_DATE, 120)), date_str(date_plus_days(START_DATE, 132)), "41", "KPMG Data Architect", "Wave 1/2/3 app list finalized; cross-wave dependencies", "No"),
    (71, 3, "Application Package Development (Wave 1 critical ~60 apps)", "18 days", date_str(date_plus_days(START_DATE, 128)), date_str(date_plus_days(START_DATE, 146)), "70", "KPMG Data Architect + KPMG Infrastructure Architect", "Deployment package; configuration; cutover script", "No"),
    (72, 3, "Custom App Adaptation Design (Re-point to Zebra domain)", "15 days", date_str(date_plus_days(START_DATE, 135)), date_str(date_plus_days(START_DATE, 150)), "71", "KPMG Data Architect", "Domain, AD, M365, ITSM configuration changes", "No"),
    
    # 2.4 Client Device Design (15 days)
    (73, 2, "2.4 Client Workplace Design & Imaging", "15 days", date_str(date_plus_days(START_DATE, 120)), date_str(date_plus_days(START_DATE, 135)), "47", "", "Device imaging, wave plan", "No"),
    (74, 3, "Standard Image Development (SCCM/Intune)", "12 days", date_str(date_plus_days(START_DATE, 120)), date_str(date_plus_days(START_DATE, 132)), "47", "KPMG Infrastructure Architect", "Standard image for Zebra domain; hardware refresh plan", "No"),
    (75, 3, "Regional Wave Plan (EMEA / APAC / Americas devices)", "8 days", date_str(date_plus_days(START_DATE, 125)), date_str(date_plus_days(START_DATE, 133)), "47,74", "KPMG Infrastructure Architect", "Wave sequencing; logistics; regional priorities", "No"),
    
    # 2.5 Cutover & TSA Design (25 days)
    (76, 2, "2.5 Cutover Strategy & TSA Finalization", "25 days", date_str(date_plus_days(START_DATE, 130)), date_str(date_plus_days(START_DATE, 155)), "41,43,66", "", "Go-live plan, TSA contracts", "No"),
    (77, 3, "Detailed Cutover Plan (37 sites, 208 apps, SAP)", "18 days", date_str(date_plus_days(START_DATE, 130)), date_str(date_plus_days(START_DATE, 148)), "41,65", "KPMG Project Manager + KPMG SAP Architect", "Parallel run approach; rollback plan; timing sequence", "No"),
    (78, 3, "TSA Service Descriptions Final (All 37 sites)", "15 days", date_str(date_plus_days(START_DATE, 135)), date_str(date_plus_days(START_DATE, 150)), "43", "KPMG Project Manager + RoboGmbH IT Manager", "Each TSA service line; exit date; acceptance criteria", "No"),
    (79, 3, "TSA SLA & KPI Framework", "10 days", date_str(date_plus_days(START_DATE, 145)), date_str(date_plus_days(START_DATE, 155)), "78", "KPMG Project Manager", "SLA levels; reporting cadence", "No"),
    (80, 3, "TSA Legal & Contracts Final Review", "12 days", date_str(date_plus_days(START_DATE, 145)), date_str(date_plus_days(START_DATE, 157)), "78,79", "KPMG Project Manager", "Contract finalization and signature readiness", "No"),
    (81, 3, "QG2/3 - Design Quality Gate", "0 days", date_str(date_plus_days(START_DATE, 158)), date_str(date_plus_days(START_DATE, 158)), "62,68,72,75,80", "KPMG PMO Lead + Buyer Sponsor", "Infrastructure designed; ERP designed; apps packaged; devices planned; TSA approved; proceed to Build", "Yes"),
    
    # === PHASE 3: DEVELOPMENT, BUILD & TEST (100 days, ~48% of labour) ===
    (82, 1, "Phase 3: Development, Build & Test", "100 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(GOLIVE_DATE, -10)), "81", "", "Infrastructure build, ERP setup, app prep, testing", "No"),
    
    # 3.1 Infrastructure Build (60 days)
    (83, 2, "3.1 Infrastructure Build (37 sites)", "60 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 219)), "81", "", "WAN, co-lo, AD, M365, security, operations setup", "No"),
    (84, 3, "WAN Circuits Delivery (EMEA batch 1, ~8 sites)", "25 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 184)), "56", "RoboGmbH IT Manager + KPMG Infrastructure Architect", "4-6 month lead completed; circuit testing", "No"),
    (85, 3, "WAN Circuits Delivery (APAC batch 1, ~6 sites)", "25 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 184)), "56", "RoboGmbH IT Manager + KPMG Infrastructure Architect", "China local provider; stagger by country", "No"),
    (86, 3, "WAN Circuits Delivery (Americas batch 1, ~4-5 sites)", "25 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 184)), "56", "RoboGmbH IT Manager + KPMG Infrastructure Architect", "Mexico legal entity confirmed before delivery", "No"),
    (87, 3, "Co-Location DC Setup & Testing", "30 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 189)), "57", "RoboGmbH IT Manager + KPMG Infrastructure Architect", "Cooling, power, connectivity; Zebra servers racked", "No"),
    (88, 3, "Active Directory Build & User Provisioning", "30 days", date_str(date_plus_days(START_DATE, 180)), date_str(date_plus_days(START_DATE, 210)), "58,84", "KPMG Infrastructure Architect", "All 3,500+ users and app access configured", "No"),
    (89, 3, "M365 Tenant & Azure Production Build", "25 days", date_str(date_plus_days(START_DATE, 185)), date_str(date_plus_days(START_DATE, 210)), "59,87", "KPMG Infrastructure Architect", "3,500+ mailboxes provisioned; OneDrive structure built", "No"),
    (90, 3, "Security & Firewall Implementation", "25 days", date_str(date_plus_days(START_DATE, 190)), date_str(date_plus_days(START_DATE, 215)), "60,84", "KPMG Infrastructure Architect", "Security appliances; DLP; endpoint protection", "No"),
    (91, 3, "IAM Platform Deployment & User Sync", "25 days", date_str(date_plus_days(START_DATE, 185)), date_str(date_plus_days(START_DATE, 210)), "61,88", "KPMG Infrastructure Architect", "SSO, provisioning automation for 3,500+ users", "No"),
    (92, 3, "CMDB & ITSM Platform Setup (Assets, CIs)", "20 days", date_str(date_plus_days(START_DATE, 205)), date_str(date_plus_days(START_DATE, 225)), "62,88", "KPMG Infrastructure Architect", "Asset inventory; service catalogue; incident management", "No"),
    (93, 3, "Telephony Platform Setup & Number Porting (37 sites)", "20 days", date_str(date_plus_days(START_DATE, 210)), date_str(date_plus_days(START_DATE, 230)), "87", "RoboGmbH IT Manager", "Phone system config; 4-8 weeks per operator per site", "No"),
    (94, 3, "File Share Infrastructure & User Documents", "18 days", date_str(date_plus_days(START_DATE, 205)), date_str(date_plus_days(START_DATE, 223)), "87,89", "KPMG Infrastructure Architect", "NAS/Cloud setup for 3,500+ users; shared directories", "No"),
    (95, 3, "Backup & Disaster Recovery Setup", "15 days", date_str(date_plus_days(START_DATE, 215)), date_str(date_plus_days(START_DATE, 230)), "87", "KPMG Infrastructure Architect", "Backup solution; RTO/RPO targets defined", "No"),
    
    # 3.2 ERP & SAP Build (70 days)
    (96, 2, "3.2 SAP & ERP Development", "70 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 229)), "81", "", "Shell copy, build, interfaces, auth, testing", "No"),
    (97, 3, "SAP System Landscape Build (Zebra-specific modules)", "20 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 179)), "64", "KPMG SAP Architect + KPMG SAP Build Lead + RoboGmbH ERP Specialist", "Base system build; module activation", "No"),
    (98, 3, "SAP Shell Copy Data Extraction & Preparation", "25 days", date_str(date_plus_days(START_DATE, 170)), date_str(date_plus_days(START_DATE, 195)), "97", "KPMG SAP Build Lead + RoboGmbH ERP Specialist", "Master data cleanup; extraction from Bosch SAP", "No"),
    (99, 3, "Data Migration Dry Run 1 (Test)", "15 days", date_str(date_plus_days(START_DATE, 180)), date_str(date_plus_days(START_DATE, 195)), "98", "KPMG Data Engineer + RoboGmbH ERP Specialist", "Extract, transform, load; validation", "No"),
    (100, 3, "Data Migration Dry Run 2 (Refinement)", "15 days", date_str(date_plus_days(START_DATE, 195)), date_str(date_plus_days(START_DATE, 210)), "99", "KPMG Data Engineer + RoboGmbH ERP Specialist", "Fix issues from Dry Run 1", "No"),
    (101, 3, "SAP Interfaces & Integration Development", "25 days", date_str(date_plus_days(START_DATE, 185)), date_str(date_plus_days(START_DATE, 210)), "66,98", "KPMG SAP Build Lead + KPMG Data Engineer", "APIs, EDI mappings, middleware configuration", "No"),
    (102, 3, "SAP Authorization & Security Hardening", "15 days", date_str(date_plus_days(START_DATE, 200)), date_str(date_plus_days(START_DATE, 215)), "67,101", "KPMG Infrastructure Architect", "User role assignments; access control verification", "No"),
    (103, 3, "SAP Test Environment Build (SIT ready)", "20 days", date_str(date_plus_days(START_DATE, 195)), date_str(date_plus_days(START_DATE, 215)), "68,98,101", "KPMG SAP Architect + RoboGmbH ERP Specialist", "SIT/UAT environment ready for integration testing", "No"),
    
    # 3.3 Application Migration Prep (60 days)
    (104, 2, "3.3 Application Migration Preparation (208 apps)", "60 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 219)), "81", "", "Wave packages, custom adaptation, integration test prep", "No"),
    (105, 3, "Wave 1 Application Packaging (Critical ~60 apps)", "25 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 184)), "71", "KPMG Data Architect + KPMG Infrastructure Architect", "Deployment packages; configuration parameters", "No"),
    (106, 3, "Wave 2 Application Packaging (Standard ~80 apps)", "25 days", date_str(date_plus_days(START_DATE, 180)), date_str(date_plus_days(START_DATE, 205)), "70", "KPMG Data Architect + KPMG Infrastructure Architect", "packaging during Wave 1 execution prep", "No"),
    (107, 3, "Wave 3 Application Packaging (Local/OT ~68 apps)", "20 days", date_str(date_plus_days(START_DATE, 195)), date_str(date_plus_days(START_DATE, 215)), "70", "KPMG Data Architect + KPMG Infrastructure Architect", "Regional and OT-specific apps", "No"),
    (108, 3, "Custom Application Adaptation for Zebra Environment", "25 days", date_str(date_plus_days(START_DATE, 180)), date_str(date_plus_days(START_DATE, 205)), "72,88", "KPMG Data Architect + RoboGmbH IT Manager", "Domain re-point, AD, M365, ITSM integration", "No"),
    (109, 3, "Application Integration Development (APIs, Data mapping)", "20 days", date_str(date_plus_days(START_DATE, 190)), date_str(date_plus_days(START_DATE, 210)), "66,101", "KPMG Data Engineer + KPMG SAP Build Lead", "Cross-app and SAP integration development", "No"),
    (110, 3, "Application Integration Testing Prep", "15 days", date_str(date_plus_days(START_DATE, 205)), date_str(date_plus_days(START_DATE, 220)), "105,108,109", "KPMG Infrastructure Architect + KPMG Test Specialist", "Test scenario prep; test data setup", "No"),
    
    # 3.4 Client Workplace & Device Migration (100 days — ALL BEFORE QG4/GoLive)
    (111, 2, "3.4 Client Workplace, Device & M365 Migration (ALL PRE-GoLive)", "100 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(GOLIVE_DATE, -8)), "81", "", "Device imaging, all 5 migration waves, M365 mailbox migration ALL before QG4", "No"),
    (112, 3, "SCCM / Intune Configuration (Standard image)", "20 days", date_str(date_plus_days(START_DATE, 159)), date_str(date_plus_days(START_DATE, 179)), "74", "KPMG Infrastructure Architect", "Device management platform setup", "No"),
    (113, 3, "Device Imaging & Standard OS Build", "20 days", date_str(date_plus_days(START_DATE, 170)), date_str(date_plus_days(START_DATE, 190)), "74,112", "KPMG Infrastructure Architect", "Standard image production; imaging tooling", "No"),
    (114, 3, "Regional Wave Preparation (EMEA / APAC / Americas logistics)", "15 days", date_str(date_plus_days(START_DATE, 180)), date_str(date_plus_days(START_DATE, 195)), "75,113", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "Imaging schedule; regional IT team coordination", "No"),
    (115, 3, "User Profile & OneDrive Migration Tooling (Setup)", "12 days", date_str(date_plus_days(START_DATE, 195)), date_str(date_plus_days(START_DATE, 207)), "89,113", "KPMG Infrastructure Architect", "Migration tools for user data; setup & testing", "No"),
    (116, 3, "Device Migration Wave 1 (EMEA hub sites, ~500 devices)", "8 days", date_str(date_plus_days(START_DATE, 200)), date_str(date_plus_days(START_DATE, 208)), "113,114", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "First regional wave; EMEA priority (pre-GoLive)", "No"),
    (117, 3, "Device Migration Wave 2 (APAC sites, ~700 devices)", "8 days", date_str(date_plus_days(START_DATE, 203)), date_str(date_plus_days(START_DATE, 211)), "114,113", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "Parallel with Wave 1; China coordination (pre-GoLive)", "No"),
    (118, 3, "Device Migration Wave 3 (EMEA remaining, ~300 devices)", "8 days", date_str(date_plus_days(START_DATE, 208)), date_str(date_plus_days(START_DATE, 216)), "116", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "Remaining EMEA sites (pre-GoLive)", "No"),
    (119, 3, "Device Migration Wave 4 (Americas, ~500 devices)", "8 days", date_str(date_plus_days(START_DATE, 210)), date_str(date_plus_days(START_DATE, 218)), "117", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "Americas sites (pre-GoLive)", "No"),
    (120, 3, "Device Migration Wave 5 (Remaining & exceptions, ~400 devices)", "8 days", date_str(date_plus_days(START_DATE, 215)), date_str(date_plus_days(GOLIVE_DATE, -10)), "118,119", "KPMG Infrastructure Architect", "Stragglers, replacements before GoLive (pre-GoLive)", "No"),
    (121, 3, "M365 Mailbox & OneDrive Migration (3,500+ users — ALL PRE-GoLive)", "30 days", date_str(date_plus_days(START_DATE, 190)), date_str(date_plus_days(GOLIVE_DATE, -10)), "89,115", "KPMG Infrastructure Architect", "Batched user mailbox & OneDrive migration BEFORE GoLive so users ready Day 1", "No"),
    
    # 3.5 Testing & UAT (50 days)
    (122, 2, "3.5 System Integration Testing & UAT", "50 days", date_str(date_plus_days(START_DATE, 195)), date_str(date_plus_days(GOLIVE_DATE, -10)), "81", "", "SIT, integration tests, UAT", "No"),
    (123, 3, "SAP System Integration Test 1 (SIT1)", "20 days", date_str(date_plus_days(START_DATE, 195)), date_str(date_plus_days(START_DATE, 215)), "103", "KPMG SAP Specialist + KPMG Test Specialist + RoboGmbH ERP Specialist", "Integration of apps + SAP; validation", "No"),
    (124, 3, "Application Integration Testing (Wave 1 + SAP)", "18 days", date_str(date_plus_days(START_DATE, 205)), date_str(date_plus_days(START_DATE, 223)), "110,123", "KPMG Test Specialist + KPMG Data Architect", "End-to-end Wave 1 app integration", "No"),
    (125, 3, "User Acceptance Testing (Business owners, 37 sites)", "25 days", date_str(date_plus_days(START_DATE, 215)), date_str(date_plus_days(GOLIVE_DATE, -15)), "124", "KPMG QA Lead + KPMG Test Specialist + RoboGmbH IT Manager", "Sign-off on functional & SAP transactions; regional testers; device reimaging validation", "No"),
    (126, 3, "Dress Rehearsal & Cutover Simulation", "8 days", date_str(date_plus_days(GOLIVE_DATE, -15)), date_str(date_plus_days(GOLIVE_DATE, -8)), "125", "KPMG Project Manager + KPMG SAP Build Lead + KPMG Test Specialist", "Full dry-run with rollback; issue log closeout; migration validation", "No"),
    (127, 3, "Ready for QG4 Gate Review", "0 days", date_str(date_plus_days(GOLIVE_DATE, -8)), date_str(date_plus_days(GOLIVE_DATE, -8)), "120,121,126", "KPMG PMO Lead", "All infrastructure, SAP, apps ready; 3,500 devices reimaged & migrated; 3,500 mailboxes migrated; UAT passed; dress rehearsal successful", "Yes"),
    
    # === PHASE 4: GOLIVE & CLOSURE (60 days, ~12% of labour) ===
    (128, 1, "Phase 4: GoLive & Closure", "60 days", date_str(GOLIVE_DATE), date_str(date_plus_days(GOLIVE_DATE, 105)), "127", "", "Cutover, hypercare, TSA exit, closure", "No"),
    
    # 4.1 GoLive Cutover (7 days)
    (129, 2, "4.1 GoLive Day 1 Activities", "7 days", date_str(GOLIVE_DATE), date_str(date_plus_days(GOLIVE_DATE, 7)), "127", "", "Final data load, cutover execution, activation", "No"),
    (130, 3, "Final Data Load & Production Cutover", "3 days", date_str(GOLIVE_DATE), date_str(date_plus_days(GOLIVE_DATE, 3)), "127", "KPMG Data Engineer + KPMG SAP Build Lead + RoboGmbH ERP Specialist", "Final shell copy data load; integrity validation", "No"),
    (131, 3, "SAP Go-Live Activation", "2 days", date_str(GOLIVE_DATE), date_str(date_plus_days(GOLIVE_DATE, 2)), "130", "KPMG SAP Specialist + RoboGmbH ERP Specialist", "SAP production activation for Zebra", "No"),
    (132, 3, "Application Go-Live Activation (Wave 1, 60 apps)", "2 days", date_str(GOLIVE_DATE), date_str(date_plus_days(GOLIVE_DATE, 2)), "130", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "Apps cut to production; validation", "No"),
    (133, 3, "Network & Domain Cutover (All 37 sites)", "2 days", date_str(GOLIVE_DATE), date_str(date_plus_days(GOLIVE_DATE, 2)), "130", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "AD cutover, DNS, WAN routing activation", "No"),
    (134, 3, "Help Desk 24x7 Activation", "1 days", date_str(GOLIVE_DATE), date_str(GOLIVE_DATE), "127", "RoboGmbH IT Manager + KPMG Test Specialist", "Multi-region support live for 3,500 reimaged+migrated users (all devices ready pre-GoLive)", "No"),
    (135, 3, "GoLive Day 1 Milestone", "0 days", date_str(GOLIVE_DATE), date_str(GOLIVE_DATE), "131,132,133", "KPMG PMO Lead + Buyer Sponsor", "Zebra business live on independent IT systems (all devices reimaged, all mailboxes migrated PRE-GoLive); TSA support begins", "Yes"),
    
    # 4.2 Hypercare (90 days — MANDATORY)
    (136, 2, "4.2 Hypercare & Stabilization (90 days)", "90 days", date_str(date_plus_days(GOLIVE_DATE, 1)), date_str(date_plus_days(GOLIVE_DATE, 90)), "135", "", "24x7 support, stabilization, issue resolution; NO migrations post-GoLive", "No"),
    (137, 3, "Daily Operations Standups & Issue Management", "90 days", date_str(date_plus_days(GOLIVE_DATE, 1)), date_str(date_plus_days(GOLIVE_DATE, 90)), "135", "KPMG Project Manager + RoboGmbH IT Manager", "Regional EMEA/APAC/Americas standups; P1/P2 hotline", "No"),
    (138, 3, "Performance Monitoring & Optimization", "90 days", date_str(date_plus_days(GOLIVE_DATE, 1)), date_str(date_plus_days(GOLIVE_DATE, 90)), "135", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "System performance tuning; infrastructure adjustments", "No"),
    (139, 3, "Data Integrity & Reconciliation (Bosch to Zebra)", "20 days", date_str(date_plus_days(GOLIVE_DATE, 1)), date_str(date_plus_days(GOLIVE_DATE, 21)), "135", "KPMG Data Engineer + RoboGmbH ERP Specialist", "Verify all data migrated correctly; fix reconciliation items", "No"),
    (140, 3, "Issue Backlog & Prioritized Fix List", "60 days", date_str(date_plus_days(GOLIVE_DATE, 10)), date_str(date_plus_days(GOLIVE_DATE, 70)), "137", "KPMG Test Specialist + RoboGmbH IT Manager", "Capture, prioritize, deploy hotfixes; spans hypercare period", "No"),
    (141, 3, "Wave 2 Application Go-Live (80 standard apps)", "5 days", date_str(date_plus_days(GOLIVE_DATE, 20)), date_str(date_plus_days(GOLIVE_DATE, 25)), "137", "KPMG Infrastructure Architect", "Phased app activation after Wave 1 stabilization", "No"),
    (142, 3, "Wave 3 Application Go-Live (68 local/OT apps)", "3 days", date_str(date_plus_days(GOLIVE_DATE, 50)), date_str(date_plus_days(GOLIVE_DATE, 53)), "141", "KPMG Infrastructure Architect", "Final app activation; OT system go-live within hypercare window", "No"),
    (143, 3, "Hypercare Close-Out Sign-Off", "0 days", date_str(date_plus_days(GOLIVE_DATE, 90)), date_str(date_plus_days(GOLIVE_DATE, 90)), "137,138,139,140,142", "KPMG PMO Lead + Buyer Sponsor", "Hypercare period officially closed (90 days complete); transition to steady state", "Yes"),
    
    # 4.3 TSA Exit (30 days — starts after 90-day hypercare closes)
    (144, 2, "4.3 TSA Transition & Service Exit", "30 days", date_str(date_plus_days(GOLIVE_DATE, 91)), date_str(date_plus_days(GOLIVE_DATE, 121)), "143", "", "TSA service wave exit; handover to Zebra IT (after hypercare closes)", "No"),
    (145, 3, "TSA Exit Planning per Service Line", "10 days", date_str(date_plus_days(GOLIVE_DATE, 91)), date_str(date_plus_days(GOLIVE_DATE, 101)), "143", "KPMG Project Manager + RoboGmbH IT Manager", "Define exit criteria; acceptance tests per service", "No"),
    (146, 3, "TSA Infrastructure Services Exit (Wave 1)", "15 days", date_str(date_plus_days(GOLIVE_DATE, 101)), date_str(date_plus_days(GOLIVE_DATE, 116)), "145", "KPMG Infrastructure Architect + RoboGmbH IT Manager", "Network, data centre, security handoff", "No"),
    (147, 3, "TSA Application & ERP Services Exit (Wave 2)", "10 days", date_str(date_plus_days(GOLIVE_DATE, 110)), date_str(date_plus_days(GOLIVE_DATE, 120)), "146", "KPMG SAP Specialist + RoboGmbH IT Manager", "SAP, apps, integration handoff", "No"),
    (148, 3, "TSA Exit Confirmation (Bosch services fully terminated)", "0 days", date_str(date_plus_days(GOLIVE_DATE, 121)), date_str(date_plus_days(GOLIVE_DATE, 121)), "147", "KPMG PMO Lead + Buyer Sponsor", "All TSA services exited; Zebra IT fully independent", "Yes"),
    
    # 4.4 Programme Closure (15 days — after hypercare & TSA exit complete)
    (149, 2, "4.4 Programme Closure", "15 days", date_str(date_plus_days(GOLIVE_DATE, 91)), date_str(date_plus_days(GOLIVE_DATE, 105)), "143,148", "", "Lessons learned, archive, final sign-off", "No"),
    (150, 3, "Lessons Learned Workshop (All regions)", "5 days", date_str(date_plus_days(GOLIVE_DATE, 91)), date_str(date_plus_days(GOLIVE_DATE, 96)), "143", "KPMG PMO Lead + All Workstream Leads", "Debrief; best practices; challenges documented", "No"),
    (151, 3, "Programme Closure Report & Knowledge Transfer", "8 days", date_str(date_plus_days(GOLIVE_DATE, 97)), date_str(date_plus_days(GOLIVE_DATE, 104)), "150", "KPMG PMO Lead + KPMG Project Manager", "Final report; operational runbooks; team debrief", "No"),
    (152, 3, "Financial Reconciliation & Archive", "10 days", date_str(date_plus_days(GOLIVE_DATE, 91)), date_str(date_plus_days(GOLIVE_DATE, 101)), "143", "KPMG PMO Lead", "Finance closeout; CAPEX/labour final reconciliation", "No"),
    (153, 3, "QG5 - Programme Closure Gate", "0 days", date_str(date_plus_days(GOLIVE_DATE, 105)), date_str(date_plus_days(GOLIVE_DATE, 105)), "150,151,152,148", "KPMG PMO Lead + Buyer Sponsor", "Hypercare closed (90 days); TSA fully exited; Zebra IT independent; programme formally closed", "Yes"),
    
    # Final milestone
    (154, 2, "Project Completion", "0 days", date_str(date_plus_days(GOLIVE_DATE, 105)), date_str(date_plus_days(GOLIVE_DATE, 105)), "153", "", "All artifacts archived; final sign-off", "Yes"),
]

# ============================================================================
# WRITE TEMP CSV FOR MS PROJECT XML GENERATION
# ============================================================================

def _write_temp_csv(temp_csv_path: Path):
    """Write task list as CSV for generate_msp_xml.py."""
    with open(temp_csv_path, "w", encoding="utf-8", newline="") as f:
        f.write("ID,Outline Level,Name,Duration,Start,Finish,Predecessors,Resource Names,Notes,Milestone\n")
        for task_id, outline, name, duration, start, finish, preds, resources, notes, milestone in TASKS:
            # Escape quotes and newlines
            name = name.replace('"', '""')
            notes = notes.replace('"', '""')
            resources = resources.replace('"', '""')
            f.write(f'{task_id},"{outline}","{name}","{duration}","{start}","{finish}","{preds}","{resources}","{notes}","{milestone}"\n')

# ============================================================================
# GENERATE XLSX
# ============================================================================

def _generate_excel():
    """Generate Zebra schedule XLSX with Bosch blue formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Header row
    headers = ["ID", "Outline Level", "Name", "Duration (days)", "Start", "Finish", "Predecessors", "Resource Names", "Notes", "Milestone"]
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="003B6E", end_color="003B6E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Add data rows
    # Pre-parse ISO dates back to datetime so openpyxl stores them as real date cells
    from datetime import datetime as _dt
    def _to_date(v):
        try:
            return _dt.strptime(v, "%Y-%m-%d")
        except Exception:
            return v

    for idx, (task_id, outline, name, duration, start, finish, preds, resources, notes, milestone) in enumerate(TASKS, start=2):
        ws.append([task_id, outline, name, duration, _to_date(start), _to_date(finish), preds, resources, notes, milestone])
        
        row = ws[idx]
        
        # Determine fill and font based on outline level and milestone
        if milestone == "Yes":
            # Amber background, black bold font
            cell_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell_font = Font(bold=True, color="000000", size=10)
        elif outline == "1":
            # Dark Bosch blue, white font
            cell_fill = PatternFill(start_color="003B6E", end_color="003B6E", fill_type="solid")
            cell_font = Font(bold=True, color="FFFFFF", size=10)
        elif outline == "2":
            # Mid blue, white font
            cell_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            cell_font = Font(bold=True, color="FFFFFF", size=10)
        else:
            # Detail row: alternating light blue / white
            if (idx - 2) % 2 == 0:
                cell_fill = PatternFill(start_color="EFF4FB", end_color="EFF4FB", fill_type="solid")
            else:
                cell_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell_font = Font(color="000000", size=9)
        
        for cell in row:
            cell.fill = cell_fill
            cell.font = cell_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    # Apply date format to Start and Finish columns
    from openpyxl.styles import numbers
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 50
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["J"].width = 10
    
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Save
    wb.save(str(XLSX_PATH))
    print(f"  XLSX written to {XLSX_PATH}")

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    print(f"[{PROJECT_NAME}] Generating comprehensive schedule with detailed workstreams...")
    
    # Generate XLSX
    _generate_excel()
    
    # Write CSV (kept permanently in output folder)
    _write_temp_csv(CSV_PATH)
    print(f"  CSV written to {CSV_PATH}")

    # Generate XML from CSV
    try:
        result = subprocess.run(
            [sys.executable, str(HERE / "generate_msp_xml.py"),
             "--csv", str(CSV_PATH),
             "--out", str(XML_PATH),
             "--project", PROJECT_NAME],
            capture_output=True,
            text=True,
            timeout=30
        )
        if result.returncode == 0:
            print(f"  XML written to {XML_PATH}")
        else:
            print(f"  XML generation failed: {result.stderr}")
            sys.exit(1)
    except subprocess.TimeoutExpired:
        print(f"  XML generation timed out")
        sys.exit(1)
    
    print(f"✓ {PROJECT_NAME} schedule generated successfully!")
    print(f"  XLSX: {XLSX_PATH}")
    print(f"  CSV:  {CSV_PATH}")
    print(f"  XML:  {XML_PATH}")
    print(f"\nSchedule Summary:")
    print(f"  Total Tasks: {len(TASKS)}")
    print(f"  Phases: 5 (Phase 0-4)")
    print(f"  Workstreams: 30+")
    print(f"  Quality Gates: QG0, QG1, QG2/3, QG4, QG5")
    print(f"  Timeline: {date_str(START_DATE)} → {date_str(COMPLETION_DATE)}")
