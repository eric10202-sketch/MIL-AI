#!/usr/bin/env python3

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas


BASE_DIR = Path(__file__).parent
PROJECT_NAME = "Trinity-CAM (GPT)"
OUTPUT_FOLDER_NAME = "Trinity-CAM (GPT) v1.1"
SELLER = "Johnson Controls International (JCI)"
BUYER = "Robert Bosch GmbH"
BUSINESS = "Air conditioning business"
MODEL = "Integration"
PMO = "KPMG"
DELIVERY_PARTNER = "Infosys"
SITE_COUNT = 48
USER_COUNT = 12000
APPLICATION_COUNT = 1800
PROJECT_DIR = BASE_DIR / "active-projects" / OUTPUT_FOLDER_NAME
SCHEDULE_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Risk_Register.xlsx"
COST_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Cost_Plan.xlsx"
LOGO_PATH = BASE_DIR / "Bosch.png"

TODAY = dt.date.today()
FILE_MONTH = TODAY.strftime("%b_%Y")
OUTPUT_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Monthly_Status_Report_{FILE_MONTH}.pdf"

IMPACT_SCORES = {"Very Low": 1, "Low": 2, "Moderate": 3, "High": 4, "Very High": 5}
PROBABILITY_SCORES = {"10%": 1, "30%": 2, "50%": 3, "70%": 4, "90%": 5}

C_NAVY = colors.HexColor("#003b6e")
C_BLUE = colors.HexColor("#0066CC")
C_GREEN = colors.HexColor("#007A33")
C_AMBER = colors.HexColor("#E8A000")
C_RED = colors.HexColor("#CC0000")
C_BG = colors.HexColor("#f4f6f9")
C_CARD = colors.white
C_LINE = colors.HexColor("#d8dde8")
C_TEXT = colors.HexColor("#1a1a1a")
C_MUTED = colors.HexColor("#5a6478")
C_LT = colors.HexColor("#EFF4FB")


def as_date(value: dt.datetime | dt.date) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    return value


def fmt_date(value: dt.datetime | dt.date) -> str:
    return as_date(value).strftime("%d %b %Y")


def fmt_money(value: int | float) -> str:
    return f"EUR {value:,.0f}"


def days_to(value: dt.datetime | dt.date) -> int:
    return (as_date(value) - TODAY).days


def read_schedule() -> dict[str, object]:
    ws = load_workbook(SCHEDULE_PATH, data_only=True)["Schedule"]
    phases = []
    milestones = {}
    upcoming_actions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        task_id, outline_level, name, duration, start, finish, predecessors, resources, notes, milestone = row
        if not name:
            continue
        item = {
            "name": str(name).strip(),
            "outline_level": int(outline_level),
            "start": start,
            "finish": finish,
            "notes": str(notes or ""),
        }
        if int(outline_level) == 1:
            phases.append(item)
        if str(milestone).strip().lower() == "yes":
            milestones[item["name"]] = item
        start_date = as_date(start)
        if int(outline_level) == 3 and TODAY <= start_date <= TODAY + dt.timedelta(days=90):
            upcoming_actions.append(item)
    return {"phases": phases, "milestones": milestones, "actions": upcoming_actions[:4]}


def read_risks() -> list[dict[str, object]]:
    ws = load_workbook(RISK_PATH, data_only=True)["Risk Register"]
    risks = []
    for row_num in range(5, ws.max_row + 1):
        risk_id = ws[f"B{row_num}"].value
        if not risk_id or not re.fullmatch(r"R\d{3}", str(risk_id).strip()):
            continue
        if str(ws[f"P{row_num}"].value or "").strip().lower() == "opportunity":
            continue
        impact = str(ws[f"L{row_num}"].value or "")
        probability = str(ws[f"N{row_num}"].value or "")
        score = IMPACT_SCORES.get(impact, 0) * PROBABILITY_SCORES.get(probability, 0)
        risks.append(
            {
                "id": str(risk_id).strip(),
                "event": str(ws[f"F{row_num}"].value or ""),
                "owner": str(ws[f"I{row_num}"].value or ""),
                "score": score,
            }
        )
    risks.sort(key=lambda item: (-int(item["score"]), item["id"]))
    return risks[:5]


def read_costs() -> dict[str, object]:
    ws = load_workbook(COST_PATH, data_only=True)["Cost Plan"]
    total = 0.0
    baseline = ""
    phase_zero = 0.0
    for row in range(1, ws.max_row + 1):
        col_a = ws.cell(row, 1).value
        col_b = ws.cell(row, 2).value
        col_f = ws.cell(row, 6).value
        if col_a == "Budget Baseline":
            baseline = str(col_b or "")
        elif col_a == "OVERALL PROJECT TOTAL":
            total = float(col_f or 0)
        elif col_a == "Phase 0: Mobilize carve-out governance and TSA baseline":
            phase_zero = float(col_f or 0)
    return {"total": total, "baseline": baseline, "phase_zero": phase_zero}


def wrap_text(text: str, limit: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current = ""
    for word in words:
        proposal = word if not current else f"{current} {word}"
        if len(proposal) <= limit:
            current = proposal
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines or [""]


def draw_pill(c: canvas.Canvas, x: float, y: float, text: str, fill: colors.Color, txt: colors.Color) -> None:
    width = max(34, 6 + len(text) * 4.4)
    c.setFillColor(fill)
    c.roundRect(x, y - 10, width, 12, 6, fill=1, stroke=0)
    c.setFillColor(txt)
    c.setFont("Helvetica-Bold", 6.5)
    c.drawCentredString(x + width / 2, y - 7, text)


def build_report() -> None:
    schedule = read_schedule()
    risks = read_risks()
    costs = read_costs()

    qg0 = schedule["milestones"]["QG0 - Programme kickoff approved"]["start"]
    qg1 = schedule["milestones"]["QG1 - Concept and transition model approved"]["start"]
    qg23 = schedule["milestones"]["QG2 and QG3 - Build complete and test entry approved"]["start"]
    qg4 = schedule["milestones"]["QG4 - Pre-GoLive gate approved"]["start"]
    golive = schedule["milestones"]["GoLive - Day 1 cutover to merger zone complete"]["start"]
    qg5 = schedule["milestones"]["QG5 - Project completion approved"]["start"]

    project_dir = OUTPUT_PATH.parent
    project_dir.mkdir(parents=True, exist_ok=True)

    width, height = A4
    left = 24
    right = width - 24
    content_width = right - left
    c = canvas.Canvas(str(OUTPUT_PATH), pagesize=A4)

    c.setFillColor(C_BG)
    c.rect(0, 0, width, height, fill=1, stroke=0)

    y = height - 26
    header_h = 62
    c.setFillColor(C_NAVY)
    c.roundRect(left, y - header_h, content_width, header_h, 8, fill=1, stroke=0)
    if LOGO_PATH.exists():
        c.setFillColor(colors.white)
        c.roundRect(left + 10, y - 49, 74, 30, 4, fill=1, stroke=0)
        c.drawImage(ImageReader(str(LOGO_PATH)), left + 16, y - 46, width=62, height=24, preserveAspectRatio=True, mask="auto")
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(left + 94, y - 24, f"{PROJECT_NAME} Monthly Status Report")
    c.setFont("Helvetica", 9)
    c.drawString(left + 94, y - 38, f"{SELLER} -> {DELIVERY_PARTNER} merger zone -> {BUYER} | {MODEL} model")
    c.drawRightString(right - 12, y - 20, TODAY.strftime("%B %Y"))
    c.drawRightString(right - 12, y - 34, f"Report date: {TODAY.strftime('%d %b %Y')}")
    y -= header_h + 8

    c.setFillColor(C_AMBER)
    c.roundRect(left, y - 24, content_width, 24, 4, fill=1, stroke=0)
    c.setFillColor(C_TEXT)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(left + 10, y - 10, "Pre-initiation reporting view")
    c.setFont("Helvetica", 8)
    c.drawString(left + 126, y - 10, "Formal programme start remains QG0 on 01 Jul 2026; SPI/CPI are held at baseline until kick-off.")
    y -= 32

    rag_labels = [
        ("Overall", "AMBER", C_AMBER),
        ("SPI", "1.00", C_AMBER),
        ("CPI", "1.00", C_AMBER),
        ("Readiness", "0%", C_BLUE),
        ("Top Risk", f"{risks[0]['id']} / {risks[0]['score']}" if risks else "None", C_RED),
        ("Days to GoLive", str(days_to(golive)), C_BLUE),
    ]
    box_w = content_width / len(rag_labels)
    for idx, (label, value, tone) in enumerate(rag_labels):
        x = left + idx * box_w
        c.setFillColor(colors.white)
        c.roundRect(x, y - 34, box_w - 4, 34, 4, fill=1, stroke=0)
        c.setStrokeColor(C_LINE)
        c.setLineWidth(0.5)
        c.roundRect(x, y - 34, box_w - 4, 34, 4, fill=0, stroke=1)
        c.setFillColor(C_MUTED)
        c.setFont("Helvetica", 7)
        c.drawString(x + 6, y - 11, label.upper())
        c.setFillColor(tone)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(x + 6, y - 25, value)
    y -= 44

    def section(title: str) -> None:
        nonlocal y
        c.setFillColor(C_BLUE)
        c.roundRect(left, y - 15, content_width, 15, 4, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(left + 8, y - 10, title.upper())
        y -= 18

    section("1. Executive Summary")
    c.setFillColor(C_CARD)
    c.roundRect(left, y - 74, content_width, 74, 4, fill=1, stroke=0)
    c.setStrokeColor(C_LINE)
    c.roundRect(left, y - 74, content_width, 74, 4, fill=0, stroke=1)
    summary_lines = [
        f"{PROJECT_NAME} remains in pre-kickoff mobilisation as of {TODAY.strftime('%B %Y')}. Change Request 1 confirms an approved JCI TSA extension through 31 Jul 2027 so users can continue to work in the legacy environment while Infosys completes merger-zone build-up.",
        f"Overall status is AMBER because the critical SAP carve-out risk ({risks[0]['id'] if risks else 'n/a'}) remains open while merger-zone design, application inventory, and legal/data controls are still driving toward QG1 on {fmt_date(qg1)}.",
        f"No schedule or cost variance is recorded yet. External labour baseline is {fmt_money(costs['total'])} with budget approval still marked as '{costs['baseline']}'.",
    ]
    text_y = y - 12
    c.setFillColor(C_TEXT)
    c.setFont("Helvetica", 8)
    for paragraph in summary_lines:
        for line in wrap_text(paragraph, 118):
            c.drawString(left + 10, text_y, line)
            text_y -= 10
        text_y -= 2
    y -= 82

    section("2. Key Facts and Gate Outlook")
    c.setFillColor(C_CARD)
    c.roundRect(left, y - 56, content_width, 56, 4, fill=1, stroke=0)
    c.setStrokeColor(C_LINE)
    c.roundRect(left, y - 56, content_width, 56, 4, fill=0, stroke=1)
    facts = [
        ("Sites / users", f"{SITE_COUNT} sites | {USER_COUNT:,} users"),
        ("Applications", f"{APPLICATION_COUNT:,}+ incl. major SAP"),
        ("PMO / delivery", f"{PMO} | {DELIVERY_PARTNER}"),
        ("QG1", f"{fmt_date(qg1)} ({days_to(qg1)} days)"),
        ("GoLive", f"{fmt_date(golive)} ({days_to(golive)} days)"),
        ("Completion", f"{fmt_date(qg5)} ({days_to(qg5)} days)"),
    ]
    fact_w = content_width / 3
    for idx, (label, value) in enumerate(facts):
        row = idx // 3
        col = idx % 3
        x = left + col * fact_w
        box_y = y - 8 - row * 24
        c.setFillColor(C_LT)
        c.roundRect(x + 6, box_y - 18, fact_w - 12, 18, 3, fill=1, stroke=0)
        c.setFillColor(C_MUTED)
        c.setFont("Helvetica", 6.5)
        c.drawString(x + 12, box_y - 6, label.upper())
        c.setFillColor(C_NAVY)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(x + 12, box_y - 15, value)
    y -= 64

    section("3. Phase, Risk, and Budget Status")
    c.setFillColor(C_CARD)
    c.roundRect(left, y - 118, content_width, 118, 4, fill=1, stroke=0)
    c.setStrokeColor(C_LINE)
    c.roundRect(left, y - 118, content_width, 118, 4, fill=0, stroke=1)

    c.setFont("Helvetica-Bold", 7.5)
    c.setFillColor(C_NAVY)
    c.drawString(left + 10, y - 12, "PHASE")
    c.drawString(left + 170, y - 12, "STATUS")
    c.drawString(left + 230, y - 12, "COMMENT")
    row_y = y - 24
    phase_comments = [
        ("Phase 0", "AMBER", "Mobilisation underway; governance, partner onboarding, and approved TSA buffer drive QG0/QG1 readiness without changing the overall programme timeline."),
        ("Top risks", "AMBER", f"{risks[0]['id'] if risks else 'R001'} and {risks[1]['id'] if len(risks) > 1 else 'R003'} remain the primary threats to SAP and final readiness."),
        ("Budget", "BLUE", f"External labour baseline {fmt_money(costs['total'])}; Phase 0 share {fmt_money(costs['phase_zero'])}; approval remains {costs['baseline'].lower()}."),
        ("Gate path", "BLUE", f"QG1 {fmt_date(qg1)}, QG2&3 {fmt_date(qg23)}, QG4 {fmt_date(qg4)}, GoLive {fmt_date(golive)}."),
    ]
    for idx, (label, status, comment) in enumerate(phase_comments):
        if idx % 2 == 0:
            c.setFillColor(C_LT)
            c.rect(left + 1, row_y - 10, content_width - 2, 18, fill=1, stroke=0)
        c.setFillColor(C_TEXT)
        c.setFont("Helvetica-Bold", 7.5)
        c.drawString(left + 10, row_y, label)
        if status == "AMBER":
            draw_pill(c, left + 170, row_y + 5, status, colors.HexColor("#fef9e7"), colors.HexColor("#b7770d"))
        else:
            draw_pill(c, left + 170, row_y + 5, status, colors.HexColor("#EFF4FB"), C_BLUE)
        c.setFillColor(C_TEXT)
        c.setFont("Helvetica", 7.2)
        for line in wrap_text(comment, 70)[:2]:
            c.drawString(left + 230, row_y, line)
            row_y -= 8
        row_y -= 10
    y -= 126

    section("4. Top Risks and Next 90 Days")
    c.setFillColor(C_CARD)
    c.roundRect(left, y - 120, content_width, 120, 4, fill=1, stroke=0)
    c.setStrokeColor(C_LINE)
    c.roundRect(left, y - 120, content_width, 120, 4, fill=0, stroke=1)
    c.setFillColor(C_NAVY)
    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(left + 10, y - 12, "TOP RISKS")
    c.drawString(left + 290, y - 12, "NEXT 90 DAYS")
    risk_y = y - 24
    for risk in risks[:3]:
        badge_color = C_RED if risk["score"] >= 15 else C_AMBER
        c.setFillColor(badge_color)
        c.circle(left + 18, risk_y + 2, 8, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 6.5)
        c.drawCentredString(left + 18, risk_y, str(risk["score"]))
        c.setFillColor(C_TEXT)
        c.setFont("Helvetica-Bold", 7)
        c.drawString(left + 32, risk_y + 2, risk["id"])
        c.setFont("Helvetica", 6.8)
        for line in wrap_text(risk["event"], 44)[:2]:
            c.drawString(left + 54, risk_y + 2, line)
            risk_y -= 8
        c.setFillColor(C_MUTED)
        c.drawString(left + 54, risk_y + 1, f"Owner: {risk['owner']}")
        risk_y -= 12

    action_y = y - 24
    c.setFillColor(C_TEXT)
    for action in schedule["actions"] or []:
        c.setFillColor(C_LT)
        c.roundRect(left + 288, action_y - 9, 58, 16, 3, fill=1, stroke=0)
        c.setFillColor(C_BLUE)
        c.setFont("Helvetica-Bold", 6.8)
        c.drawCentredString(left + 317, action_y - 3, fmt_date(action["start"]))
        c.setFillColor(C_TEXT)
        c.setFont("Helvetica-Bold", 7)
        c.drawString(left + 356, action_y + 2, action["name"])
        c.setFont("Helvetica", 6.8)
        c.drawString(left + 356, action_y - 7, action["notes"][:72])
        action_y -= 24

    footer_text = f"{PROJECT_NAME} | {TODAY.strftime('%d %b %Y')} | Sources: {SCHEDULE_PATH.name}, {RISK_PATH.name}, {COST_PATH.name} | CONFIDENTIAL"
    c.setFillColor(C_NAVY)
    c.roundRect(left, 18, content_width, 14, 4, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica", 6.8)
    c.drawCentredString(left + content_width / 2, 22, footer_text)

    c.save()


def main() -> None:
    build_report()
    print(f"[{PROJECT_NAME}] Monthly Status Report: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()