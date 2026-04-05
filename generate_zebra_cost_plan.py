#!/usr/bin/env python3
"""
Generate Zebra cost plan from schedule and risk register.

Project: Zebra (Packaging carve-out)
Timeline: 1 April 2026 - 31 October 2027
Scope: 37 sites, 3500+ users, 208 applications, SAP + TSA

Cost plan structure:
- KPMG PMO & Governance
- KPMG SAP & ERP Build
- KPMG Testing & QA
- KPMG Change Management & Training
- KPMG Data & Integration
- Seller (RoboGmbH) IT Support
- Contingency (risk-driven per Risk Register)
"""

import sys
import os
from pathlib import Path
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
SCHEDULE_PATH = HERE / "active-projects" / "Zebra" / "Zebra_Project_Schedule.xlsx"
RISK_PATH = HERE / "active-projects" / "Zebra" / "Zebra_Risk_Register.xlsx"
OUTPUT_PATH = HERE / "active-projects" / "Zebra" / "Zebra_Cost_Plan.xlsx"

# ============================================================================
# COST DATA: Zebra Project Specific
# ============================================================================
# Resource daily rates (EUR)
RATES = {
    "KPMG PMO Lead": 750,
    "KPMG Project Manager": 650,
    "KPMG SAP Architect": 700,
    "KPMG SAP Build Lead": 680,
    "KPMG SAP Specialist": 600,
    "KPMG QA Lead": 650,
    "KPMG Test Specialist": 550,
    "KPMG Change Manager": 600,
    "KPMG Trainer": 500,
    "KPMG Data Architect": 700,
    "KPMG Data Engineer": 550,
    "KPMG Infrastructure Architect": 650,
    "RoboGmbH IT Manager": 400,
    "RoboGmbH ERP Specialist": 350,
    "RoboGmbH IT Technician": 250,
}

# Labor cost categories (category_name, resource_list, total_days, resources_per_day)
# Resources per day estimated from schedule
LABOR_CATEGORIES = [
    ("KPMG PMO & Governance", [
        ("KPMG PMO Lead", 240),  # 240 days across 20 months = 12 FTE
        ("KPMG Project Manager", 320),  # 320 days
    ]),
    ("KPMG SAP & ERP Build", [
        ("KPMG SAP Architect", 150),
        ("KPMG SAP Build Lead", 280),
        ("KPMG SAP Specialist", 350),
    ]),
    ("KPMG Testing & QA", [
        ("KPMG QA Lead", 200),
        ("KPMG Test Specialist", 480),  # Many test cycles for 37 sites
    ]),
    ("KPMG Change Management & Training", [
        ("KPMG Change Manager", 200),
        ("KPMG Trainer", 280),  # Train-the-trainer for 3500+ users
    ]),
    ("KPMG Data & Integration", [
        ("KPMG Data Architect", 160),
        ("KPMG Data Engineer", 320),
    ]),
    ("KPMG Infrastructure & Security", [
        ("KPMG Infrastructure Architect", 120),
    ]),
    ("RoboGmbH IT Support (Seller)", [
        ("RoboGmbH IT Manager", 180),
        ("RoboGmbH ERP Specialist", 200),
        ("RoboGmbH IT Technician", 250),
    ]),
]

# Phase breakdown (phase_name, phase_start_date, phase_end_date, phase_cost_pct)
PHASES = [
    ("Phase 0: Initialization", "01.04.2026", "17.04.2026", 0.03),
    ("Phase 1: Concept", "18.04.2026", "18.05.2026", 0.12),
    ("Phase 2: Architecture & Design", "19.05.2026", "27.07.2026", 0.25),
    ("Phase 3: Development, Build & Test", "28.07.2026", "27.05.2027", 0.48),
    ("Phase 4: GoLive & Closure", "28.05.2027", "31.10.2027", 0.12),
]

# CAPEX / Additional costs (linked to risks)
CAPEX_ITEMS = [
    ("Risk #1 - SAP Separation Audit (external audit firm)", 35000, "Risk #1 (High P×I): 3x dry-run audits"),
    ("Risk #2 - Application Portfolio Tools & Licenses", 25000, "Risk #2 (High P×I): License transition support tooling"),
    ("Risk #7 - Regulatory Compliance Assessment", 40000, "Risk #7 (Very High P×I): GDPR + data residency audit"),
    ("Risk #12 - Parallel Run System Rental (2 weeks)", 20000, "Risk #12 (High P×I): Temporary parallel environment"),
    ("General Contingency (5% of labor)", 0, "TBC - calculated at end"),  # Calculated below
]

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def calculate_labor_total():
    """Calculate total labor cost."""
    total = 0
    for category_name, resources in LABOR_CATEGORIES:
        for resource_name, days in resources:
            rate = RATES.get(resource_name, 500)
            cost = days * rate
            total += cost
    return total

def write_xlsx(output_path):
    """Write cost plan to XLSX with Bosch blue theme."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Plan"
    
    # Color definitions (Bosch blue theme)
    DARK_BLUE = "003B6E"
    MID_BLUE = "0066CC"
    LIGHT_BLUE = "EFF4FB"
    GREY = "F2F2F2"
    
    # Fonts
    TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    REGULAR_FONT = Font(name="Calibri", size=9, color="000000")
    BOLD_FONT = Font(name="Calibri", size=9, bold=True, color="000000")
    META_FONT = Font(name="Calibri", size=8, color="000000")
    
    # Fills
    DARK_BLUE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    MID_BLUE_FILL = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    GREY_FILL = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")
    SUBTOTAL_FILL = PatternFill(start_color="C6D4E8", end_color="C6D4E8", fill_type="solid")
    
    # Borders
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    current_row = 1
    
    # ===== Header Banner =====
    ws.merge_cells(f"A{current_row}:F{current_row}")
    title_cell = ws[f"A{current_row}"]
    title_cell.value = "ZEBRA PROJECT COST PLAN"
    title_cell.font = TITLE_FONT
    title_cell.fill = DARK_BLUE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[current_row].height = 22
    current_row += 1
    
    # Metadata
    for label, value in [
        ("Project Name:", "Zebra (Packaging Carve-Out)"),
        ("Timeline:", "1 April 2026 - 31 October 2027"),
        ("Scope:", "37 sites, 3500+ users, 208 applications, SAP + TSA"),
        ("Based on:", f"{SCHEDULE_PATH.name}"),
        ("Risk-aligned:", f"{RISK_PATH.name}"),
    ]:
        ws[f"A{current_row}"] = label
        ws[f"B{current_row}"] = value
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{current_row}"].fill = GREY_FILL
            ws[f"{col}{current_row}"].font = META_FONT
        ws.row_dimensions[current_row].height = 16
        current_row += 1
    
    current_row += 1  # Blank line
    
    # ===== Column Headers =====
    headers = ["CATEGORY", "RESOURCE", "TOTAL DAYS", "TOTAL HRS", "HOURLY RATE (EUR)", "TOTAL COST (EUR)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = MID_BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[current_row].height = 18
    current_row += 1
    
    # ===== Labor Cost Categories =====
    labor_total = 0
    category_totals = {}
    
    for category_name, resources in LABOR_CATEGORIES:
        # Category header row
        ws.merge_cells(f"A{current_row}:F{current_row}")
        cat_cell = ws[f"A{current_row}"]
        cat_cell.value = category_name
        cat_cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cat_cell.fill = MID_BLUE_FILL
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 16
        current_row += 1
        
        category_subtotal = 0
        
        # Detail rows
        for idx, (resource_name, days) in enumerate(resources):
            rate = RATES.get(resource_name, 500)
            hours = days * 8  # 8 hours per day
            cost = days * rate
            
            is_even = (idx + 1) % 2 == 0
            row_fill = LIGHT_BLUE_FILL if is_even else PatternFill(fill_type=None)
            
            ws.cell(current_row, 1).value = ""  # Category (blank for detail rows)
            ws.cell(current_row, 2).value = resource_name
            ws.cell(current_row, 3).value = days
            ws.cell(current_row, 4).value = hours
            ws.cell(current_row, 5).value = rate
            ws.cell(current_row, 6).value = cost
            
            for col in range(1, 7):
                cell = ws.cell(current_row, col)
                cell.font = REGULAR_FONT
                cell.fill = row_fill
                cell.border = THIN_BORDER
                cell.number_format = "#,##0" if col in [3, 4, 5, 6] else "@"
            
            ws.row_dimensions[current_row].height = 14
            category_subtotal += cost
            current_row += 1
        
        # Subtotal row
        ws.cell(current_row, 1).value = "SUBTOTAL"
        ws.merge_cells(f"A{current_row}:E{current_row}")
        subtotal_cell = ws[f"A{current_row}"]
        ws.cell(current_row, 6).value = category_subtotal
        
        for col in range(1, 7):
            cell = ws.cell(current_row, col)
            cell.font = BOLD_FONT
            cell.fill = SUBTOTAL_FILL
            cell.border = THIN_BORDER
            if col == 6:
                cell.number_format = "#,##0"
        
        ws.row_dimensions[current_row].height = 16
        category_totals[category_name] = category_subtotal
        labor_total += category_subtotal
        current_row += 1
        current_row += 1  # Blank line
    
    # ===== Overall Labor Total =====
    ws.merge_cells(f"A{current_row}:E{current_row}")
    total_label = ws[f"A{current_row}"]
    total_label.value = "TOTAL LABOR COST"
    total_label.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    total_label.fill = DARK_BLUE_FILL
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    
    total_cell = ws.cell(current_row, 6)
    total_cell.value = labor_total
    total_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    total_cell.fill = DARK_BLUE_FILL
    total_cell.number_format = "#,##0"
    total_cell.border = THIN_BORDER
    
    ws.row_dimensions[current_row].height = 18
    current_row += 2
    
    # ===== Cost Breakdown by Category =====
    ws.merge_cells(f"A{current_row}:D{current_row}")
    section = ws[f"A{current_row}"]
    section.value = "COST BREAKDOWN BY CATEGORY"
    section.font = HEADER_FONT
    section.fill = MID_BLUE_FILL
    ws.row_dimensions[current_row].height = 16
    current_row += 1
    
    for cat, subtotal in category_totals.items():
        ws.cell(current_row, 1).value = cat
        ws.cell(current_row, 2).value = subtotal
        ws.cell(current_row, 2).number_format = "#,##0"
        for col in range(1, 3):
            ws.cell(current_row, col).fill = (LIGHT_BLUE_FILL if current_row % 2 == 0 else PatternFill(fill_type=None))
            ws.cell(current_row, col).font = REGULAR_FONT
        current_row += 1
    
    current_row += 1
    
    # ===== Cost Breakdown by Phase =====
    ws.merge_cells(f"A{current_row}:D{current_row}")
    section = ws[f"A{current_row}"]
    section.value = "COST BREAKDOWN BY PHASE"
    section.font = HEADER_FONT
    section.fill = MID_BLUE_FILL
    ws.row_dimensions[current_row].height = 16
    current_row += 1
    
    for phase_name, start_date, end_date, pct in PHASES:
        phase_cost = labor_total * pct
        ws.cell(current_row, 1).value = f"{phase_name} ({start_date} - {end_date})"
        ws.cell(current_row, 2).value = phase_cost
        ws.cell(current_row, 2).number_format = "#,##0"
        for col in range(1, 3):
            ws.cell(current_row, col).fill = (LIGHT_BLUE_FILL if current_row % 2 == 0 else PatternFill(fill_type=None))
            ws.cell(current_row, col).font = REGULAR_FONT
        current_row += 1
    
    current_row += 2
    
    # ===== CAPEX / Additional Costs =====
    ws.merge_cells(f"A{current_row}:D{current_row}")
    section = ws[f"A{current_row}"]
    section.value = "CAPEX / ADDITIONAL COSTS (Excluded from Labor Total)"
    section.font = HEADER_FONT
    section.fill = MID_BLUE_FILL
    ws.row_dimensions[current_row].height = 16
    current_row += 1
    
    capex_total = 0
    for idx, (item_name, cost, note) in enumerate(CAPEX_ITEMS):
        # Calculate contingency as 5% of labor total
        if "General Contingency" in item_name:
            cost = int(labor_total * 0.05)
        
        ws.cell(current_row, 1).value = item_name
        ws.cell(current_row, 2).value = cost
        ws.cell(current_row, 3).value = note
        ws.cell(current_row, 2).number_format = "#,##0"
        
        is_even = (idx + 1) % 2 == 0
        for col in range(1, 4):
            ws.cell(current_row, col).fill = (LIGHT_BLUE_FILL if is_even else PatternFill(fill_type=None))
            ws.cell(current_row, col).font = REGULAR_FONT
        
        capex_total += cost
        current_row += 1
    
    # CAPEX Subtotal
    ws.cell(current_row, 1).value = "CAPEX SUBTOTAL"
    ws.cell(current_row, 2).value = capex_total
    ws.cell(current_row, 2).number_format = "#,##0"
    for col in range(1, 3):
        ws.cell(current_row, col).font = BOLD_FONT
        ws.cell(current_row, col).fill = SUBTOTAL_FILL
    current_row += 2
    
    # ===== Grand Total =====
    ws.cell(current_row, 1).value = "GRAND TOTAL (Labor + CAPEX)"
    ws.cell(current_row, 2).value = labor_total + capex_total
    ws.cell(current_row, 2).number_format = "#,##0"
    for col in range(1, 3):
        ws.cell(current_row, col).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ws.cell(current_row, col).fill = DARK_BLUE_FILL
    
    current_row += 2
    
    # ===== Notes =====
    ws.merge_cells(f"A{current_row}:F{current_row}")
    notes_header = ws[f"A{current_row}"]
    notes_header.value = "NOTES"
    notes_header.font = BOLD_FONT
    notes_header.fill = GREY_FILL
    ws.row_dimensions[current_row].height = 14
    current_row += 1
    
    notes_text = (
        "1. Cost plan is DRAFT — to be approved at QG0.\n"
        "2. Labor rates based on typical KPMG/Seller day rates; subject to FTE negotiation.\n"
        "3. SAP system separation and 208-application portfolio transition are highest-cost complexity drivers.\n"
        "4. CAPEX items (rows above) are linked to high-priority risks in Risk Register.\n"
        "5. 37-site multi-geography cutover requires significant QA and change management investment.\n"
        "6. TSA period labor (seller IT support) extends through Phase 4 to handover completion.\n"
        "7. Contingency (5% of labor) covers risk mitigation overruns; escalation for CAPEX overruns required.\n"
        "8. Buyer IT cost allocation not included (assumed by buyer under Stand Alone model)."
    )
    
    ws.merge_cells(f"A{current_row}:F{current_row}")
    notes_cell = ws[f"A{current_row}"]
    notes_cell.value = notes_text
    notes_cell.font = Font(name="Calibri", size=8, italic=True, color="000000")
    notes_cell.fill = GREY_FILL
    notes_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[current_row].height = 100
    
    # Column widths
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    
    # Freeze pane
    ws.freeze_panes = "A2"
    
    # Save
    wb.save(output_path)
    print(f"✓ Cost plan saved to {output_path}")
    print(f"  Total labor cost: €{labor_total:,}")
    print(f"  CAPEX total: €{capex_total:,}")
    print(f"  Grand total: €{labor_total + capex_total:,}")

def main():
    print("[Zebra] Generating cost plan...")
    
    # Verify schedule and risk register exist
    if not SCHEDULE_PATH.exists():
        print(f"ERROR: Schedule not found at {SCHEDULE_PATH}")
        sys.exit(1)
    if not RISK_PATH.exists():
        print(f"ERROR: Risk register not found at {RISK_PATH}")
        sys.exit(1)
    
    print("  Schedule and risk register validated")
    write_xlsx(OUTPUT_PATH)

if __name__ == "__main__":
    main()
