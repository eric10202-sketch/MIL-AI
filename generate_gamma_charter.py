#!/usr/bin/env python3
"""Generate the Gamma project charter as a Bosch-branded HTML file."""

from __future__ import annotations

import base64
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


HERE = Path(__file__).parent
PROJECT_NAME = "Gamma"
OUTPUT_FOLDER_NAME = "Gamma v1.0"
DOCUMENT_VERSION = "Version 1.0 - Initial Baseline"
SELLER = "Robert Bosch China"
BUYER = "Alibaba"
BUSINESS = "Bosch Cloud business"
CARVEOUT_MODEL = "Combination"
PMO_LEAD = "EY"
SITES = 5
USERS = 250
APPLICATIONS = 20
START_DATE = "2026-08-01"
GOLIVE_DATE = "2027-02-01"
COMPLETION_DATE = "2027-05-31"

SCHEDULE_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Risk_Register.xlsx"
COST_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Cost_Plan.xlsx"
LOGO_PATH = HERE / "Bosch.png"
OUTPUT_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Project_Charter.html"

IMPACT_VALUES = {"Very Low": 1, "Low": 2, "Moderate": 3, "High": 4, "Very High": 5}
PROBABILITY_VALUES = {"10%": 1, "30%": 2, "50%": 3, "70%": 4, "90%": 5}


def format_date(value) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%d %b %Y")
    return str(value)


def load_schedule_summary() -> tuple[list[tuple[str, str, str]], list[tuple[str, str]]]:
    wb = load_workbook(SCHEDULE_PATH, data_only=False)
    ws = wb["Schedule"]
    phases = []
    milestones = []
    for row in range(2, ws.max_row + 1):
        level = ws.cell(row, 2).value
        if level is None:
            continue
        level = int(level)
        name = str(ws.cell(row, 3).value or "").strip()
        start = format_date(ws.cell(row, 5).value)
        if level == 1:
            finish = format_date(ws.cell(row, 6).value)
            phases.append((name, start, finish))
        if str(ws.cell(row, 10).value) == "Yes":
            if "QG" in name or "GoLive" in name or "Closure" in name:
                milestones.append((name, start))
    return phases, milestones


def load_risk_summary() -> tuple[int, int, list[dict]]:
    wb = load_workbook(RISK_PATH, data_only=False)
    ws = wb["Risk Register"]
    risks = []
    opportunities = 0
    for row in range(5, 140):
        risk_id = ws.cell(row, 2).value
        if risk_id is None:
            continue
        impact = ws.cell(row, 12).value
        probability = ws.cell(row, 14).value
        score = IMPACT_VALUES.get(impact, 0) * PROBABILITY_VALUES.get(probability, 0)
        risk_type = ws.cell(row, 16).value
        if risk_type == "opportunity":
            opportunities += 1
        risks.append(
            {
                "id": int(risk_id),
                "category": ws.cell(row, 4).value,
                "event": ws.cell(row, 6).value,
                "owner": ws.cell(row, 9).value,
                "impact": impact,
                "probability": probability,
                "score": score,
                "type": risk_type,
            }
        )
    risks.sort(key=lambda item: (item["score"], item["type"] != "threat"), reverse=True)
    return len(risks), opportunities, risks[:5]


def load_cost_summary() -> tuple[int, int]:
    wb = load_workbook(COST_PATH, data_only=False)
    ws = wb["Cost Plan"]
    labour_total = 0
    capex_total = 0
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label == "OVERALL PROJECT TOTAL - LABOUR ONLY":
            labour_total = int(ws.cell(row, 6).value or 0)
        if label == "TOTAL CAPEX / ADDITIONAL COSTS":
            capex_total = int(ws.cell(row, 6).value or 0)
    return labour_total, capex_total


def euro(amount: int) -> str:
    return f"EUR {amount:,.0f}"


def main() -> None:
    if not SCHEDULE_PATH.exists() or not RISK_PATH.exists() or not COST_PATH.exists():
        raise FileNotFoundError("Schedule, risk register, and cost plan must exist before generating the charter.")

    phases, milestones = load_schedule_summary()
    risk_count, opportunity_count, top_risks = load_risk_summary()
    labour_total, capex_total = load_cost_summary()
    logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode()
    report_date = datetime.now().strftime("%d %B %Y")

    milestone_rows = "\n".join(
        f"<tr><td>{name}</td><td>{date}</td></tr>" for name, date in milestones
    )
    phase_rows = "\n".join(
        f"<tr><td>{name}</td><td>{start}</td><td>{finish}</td></tr>" for name, start, finish in phases
    )
    risk_rows = "\n".join(
        f"<tr><td>#{risk['id']}</td><td>{risk['category']}</td><td>{risk['score']}</td><td>{risk['owner']}</td><td>{risk['event']}</td></tr>"
        for risk in top_risks
    )

    html = f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"UTF-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\" />
  <title>{PROJECT_NAME} Project Charter</title>
  <style>
    :root {{
      --bosch-blue: #003b6e;
      --bosch-mid: #0066cc;
      --bosch-light: #eff4fb;
      --bosch-ink: #102033;
      --bosch-muted: #5f6f86;
      --bosch-line: #d7deea;
      --bosch-bg: #f3f6fb;
      --card: #ffffff;
      --accent: #dbe8fb;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", Arial, sans-serif; background: linear-gradient(180deg, #eaf0f9 0%, #f7f9fc 100%); color: var(--bosch-ink); }}
    .page {{ max-width: 1120px; margin: 0 auto; padding: 28px 24px 48px; }}
    .cover {{ background: linear-gradient(140deg, #002147 0%, #003b6e 58%, #0f7dd8 100%); color: #fff; border-radius: 20px; padding: 34px 36px; box-shadow: 0 18px 48px rgba(0, 33, 71, 0.20); }}
    .bosch-logo {{ display:flex; align-items:center; background:#fff; padding:6px 10px; border-radius:6px; width:fit-content; margin-bottom:18px; }}
    .eyebrow {{ font-size: 11px; letter-spacing: 1.6px; text-transform: uppercase; opacity: 0.84; margin-bottom: 12px; }}
    h1 {{ margin: 0 0 8px; font-size: 36px; }}
    .subtitle {{ margin: 0; font-size: 16px; line-height: 1.5; max-width: 760px; opacity: 0.95; }}
    .cover-grid {{ display:grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-top: 26px; }}
    .cover-card {{ background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.14); border-radius: 14px; padding: 14px; }}
    .cover-card .label {{ font-size: 11px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.78; margin-bottom: 6px; }}
    .cover-card .value {{ font-size: 14px; font-weight: 700; line-height: 1.4; }}
    .section {{ background: var(--card); margin-top: 18px; padding: 24px; border-radius: 16px; box-shadow: 0 10px 30px rgba(16, 32, 51, 0.06); }}
    .section h2 {{ margin: 0 0 14px; font-size: 19px; color: var(--bosch-blue); }}
    .lead {{ font-size: 14px; line-height: 1.7; color: var(--bosch-ink); }}
    .grid {{ display:grid; grid-template-columns: repeat(2, 1fr); gap: 16px; }}
    .panel {{ background: var(--bosch-light); border-radius: 12px; padding: 16px; border: 1px solid var(--bosch-line); }}
    .panel h3 {{ margin: 0 0 10px; font-size: 14px; color: var(--bosch-blue); }}
    ul {{ margin: 0; padding-left: 18px; }}
    li {{ margin-bottom: 8px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
    th, td {{ text-align: left; padding: 10px 12px; border-bottom: 1px solid var(--bosch-line); vertical-align: top; }}
    th {{ background: var(--accent); color: var(--bosch-blue); font-weight: 700; }}
    .kpis {{ display:grid; grid-template-columns: repeat(3, 1fr); gap: 14px; }}
    .kpi {{ background: linear-gradient(180deg, #ffffff 0%, #f6f9fe 100%); border: 1px solid var(--bosch-line); border-radius: 14px; padding: 16px; }}
    .kpi .label {{ color: var(--bosch-muted); font-size: 11px; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px; }}
    .kpi .value {{ font-size: 24px; font-weight: 700; color: var(--bosch-blue); }}
    .foot {{ margin-top: 22px; font-size: 11px; color: var(--bosch-muted); text-align: center; }}
  </style>
</head>
<body>
  <div class=\"page\">
    <section class=\"cover\">
      <div class=\"bosch-logo\"><img src=\"data:image/png;base64,{logo_b64}\" alt=\"Bosch\" style=\"height:36px;display:block;\" /></div>
      <div class=\"eyebrow\">Project Charter | {DOCUMENT_VERSION}</div>
      <h1>{PROJECT_NAME}</h1>
      <p class=\"subtitle\">Carve-out of the Bosch Cloud business into a jointly managed 50/50 JV with Alibaba. The programme is infrastructure-led, legally sensitive, and executed under a restricted clean-team planning model until antitrust and disclosure decisions mature.</p>
      <div class=\"cover-grid\">
        <div class=\"cover-card\"><div class=\"label\">Seller</div><div class=\"value\">{SELLER}</div></div>
        <div class=\"cover-card\"><div class=\"label\">Buyer / JV Partner</div><div class=\"value\">{BUYER}</div></div>
        <div class=\"cover-card\"><div class=\"label\">Model</div><div class=\"value\">{CARVEOUT_MODEL}</div></div>
        <div class=\"cover-card\"><div class=\"label\">Report Date</div><div class=\"value\">{report_date}</div></div>
      </div>
    </section>

    <section class=\"section\">
      <h2>Executive Summary</h2>
      <p class=\"lead\">Gamma establishes the formal mandate for separating the Bosch Cloud business from Robert Bosch China and standing up the Day 1 operating environment for a new 50/50 JV with Alibaba. The programme scope covers 5 sites, 250 IT users, and 20 applications, with no SAP in scope and a strong emphasis on infrastructure, identity, service management, and operational readiness. The project is materially shaped by unresolved antitrust due diligence, high confidentiality, and a limited aware stakeholder group, so the delivery model relies on phased disclosure, legal dependency control, and tightly governed approvals.</p>
      <p class=\"lead\">The baseline plan targets GoLive on 01 February 2027, followed by a 90-day stabilization window and formal closure by 31 May 2027. Upstream baselines already created for Gamma are the schedule, risk register, and cost plan, and this charter reconciles to those current deliverables only.</p>
    </section>

    <section class=\"section\">
      <h2>Scope And Objectives</h2>
      <div class=\"grid\">
        <div class=\"panel\">
          <h3>In Scope</h3>
          <ul>
            <li>Infrastructure separation and target environment setup for the new JV.</li>
            <li>Identity, access, endpoint, service management, monitoring, and support readiness.</li>
            <li>Separation and transition of approximately {APPLICATIONS} applications.</li>
            <li>Transition planning for {SITES} sites and {USERS} users.</li>
            <li>Controlled interim service continuity where seller support is still required.</li>
            <li>GoLive readiness, cutover, 90-day hypercare, and operational handover.</li>
          </ul>
        </div>
        <div class=\"panel\">
          <h3>Out Of Scope</h3>
          <ul>
            <li>SAP separation or ERP transformation.</li>
            <li>Manufacturing or OT scope.</li>
            <li>Broad Day 2 transformation beyond the Day 1 JV operating baseline.</li>
            <li>Post-closure strategic changes not required for Day 1 readiness.</li>
            <li>Non-approved stakeholder disclosure outside the legal clean-team model.</li>
          </ul>
        </div>
      </div>
    </section>

    <section class=\"section\">
      <h2>Timeline And Quality Gates</h2>
      <div class=\"grid\">
        <div>
          <table>
            <tr><th>Phase</th><th>Start</th><th>Finish</th></tr>
            {phase_rows}
          </table>
        </div>
        <div>
          <table>
            <tr><th>Milestone</th><th>Date</th></tr>
            {milestone_rows}
          </table>
        </div>
      </div>
      <p class=\"lead\" style=\"margin-top:14px;\">The baseline enforces the mandatory Bosch quality-gate sequence QG0 to QG5. QG4 is on 25 January 2027 and GoLive remains on a separate date, 01 February 2027, with a five-day final-readiness window between them. Hypercare runs for exactly 90 calendar days from 02 February 2027 through 02 May 2027.</p>
    </section>

    <section class=\"section\">
      <h2>Governance And Delivery Model</h2>
      <div class=\"grid\">
        <div class=\"panel\">
          <h3>Governance Structure</h3>
          <ul>
            <li>Sponsor customer: {BUYER}</li>
            <li>Sponsor contractor: {SELLER}</li>
            <li>PMO and methodology lead: {PMO_LEAD}</li>
            <li>Joint steering decisions aligned to the 50/50 JV governance path.</li>
            <li>Restricted clean-team model until legal allows broader disclosure.</li>
          </ul>
        </div>
        <div class=\"panel\">
          <h3>Operating Model Note</h3>
          <p class=\"lead\" style=\"font-size:13px; margin:0;\">Gamma uses the Combination carve-out model. During transition, interim service continuity can remain in place where required, but the target state is a jointly managed JV environment rather than a long-term merger-zone platform. This matters because legal, security, and decision-right design must support shared management without leaving Bosch China dependencies unresolved at Day 1.</p>
        </div>
      </div>
    </section>

    <section class=\"section\">
      <h2>Risk Summary</h2>
      <div class=\"kpis\">
        <div class=\"kpi\"><div class=\"label\">Total Risks</div><div class=\"value\">{risk_count}</div></div>
        <div class=\"kpi\"><div class=\"label\">Opportunities</div><div class=\"value\">{opportunity_count}</div></div>
        <div class=\"kpi\"><div class=\"label\">Primary Exposure</div><div class=\"value\">Legal + Infrastructure</div></div>
      </div>
      <p class=\"lead\" style=\"margin-top:14px;\">The Gamma risk baseline is dominated by legal timing, confidentiality control, shared infrastructure dependencies, data protection, and QG4 readiness. Two positive risks are also tracked to capture upside from the limited application footprint and potential post-GoLive cost simplification.</p>
      <table>
        <tr><th>Risk</th><th>Category</th><th>Score</th><th>Owner</th><th>Event</th></tr>
        {risk_rows}
      </table>
    </section>

    <section class=\"section\">
      <h2>Budget Summary</h2>
      <div class=\"kpis\">
        <div class=\"kpi\"><div class=\"label\">Labour Baseline</div><div class=\"value\">{euro(labour_total)}</div></div>
        <div class=\"kpi\"><div class=\"label\">CAPEX / Additional</div><div class=\"value\">{euro(capex_total)}</div></div>
        <div class=\"kpi\"><div class=\"label\">Budget Status</div><div class=\"value\">Draft for QG1</div></div>
      </div>
      <p class=\"lead\" style=\"margin-top:14px;\">The current Gamma cost plan establishes a labour baseline derived from schedule resources and separates risk-driven CAPEX and contingency items from the labour total. Cost exposure is highest where legal timing, hidden shared-service dependencies, provider lead times, or QG4 readiness require external support or temporary duplicate services.</p>
    </section>

    <section class=\"section\">
      <h2>Assumptions, Dependencies, And Success Criteria</h2>
      <div class=\"grid\">
        <div class=\"panel\">
          <h3>Key Assumptions And Dependencies</h3>
          <ul>
            <li>Legal provides timely guidance on antitrust due diligence and permitted disclosure scope.</li>
            <li>The limited clean-team model can be expanded in phases without causing rework to the operating baseline.</li>
            <li>Shared Bosch China infrastructure dependencies are identified before Phase 2 build materially advances.</li>
            <li>Third-party approvals for network, certificates, licences, and contract novation are secured before QG4.</li>
            <li>Seller support remains available where required until the JV steady-state model is accepted.</li>
          </ul>
        </div>
        <div class=\"panel\">
          <h3>Success Criteria</h3>
          <ul>
            <li>QG0 to QG5 milestones are passed on the approved baseline dates.</li>
            <li>GoLive on 01 February 2027 occurs without critical service-continuity failure.</li>
            <li>All 5 sites, 250 users, and 20 applications operate in the JV environment on Day 1.</li>
            <li>No critical security or data-segregation issue remains open at QG4 or GoLive.</li>
            <li>Hypercare completes in 90 days and the project closes on 31 May 2027.</li>
          </ul>
        </div>
      </div>
    </section>

    <div class=\"foot\">Generated from the current Gamma schedule, risk register, and cost plan baselines only.</div>
  </div>
</body>
</html>
"""

    OUTPUT_PATH.write_text(html, encoding="utf-8")
    print(f"[{PROJECT_NAME}] Charter written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()