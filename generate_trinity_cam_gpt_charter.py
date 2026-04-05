#!/usr/bin/env python3

from __future__ import annotations

import base64
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent
PROJECT_NAME = "Trinity-CAM (GPT)"
OUTPUT_FOLDER_NAME = "Trinity-CAM (GPT) v1.1"
DOCUMENT_VERSION = "Version 1.1 - Change Request 1"
PROJECT_CODE = "TCMGPT-2026"
SELLER = "Johnson Controls International (JCI)"
BUYER = "Robert Bosch GmbH"
BUYER_SHORT = "Bosch"
BUSINESS = "Air conditioning business"
MODEL = "Integration"
PMO_LEAD = "KPMG"
IT_DELIVERY_PARTNER = "Infosys"
SITE_COUNT = 48
USER_COUNT = 12000
APPLICATION_COUNT_TEXT = "1,800+"
START_DATE = "01.07.2026"
GOLIVE_DATE = "01.01.2028"
COMPLETION_DATE = "01.04.2028"

PROJECT_DIR = BASE_DIR / "active-projects" / OUTPUT_FOLDER_NAME
SCHEDULE_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Risk_Register.xlsx"
COST_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Cost_Plan.xlsx"
OUTPUT_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Project_Charter.html"
LOGO_PATH = BASE_DIR / "Bosch.png"

IMPACT_SCORES = {
    "Very Low": 1,
    "Low": 2,
    "Moderate": 3,
    "High": 4,
    "Very High": 5,
}

PROBABILITY_SCORES = {
    "10%": 1,
    "30%": 2,
    "50%": 3,
    "70%": 4,
    "90%": 5,
}


def fmt_date(value: datetime | date | str) -> str:
  if isinstance(value, datetime):
    return value.strftime("%d %b %Y")
  if isinstance(value, date):
    return value.strftime("%d %b %Y")
  return str(value)


def fmt_eur(value: int | float) -> str:
    return f"EUR {value:,.0f}"


def html_escape(text: object) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def load_schedule() -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    ws = load_workbook(SCHEDULE_PATH, data_only=True)["Schedule"]
    phases: list[dict[str, object]] = []
    milestones: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        task_id, outline_level, name, duration, start, finish, predecessors, resources, notes, milestone = row
        if not name:
            continue
        item = {
            "id": task_id,
            "outline_level": outline_level,
            "name": str(name).strip(),
            "start": start,
            "finish": finish,
            "predecessors": predecessors,
            "resources": resources,
            "notes": notes,
        }
        if outline_level == 1:
            phases.append(item)
        if str(milestone).strip().lower() == "yes":
            milestones.append(item)
    return phases, milestones

def load_risks() -> list[dict[str, object]]:
    ws = load_workbook(RISK_PATH, data_only=True)["Risk Register"]
    risks: list[dict[str, object]] = []
    row_num = 5
    while True:
        risk_id = ws[f"B{row_num}"].value
        if not risk_id:
            row_num += 1
            if row_num > ws.max_row:
                break
            continue
        risk_id_text = str(risk_id).strip()
        if not re.fullmatch(r"R\d{3}", risk_id_text):
            row_num += 1
            if row_num > ws.max_row:
                break
            continue
        impact_label = ws[f"L{row_num}"].value
        probability_label = ws[f"N{row_num}"].value
        risk_type = str(ws[f"P{row_num}"].value or "").strip().lower()
        score = IMPACT_SCORES.get(str(impact_label), 0) * PROBABILITY_SCORES.get(str(probability_label), 0)
        risks.append(
            {
                "id": risk_id_text,
                "category": str(ws[f"D{row_num}"].value or ""),
                "event": str(ws[f"F{row_num}"].value or ""),
                "owner": str(ws[f"I{row_num}"].value or ""),
                "impact_label": str(impact_label or ""),
                "probability_label": str(probability_label or ""),
                "type": risk_type,
                "score": score,
                "measure": str(ws[f"W{row_num}"].value or ""),
                "notes": str(ws[f"AJ{row_num}"].value or ""),
            }
        )
        row_num += 1
        if row_num > ws.max_row:
            break
    return risks


def load_costs() -> dict[str, object]:
    ws = load_workbook(COST_PATH, data_only=True)["Cost Plan"]
    total_labour = 0
    baseline = ""
    note = ""
    categories: list[tuple[str, float]] = []
    capex: list[tuple[str, str, str]] = []
    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value
        e_val = ws.cell(row, 5).value
        f_val = ws.cell(row, 6).value
        if a_val == "Budget Baseline":
            baseline = str(b_val or "")
        elif a_val == "Note":
            note = str(b_val or "")
        elif a_val == "OVERALL PROJECT TOTAL":
            total_labour = float(f_val or 0)
        elif isinstance(a_val, str) and row >= 81 and row <= 90 and f_val is not None:
            categories.append((a_val, float(f_val or 0)))
        elif row >= 101 and row <= 107 and a_val:
            capex.append((str(a_val), str(b_val or ""), str(e_val or "")))
    infosys_total = sum(value for name, value in categories if name.startswith("Infosys"))
    kpmg_total = sum(value for name, value in categories if name.startswith("KPMG"))
    return {
        "total_labour": total_labour,
        "baseline": baseline,
        "note": note,
        "categories": categories,
        "infosys_total": infosys_total,
        "kpmg_total": kpmg_total,
        "capex": capex,
    }


def risk_class(score: int) -> str:
    if score >= 15:
        return "high"
    if score >= 9:
        return "medium"
    return "low"


def build_html() -> str:
    phases, milestones = load_schedule()
    risks = load_risks()
    costs = load_costs()

    top_risks = sorted(
        [risk for risk in risks if risk["type"] == "threat"],
        key=lambda risk: (-int(risk["score"]), risk["id"]),
    )[:6]
    opportunities = [risk for risk in risks if risk["type"] == "opportunity"]

    logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii") if LOGO_PATH.exists() else ""
    logo_tag = (
        f'<img src="data:image/png;base64,{logo_b64}" style="height:36px;" alt="Bosch">'
        if logo_b64
        else ""
    )
    report_date = date.today().strftime("%d %B %Y")

    milestone_cards = "".join(
        f"<div class=\"milestone-card\"><span class=\"milestone-name\">{html_escape(item['name'])}</span>"
        f"<span class=\"milestone-date\">{html_escape(fmt_date(item['start']))}</span></div>"
        for item in milestones
    )

    phase_rows = "".join(
        "<tr>"
        f"<td>{html_escape(phase['name'])}</td>"
        f"<td>{html_escape(fmt_date(phase['start']))}</td>"
        f"<td>{html_escape(fmt_date(phase['finish']))}</td>"
        f"<td>{html_escape(phase['notes'])}</td>"
        "</tr>"
        for phase in phases
    )

    risk_rows = "".join(
        "<tr>"
        f"<td><strong>{html_escape(risk['id'])}</strong></td>"
        f"<td>{html_escape(risk['category'])}</td>"
        f"<td>{html_escape(risk['event'])}</td>"
        f"<td><span class=\"risk-badge {risk_class(int(risk['score']))}\">{html_escape(risk['score'])}</span></td>"
        f"<td>{html_escape(risk['owner'])}</td>"
        "</tr>"
        for risk in top_risks
    )

    category_rows = "".join(
        "<tr>"
        f"<td>{html_escape(name)}</td>"
        f"<td>{html_escape(fmt_eur(value))}</td>"
        "</tr>"
        for name, value in costs["categories"]
        if value
    )

    capex_rows = "".join(
        "<tr>"
        f"<td>{html_escape(name)}</td>"
        f"<td>{html_escape(desc)}</td>"
        f"<td>{html_escape(value)}</td>"
        "</tr>"
        for name, desc, value in costs["capex"]
    )

    opportunity_html = ""
    if opportunities:
        opportunity = opportunities[0]
        opportunity_html = (
            "<div class=\"callout\">"
            f"<strong>Opportunity in register:</strong> {html_escape(opportunity['id'])} - "
            f"{html_escape(opportunity['event'])}"
            "</div>"
        )

    return f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
<meta charset=\"UTF-8\">
<meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">
<title>{html_escape(PROJECT_NAME)} Project Charter</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; font-family: 'Segoe UI', Arial, sans-serif; color: #122033; background: linear-gradient(180deg, #eef4fb 0%, #f8fafc 260px, #eef2f7 100%); }}
  .hero {{ background: linear-gradient(135deg, #003b6e 0%, #00569b 58%, #0f7ac7 100%); color: #fff; padding: 28px 34px 32px; }}
  .bosch-logo {{ display: flex; align-items: center; }}
  .hero-top {{ display: flex; align-items: center; justify-content: space-between; gap: 24px; }}
  .hero h1 {{ margin: 18px 0 8px; font-size: 34px; line-height: 1.1; }}
  .hero p {{ margin: 0; max-width: 980px; font-size: 15px; line-height: 1.6; opacity: 0.95; }}
  .hero-meta {{ display: grid; grid-template-columns: repeat(4, minmax(120px, 1fr)); gap: 12px; margin-top: 22px; }}
  .hero-meta div {{ background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.16); border-radius: 10px; padding: 12px 14px; }}
  .hero-meta span {{ display: block; }}
  .hero-meta .k {{ font-size: 11px; text-transform: uppercase; letter-spacing: 0.08em; opacity: 0.78; margin-bottom: 5px; }}
  .hero-meta .v {{ font-size: 16px; font-weight: 700; }}
  .page {{ max-width: 1180px; margin: 0 auto; padding: 24px 24px 48px; }}
  .section {{ background: rgba(255,255,255,0.94); border: 1px solid #d8e3ef; border-radius: 16px; box-shadow: 0 12px 30px rgba(7, 37, 73, 0.07); margin-bottom: 18px; overflow: hidden; }}
  .section h2 {{ margin: 0; padding: 15px 18px; background: #e7f0fb; color: #003b6e; font-size: 14px; text-transform: uppercase; letter-spacing: 0.08em; }}
  .section-body {{ padding: 18px; }}
  .grid-2 {{ display: grid; grid-template-columns: 1.25fr 1fr; gap: 18px; }}
  .facts {{ display: grid; grid-template-columns: 210px 1fr; gap: 8px 14px; font-size: 13px; }}
  .facts .k {{ color: #54657a; font-weight: 700; }}
  .facts .v {{ color: #122033; }}
  .callout {{ background: #eef6ff; border-left: 4px solid #0f7ac7; border-radius: 8px; padding: 12px 14px; font-size: 13px; line-height: 1.6; }}
  .milestones {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; }}
  .milestone-card {{ background: linear-gradient(180deg, #0d5c99 0%, #003b6e 100%); color: #fff; border-radius: 12px; padding: 14px; min-height: 94px; display: flex; flex-direction: column; justify-content: space-between; }}
  .milestone-name {{ font-size: 13px; font-weight: 700; line-height: 1.35; }}
  .milestone-date {{ font-size: 12px; opacity: 0.88; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  th {{ background: #003b6e; color: #fff; text-align: left; padding: 10px 12px; font-size: 11px; letter-spacing: 0.04em; text-transform: uppercase; }}
  td {{ border-bottom: 1px solid #e2ebf4; padding: 10px 12px; vertical-align: top; line-height: 1.5; }}
  tr:nth-child(even) td {{ background: #f7fbff; }}
  ul {{ margin: 0; padding-left: 20px; }}
  li {{ margin: 0 0 8px; line-height: 1.55; }}
  .budget-strip {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 12px; margin-bottom: 16px; }}
  .budget-strip div {{ background: linear-gradient(180deg, #f3f8fd 0%, #e8f1fb 100%); border: 1px solid #d6e4f2; border-radius: 12px; padding: 12px 14px; }}
  .budget-strip .k {{ display: block; color: #5b6d83; font-size: 11px; text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 4px; }}
  .budget-strip .v {{ display: block; color: #003b6e; font-size: 18px; font-weight: 800; }}
  .risk-badge {{ display: inline-flex; width: 34px; height: 34px; border-radius: 999px; align-items: center; justify-content: center; color: #fff; font-weight: 800; }}
  .risk-badge.high {{ background: #c53d2c; }}
  .risk-badge.medium {{ background: #d28a12; }}
  .risk-badge.low {{ background: #2c8a53; }}
  .signoff {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 14px; }}
  .signoff-card {{ border: 1px solid #d8e3ef; border-radius: 12px; padding: 14px; background: #fbfdff; }}
  .signoff-card .role {{ color: #5b6d83; font-size: 11px; text-transform: uppercase; letter-spacing: 0.08em; }}
  .signoff-card .name {{ margin-top: 8px; font-weight: 700; min-height: 42px; }}
  .signoff-card .line {{ margin-top: 22px; border-top: 1px solid #b8c8da; }}
  .signoff-card .date {{ margin-top: 8px; color: #6b7b90; font-size: 11px; }}
  @media (max-width: 900px) {{
    .hero-meta, .budget-strip, .milestones, .signoff, .grid-2 {{ grid-template-columns: 1fr; }}
    .facts {{ grid-template-columns: 1fr; }}
  }}
</style>
</head>
<body>
  <div class=\"hero\">
    <div class=\"hero-top\">
      <div class=\"bosch-logo\">{logo_tag}</div>
      <div>Project Charter | {html_escape(DOCUMENT_VERSION)} | {html_escape(report_date)}</div>
    </div>
    <h1>{html_escape(PROJECT_NAME)} - IT Carve-out Charter</h1>
    <p>{html_escape(SELLER)} is divesting the {html_escape(BUSINESS)} to {html_escape(BUYER)}. Change Request 1 records the approved JCI TSA extension through 31 Jul 2027. This does not change programme progress, milestone dates, or GoLive; it gives Infosys more merger-zone build-up time while users continue to work in the legacy JCI environment and therefore reduces pressure on Bosch.</p>
    <div class=\"hero-meta\">
      <div><span class=\"k\">Model</span><span class=\"v\">{html_escape(MODEL)}</span></div>
      <div><span class=\"k\">Sites</span><span class=\"v\">{SITE_COUNT}</span></div>
      <div><span class=\"k\">Users</span><span class=\"v\">{USER_COUNT:,}</span></div>
      <div><span class=\"k\">Applications</span><span class=\"v\">{html_escape(APPLICATION_COUNT_TEXT)}</span></div>
    </div>
  </div>
  <div class=\"page\">
    <div class=\"section\">
      <h2>Project Frame</h2>
      <div class=\"section-body grid-2\">
        <div class=\"facts\">
          <div class=\"k\">Project Name</div><div class=\"v\">{html_escape(PROJECT_NAME)}</div>
          <div class=\"k\">Project Code</div><div class=\"v\">{html_escape(PROJECT_CODE)}</div>
          <div class=\"k\">Seller</div><div class=\"v\">{html_escape(SELLER)}</div>
          <div class=\"k\">Buyer</div><div class=\"v\">{html_escape(BUYER)}</div>
          <div class=\"k\">Business</div><div class=\"v\">{html_escape(BUSINESS)}</div>
          <div class=\"k\">PMO / Methodology Lead</div><div class=\"v\">{html_escape(PMO_LEAD)}</div>
          <div class=\"k\">IT Delivery Partner</div><div class=\"v\">{html_escape(IT_DELIVERY_PARTNER)}</div>
          <div class=\"k\">Project Start</div><div class=\"v\">{html_escape(START_DATE)}</div>
          <div class=\"k\">GoLive</div><div class=\"v\">{html_escape(GOLIVE_DATE)}</div>
          <div class=\"k\">Completion</div><div class=\"v\">{html_escape(COMPLETION_DATE)}</div>
        </div>
        <div>
          <div class=\"callout\"><strong>Delivery model:</strong> Seller IT -> Merger Zone -> Buyer IT. The merger zone is a temporary operating bridge built and run by Infosys so that user, application, data, and service migrations can be sequenced without forcing direct cutover from JCI into Bosch.</div>
          <div class=\"callout\" style=\"margin-top:12px;\"><strong>TSA position:</strong> JCI has approved a TSA extension through 31 July 2027 because the merger zone is not yet ready and existing-user migration cannot start earlier. Users remain on the legacy JCI environment during this buffer period while Infosys continues the merger-zone build, and the overall programme milestones remain unchanged.</div>
        </div>
      </div>
    </div>

    <div class=\"section\">
      <h2>Objectives</h2>
      <div class=\"section-body grid-2\">
        <div>
          <ul>
            <li>Establish a controlled separation path for {USER_COUNT:,} users currently operating on JCI infrastructure across {SITE_COUNT} sites.</li>
            <li>Prepare and migrate {APPLICATION_COUNT_TEXT} applications, including the major SAP estate, through the merger zone into Bosch-ready landing patterns.</li>
            <li>Reach QG4 with completed user migration, stable application hosting, validated security controls, and no unresolved blocking defects.</li>
            <li>Execute GoLive on {html_escape(GOLIVE_DATE)} and close all remaining TSA dependencies during post-GoLive stabilisation, using the approved seller-side buffer to remove avoidable pre-GoLive pressure on Bosch.</li>
          </ul>
        </div>
        <div>
          <ul>
            <li>Maintain business continuity for manufacturing, supply chain, finance, workplace, and service operations throughout build, migration, and cutover.</li>
            <li>Keep governance, cost, and risk control aligned across Bosch, JCI, KPMG, and Infosys for the full programme duration.</li>
            <li>Hand over a stable operating model to Bosch by {html_escape(COMPLETION_DATE)} after 90 days of hypercare.</li>
          </ul>
          {opportunity_html}
        </div>
      </div>
    </div>

    <div class=\"section\">
      <h2>Scope</h2>
      <div class=\"section-body\">
        <table>
          <tr><th>Domain</th><th>In Scope</th><th>Out of Scope</th></tr>
          <tr><td>Applications</td><td>Full JCI Air Conditioning application estate, SAP carve-out build, interface rewiring, migration waves, and merger-zone landing.</td><td>Post-hypercare Bosch optimisation and long-term application rationalisation.</td></tr>
          <tr><td>Infrastructure</td><td>Merger-zone hosting, network connectivity, identity services, workplace services, monitoring, and security controls required for migration and stabilisation.</td><td>Broader Bosch infrastructure transformation outside the carve-out landing scope.</td></tr>
          <tr><td>Users and Devices</td><td>User move planning, collaboration migration, access transition, support readiness, and site-by-site deployment execution.</td><td>Large-scale hardware refresh not required for separation readiness.</td></tr>
          <tr><td>Data and Compliance</td><td>Data segregation, transfer controls, legal and works-council coordination, and cross-border compliance handling for the migration path.</td><td>Long-term data archival redesign after formal handover.</td></tr>
          <tr><td>Hypercare and Exit</td><td>Stabilisation, incident handling, knowledge transfer, TSA exit closure, and Bosch operational handover through QG5.</td><td>Business-as-usual support after programme closure.</td></tr>
        </table>
      </div>
    </div>

    <div class=\"section\">
      <h2>Milestones and Phases</h2>
      <div class=\"section-body\">
        <div class=\"milestones\">{milestone_cards}</div>
        <table style=\"margin-top:16px;\">
          <tr><th>Phase</th><th>Start</th><th>Finish</th><th>Purpose</th></tr>
          {phase_rows}
        </table>
      </div>
    </div>

    <div class=\"section\">
      <h2>Governance</h2>
      <div class=\"section-body\">
        <table>
          <tr><th>Role</th><th>Organisation</th><th>Accountability</th></tr>
          <tr><td>Sponsor Customer</td><td>{html_escape(BUYER)}</td><td>Buyer governance, budget sponsorship, GoLive approval, and post-cutover operating acceptance.</td></tr>
          <tr><td>Sponsor Contractor</td><td>{html_escape(SELLER)}</td><td>Seller transition support, source-environment access, data quality cooperation, and TSA adherence.</td></tr>
          <tr><td>PMO Lead</td><td>{html_escape(PMO_LEAD)}</td><td>Integrated plan control, RAID governance, steering cadence, and workstream coordination.</td></tr>
          <tr><td>IT Delivery Partner</td><td>{html_escape(IT_DELIVERY_PARTNER)}</td><td>Merger-zone build and operation, migration execution, test support, cutover readiness, and hypercare delivery.</td></tr>
          <tr><td>Steering Committee</td><td>{html_escape(BUYER_SHORT)} + JCI + {html_escape(PMO_LEAD)}</td><td>Decision-making above PMO authority, gate approvals, funding decisions, and escalation handling.</td></tr>
        </table>
      </div>
    </div>

    <div class=\"section\">
      <h2>Risk Position</h2>
      <div class=\"section-body\">
        <table>
          <tr><th>ID</th><th>Category</th><th>Risk Event</th><th>Score</th><th>Owner</th></tr>
          {risk_rows}
        </table>
        <div class=\"callout\" style=\"margin-top:14px;\"><strong>Register summary:</strong> {len(risks)} total entries, including {len(opportunities)} opportunity item(s). Highest current pressure remains on SAP carve-out complexity, merger-zone readiness, and late-stage QG4 readiness control.</div>
      </div>
    </div>

    <div class=\"section\">
      <h2>Budget Summary</h2>
      <div class=\"section-body\">
        <div class=\"budget-strip\">
          <div><span class=\"k\">External Labour Total</span><span class=\"v\">{html_escape(fmt_eur(costs['total_labour']))}</span></div>
          <div><span class=\"k\">Infosys Labour</span><span class=\"v\">{html_escape(fmt_eur(costs['infosys_total']))}</span></div>
          <div><span class=\"k\">KPMG Labour</span><span class=\"v\">{html_escape(fmt_eur(costs['kpmg_total']))}</span></div>
          <div><span class=\"k\">Baseline Status</span><span class=\"v\">{html_escape(costs['baseline'])}</span></div>
        </div>
        <table>
          <tr><th>Category</th><th>Labour Cost</th></tr>
          {category_rows}
        </table>
        <table style=\"margin-top:16px;\">
          <tr><th>CAPEX / Additional Cost</th><th>Link to Risk</th><th>Status / Range</th></tr>
          {capex_rows}
        </table>
        <div class=\"callout\" style=\"margin-top:14px;\">{html_escape(costs['note'])}</div>
      </div>
    </div>

    <div class=\"section\">
      <h2>Assumptions and Constraints</h2>
      <div class=\"section-body grid-2\">
        <div>
          <ul>
            <li>Infosys remains the primary delivery partner for merger-zone setup, operation, and migration execution.</li>
            <li>JCI continues to provide timely source-system access, SME availability, and TSA support through the approved extension date of 31 Jul 2027 while users remain on the legacy environment.</li>
            <li>Major application and SAP dependency mapping reaches sufficient fidelity during early discovery to maintain the current phase plan.</li>
            <li>Buyer-side landing patterns remain stable enough to avoid repeated redesign of merger-zone controls.</li>
          </ul>
        </div>
        <div>
          <ul>
            <li>{html_escape(GOLIVE_DATE)} remains a hard business target and cannot absorb broad late-phase scope growth.</li>
            <li>Cross-border data transfer and local labour-law constraints may force country-specific sequencing decisions.</li>
            <li>Budget baseline and non-labour contingency funding remain subject to QG1 approval.</li>
            <li>Post-GoLive work is limited to stabilisation, TSA exit, and Bosch handover, not new transformation scope.</li>
          </ul>
        </div>
      </div>
    </div>

    <div class=\"section\">
      <h2>Approval</h2>
      <div class=\"section-body\">
        <div class=\"signoff\">
          <div class=\"signoff-card\"><div class=\"role\">Sponsor Customer</div><div class=\"name\">{html_escape(BUYER)}</div><div class=\"line\"></div><div class=\"date\">Date: ____________________</div></div>
          <div class=\"signoff-card\"><div class=\"role\">Sponsor Contractor</div><div class=\"name\">{html_escape(SELLER)}</div><div class=\"line\"></div><div class=\"date\">Date: ____________________</div></div>
          <div class=\"signoff-card\"><div class=\"role\">PMO Lead</div><div class=\"name\">{html_escape(PMO_LEAD)}</div><div class=\"line\"></div><div class=\"date\">Date: ____________________</div></div>
          <div class=\"signoff-card\"><div class=\"role\">IT Delivery Partner</div><div class=\"name\">{html_escape(IT_DELIVERY_PARTNER)}</div><div class=\"line\"></div><div class=\"date\">Date: ____________________</div></div>
        </div>
      </div>
    </div>
  </div>
</body>
</html>
"""


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(build_html(), encoding="utf-8")
    print(f"[Trinity-CAM (GPT)] Project Charter: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()