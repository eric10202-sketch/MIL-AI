"""
generate_risk_register.py
Generates Trinity_Risk_Register.xlsx by populating Risk_analysis_template.xlsx
with a full risk assessment derived from the Trinity project schedule and
the FRAME reference risk register.

Usage:
    python generate_risk_register.py
    python generate_risk_register.py --template path/to/template.xlsx
                                     --schedule path/to/Trinity_Project_Schedule.csv
                                     --out      path/to/output.xlsx

Requirements: pip install openpyxl
"""

import argparse
import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─── Risk data ────────────────────────────────────────────────────────────────
# Columns: sub_project, entry_date, category, description, effects, causes,
#          probability (W 1-5), impact (T 1-5), actions, responsible, deadline,
#          status, remarks
# Risk Rating (RZ) = W × T — kept as a formula in the spreadsheet.
#
# Categories (FRAME convention):
#   ScR = Schedule   SR = Scope      RR = Resource   BtR = Budget
#   QR  = Quality    BR = Business   LR = Legality   CR  = Customer Satisfaction
# ─────────────────────────────────────────────────────────────────────────────

RISKS = [
    # ── Phase 1 / Initialization ─────────────────────────────────────────────
    (
        "Phase 1 – Governance",
        "2026-07-01",
        "ScR = Schedule",
        "IT project leadership not appointed on time by Bosch and JCI",
        "Delayed project kick-off; governance vacuum; Phase 1 milestones slip",
        "Executive bandwidth constraints; unclear mandate between Bosch CIO and JCI CIO",
        2, 4,
        "Escalate appointment to ExCo by project start date; confirm names in SPA pre-conditions",
        "Bosch CIO + JCI CIO",
        "2026-07-14",
        "Open",
        "Prerequisite for all Phase 1 activities",
    ),
    (
        "Phase 1 – Governance",
        "2026-07-01",
        "SR = Scope",
        "Governance model and RACI not agreed between Bosch and JCI within Initialization",
        "Unclear decision authority; stalled workstream progress; Steering Committee ineffective",
        "Differing governance expectations; prior relationship between JCI and Bosch not yet formalised",
        2, 3,
        "RACI workshop facilitated by KPMG by week 2; escalation matrix agreed before QG1",
        "KPMG + IT PM",
        "2026-07-20",
        "Open",
        "FRAME precedent: RACI agreed in first 4 weeks",
    ),
    (
        "Phase 1 – Scope",
        "2026-07-01",
        "SR = Scope",
        "180-site location list incomplete or inaccurate",
        "Infrastructure design based on wrong site count; wave planning errors; budget miscalculation",
        "JCI facilities data fragmented across regions; AP/AM site data not centrally maintained",
        3, 3,
        "Mandate JCI Facilities to submit full site list with addresses and types by week 3; "
        "validate against HR headcount data",
        "JCI IT + Facilities",
        "2026-08-24",
        "Open",
        "180 sites confirmed as baseline; AP and AM are largest regions",
    ),
    (
        "Phase 1 – Legal",
        "2026-07-01",
        "LR = Legality",
        "Legal entity identification delayed — especially Mexico, China, India",
        "ERP country go-live excluded; manual compliance workarounds; SPA conditions not met",
        "Complex local registration processes; bank account setup delays; tax authority lead times",
        4, 3,
        "Engage local legal counsel per country by Phase 1; Mexico contingency plan (phased go-live); "
        "India RBIN exception initiated with Bosch CISO",
        "Legal + Finance",
        "2026-08-28",
        "Open",
        "Brazil/China/India/Mexico flagged as high-complexity in CLAUDE.md",
    ),
    (
        "Phase 1 – TSA",
        "2026-07-01",
        "SR = Scope",
        "TSA framework not agreed before QG1; 18-month scope and service catalogue disputed",
        "TSA cannot be operationalised; service gaps on Day 1; JCI obligations unclear",
        "JCI and Bosch not aligned on which services are included; pricing model disputed",
        2, 4,
        "TSA preliminary framework defined by Legal and IT PM before QG1; catalogue refined in Phase 2",
        "Legal + IT PM",
        "2026-09-12",
        "Open",
        "JCI operates all IT for Bosch for 18 months from 01 Jul 2028",
    ),
    # ── Phase 2 / Concept ────────────────────────────────────────────────────
    (
        "2.1 – As-Is Analysis",
        "2026-10-01",
        "SR = Scope",
        "Application inventory incomplete — not all ~500 apps captured in LeanIX / CMDB",
        "Unplanned apps discovered during migration; scope inflation; wave plan disruption",
        "Shadow IT; regional/local apps not registered centrally; incomplete CMDB entries",
        4, 3,
        "Mandate all JCI business units to submit full app list; LeanIX discovery scan; "
        "app inventory freeze after QG2; regional validation workshops in AP/AM/EMEA",
        "JCI IT Architects + KPMG",
        "2026-11-26",
        "Open",
        "FRAME comparable: ~500 apps; use LeanIX as master landscape tool",
    ),
    (
        "2.1 – As-Is Analysis",
        "2026-10-01",
        "RR = Resource",
        "JCI non-cooperative or slow in providing as-is IT data",
        "Incomplete landscape assessment; concept phase delayed; TSA service catalogue gaps",
        "JCI data confidentiality concerns pre-SPA; competing JCI operational priorities; unclear SPA obligations",
        3, 4,
        "Establish data access protocol in SPA; weekly data request tracker; "
        "Steering Committee escalation path; KPMG facilitates JCI data sessions",
        "Bosch IT PM + KPMG",
        "2026-09-30",
        "Open",
        "Critical dependency: all Phase 2 analysis relies on JCI data provision",
    ),
    (
        "2.2 – Architecture",
        "2026-10-01",
        "ScR = Schedule",
        "Co-locator data centre procurement delayed across 3 regions (AP / AM / EMEA)",
        "Regional hub not ready; all Phase 3 infrastructure build activities blocked for affected region",
        "Co-locator RFP takes longer than planned; AP/AM data centre capacity constraints; contract negotiation",
        2, 4,
        "Co-locator selection decision by QG2; parallel RFP for all 3 regions; "
        "shortlist of 2 providers per region as contingency",
        "Bosch Procurement + Bosch Infra",
        "2027-01-13",
        "Open",
        "3 regional co-locators required: AP hub, AM hub, EMEA hub",
    ),
    (
        "2.3 – Migration Strategy",
        "2026-10-01",
        "QR = Quality",
        "Data separation rules not finalised before migration begins",
        "Data leakage risk; compliance breach; migration rework; legal disputes between Bosch and JCI",
        "Unclear data ownership; shared data between Bosch-retained and JCI carved-out entities",
        3, 4,
        "Data ownership working group with Legal, JCI, and KPMG; non-selective migration approach "
        "(FRAME precedent); finalise rules and obtain sign-off before QG2",
        "KPMG + Legal",
        "2027-01-29",
        "Open",
        "FRAME used non-selective migration approach — recommended for Trinity",
    ),
    (
        "2.1 – Contracts",
        "2026-10-01",
        "LR = Legality",
        "Change-of-control clauses triggered on key vendor contracts (~6000 device licences)",
        "Vendor contract termination; service disruption Day 1; unplanned renegotiation costs",
        "~6000 device licences and software contracts contain change-of-control clauses not yet reviewed",
        3, 3,
        "Full contract audit completed by QG2; proactive vendor engagement initiated early; "
        "licence transfer agreements executed before SPA signing",
        "JCI Procurement + Legal",
        "2028-04-01",
        "Open",
        "~6000 device licences in scope; change of control notifications to vendors in Phase 4",
    ),
    (
        "2.3 – ERP Strategy",
        "2026-10-01",
        "SR = Scope",
        "ERP migration strategy (relocation vs split vs greenfield) not decided by QG2",
        "Phase 3 SAP development cannot start; migration partner onboarding delayed; cost unknown",
        "Bosch and JCI management not aligned on ERP carve-out model; greenfield cost estimates pending",
        2, 4,
        "ERP strategy workshop facilitated by KPMG by Dec 2026; decision escalated to Steering Committee; "
        "FRAME precedent (relocation/shell copy) used as default baseline",
        "ERP Architects + KPMG",
        "2026-12-25",
        "Open",
        "FRAME used shell copy approach; recommended for Trinity SAP landscape",
    ),
    (
        "2.1 – Regional",
        "2026-10-01",
        "SR = Scope",
        "Brazil ERP tax complexity causes scope inflation and potential Day 1 exclusion",
        "Brazil CSO excluded from Day 1 go-live; manual tax compliance process; additional ERP development cost",
        "Brazil nota fiscal / SPED requirements unique; JCI Brazil SAP heavily customised for local tax law",
        4, 3,
        "Engage Brazilian tax ERP specialist by Phase 2; assess scope of Brazilian localisations; "
        "develop Brazil-specific phased go-live contingency; timeline buffer for Brazil",
        "JCI ERP Team + KPMG",
        "2027-01-22",
        "Open",
        "Brazil flagged as extreme tax complexity in project guidelines",
    ),
    (
        "2.1 – Regional",
        "2026-10-01",
        "LR = Legality",
        "China local FTS / customs system not available or not compliant in Merger Zone",
        "China operations non-compliant on Day 1; customs clearance disruption; regulatory penalties",
        "China FTS requires local hosting; Great Firewall constraints; cloud hosting restrictions",
        3, 4,
        "China IT requirements scoped separately from Phase 2; local IT partner identified; "
        "China-specific Merger Zone architecture reviewed with local regulators",
        "Regional IT Leads + Legal",
        "2027-08-06",
        "Open",
        "China local FTS flagged as high-complexity requirement",
    ),
    (
        "2.1 – OT",
        "2026-10-01",
        "SR = Scope",
        "OT / production IT scope at 180 sites significantly underestimated",
        "OT systems not migrated Day 1; production continuity risk; safety risk at manufacturing sites",
        "OT systems poorly documented; vendor-specific protocols (Siemens, Rockwell); safety certifications",
        3, 3,
        "OT assessment including site visits at top 20 production sites; OT security specialist engaged; "
        "OT systems catalogued separately from IT applications",
        "JCI OT + Production IT",
        "2026-11-25",
        "Open",
        "OT scope separate from IT; production continuity is critical",
    ),
    (
        "2.2 – Architecture",
        "2026-10-01",
        "SR = Scope",
        "Merger Zone architecture requires significant rework after QG2 approval",
        "Phase 3 build delayed; sunk cost in rework; cascading schedule impacts",
        "New security requirements identified post-QG2; JCI landscape findings conflict with initial design",
        2, 4,
        "Architecture design reviewed with all sub-workstream leads before QG2; "
        "architecture freeze after QG2 with formal change control process",
        "Bosch IT Architects + KPMG",
        "2027-01-11",
        "Open",
        "Architecture freeze = Point of No Return for Merger Zone design",
    ),
    # ── Phase 3 / Development ─────────────────────────────────────────────────
    (
        "3.1 – Infrastructure",
        "2027-04-01",
        "ScR = Schedule",
        "WAN ordering delayed across 180 sites — 4-6 month lead time not met",
        "WAN not available for Phase 4 cutover; critical path delay; regional wave plan collapses",
        "Late QG2 approval; vendor capacity constraints; regulatory delays in AP/AM regions",
        3, 5,
        "Order WAN immediately after QG3 (April 2027); dedicated WAN project manager per region; "
        "weekly status tracking with all 3 regional WAN providers; pre-qualify vendors in Phase 2",
        "Bosch Procurement + WAN Provider",
        "2027-04-28",
        "Open",
        "FRAME benchmark: 4-6 month minimum lead time — no buffer if ordering late",
    ),
    (
        "3.1 – Infrastructure",
        "2027-04-01",
        "ScR = Schedule",
        "Active Directory (Merger Zone) build delayed — cascades to all app and client migrations",
        "All app and 6000 client migrations blocked; Day 1 deadline missed",
        "AD design rework post-QG2; Bosch AD team resource constraints; JCI IdM integration complexity",
        2, 5,
        "AD design workshop completed before QG2; dedicated Bosch AD team from Phase 3 start; "
        "AD build dry-run with 100 test accounts before production go-live",
        "Bosch AD Team",
        "2027-06-10",
        "Open",
        "FRAME benchmark: AD build and go-live takes ~6 months — no slack",
    ),
    (
        "3.2 – ERP",
        "2027-04-01",
        "ScR = Schedule",
        "SAP shell copy preparation takes significantly longer than 40 days planned",
        "ERP not ready for SIT1; SIT1/SIT2/UAT chain delayed; Day 1 ERP go-live at risk",
        "SAP system complexity; JCI legacy customisations identified late; migration partner discovery findings",
        3, 4,
        "Early SAP landscape analysis completed Phase 2; migration partner contracted by QG2; "
        "technical dress rehearsal (test shell copy) before production copy",
        "Migration Partner + Bosch SAP",
        "2027-06-25",
        "Open",
        "FRAME benchmark: SAP migration (shell copy + testing) = 9-12 months",
    ),
    (
        "3.2 – ERP",
        "2027-04-01",
        "RR = Resource",
        "Migration partner (ERP / SAP) capacity not available per agreed schedule",
        "SAP shell copy delayed; SIT1/SIT2 milestones missed; UAT sign-off deadline at risk",
        "Migration partner engaged on competing programmes; exclusive SAP DMLT/SNP skill sets",
        3, 4,
        "Contractual capacity commitments from migration partner signed by QG2; "
        "penalty clauses for schedule deviation; backup migration partner identified",
        "ERP Team + Procurement",
        "2027-03-31",
        "Open",
        "Single-source risk — specialist SAP migration skills market is tight",
    ),
    (
        "3.5 – Security / IAM",
        "2027-04-01",
        "QR = Quality",
        "IAM / Identity Provider deployment delayed or defective",
        "User provisioning fails; application access blocked on Day 1; security compliance breach",
        "IAM integration complexity; Saviynt SC2 connector restricted to 10 days before closing; "
        "late identity data from JCI",
        3, 4,
        "IAM deployment completed before QG3; connector configuration frozen 10 days pre-closing; "
        "fallback manual provisioning plan defined for Day 1",
        "Bosch IAM Team",
        "2027-05-28",
        "Open",
        "FRAME benchmark: IAM/IdM implementation = 6-9 months; Saviynt SC2 10-day constraint is hard",
    ),
    (
        "3.1 – Infrastructure",
        "2027-04-01",
        "CR = Customer Satisfaction",
        "Phone number migration incomplete on Day 1 at multiple sites",
        "External calls not possible from affected sites; customer-facing communication failure",
        "4-8 week per-operator migration lead time; state-owned operators in AP/EMEA unresponsive",
        4, 3,
        "Start all phone migration requests immediately after WAN cutover per site; "
        "temporary external numbers as fallback for operators with >8-week lead time; "
        "track per operator per site",
        "Bosch Telecom",
        "2028-03-14",
        "Open",
        "FRAME precedent: phone number migration was a realised risk — plan mitigations early",
    ),
    (
        "3.3 – Applications",
        "2027-04-01",
        "SR = Scope",
        "Application wave planning disrupted by late-discovered apps (~500 apps)",
        "Wave 1 critical apps re-prioritised; migration packages need rework; schedule slip",
        "Shadow IT discovered during Phase 3; regional/local apps not in LeanIX; app owners unresponsive",
        3, 3,
        "App inventory frozen after QG2; any new apps after QG2 require change control; "
        "wave plan buffer of 10% capacity for late discoveries",
        "KPMG + WS Leads",
        "2027-05-19",
        "Open",
        "FRAME had comparable scope inflation risk with ~500 apps",
    ),
    (
        "Phase 3 – Resources",
        "2027-04-01",
        "RR = Resource",
        "Bosch IT team overloaded across 8 parallel sub-workstreams in Phase 3",
        "Task delays; quality degradation; key person burnout; critical path slippage",
        "8 concurrent Phase 3 workstreams; same resources shared across infra, SAP, M365, CWP, IAM",
        4, 3,
        "BigRoom planning to surface resource conflicts by QG2; resource plan validated monthly; "
        "external contractors for peak load (KPMG + specialist firms); "
        "IT PM tracks utilisation weekly",
        "IT PM + All WS Leads",
        "2027-11-28",
        "Open",
        "Phase 3 is the peak resource demand period for the entire programme",
    ),
    (
        "3.3 – Applications",
        "2027-04-01",
        "QR = Quality",
        "Custom application adaptation not completed before end of Phase 3",
        "Wave 1 migration delayed; critical business apps not ready for Merger Zone",
        "Dev team capacity; under-scoped custom app modifications; integration dependency complexity",
        3, 3,
        "Custom app adaptation started immediately after QG3; dev team capacity confirmed; "
        "critical apps prioritised in Wave 1",
        "Dev Teams + SWS3",
        "2027-07-25",
        "Open",
        "~500 apps — even small percentage of custom apps can create significant effort",
    ),
    (
        "3.6 – IT Org",
        "2027-04-01",
        "QR = Quality",
        "Help desk not adequately staffed for multilingual 24×7 support across AP / AM / EMEA",
        "User incidents unresolved post-Day 1; SLA breach; business disruption; management escalations",
        "Multilingual staffing across 3 time zones; MSP/ITO contracting delays; training lag",
        2, 4,
        "Help desk staffing plan approved by QG3; MSP contract signed by QG3; "
        "multilingual training programme (AP / AM / EMEA languages); IVR routing by region",
        "Bosch IT Ops + Bosch Procurement",
        "2027-07-23",
        "Open",
        "24×5 multilingual support required at minimum; 24×7 for Day 1 hypercare",
    ),
    (
        "3.5 – Security",
        "2027-04-01",
        "QR = Quality",
        "Cybersecurity breach during migration transition window",
        "Data exfiltration; production environment compromised; regulatory breach; project suspension",
        "Expanded attack surface during TSA; migration tooling access rights; weakened controls in transition",
        2, 5,
        "Security architecture reviewed by CISO before QG3; penetration testing of Merger Zone pre-go-live; "
        "SOC monitoring active from Day 1; Zero Trust architecture for Merger Zone connectivity",
        "Bosch CISO",
        "2027-07-19",
        "Open",
        "Transition windows are prime targets; FRAME also had security risk in this phase",
    ),
    (
        "3.5 – ITSM",
        "2027-04-01",
        "QR = Quality",
        "CMDB not fully populated before Day 1 (6000 assets / 180 sites)",
        "Incident management degraded post-Day 1; asset tracking gaps; ITSM SLA reporting unreliable",
        "Asset discovery at 180 sites takes longer than planned; JCI CMDB data quality poor",
        3, 2,
        "CMDB build starts Phase 3; discovery tools deployed at top 30 sites by QG4; "
        "CMDB accuracy threshold defined (>80% by Day 1); manual data import for remaining sites",
        "Bosch ITSM Team",
        "2027-07-09",
        "Open",
        "ServiceNow CMDB target for 6000 assets; AP/AM sites = largest data volumes",
    ),
    (
        "3.7 – TSA",
        "2027-04-01",
        "SR = Scope",
        "TSA service catalogue scope not agreed before QG2",
        "TSA SLA gaps; disputed service ownership; Day 1 service gaps; legal disputes",
        "JCI and Bosch not aligned on service scope; unclear service exit criteria; pricing disputes",
        2, 4,
        "TSA catalogue working group from Phase 2; service descriptions approved by both parties "
        "before QG2; legal review of all service descriptions; exit criteria defined per service",
        "JCI IT + Bosch IT + Legal",
        "2027-05-13",
        "Open",
        "18-month TSA is the primary operational risk vehicle — catalogue must be watertight",
    ),
    # ── Phase 4 / Implementation ──────────────────────────────────────────────
    (
        "4.3 – ERP Migration",
        "2027-12-01",
        "QR = Quality",
        "SAP SIT1 or SIT2 reveals critical defects requiring extended remediation",
        "Remediation delay; UAT pushed out; Day 1 ERP go-live at risk; cascades to QG4",
        "Incomplete interface mapping in Phase 3; SAP customisation conflicts; data quality issues",
        3, 4,
        "Comprehensive SIT test plan agreed with business key users before SIT1; "
        "dedicated defect triage and remediation sprints; remediation schedule buffer of 15 days",
        "ERP Team + Business",
        "2028-04-05",
        "Open",
        "SIT1 target: Jan 2028; SIT2 target: Mar 2028 — both on critical path",
    ),
    (
        "4.3 – ERP Migration",
        "2027-12-01",
        "BR = Business",
        "SAP interface breakage during shell copy — critical SAP-to-non-SAP interfaces",
        "Critical business processes fail Day 1; EDI down; supply chain disruption",
        "Incomplete interface catalogue; JCI interface customisations not captured in Phase 2",
        3, 4,
        "Full interface inventory completed by QG2; interface adaptation plan per interface; "
        "integration test plan covering all critical interfaces including EDI",
        "ERP Team + Integration Team",
        "2027-06-11",
        "Open",
        "FRAME interface mapping was a major Phase 3 effort — do not underestimate for Trinity",
    ),
    (
        "4.2 – Client Device",
        "2027-12-01",
        "RR = Resource",
        "Client device migration wave behind schedule — 6000 devices / 180 sites",
        "Users unable to work on Day 1; help desk overwhelmed; productivity loss",
        "Remote site logistics; device age and hardware failure; network connectivity during reimaging",
        3, 3,
        "Wave-based migration with AP/AM hubs first; pre-stage devices at regional hubs; "
        "regional IT team leads; 10% device buffer per wave for stragglers",
        "Bosch CWP + Regional IT",
        "2028-03-22",
        "Open",
        "6000 devices across 180 sites — logistics risk is high, especially in AP",
    ),
    (
        "4.4 – App Migration",
        "2027-12-01",
        "QR = Quality",
        "Regression testing sign-off delayed due to Wave 3 local/regional apps",
        "QG4 not met; Day 1 postponed; application defects reach production environment",
        "Wave 3 apps poorly documented; test team capacity; local app owners unavailable for testing",
        3, 3,
        "Test planning starts at QG3; risk-based testing prioritises critical apps; "
        "Test Manager appointed as formal sign-off authority; regression automation for core suite",
        "Test Team + App Teams",
        "2028-04-04",
        "Open",
        "Wave 3 = regional/local apps — highest uncertainty in test scope",
    ),
    (
        "4.2 – M365",
        "2027-12-01",
        "QR = Quality",
        "Email / M365 mailbox migration data loss or corruption — 6000+ mailboxes",
        "Email data loss; non-compliance; business disruption; end-user trust damage",
        "Large mailbox sizes; PST import errors; calendar item corruption; migration tool limitations",
        2, 4,
        "Pre-migration mailbox audit; pilot migration of 100 mailboxes before Wave 1; "
        "full backup before each wave; data integrity validation post-migration per user",
        "Bosch Azure Team",
        "2028-02-29",
        "Open",
        "6000+ mailboxes; FRAME benchmark: M365 tenant build + cutover ~6 months",
    ),
    (
        "4.2 – M365",
        "2027-12-01",
        "RR = Resource",
        "Business key users not available for ERP UAT and app testing",
        "UAT incomplete; critical defects not caught; Day 1 production issues; delayed UAT sign-off",
        "Business operational commitments; line managers deprioritise UAT; UAT scope too large",
        3, 3,
        "Formal business UAT commitment signed by business leads at QG2; "
        "UAT plan agreed 6 months in advance; dedicated business UAT champions per region",
        "Business Leads + IT PM",
        "2028-04-05",
        "Open",
        "FRAME risk #19: Business does not have enough capacity for testing — realised risk",
    ),
    (
        "4.5 – Signing",
        "2028-02-12",
        "LR = Legality",
        "SPA / APA execution delayed beyond planned Q2 2028 date",
        "Project closure delayed; Phase 5 activities postponed; continued dual-operating cost",
        "Regulatory approval delays; final legal entity setup outstanding; SPA conditions precedent",
        2, 5,
        "Legal entity finalisation tracked from Phase 2; all SPA conditions precedent reviewed monthly; "
        "contingency signing date buffer of 4 weeks built into plan",
        "Legal + Executive Leadership",
        "2028-04-19",
        "Open",
        "Signing (SPA/APA) = contractual close; Day 1 = operational go-live on 01 Jul 2028",
    ),
    (
        "4.5 – Legal Closeout",
        "2028-02-12",
        "BR = Business",
        "Frozen Zone violations — production changes made after Signing without SteerCo approval",
        "Production environment instability; untested changes on Day 1; incident risk",
        "Business pressure for last-minute fixes; inadequate change governance post-signing",
        2, 3,
        "Frozen Zone policy communicated to all teams at Signing; emergency change process "
        "requires SteerCo approval; automated change detection alerts in CMDB",
        "IT PM + All WS Leads",
        "2028-07-01",
        "Open",
        "Frozen Zone begins 19 Apr 2028 — no production changes without SteerCo approval",
    ),
    (
        "4.6 – Cutover",
        "2028-04-22",
        "ScR = Schedule",
        "Go/No-Go criteria not met — Day 1 cutover delayed",
        "SPA closing date missed; financial and legal penalties; customer and stakeholder disruption",
        "Open P1/P2 defects at cutover; critical migration wave not completed; business sign-off withheld",
        2, 5,
        "Go/No-Go criteria defined and agreed by QG3; weekly readiness dashboard from Phase 4; "
        "SteerCo escalation process for blocking issues; minimum viability threshold defined",
        "Steering Committee + IT PM",
        "2028-06-14",
        "Open",
        "Go/No-Go = 14 Jun 2028; Day 1 = 01 Jul 2028 — minimal buffer",
    ),
    # ── Phase 5 / GoLive & Hypercare ─────────────────────────────────────────
    (
        "Phase 5 – GoLive",
        "2028-07-01",
        "BR = Business",
        "Day 1 critical IT incident causing business operations failure",
        "Revenue loss; regulatory breach; reputational damage; SPA penalty clauses triggered",
        "Untested integration paths; high complexity simultaneous cutover; insufficient dress rehearsal",
        2, 5,
        "Dress rehearsal and cutover simulation completed by 07 Jun 2028; "
        "comprehensive go/no-go checklist; 24×7 hypercare P1 war room (Bosch + JCI) for 90 days",
        "IT PM + All WS Leads",
        "2028-07-05",
        "Open",
        "Hypercare = 90 calendar days from 01 Jul 2028 to 30 Sep 2028",
    ),
    (
        "Phase 5 – TSA",
        "2028-07-01",
        "BR = Business",
        "JCI TSA service degradation in first 90 days post Day 1",
        "Bosch business operations impaired; SLA breaches; financial penalties; stakeholder confidence loss",
        "JCI staff attrition post-deal; competing JCI priorities; unclear SLA governance",
        3, 4,
        "Strong TSA SLA framework with KPIs; monthly service reviews from Day 1; "
        "contractual penalty clauses for SLA breach; escalation path to JCI senior management",
        "IT PM + JCI IT",
        "2028-09-30",
        "Open",
        "18-month TSA clock starts 01 Jul 2028; expires 31 Dec 2029",
    ),
    (
        "Phase 5 – Communications",
        "2028-07-01",
        "CR = Customer Satisfaction",
        "End-user adoption failure post Day 1 — 8000 users across 180 sites",
        "Productivity loss; help desk overwhelmed; management escalations; user workarounds",
        "Insufficient change management; multi-language communication gap; no end-user training plan",
        2, 3,
        "End-user communication plan issued 60 days pre-cutover (multilingual: AP/AM/EMEA); "
        "regional IT champions; how-to guides per application; hypercare daily standup per region",
        "Comms + Regional IT",
        "2028-06-07",
        "Open",
        "8000 users across 180 sites; multi-language comms critical for AP region",
    ),
    # ── Phase 6 / Stabilisation & TSA Exit ───────────────────────────────────
    (
        "6.1 – Stabilisation",
        "2028-10-01",
        "RR = Resource",
        "Key JCI IT personnel leave before completing TSA obligations",
        "Knowledge gaps; service degradation; runbook coverage incomplete; TSA exit delayed",
        "JCI staff attrition post-deal announcement; alternative employment; JCI restructuring",
        3, 4,
        "Key person retention plan agreed with JCI in TSA contract; "
        "knowledge transfer programme starts Day 1; documentation quality reviews quarterly",
        "JCI IT + IT PM",
        "2029-12-31",
        "Open",
        "TSA = 18 months of JCI operating IT for Bosch — staff retention is critical",
    ),
    (
        "6.1 – Knowledge Transfer",
        "2028-10-01",
        "QR = Quality",
        "Insufficient knowledge transfer from JCI to Bosch — runbooks, processes, documentation",
        "Bosch unable to independently operate IT post-TSA; service gaps after Dec 2029",
        "JCI staff attrition; runbook quality low; KT programme not structured; time pressure",
        3, 4,
        "Formal KT plan with acceptance criteria per service by Day 1; "
        "structured runbook review (all 180 sites); KT completion sign-off per service required "
        "before TSA exit is approved",
        "JCI IT + Bosch IT",
        "2029-02-07",
        "Open",
        "KT window = Oct 2028 – Feb 2029; AP/AM sites are highest priority",
    ),
    (
        "6.2 – MZ to Bosch Migration",
        "2029-01-15",
        "ScR = Schedule",
        "Bosch Active Directory integration (Merger Zone → Bosch) delayed in Phase 6",
        "Client re-migration to Bosch domain blocked; TSA exit delayed; user access issues",
        "Bosch AD forest complexity; identity federation technical issues; naming convention conflicts",
        3, 3,
        "Bosch AD integration design started in Phase 5; pilot with 100 test accounts in Oct 2028; "
        "rollback plan if federation fails; Bosch IAM lead dedicated resource from Phase 6 start",
        "Bosch IAM + Bosch AD Team",
        "2029-02-25",
        "Open",
        "AD integration = foundation for all Phase 6 client and app re-migrations",
    ),
    (
        "6.2 – ERP Integration",
        "2029-02-26",
        "BR = Business",
        "ERP integration with Bosch parent systems fails in Phase 6",
        "Bosch financial consolidation broken; audit trail gaps; manual reconciliation required",
        "Bosch ERP standard not compatible with Merger Zone SAP customisations; interface gaps",
        2, 4,
        "Bosch SAP integration design from Phase 5; interface harmonisation plan agreed; "
        "reconciliation testing before production cutover; Bosch BW team prepares consolidated reporting",
        "ERP Team + Bosch IT",
        "2029-04-22",
        "Open",
        "ERP integration is on critical path for TSA exit of commercial IT services",
    ),
    (
        "6.2 – App Migration",
        "2029-01-15",
        "SR = Scope",
        "Phase 6 app re-migration waves discover compatibility issues with Bosch environment",
        "Applications cannot migrate; extended TSA for affected apps; rework cost",
        "Bosch standard application portfolio differs; version incompatibilities; "
        "security policy conflicts",
        3, 3,
        "Bosch application compatibility assessment in Phase 5; "
        "wave planning accounts for compatibility testing buffer; "
        "remediation sprint built into each wave timeline",
        "App Teams + Bosch IT",
        "2029-07-01",
        "Open",
        "~500 apps need re-migration from Merger Zone to Bosch; compatibility unknown until Phase 5",
    ),
    (
        "6.3 – TSA Exit",
        "2029-07-02",
        "ScR = Schedule",
        "TSA exit delayed beyond contractual 31 Dec 2029 deadline",
        "Post-TSA cost exposure; JCI relationship strain; operational dependency continuation",
        "Phase 6 migration waves delayed; Bosch AD integration issues; app re-migration complexity",
        3, 3,
        "TSA exit milestones tracked monthly from Day 1; service exit criteria defined per service "
        "from Phase 3; exit readiness reviews from Q3 2029; contractual penalties for JCI if TSA "
        "extended beyond Dec 2029 due to JCI failure",
        "IT PM + JCI TSA Team",
        "2029-11-14",
        "Open",
        "TSA hard deadline = 31 Dec 2029 (18 months from 01 Jul 2028 Day 1)",
    ),
    # ── Cross-cutting ─────────────────────────────────────────────────────────
    (
        "Overall – PMO",
        "2026-07-01",
        "BtR = Budget",
        "Overall project budget overrun due to scope growth, resource shortage, or FX exposure",
        "Budget cap exceeded; delayed approvals; programme decisions delayed; quality compromised",
        "180-site complexity; unexpected as-is findings; additional migration waves; EUR/USD FX exposure",
        3, 4,
        "Monthly budget-to-complete tracking; change control process for all scope additions; "
        "15% contingency reserve; Steering Committee budget review at every QG",
        "Finance + IT PM",
        "Permanent",
        "Open",
        "Trinity budget is multi-year (~3.5 years); FX and scope are primary drivers",
    ),
    (
        "Overall – PMO",
        "2026-07-01",
        "RR = Resource",
        "KPMG consulting capacity reduced or key consultant departs",
        "Methodology gaps; delayed concept deliverables; loss of institutional project knowledge",
        "KPMG reallocation to other engagements; consulting market demand; key staff turnover",
        2, 3,
        "KPMG staffing plan contractually committed by project start; "
        "knowledge management in SharePoint updated weekly; backup consultants briefed from Phase 2",
        "IT PM + KPMG",
        "Permanent",
        "Open",
        "KPMG is lead consulting partner across all workstreams in Phases 1-4",
    ),
    (
        "Overall – Legal",
        "2026-07-01",
        "LR = Legality",
        "GDPR / data privacy compliance gaps during cross-border data migration",
        "Regulatory fines; reputational damage; personal data breach notification obligation",
        "Data transferred across AP/EMEA/AM without adequate controls; "
        "migration tooling not GDPR-validated",
        2, 4,
        "GDPR compliance review of all migration tools before Phase 3; "
        "DPIA for cross-border transfers; GDPR representative in Legal workstream; "
        "data minimisation approach adopted",
        "Legal + Bosch CISO",
        "2028-06-30",
        "Open",
        "GDPR applies across EMEA; Brazil LGPD applies in AM; China PIPL applies in AP",
    ),
    (
        "Overall – Governance",
        "2026-07-01",
        "RR = Resource",
        "Steering Committee engagement insufficient — decision velocity too slow",
        "Delayed decisions on critical issues; QG approvals delayed; workstream blockers unresolved",
        "SteerCo member availability (Bosch + JCI executives); competing strategic priorities",
        2, 4,
        "Monthly SteerCo cadence from Phase 1; pre-aligned recommendations before each session; "
        "documented decision log; escalation tracker for blocking issues",
        "IT PM + Steering Committee",
        "Permanent",
        "Open",
        "FRAME lesson: early governance alignment is critical for speed of decision-making",
    ),
    (
        "Overall – Dependency",
        "2026-07-01",
        "BR = Business",
        "JCI cooperation deteriorates over the 3.5-year programme timeline",
        "Data provision delays; TSA service degradation; knowledge transfer gaps; exit disputes",
        "JCI management changes; competing JCI strategic priorities; relationship friction post-deal",
        2, 4,
        "Joint governance forums (monthly); clear contractual obligations in SPA and TSA; "
        "escalation path to both CEOs; relationship management plan",
        "IT PM + Executive Leadership",
        "Permanent",
        "Open",
        "JCI cooperation is a dependency across all 6 phases — critical to maintain",
    ),
    (
        "Overall – PMO",
        "2026-07-01",
        "CR = Customer Satisfaction",
        "Stakeholder confidence erodes due to poor programme communications over 3.5 years",
        "Business disengagement; management resistance; press/media risk; programme credibility loss",
        "Complex multi-year programme; frequent design changes; multi-region stakeholder base",
        2, 3,
        "Quarterly executive briefing pack; QG milestone communications; "
        "monthly stakeholder newsletter; Day 1 communication plan for all 8000 users",
        "IT PM + Comms",
        "Permanent",
        "Open",
        "Programme spans Jul 2026 – Dec 2029; stakeholder fatigue is a genuine risk",
    ),
    (
        "Overall – Regional",
        "2026-07-01",
        "ScR = Schedule",
        "AP region migration (largest) delayed — hub sites fall behind plan",
        "AP users not migrated on time; Merger Zone AP hub unstable; regional wave plan collapses",
        "Largest region by device count; complex country mix; WAN lead time variance across AP countries",
        3, 3,
        "AP hub-first approach; dedicated AP IT Team; weekly AP migration status report; "
        "Steering Committee escalation if >10% behind wave plan",
        "AP IT Team + Bosch Infra",
        "2028-02-23",
        "Open",
        "AP is largest region by sites and device count; highest logistical complexity",
    ),
    (
        "Overall – India",
        "2026-07-01",
        "SR = Scope",
        "India RBIN (Restricted Bosch Internal Network) compliance conflicts with Merger Zone architecture",
        "India sites cannot connect to Merger Zone; workaround required; user access disrupted",
        "RBIN restricts external network connectivity; India-specific Bosch security policy",
        3, 3,
        "India IT requirements scoped in Phase 2; RBIN exception process initiated with Bosch CISO; "
        "India-specific network design approved before QG2",
        "Regional IT Leads + Bosch CISO",
        "2026-11-20",
        "Open",
        "India RBIN flagged as country-specific requirement in project guidelines",
    ),
]


# ─── Helper: copy cell style from a source cell ───────────────────────────────

def copy_style(src, dst):
    """Copy font, alignment, fill, border from src cell to dst cell."""
    if src.has_style:
        dst.font      = copy.copy(src.font)
        dst.alignment = copy.copy(src.alignment)
        dst.fill      = copy.copy(src.fill)
        dst.border    = copy.copy(src.border)
        dst.number_format = src.number_format


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate Trinity_Risk_Register.xlsx from the risk analysis template"
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=Path(__file__).parent / "Risk_analysis_template.xlsx",
    )
    parser.add_argument(
        "--schedule",
        type=Path,
        default=Path(__file__).parent / "Trinity_Project_Schedule.csv",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).parent / "Trinity_Risk_Register.xlsx",
    )
    args = parser.parse_args()

    print(f"Reading template : {args.template}")
    wb = load_workbook(args.template)

    # ── Cover Sheet ───────────────────────────────────────────────────────────
    wc = wb["Cover sheet"]
    wc["D1"] = "Bosch CIO / JCI CIO"
    wc["D2"] = "IT PM (Bosch-led)"
    wc["D3"] = "KPMG"

    # ── Analysis Sheet ────────────────────────────────────────────────────────
    # Columns:
    #  A=No.  B=Sub-project  C=Entry Date  D=Risk Category  E=Risk Description
    #  F=Effects  G=Causes  H=W(Prob)  I=T(Impact)  J=RZ(formula)
    #  K=Actions  L=Responsible  M=Deadline  N=Status  O=Remarks
    ws = wb["Analysis of project risks"]

    DATA_START_ROW = 9       # first data row in template
    TEMPLATE_ROWS  = 43      # pre-built rows in template (rows 9–51)

    # Style reference row (row 9 from template — copy its alignment for new rows)
    ref_row = DATA_START_ROW

    wrap = Alignment(wrap_text=True, vertical="top")

    for idx, risk in enumerate(RISKS):
        row_num = DATA_START_ROW + idx

        (sub, entry_date, category, description, effects, causes,
         prob, impact, actions, responsible, deadline, status, remarks) = risk

        # If beyond template rows, set the RZ formula (column J = col 10)
        if idx >= TEMPLATE_ROWS:
            ws.cell(row_num, 10).value = f"=$H{row_num}*$I{row_num}"

        ws.cell(row_num,  1).value = idx + 1
        ws.cell(row_num,  2).value = sub
        ws.cell(row_num,  3).value = entry_date
        ws.cell(row_num,  4).value = category
        ws.cell(row_num,  5).value = description
        ws.cell(row_num,  6).value = effects
        ws.cell(row_num,  7).value = causes
        ws.cell(row_num,  8).value = prob
        ws.cell(row_num,  9).value = impact
        # J (col 10) already has formula from template; leave it for template rows
        ws.cell(row_num, 11).value = actions
        ws.cell(row_num, 12).value = responsible
        ws.cell(row_num, 13).value = deadline
        ws.cell(row_num, 14).value = status
        ws.cell(row_num, 15).value = remarks

        # Apply wrap text alignment to all populated cells
        for col in range(1, 16):
            ws.cell(row_num, col).alignment = wrap

    # ── Adjust row heights for readability ────────────────────────────────────
    for i in range(DATA_START_ROW, DATA_START_ROW + len(RISKS)):
        ws.row_dimensions[i].height = 75

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(args.out)

    # Summary
    high_risks = [(i+1, r[6]*r[7]) for i, r in enumerate(RISKS) if r[6]*r[7] >= 12]
    print(f"Written  : {args.out}")
    print(f"  Total risks    : {len(RISKS)}")
    print(f"  High (RZ ≥ 12) : {len(high_risks)} — risks: {[r[0] for r in high_risks]}")
    print(f"  Medium (RZ 6-11): "
          f"{sum(1 for r in RISKS if 6 <= r[6]*r[7] < 12)}")
    print(f"  Low (RZ < 6)   : {sum(1 for r in RISKS if r[6]*r[7] < 6)}")
    print()
    print("  Top 5 risks by RZ:")
    sorted_risks = sorted(enumerate(RISKS), key=lambda x: x[1][6]*x[1][7], reverse=True)
    for rank, (i, r) in enumerate(sorted_risks[:5], 1):
        print(f"    {rank}. Risk #{i+1} [{r[2]}] RZ={r[6]*r[7]} — {r[3][:60]}")


if __name__ == "__main__":
    main()
