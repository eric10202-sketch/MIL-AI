#!/usr/bin/env python3
"""
Generate Zebra risk register using the Bosch template.

Project: Zebra (Packaging carve-out to Undisclosed buyer)
Start: 1 April 2026 | GoLive: 1 June 2027 | Completion: 31 October 2027
37 sites, 3500+ users, SAP + 208 applications, Stand Alone, TSA-relevant.
"""

import sys
import os
from pathlib import Path
from datetime import datetime

# Ensure openpyxl can be imported from site-packages
sys.path.insert(0, os.path.join(os.path.expanduser("~"), "py_packages"))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

HERE = Path(__file__).parent
# Template can be in Reference or root folder
TEMPLATE_PATH = HERE / "Reference" / "BD_Risk-Register_template_en_V1.0_Dec2023.xlsx"
if not TEMPLATE_PATH.exists():
    TEMPLATE_PATH = HERE / "BD_Risk-Register_template_en_V1.0_Dec2023.xlsx"
OUTPUT_PATH = HERE / "active-projects" / "Zebra" / "Zebra_Risk_Register.xlsx"

# Ensure output directory exists
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# ============================================================================
# RISK DATA - Zebra Project Specific
# ============================================================================
# Format: (ID, creation_date, category, cause, event, effect, event_date,
#          owner, source, impact_text, probability_text, response_strategy, 
#          measure, due_date, status, notes)
#
# impact_text: Very Low | Low | Moderate | High | Very High
# probability_text: 10% | 30% | 50% | 70% | 90%
# response_strategy: Avoid | Transfer | Mitigate | Accept | Exploit | Enhance | Share

RISKS = [
    # Risk 1: SAP Separation Complexity
    (1, "01/04/2026", "Technology, R&D",
     "37 Packaging sites using shared SAP instance with multiple business units",
     "SAP data separation incomplete or incorrect, affecting Packaging vs other business unit transactions",
     "Post-GoLive data leakage; regulatory exposure; loss of buyer confidence in data integrity",
     "05/06/2027",
     "KPMG SAP CoE + RoboGmbH ERP",
     "Formal Risk Review",
     "High",
     "70%",
     "Mitigate",
     "3x SAP dry-run data extractions; custom ABAP separation logic; independent audit pre-cutover",
     "15/05/2027",
     "in progress",
     "Critical path item; 37-site footprint magnifies impact"),
    
    # Risk 2: 208 Application Portfolio Transition
    (2, "01/04/2026", "Technology, R&D",
     "208 legacy applications with unclear owner responsibility; mixed cloud/on-prem/mainframe",
     "Application transition delay; licensing disputes; integration failures on buyer infrastructure",
     "GoLive delay; operational disruption; additional cost recovery from buyer",
     "15/05/2027",
     "KPMG IT Architecture + App Owners",
     "Status Meeting",
     "High",
     "50%",
     "Mitigate",
     "Detailed app portfolio triage by week 8; vendor engagement for license transfer; integration labs",
     "30/03/2027",
     "not started",
     "208 apps is high-complexity.; expect vendor lock-in on 15 - 20 apps"),
    
    # Risk 3: Multi-Site Cutover Coordination (37 sites)
    (3, "01/04/2026", "Schedule",
     "37 geographically distributed sites with varying IT maturity and network readiness",
     "Parallel cutover rollout fails at 1 - 2 sites due to network latency, local infra gaps, or coordination breakdown",
     "Partial GoLive; islands of downtime; customer impact in Packaging operations; regulatory breach if supply-chain systems offline",
     "20/05/2027",
     "KPMG Program Manager + RoboGmbH Ops + Buyer Ops",
     "Formal Risk Review",
     "Very High",
     "50%",
     "Mitigate",
     "Site readiness checklist (week 16); network validation by week 12; regional coordinators; rollback playbooks per site",
     "10/05/2027",
     "in progress",
     "37-site handover is single highest operational complexity; site variance assessment due week 8"),
    
    # Risk 4: Data Migration Scope (3500+ users, large datasets)
    (4, "05/04/2026", "Technology, R&D",
     "Packaging master data (materials, orders, supply-chain records) spans 15 years; data quality varies by source system",
     "Data migration validation reveals anomalies; rework required post-GoLive; inventory/order discrepancies",
     "Hypercare extensions; buyer unable to reconcile; supply-chain disruption; cost overrun",
     "25/05/2027",
     "KPMG Data Team + DB Admins",
     "Status Meeting",
     "Moderate",
     "70%",
     "Mitigate",
     "Data audit week 6; transformation rules validated week 12; reconciliation tooling; buyer sign-off pre-cutover",
     "20/05/2027",
     "not started",
     "3500+ users depend on data accuracy; legacy system silos expected"),
    
    # Risk 5: TSA Service Level Disputes
    (5, "08/04/2026", "Legal & Compliance",
     "Seller (RoboGmbH) IT and Buyer IT teams not yet aligned on TSA SLAs; hand-off criteria undefined",
     "TSA exit delayed; ambiguous accountability; seller IT unable to wind down Packaging operations on time",
     "Extended TSA costs; vendor lock-in; reputational damage to seller-buyer relationship",
     "31/05/2027",
     "KPMG + RoboGmbH Legal + IT + Buyer Legal & IT",
     "Formal Risk Review",
     "High",
     "30%",
     "Mitigate",
     "TSA workstream kickoff week 5; co-define SLA schedule week 8; legal sign-off week 12; exit milestones locked week 16",
     "31/03/2027",
     "in progress",
     "Stand Alone model with TSA; separation of concerns critical"),
    
    # Risk 6: Buyer Infrastructure Readiness
    (6, "10/04/2026", "Resources",
     "Buyer IT staffing plan not finalized; infrastructure capacity for 3500+ users unvalidated",
     "Buyer unable to provision network, storage, or compute by GoLive; infrastructure gaps discovered in week 20",
     "Cutover delay; temporary rehost on seller; additional licensing; buyer penalties",
     "10/05/2027",
     "KPMG + Buyer IT Lead + Buyer Infrastructure",
     "Stakeholder",
     "High",
     "40%",
     "Mitigate",
     "Buyer infrastructure audit week 6; capacity plan week 10; resource hiring plan week 8; monthly validation gates",
     "31/03/2027",
     "not started",
     "Buyer unknown; limited visibility; critical dependency on buyer execution"),
    
    # Risk 7: Regulatory & Data Residency Compliance
    (7, "12/04/2026", "Legal & Compliance",
     "Packaging data spans EU + global sites; GDPR, local data residency, export-control rules not yet mapped to new buyer infrastructure",
     "Post-GoLive compliance breach; data residency violation; regulatory fine or injunction",
     "Operational restrictions; forced data replication; supply-chain disruption; reputational risk",
     "01/06/2027",
     "KPMG + Information Security + Buyer Legal",
     "Formal Risk Review",
     "Very High",
     "30%",
     "Mitigate",
     "Data classification by week 6; residency rules mapped week 10; buyer infra audit week 12; legal validation week 16",
     "20/05/2027",
     "not started",
     "Global footprint (37 sites) = multi-jurisdiction complexity; legal hold may block buyer if unresolved"),
    
    # Risk 8: Packaging Business Continuity (Go/No-Go)
    (8, "15/04/2026", "Schedule",
     "Packaging is revenue-generating business unit; minimal tolerance for downtime; no detailed contingency for extended issues",
     "QG4 (pre-GoLive) escalates due to unresolved critical issues; cutover delayed; business impact",
     "Packaging sales disruption; customer contracts at risk; financial penalty from buyer",
     "15/05/2027",
     "KPMG + PMO + RoboGmbH Ops + Buyer Exec",
     "Formal Risk Review",
     "Very High",
     "20%",
     "Mitigate",
     "Business continuity playbook week 10; QG4 criteria locked week 14; dress rehearsal week 18; go/no-go calls daily week 20-22",
     "10/05/2027",
     "not started",
     "Packaging revenue sensitivity = commercial risk; escalation path to buyer exec required"),
    
    # Risk 9: Third-Party Vendor Lock-In (ISV/SaaS)
    (9, "18/04/2026", "Budget",
     "Among 208 apps, 15 - 20 are ISV/SaaS solutions requiring vendor re-licensing or contract renegotiation for new buyer",
     "Vendor re-contract delays; licensing costs spike; buyer refuses cost; cutover blocked; app decommission unplanned",
     "Budget overrun; scope reduction; loss of functionality; go-live delay",
     "01/05/2027",
     "KPMG + IT Procurement + Vendor Management",
     "Status Meeting",
     "Moderate",
     "70%",
     "Mitigate",
     "ISV app audit week 5; vendor contact + license check week 8; cost negotiation week 12; buyer approval week 14",
     "30/04/2027",
     "not started",
     "Vendor surprises common in carve-outs; budget TBC if vendors non-cooperative"),
    
    # Risk 10: Knowledge Transfer & Skills Gap
    (10, "20/04/2026", "Resources",
     "Seller IT teams on Packaging systems are stretched across other carve-out projects; buyer hiring lag; knowledge capture informal",
     "Hypercare extended; seller IT unavailable for issue resolution; buyer staff lack Packaging-specific system knowledge",
     "Post-GoLive support costs; issue resolution SLA breaches; staff turnover risk (buyer frustration)",
     "30/05/2027",
     "KPMG + Training + Seller IT + Buyer IT",
     "Status Meeting",
     "Moderate",
     "50%",
     "Mitigate",
     "Knowledge audit week 8; train-the-trainer program week 12; documentation frozen week 18; buyers sign-off on readiness week 20",
     "25/05/2027",
     "not started",
     "Buyer unknown; timing of buyer hiring unknown; hypercare cost exposure high"),
    
    # Risk 11: Network Interconnect & Security
    (11, "25/04/2026", "Technology, R&D",
     "37 sites require persistent network connectivity from buyer infrastructure back to residual RoboGmbH systems (interfaces, shared services)",
     "Network latency or bandwidth bottleneck discovered post-GoLive; interface performance degradation; security audit blocks traffic",
     "Transaction delays; batch job failures; regulatory compliance issue; data breach risk",
     "10/05/2027",
     "KPMG + Infrastructure Architect + Information Security + Buyer IT",
     "Formal Risk Review",
     "Moderate",
     "40%",
     "Mitigate",
     "Network design review week 10; DMZ architecture week 12; security validation week 16; load testing week 18",
     "01/05/2027",
     "not started",
     "Stand Alone model = buyer must own network; residual interconnects = operational risk"),
    
    # Risk 12: Parallel Running & Cutover Window
    (12, "28/04/2026", "Schedule",
     "Parallel run duration (2 - 4 weeks) underfunded; business pressure to shorten; dual-system reconciliation unvalidated",
     "Parallel run cut short prematurely; reconciliation gaps undetected; go/no-go decision made on incomplete data",
     "Post-GoLive data inconsistencies; transaction loss in Packaging orders/inventory; buyer escalation",
     "20/05/2027",
     "KPMG + PMO + Packaging Operations",
     "Status Meeting",
     "High",
     "40%",
     "Mitigate",
     "Parallel run protocol locked week 14; reconciliation tools validated week 16; daily go/no-go reviews week 20 onwards",
     "15/05/2027",
     "not started",
     "Business continuity critical; cutover window = highest-risk period"),
    
    # Risk 13: Change Management & User Adoption (3500+ users)
    (13, "30/04/2026", "Stakeholder Relations & Public Affairs",
     "3500+ Packaging users across 37 sites; limited change readiness; some sites in low-maturity markets",
     "User error spike post-GoLive; helpdesk overwhelmed; low adoption; workarounds undermine controls",
     "Support overrun; hidden business process failures; regulatory exposure (if controls bypassed); staff retention risk",
     "25/05/2027",
     "KPMG + Training + Buyer Change Manager + Packaging Leadership",
     "Stakeholder",
     "Moderate",
     "60%",
     "Mitigate",
     "Change readiness assessment week 6; communications plan week 8; train-the-trainer week 12; site-level kick-offs week 19",
     "20/05/2027",
     "not started",
     "3500+ users = change management complexity; low-maturity sites = higher risk"),
    
    # Risk 14: Legal Hold / Buyer Disclosure
    (14, "02/05/2026", "Legal & Compliance",
     "Buyer legally confidential; legal team blocking normal governance cadence; buyer exec visibility limited; buyer IT access delayed",
     "Project decisions delayed; buyer IT prep incomplete; confidentiality breach if buyer details leak",
     "Shortened runway for buyer ramp-up; QG4 risks escalate; GoLive delay",
     "10/05/2027",
     "KPMG + RoboGmbH Legal + Project Manager",
     "Formal Risk Review",
     "Moderate",
     "30%",
     "Accept",
     "Confidentiality structure established week 1; buyer IT point-of-contact assigned week 3; encrypted comms from week 3",
     "15/04/2026",
     "implemented",
     "Buyer unknown until deal finalized; accept legal risk as project constraint"),
    
    # Risk 15: SAP Support & Upgrade Path
    (15, "05/05/2026", "Budget",
     "Seller running SAP on extended support; buyer must plan upgrade; upgrade path for Packaging suite unclear",
     "Buyer inherits unsupported SAP version; licensing disputes; emergency upgrade required 6 - 12 months post-GoLive",
     "Unplanned SAP upgrade cost; operational disruption 6 - 12 months post-GoLive; buyer cost recovery claim",
     "30/05/2027",
     "KPMG + SAP CoE + RoboGmbH ERP + Buyer ERP Lead",
     "Formal Risk Review",
     "Moderate",
     "50%",
     "Mitigate",
     "SAP roadmap finalized week 8; upgrade decision locked week 12; buyer agrees to roadmap week 14",
     "30/04/2027",
     "not started",
     "Seller may shift upgrade burden to buyer; contractual clarity essential"),
    
]

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def populate_risk_register(ws, risks):
    """Populate Risk Register sheet with risk data."""
    start_row = 5  # Data starts at row 5
    
    for idx, risk in enumerate(risks, start=0):
        row = start_row + idx
        
        task_id, creation_date, category, cause, event, effect, event_date, \
        owner, source, impact_text, probability_text, response_strategy, \
        measure, due_date, status, notes = risk
        
        # Write cell values
        ws.cell(row, 2).value = task_id                # B
        ws.cell(row, 3).value = creation_date          # C
        ws.cell(row, 4).value = category               # D
        ws.cell(row, 5).value = cause                  # E
        ws.cell(row, 6).value = event                  # F
        ws.cell(row, 7).value = effect                 # G
        ws.cell(row, 8).value = event_date             # H
        ws.cell(row, 9).value = owner                  # I
        ws.cell(row, 10).value = source                # J
        ws.cell(row, 12).value = impact_text           # L
        ws.cell(row, 14).value = probability_text      # N
        ws.cell(row, 16).value = "threat"              # P
        ws.cell(row, 18).value = ""                    # R - qualitative impact (optional)
        ws.cell(row, 19).value = ""                    # S - monetary impact current year
        ws.cell(row, 20).value = ""                    # T - monetary impact 3 years
        ws.cell(row, 22).value = response_strategy     # V
        ws.cell(row, 23).value = measure               # W
        ws.cell(row, 24).value = due_date              # X
        ws.cell(row, 26).value = status                # Z
        ws.cell(row, 27).value = datetime.now().strftime("%d/%m/%Y")  # AA - reporting date
        ws.cell(row, 28).value = impact_text           # AB - impact actual
        ws.cell(row, 30).value = probability_text      # AD - probability actual
        ws.cell(row, 35).value = notes                 # AI
        
        # Write VLOOKUP formulas
        ws.cell(row, 13).value = f'=_xlfn.IFNA(VLOOKUP(L{row},$D$182:$E$186,2,FALSE),"")'  # M
        ws.cell(row, 15).value = f'=_xlfn.IFNA(VLOOKUP(N{row},$D$189:$E$193,2,FALSE),"")'  # O
        ws.cell(row, 17).value = f'=M{row}*O{row}'                                          # Q
        ws.cell(row, 29).value = f'=_xlfn.IFNA(VLOOKUP(AB{row},$D$182:$E$186,2,FALSE),"")'  # AC
        ws.cell(row, 31).value = f'=_xlfn.IFNA(VLOOKUP(AD{row},$D$189:$E$193,2,FALSE),"")'  # AE
        ws.cell(row, 32).value = f'=AC{row}'                                                # AF
        ws.cell(row, 33).value = f'=AE{row}'                                                # AG
        ws.cell(row, 34).value = f'=AF{row}*AG{row}'                                        # AH

def main():
    print("[Zebra] Generating risk register...")
    
    # Load template
    try:
        wb = load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"ERROR: Template not found at {TEMPLATE_PATH}")
        sys.exit(1)
    
    # Update Info sheet
    info = wb["Info"]
    info["C4"].value = f"ZBR-RR-{datetime.now().strftime('%Y%m%d')}"  # Document ID
    info["C5"].value = "Zebra_Risk_Register.xlsx"
    info["C6"].value = "Zebra (Packaging Carve-Out)"
    info["C7"].value = "ZBR"
    info["C8"].value = "KPMG PMO"
    
    # Populate Risk Register sheet
    rr_sheet = wb["Risk Register"]
    populate_risk_register(rr_sheet, RISKS)
    
    # Fix Matrix sheet formatting (ensure yellow cells are readable)
    matrix_sheet = wb["Matrix "]
    for row in matrix_sheet.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.start_color:
                # Check if cell has yellow fill
                fill_color = str(cell.fill.start_color.rgb)
                if fill_color in ("00FFFFFF00", "00FFFFCC"):  # Yellow fills
                    # Force black font
                    cell.font = Font(color="000000", bold=cell.font.bold if cell.font else False)
    
    # Save workbook
    wb.save(OUTPUT_PATH)
    print(f"✓ Risk register saved to {OUTPUT_PATH}")
    print(f"  {len(RISKS)} risks populated (top risks: P×I >= 12)")

if __name__ == "__main__":
    main()
