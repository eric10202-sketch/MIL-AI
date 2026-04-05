#!/usr/bin/env python3

from __future__ import annotations

import base64
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent
PROJECT_NAME = "Trinity-CAM (GPT)"
OUTPUT_FOLDER_NAME = "Trinity-CAM (GPT) v1.1"
SELLER = "Johnson Controls International (JCI)"
BUYER = "Robert Bosch GmbH"
BUSINESS = "Air conditioning business"
MODEL = "Integration"
PMO = "KPMG"
DELIVERY_PARTNER = "Infosys"
PROJECT_DIR = BASE_DIR / "active-projects" / OUTPUT_FOLDER_NAME
SCHEDULE_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Risk_Register.xlsx"
COST_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Cost_Plan.xlsx"
OUTPUT_PATH = PROJECT_DIR / f"{PROJECT_NAME}_Executive_Dashboard.html"
LOGO_PATH = BASE_DIR / "Bosch.png"

SITE_COUNT = 48
USER_COUNT = 12000
APPLICATION_COUNT = 1800
DEVICE_COUNT = 12000
REGIONAL_DISTRIBUTION = [("EMEA", 22), ("APAC", 14), ("Americas", 12)]
COUNTRY_HOTSPOTS = [
    (
        "Germany",
        "6 sites",
        [
            "Largest EMEA labour-law and works-council exposure for user transition windows.",
            "High SAP and shared-service dependency concentration raises cutover sensitivity.",
            "Major governance focus for consultation timing before mass migration waves.",
        ],
    ),
    (
        "France",
        "3 sites",
        [
            "Local consultation requirements may delay workplace and access changes if sequencing slips.",
            "Customer-facing service continuity remains sensitive during late migration waves.",
            "Data-transfer controls need explicit approval before merger-zone landing activities.",
        ],
    ),
    (
        "USA",
        "8 sites",
        [
            "Largest Americas cluster and major Wave 2 business-volume concentration.",
            "Manufacturing and customer-service dependencies make downtime tolerance very low.",
            "Operational readiness and local support coverage must scale for peak user moves.",
        ],
    ),
    (
        "China",
        "5 sites",
        [
            "Regional application and connectivity constraints increase merger-zone compatibility risk.",
            "Country-specific data-handling and network controls require early design validation.",
            "Late remediation would compress Wave 3 closure and final readiness buffer.",
        ],
    ),
    (
        "Japan",
        "3 sites",
        [
            "Time-zone and local support-window constraints affect rehearsal and floorwalking coverage.",
            "Local service continuity expectations tighten acceptable migration-window duration.",
            "Regional business validation must complete early enough to avoid QG4 drag.",
        ],
    ),
    (
        "India",
        "3 sites",
        [
            "Strong delivery overlap with Infosys execution teams supports pilot and test throughput.",
            "Useful early-wave location for proving workplace, service desk, and runbook patterns.",
            "Needs disciplined separation of delivery access from business-user migration access paths.",
        ],
    ),
]
WORKSTREAMS = [
    (
        "WS1 - PMO and Governance",
        [
            "Integrated planning, gate control, RAID cadence, and steering governance.",
            "Cross-party dependency control across Bosch, JCI, KPMG, and Infosys.",
        ],
        "HIGH",
    ),
    (
        "WS2 - Merger Zone Platform",
        [
            "Hosting, network, monitoring, backup, and operational control plane build.",
            "Temporary landing platform between seller IT and buyer IT.",
        ],
        "HIGH",
    ),
    (
        "WS3 - SAP Carve-out",
        [
            "System copy, interface rewiring, role redesign, rehearsal cutovers, and rollback proof.",
            "Direct critical-path dependency to QG2/QG3 and QG4 readiness.",
        ],
        "MEDIUM",
    ),
    (
        "WS4 - Application Migration",
        [
            "Wave-based movement of 1,800+ applications plus validation and remediation handling.",
            "Critical app cohort, standard cohort, and long-tail closure managed separately.",
        ],
        "MEDIUM",
    ),
    (
        "WS5 - Identity and Workplace",
        [
            "AD trust, privileged access, M365, endpoint management, and site support readiness.",
            "User migration path for 12,000 people across four user waves.",
        ],
        "HIGH",
    ),
    (
        "WS6 - Security and Compliance",
        [
            "Control uplift for merger-zone access, logging, DLP, and security review.",
            "Cross-border privacy, labour-law, and data-transfer constraints built into sequencing.",
        ],
        "MEDIUM",
    ),
    (
        "WS7 - Data Migration",
        [
            "Segregation, movement, reconciliation, and evidence across SAP and non-SAP data domains.",
            "Audit-ready validation for each migration wave and cutover rehearsal.",
        ],
        "HIGH",
    ),
    (
        "WS8 - Regional Change and Deployment",
        [
            "Site playbooks, communications, floorwalking, and local readiness across 48 sites.",
            "Country-specific sequencing driven by legal and operational constraints.",
        ],
        "MEDIUM",
    ),
    (
        "WS9 - Hypercare and Handover",
        [
            "24x7 command-center support, hotfix-only stabilisation, TSA exit, and Bosch handover.",
            "No new transformation scope after GoLive; stabilisation only.",
        ],
        "HIGH",
    ),
]
APP_WAVES = [
    ("SAP Track", "Approx. 50 SAP systems", "Apr 2027 -> Nov 2027", 380, "#003b6e"),
    ("Wave 1 - Critical apps", "Approx. 450 apps", "Aug 2027 -> Sep 2027", 180, "#0066CC"),
    ("Wave 2 - Core business apps", "Approx. 800 apps", "Sep 2027 -> Nov 2027", 280, "#357ab7"),
    ("Wave 3 - Long tail and local apps", "Approx. 500 apps", "Oct 2027 -> Dec 2027", 220, "#5b9bd5"),
]

IMPACT_SCORES = {"Very Low": 1, "Low": 2, "Moderate": 3, "High": 4, "Very High": 5}
PROBABILITY_SCORES = {"10%": 1, "30%": 2, "50%": 3, "70%": 4, "90%": 5}


def html_escape(text: object) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def fmt_date(value: datetime | date | str) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d %b %Y")
    if isinstance(value, date):
        return value.strftime("%d %b %Y")
    return str(value)


def fmt_eur(value: int | float) -> str:
    return f"EUR {value:,.0f}"


def as_date(value: datetime | date) -> date:
    if isinstance(value, datetime):
        return value.date()
    return value


def days_to(target: date) -> int:
    return (as_date(target) - date.today()).days


def countdown_label(target: date) -> str:
    days = days_to(target)
    if days > 0:
        return f"+{days} days"
    if days < 0:
        return f"{abs(days)} days ago"
    return "TODAY"


def pill_class(target: date, target_label: str = "PLANNED") -> tuple[str, str]:
    days = days_to(target)
    if days < 0:
        return "pill-done", "DONE"
    if days == 0:
        return "pill-active", "TODAY"
    return "pill-active" if target_label == "TARGET" else "pill-future", target_label


def load_schedule() -> dict[str, object]:
    ws = load_workbook(SCHEDULE_PATH, data_only=True)["Schedule"]
    phases: list[dict[str, object]] = []
    milestones: list[dict[str, object]] = []
    task_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        task_id, outline_level, name, duration, start, finish, predecessors, resources, notes, milestone = row
        if not name:
            continue
        task_count += 1
        item = {
            "id": task_id,
            "outline_level": outline_level,
            "name": str(name).strip(),
            "start": start,
            "finish": finish,
            "notes": str(notes or ""),
        }
        if outline_level == 1:
            phases.append(item)
        if str(milestone).strip().lower() == "yes":
            milestones.append(item)
    start_date = phases[0]["start"]
    finish_date = phases[-1]["finish"]
    duration_months = ((finish_date.year - start_date.year) * 12) + finish_date.month - start_date.month
    return {
        "task_count": task_count,
        "phases": phases,
        "milestones": milestones,
        "start_date": start_date,
        "finish_date": finish_date,
        "duration_months": duration_months,
    }


def load_risks() -> dict[str, object]:
    ws = load_workbook(RISK_PATH, data_only=True)["Risk Register"]
    threats: list[dict[str, object]] = []
    opportunities = 0
    high = 0
    medium = 0
    low = 0
    for row_num in range(5, ws.max_row + 1):
        risk_id = ws[f"B{row_num}"].value
        if not risk_id or not re.fullmatch(r"R\d{3}", str(risk_id).strip()):
            continue
        risk_type = str(ws[f"P{row_num}"].value or "").strip().lower()
        impact_label = str(ws[f"L{row_num}"].value or "")
        probability_label = str(ws[f"N{row_num}"].value or "")
        score = IMPACT_SCORES.get(impact_label, 0) * PROBABILITY_SCORES.get(probability_label, 0)
        risk = {
            "id": str(risk_id).strip(),
            "category": str(ws[f"D{row_num}"].value or ""),
            "event": str(ws[f"F{row_num}"].value or ""),
            "owner": str(ws[f"I{row_num}"].value or ""),
            "impact_label": impact_label,
            "probability_label": probability_label,
            "score": score,
            "type": risk_type,
        }
        if risk_type == "opportunity":
            opportunities += 1
            continue
        threats.append(risk)
        if score >= 15:
            high += 1
        elif score >= 9:
            medium += 1
        else:
            low += 1
    threats.sort(key=lambda item: (-int(item["score"]), item["id"]))
    return {
        "threats": threats,
        "opportunities": opportunities,
        "high": high,
        "medium": medium,
        "low": low,
    }


def load_costs() -> dict[str, object]:
    ws = load_workbook(COST_PATH, data_only=True)["Cost Plan"]
    total_labour = 0.0
    baseline = ""
    note = ""
    category_rows: list[tuple[str, float]] = []
    phase_rows: list[tuple[str, str, float]] = []
    total_hours = 0.0
    resource_lines = 0
    for row in range(1, ws.max_row + 1):
        col_a = ws.cell(row, 1).value
        col_b = ws.cell(row, 2).value
        col_d = ws.cell(row, 4).value
        col_f = ws.cell(row, 6).value
        if col_a == "Budget Baseline":
            baseline = str(col_b or "")
        elif col_a == "Note":
            note = str(col_b or "")
        elif col_a == "OVERALL PROJECT TOTAL":
            total_labour = float(col_f or 0)
        elif 13 <= row <= 76 and col_b and col_d:
            resource_lines += 1
            total_hours += float(col_d or 0)
        elif 81 <= row <= 90 and col_a and col_f:
            category_rows.append((str(col_a), float(col_f or 0)))
        elif 93 <= row <= 98 and col_a and col_f:
            phase_rows.append((str(col_a), str(col_b or ""), float(col_f or 0)))
    return {
        "total_labour": total_labour,
        "baseline": baseline,
        "note": note,
        "categories": [(name, value) for name, value in category_rows if value > 0],
        "phases": phase_rows,
        "total_hours": int(total_hours),
        "resource_lines": resource_lines,
    }


def budget_segments(categories: list[tuple[str, float]]) -> list[dict[str, object]]:
    colors = ["#003b6e", "#005199", "#0066CC", "#357ab7", "#5b9bd5", "#7db4de", "#9ccce8", "#b8dff2"]
    total = sum(value for _, value in categories) or 1
    segments = []
    for index, (name, value) in enumerate(categories):
        percent = (value / total) * 100
        segments.append(
            {
                "name": name,
                "value": value,
                "percent": percent,
                "color": colors[index % len(colors)],
            }
        )
    return segments


def risk_class(score: int) -> str:
    if score >= 15:
        return "risk-high"
    if score >= 9:
        return "risk-med"
    return "risk-low"


def confidence_class(level: str) -> str:
    return {"HIGH": "conf-g", "MEDIUM": "conf-a", "LOW": "conf-r"}[level]


def build_html() -> str:
    schedule = load_schedule()
    risks = load_risks()
    costs = load_costs()
    segments = budget_segments(costs["categories"])

    logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii") if LOGO_PATH.exists() else ""
    logo_tag = (
        f'<img src="data:image/png;base64,{logo_b64}" alt="Bosch" style="height:36px;display:block;" />'
        if logo_b64
        else ""
    )
    today_label = date.today().strftime("%d %B %Y")
    milestones = {item["name"]: item for item in schedule["milestones"]}
    qg0 = milestones["QG0 - Programme kickoff approved"]["start"]
    qg1 = milestones["QG1 - Concept and transition model approved"]["start"]
    qg23 = milestones["QG2 and QG3 - Build complete and test entry approved"]["start"]
    qg4 = milestones["QG4 - Pre-GoLive gate approved"]["start"]
    golive = milestones["GoLive - Day 1 cutover to merger zone complete"]["start"]
    qg5 = milestones["QG5 - Project completion approved"]["start"]
    tsa_exit = schedule["start_date"]

    max_region = max(count for _, count in REGIONAL_DISTRIBUTION)
    region_rows = "".join(
        f'<div class="region-row"><span class="region-name">{html_escape(region)}</span><div class="region-bar" style="width:{int((count / max_region) * 240)}px;"></div><span class="region-n">{count} sites</span></div>'
        for region, count in REGIONAL_DISTRIBUTION
    )
    hotspot_rows = "".join(
        "<div class=\"hotspot-card\">"
        f"<div class=\"hs-country\">{html_escape(country)} ({html_escape(site_text)})</div>"
        + "".join(f"<div>{html_escape(item)}</div>" for item in items)
        + "</div>"
        for country, site_text, items in COUNTRY_HOTSPOTS
    )
    workstream_rows = "".join(
        "<div class=\"ws-card\">"
        f"<div class=\"ws-title\">{html_escape(title)}</div>"
        f"<ul class=\"ws-bullets\">{''.join(f'<li>{html_escape(item)}</li>' for item in bullets)}</ul>"
        f"<span class=\"ws-conf {confidence_class(confidence)}\">{html_escape(confidence)} CONFIDENCE</span>"
        "</div>"
        for title, bullets, confidence in WORKSTREAMS
    )
    milestone_rows = "".join(
        (
            "<tr>"
            f"<td><strong>{html_escape(label)}</strong></td>"
            f"<td>{html_escape(description)}</td>"
            f"<td>{html_escape(fmt_date(target))}</td>"
            f"<td style=\"font-size:10px;color:#0066CC;\">{html_escape(countdown_label(target))}</td>"
            f"<td><span class=\"pill {pill_class(target, status_label)[0]}\">{pill_class(target, status_label)[1]}</span></td>"
            "</tr>"
        )
        for label, description, target, status_label in [
            ("QG0", "Programme kickoff and partner mobilisation confirmed", qg0, "PLANNED"),
            ("QG1", "Concept sign-off and transition model approval", qg1, "PLANNED"),
            ("QG2&3", "Build complete, test entry, and migration factory readiness", qg23, "PLANNED"),
            ("QG4", "Pre-GoLive gate with evidence, defects, and rollback closure", qg4, "PLANNED"),
            ("GoLive", "Day 1 cutover to merger zone operating model", golive, "TARGET"),
            ("QG5", "Hypercare complete, TSA exit closed, and Bosch handover done", qg5, "PLANNED"),
        ]
    )
    phase_breakdown_rows = "".join(
        "<tr>"
        f"<td style=\"padding:3px 6px;border-bottom:1px solid #e8ecf2;\">{html_escape(name)}</td>"
        f"<td style=\"padding:3px 6px;text-align:right;border-bottom:1px solid #e8ecf2;\">{value:,.0f}</td>"
        f"<td style=\"padding:3px 6px;text-align:right;border-bottom:1px solid #e8ecf2;\">{(value / costs['total_labour']) * 100:.0f}%</td>"
        "</tr>"
        for name, _, value in costs["phases"]
    )
    budget_bar = "".join(
        f'<div class="bb-seg" style="flex:{segment["percent"]:.2f};background:{segment["color"]};" title="{html_escape(segment["name"])} {segment["percent"]:.1f}%"></div>'
        for segment in segments
    )
    budget_legend = "".join(
        f'<div><span class="bl-dot" style="background:{segment["color"]};"></span>{html_escape(segment["name"])} {segment["percent"]:.1f}%</div>'
        for segment in segments
    )
    qg_rows = "".join(
        "<div class=\"qg-row\">"
        f"<div class=\"qg-date\">{html_escape(fmt_date(target))}</div>"
        f"<div style=\"flex:1;\"><div class=\"qg-name\">{html_escape(title)}</div><div class=\"qg-sub\">{html_escape(desc)}</div></div>"
        f"<div class=\"qg-cd\">{html_escape(countdown_label(target))}</div>"
        f"<div style=\"min-width:70px;text-align:right;\"><span class=\"pill {pill_class(target, status_label)[0]}\">{pill_class(target, status_label)[1]}</span></div>"
        "</div>"
        for title, desc, target, status_label in [
            ("QG0 - Programme Kickoff", "PMO governance active, Infosys mobilised, TSA baseline fixed, and steering cadence launched.", qg0, "PLANNED"),
            ("QG1 - Concept Approved", "Application inventory, merger-zone architecture, wave plan, and baseline controls approved for build.", qg1, "PLANNED"),
            ("QG2&3 - Build Complete and Test Entry", "Core platform, SAP build path, pilot readiness, and migration factory evidence accepted.", qg23, "PLANNED"),
            ("QG4 - Pre-GoLive Gate", "All users, applications, readiness evidence, and defect closure complete before final buffer period.", qg4, "PLANNED"),
            ("GoLive - Day 1", "Cutover command center active and all in-scope services running from the merger zone.", golive, "TARGET"),
            ("QG5 - Project Completion", "Stabilisation, TSA exit, and Bosch operations handover complete after 90 days of hypercare.", qg5, "PLANNED"),
        ]
    )
    top_risks = "".join(
        (
            f"<div class=\"risk-card {risk_class(int(risk['score']))}\">"
            f"<div class=\"risk-id\">{html_escape(risk['id'])} - {html_escape(risk['category'])}</div>"
            f"<div class=\"risk-title\">{html_escape(risk['event'])}</div>"
            f"<div class=\"risk-score\">P {html_escape(risk['probability_label'])} x I {html_escape(risk['impact_label'])} | Score: <strong>{html_escape(risk['score'])}</strong> | Owner: {html_escape(risk['owner'])}</div>"
            "</div>"
        )
        for risk in risks["threats"][:5]
    )
    wave_rows = "".join(
        "<div class=\"wave-row\">"
        f"<div class=\"wave-label\">{html_escape(name)}</div>"
        f"<div class=\"wave-bar\" style=\"flex:0 0 {width}px;background:{color};\">{html_escape(window)}</div>"
        f"<div class=\"wave-n\">{html_escape(apps)}</div>"
        "</div>"
        for name, apps, window, width, color in APP_WAVES
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{html_escape(PROJECT_NAME)} Executive Dashboard</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Segoe UI',Arial,sans-serif;font-size:12px;color:#1a1a1a;background:#f4f6f9;}}
.hdr{{background:linear-gradient(135deg,#003b6e 0%,#005199 100%);color:#fff;padding:16px 28px;display:flex;align-items:center;gap:20px;}}
.bosch-logo{{display:flex;align-items:center;background:#fff;padding:4px 8px;border-radius:4px;}}
.hdr-center{{flex:1;}}
.hdr-center h1{{font-size:18px;font-weight:700;letter-spacing:.4px;}}
.hdr-center h2{{font-size:11px;font-weight:400;opacity:.85;margin-top:3px;}}
.hdr-right{{text-align:right;font-size:11px;opacity:.85;white-space:nowrap;}}
.hdr-right strong{{font-size:14px;display:block;}}
.countdown-strip{{background:#003b6e;display:flex;gap:0;}}
.cd-box{{flex:1;border-right:1px solid rgba(255,255,255,.15);padding:8px 14px;color:#fff;}}
.cd-box:last-child{{border-right:none;}}
.cd-label{{font-size:9px;text-transform:uppercase;letter-spacing:.5px;opacity:.75;}}
.cd-n{{font-size:22px;font-weight:700;}}
.cd-sub{{font-size:9px;opacity:.7;}}
.page{{max-width:1140px;margin:0 auto;padding:16px 20px 8px;}}
.page-break{{page-break-before:always;border-top:3px solid #003b6e;margin:24px 0 0;}}
.card{{background:#fff;border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:14px;overflow:hidden;}}
.card-title{{background:#005199;color:#fff;font-size:11px;font-weight:700;padding:7px 14px;letter-spacing:.4px;text-transform:uppercase;}}
.card-body{{padding:12px 14px;}}
.overview-grid{{display:grid;grid-template-columns:1fr 340px;gap:12px;}}
.overview-text p{{font-size:12px;line-height:1.55;margin:0 0 8px;}}
.fact-box{{background:#EFF4FB;border-radius:4px;padding:10px;font-size:11px;}}
.fact-box .fact-label{{font-size:9px;text-transform:uppercase;letter-spacing:.4px;color:#555;margin-bottom:2px;}}
.fact-box .fact-val{{font-weight:700;font-size:13px;color:#003b6e;}}
.fact-sep{{border-top:1px solid #ccd8ed;margin:6px 0;}}
.stat-strip{{display:grid;grid-template-columns:repeat(6,1fr);gap:8px;}}
.stat-tile{{background:#fff;border-radius:5px;box-shadow:0 1px 3px rgba(0,0,0,.07);padding:10px 8px;text-align:center;}}
.stat-tile .si{{font-size:22px;}}
.stat-tile .sn{{font-size:16px;font-weight:700;color:#003b6e;}}
.stat-tile .sl{{font-size:9px;color:#666;text-transform:uppercase;letter-spacing:.4px;margin-top:2px;}}
.phase-bar{{display:flex;height:36px;border-radius:4px;overflow:hidden;}}
.pb-seg{{display:flex;align-items:center;justify-content:center;color:#fff;font-size:9px;font-weight:700;text-align:center;white-space:nowrap;overflow:hidden;padding:2px;}}
.phase-labels{{display:flex;font-size:9px;color:#555;margin-top:3px;}}
.phase-labels span{{flex:1;white-space:nowrap;}}
.two-col{{display:grid;grid-template-columns:1fr 1fr;gap:12px;}}
.three-col{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;}}
.four-col{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;}}
.ms-table{{width:100%;border-collapse:collapse;font-size:11px;}}
.ms-table th{{background:#003b6e;color:#fff;padding:6px 8px;text-align:left;font-size:10px;}}
.ms-table td{{padding:6px 8px;border-bottom:1px solid #e8ecf2;vertical-align:middle;}}
.ms-table tr:nth-child(even){{background:#EFF4FB;}}
.pill{{display:inline-block;padding:2px 7px;border-radius:10px;font-size:9px;font-weight:700;}}
.pill-future{{background:#EFF4FB;color:#005199;}}
.pill-active{{background:#FFF2CC;color:#c07000;}}
.pill-done{{background:#d5f5e3;color:#1e8449;}}
.budget-summary{{text-align:center;padding:6px 0;}}
.budget-total{{font-size:22px;font-weight:700;color:#003b6e;}}
.budget-sub{{font-size:10px;color:#666;margin-bottom:8px;}}
.budget-bar-wrap{{margin:6px 0;}}
.budget-bar{{height:16px;border-radius:8px;overflow:hidden;display:flex;}}
.bb-seg{{height:100%;}}
.budget-legend{{display:flex;flex-wrap:wrap;gap:6px;justify-content:center;margin-top:6px;font-size:9px;}}
.bl-dot{{display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:3px;}}
.ws-card{{background:#EFF4FB;border-radius:4px;padding:8px 10px;border-left:4px solid #0066CC;}}
.ws-title{{font-weight:700;font-size:11px;color:#003b6e;margin-bottom:4px;}}
.ws-bullets{{font-size:10px;color:#333;line-height:1.5;padding-left:16px;}}
.ws-bullets li{{margin:2px 0;}}
.ws-conf{{display:inline-block;margin-top:6px;padding:2px 7px;border-radius:10px;font-size:9px;font-weight:700;}}
.conf-g{{background:#27ae60;color:#fff;}}
.conf-a{{background:#f39c12;color:#fff;}}
.conf-r{{background:#e74c3c;color:#fff;}}
.qg-row{{display:flex;gap:10px;padding:7px 0;border-bottom:1px solid #e8ecf2;align-items:flex-start;}}
.qg-row:last-child{{border-bottom:none;}}
.qg-date{{min-width:90px;font-size:10px;font-weight:600;color:#003b6e;}}
.qg-name{{font-weight:700;font-size:11px;}}
.qg-sub{{font-size:10px;color:#666;margin-top:2px;}}
.qg-cd{{min-width:70px;text-align:right;font-size:10px;font-weight:600;color:#0066CC;}}
.risk-card{{border-radius:4px;padding:8px 10px;}}
.risk-high{{background:#fdecea;border-left:4px solid #e74c3c;}}
.risk-med{{background:#fff9ec;border-left:4px solid #f39c12;}}
.risk-low{{background:#eafbf1;border-left:4px solid #27ae60;}}
.risk-id{{font-weight:700;font-size:10px;}}
.risk-title{{font-size:11px;font-weight:600;margin:2px 0;}}
.risk-score{{font-size:9px;color:#666;}}
.wave-row{{display:flex;align-items:center;gap:8px;margin:5px 0;}}
.wave-label{{min-width:150px;font-size:10px;font-weight:600;}}
.wave-bar{{height:16px;border-radius:3px;display:flex;align-items:center;padding-left:6px;color:#fff;font-size:9px;font-weight:700;}}
.wave-n{{min-width:120px;font-size:10px;color:#555;}}
.hotspot-card{{background:#EFF4FB;border-radius:4px;padding:8px 10px;}}
.hs-country{{font-weight:700;font-size:12px;margin-bottom:4px;color:#003b6e;}}
.hotspot-card div{{font-size:10px;color:#333;line-height:1.5;}}
.cp-cell{{background:#EFF4FB;border-radius:4px;padding:8px 10px;}}
.cp-title{{font-weight:700;font-size:11px;color:#003b6e;margin-bottom:5px;border-bottom:2px solid #0066CC;padding-bottom:3px;}}
.cp-item{{font-size:10px;line-height:1.5;padding:2px 0;border-bottom:1px dotted #c8d4e8;}}
.cp-item:last-child{{border-bottom:none;}}
.region-row{{display:flex;align-items:center;gap:8px;margin:4px 0;font-size:11px;}}
.region-name{{min-width:80px;color:#333;}}
.region-bar{{height:14px;background:#0066CC;border-radius:2px;}}
.region-n{{font-size:10px;color:#555;min-width:50px;}}
.footer{{text-align:center;font-size:9px;color:#aaa;padding:10px 0 20px;}}
.stats-dark{{background:#003b6e;display:grid;grid-template-columns:repeat(4,1fr);gap:0;border-radius:5px;overflow:hidden;margin-bottom:12px;}}
.sd-cell{{padding:10px 12px;border-right:1px solid rgba(255,255,255,.12);color:#fff;text-align:center;}}
.sd-cell:last-child{{border-right:none;}}
.sd-n{{font-size:22px;font-weight:700;}}
.sd-l{{font-size:9px;opacity:.75;text-transform:uppercase;letter-spacing:.4px;margin-top:2px;}}
@media print {{ .page-break {{ page-break-before: always; }} }}
</style>
</head>
<body>
<div class="hdr">
  <div class="bosch-logo">{logo_tag}</div>
  <div class="hdr-center">
    <h1>{html_escape(PROJECT_NAME)} - IT Carve-out Executive Dashboard</h1>
    <h2>{html_escape(SELLER)} -> {html_escape(BUYER)} | {html_escape(BUSINESS)} | {html_escape(MODEL)} Model</h2>
  </div>
  <div class="hdr-right">
    <strong>{html_escape(today_label)}</strong>
    GoLive: {html_escape(fmt_date(golive))} | {html_escape(countdown_label(golive))}
  </div>
</div>

<div class="countdown-strip">
  <div class="cd-box"><div class="cd-label">Days to Kickoff</div><div class="cd-n">{days_to(qg0)}</div><div class="cd-sub">{html_escape(fmt_date(qg0))} - QG0</div></div>
  <div class="cd-box"><div class="cd-label">Days to Day 1</div><div class="cd-n">{days_to(golive)}</div><div class="cd-sub">{html_escape(fmt_date(golive))} - GoLive</div></div>
  <div class="cd-box"><div class="cd-label">Days to TSA Exit Close</div><div class="cd-n">{days_to(qg5)}</div><div class="cd-sub">{html_escape(fmt_date(qg5))} - QG5</div></div>
  <div class="cd-box"><div class="cd-label">Programme Duration</div><div class="cd-n">{schedule['duration_months']}</div><div class="cd-sub">months from {html_escape(fmt_date(schedule['start_date']))}</div></div>
</div>

<div class="page">
<div class="card">
  <div class="card-title">Project Overview</div>
  <div class="card-body overview-grid">
    <div class="overview-text">
    <p>{html_escape(PROJECT_NAME)} is the IT carve-out programme for the sale of the air conditioning business from JCI to Bosch. Change Request 1 updates the baseline to reflect JCI approval of a TSA extension through 31 Jul 2027 so users can continue to work in the legacy JCI environment while Infosys completes merger-zone build-up.</p>
    <p>The delivery model remains a true integration carve-out: JCI IT services remain the source estate, Infosys builds and operates the merger zone as the temporary landing environment, and Bosch accepts the stabilised operating model after GoLive and hypercare. Programme progress and milestone dates remain unchanged; the extension simply adds buffer and removes near-term pressure from Bosch.</p>
    <p>The largest execution pressure remains on SAP carve-out complexity, merger-zone platform readiness, and final QG4 evidence closure ahead of the hard GoLive target on {html_escape(fmt_date(golive))}, but the approved TSA buffer materially improves the run-up to migration start.</p>
    </div>
    <div>
      <div class="fact-box">
        <div class="fact-label">Carve-out Model</div>
        <div class="fact-val">{html_escape(MODEL)}</div>
        <div style="font-size:10px;color:#444;margin-top:3px;">Seller IT -> Infosys merger zone -> Buyer IT</div>
        <div class="fact-sep"></div>
        <div class="fact-label">Buyer / Sponsor Customer</div>
        <div class="fact-val">{html_escape(BUYER)}</div>
        <div class="fact-sep"></div>
        <div class="fact-label">Seller / Sponsor Contractor</div>
        <div class="fact-val">{html_escape(SELLER)}</div>
        <div class="fact-sep"></div>
        <div class="fact-label">PMO / Method Lead</div>
        <div class="fact-val">{html_escape(PMO)}</div>
        <div class="fact-label" style="margin-top:4px;">IT Delivery Partner</div>
        <div class="fact-val">{html_escape(DELIVERY_PARTNER)}</div>
        <div class="fact-sep"></div>
        <div class="fact-label">Programme Labour</div>
        <div class="fact-val">{html_escape(fmt_eur(costs['total_labour']))}</div>
        <div style="font-size:9px;color:#888;">{html_escape(costs['baseline'])} | non-labour contingencies tracked separately</div>
      </div>
    </div>
  </div>
</div>

<div class="stat-strip" style="margin-bottom:14px;">
  <div class="stat-tile"><div class="si">&#127758;</div><div class="sn">{SITE_COUNT}</div><div class="sl">Global Sites</div></div>
  <div class="stat-tile"><div class="si">&#128100;</div><div class="sn">{USER_COUNT:,}</div><div class="sl">IT Users</div></div>
  <div class="stat-tile"><div class="si">&#128187;</div><div class="sn">{DEVICE_COUNT:,}</div><div class="sl">Client Devices</div></div>
  <div class="stat-tile"><div class="si">&#9881;</div><div class="sn">{APPLICATION_COUNT:,}+</div><div class="sl">Applications</div></div>
  <div class="stat-tile"><div class="si">&#128197;</div><div class="sn">{schedule['duration_months']} mo</div><div class="sl">Programme Duration</div></div>
    <div class="stat-tile"><div class="si">&#127975;</div><div class="sn">31 Jul 2027</div><div class="sl">Approved JCI TSA End</div></div>
</div>

<div class="card">
  <div class="card-title">Programme Phase Timeline</div>
  <div class="card-body">
    <div class="phase-bar">
      <div class="pb-seg" style="flex:92;background:#357ab7;">Phase 0<br>Mobilise</div>
      <div class="pb-seg" style="flex:121;background:#0066CC;">Phase 1<br>Discover and Design</div>
      <div class="pb-seg" style="flex:180;background:#005199;">Phase 2<br>Build</div>
      <div class="pb-seg" style="flex:131;background:#003b6e;">Phase 3<br>Test and Migrate</div>
      <div class="pb-seg" style="flex:21;background:#1a1a6e;">Phase 4<br>Readiness</div>
      <div class="pb-seg" style="flex:90;background:#4a4a4a;">Phase 5<br>Hypercare</div>
    </div>
    <div class="phase-labels">
      <span>{html_escape(fmt_date(qg0))}</span>
      <span>{html_escape(fmt_date(qg1))}</span>
      <span>{html_escape(fmt_date(qg23))}</span>
      <span>{html_escape(fmt_date(qg4))}</span>
      <span>{html_escape(fmt_date(golive))}</span>
      <span style="text-align:right;">{html_escape(fmt_date(qg5))}</span>
    </div>
  </div>
</div>

<div class="two-col">
  <div class="card">
    <div class="card-title">Key Milestones and Quality Gates</div>
    <div class="card-body" style="padding:8px;">
      <table class="ms-table">
        <tr><th>Gate</th><th>Description</th><th>Date</th><th>Countdown</th><th>Status</th></tr>
        {milestone_rows}
      </table>
    </div>
  </div>

  <div class="card">
    <div class="card-title">Budget Distribution</div>
    <div class="card-body">
      <div class="budget-summary">
        <div class="budget-total">{html_escape(fmt_eur(costs['total_labour']))}</div>
        <div class="budget-sub">External labour baseline only | {html_escape(costs['baseline'])}</div>
      </div>
      <div class="budget-bar-wrap"><div class="budget-bar">{budget_bar}</div></div>
      <div class="budget-legend">{budget_legend}</div>
      <table style="width:100%;border-collapse:collapse;font-size:10px;margin-top:10px;">
        <tr><th style="background:#003b6e;color:#fff;padding:4px 6px;text-align:left;">Phase</th><th style="background:#003b6e;color:#fff;padding:4px 6px;text-align:right;">EUR</th><th style="background:#003b6e;color:#fff;padding:4px 6px;text-align:right;">%</th></tr>
        {phase_breakdown_rows}
        <tr style="font-weight:700;background:#C6D4E8;"><td style="padding:4px 6px;">Total Labour</td><td style="padding:4px 6px;text-align:right;">{costs['total_labour']:,.0f}</td><td style="padding:4px 6px;text-align:right;">100%</td></tr>
      </table>
    </div>
  </div>
</div>

<div class="page-break"></div>

<div class="card">
  <div class="card-title">IT Workstream Coverage - Confidence Overview</div>
  <div class="card-body"><div class="three-col">{workstream_rows}</div></div>
</div>

<div class="card">
  <div class="card-title">Quality Gate Tracker</div>
  <div class="card-body">{qg_rows}</div>
</div>

<div class="two-col">
  <div class="card">
    <div class="card-title">Regional Site Distribution</div>
    <div class="card-body">
      {region_rows}
      <div style="margin-top:12px;font-size:10px;color:#555;">Planning geography for dashboard tracking covers the main concentration points used in regional sequencing and local readiness management.</div>
    </div>
  </div>
  <div class="card">
    <div class="card-title">Key Risk Indicators</div>
    <div class="card-body" style="display:flex;flex-direction:column;gap:6px;">
      {top_risks}
      <div style="font-size:10px;color:#666;margin-top:4px;">Threat profile: {risks['high']} high, {risks['medium']} medium, {risks['low']} lower-score threats, plus {risks['opportunities']} opportunity entries in the register.</div>
    </div>
  </div>
</div>

<div class="page-break"></div>

<div class="card">
  <div class="card-title">Application Migration Waves</div>
  <div class="card-body">
    <div style="margin-bottom:8px;font-size:11px;color:#555;">Wave planning follows the generated schedule: critical applications first, then core business services, then regional and long-tail closure, with SAP running as a dedicated parallel track.</div>
    {wave_rows}
  </div>
</div>

<div class="card">
  <div class="card-title">Country-Specific Complexity Hotspots</div>
  <div class="card-body"><div class="three-col">{hotspot_rows}</div></div>
</div>

<div class="stats-dark">
  <div class="sd-cell"><div class="sd-n">{schedule['task_count']}</div><div class="sd-l">Schedule Tasks</div></div>
  <div class="sd-cell"><div class="sd-n">{len(costs['categories'])}</div><div class="sd-l">Resource Categories</div></div>
  <div class="sd-cell"><div class="sd-n">{costs['total_hours']:,}</div><div class="sd-l">Person-Hours</div></div>
  <div class="sd-cell"><div class="sd-n">3</div><div class="sd-l">Regional Clusters</div></div>
</div>

<div class="card">
  <div class="card-title">Critical Path and Guiding Principles</div>
  <div class="card-body">
    <div class="four-col">
      <div class="cp-cell">
        <div class="cp-title">Merger Zone Platform</div>
        <div class="cp-item">Hosting, identity, and network stack must be stable before broad migration starts.</div>
        <div class="cp-item">Operational monitoring and DR controls need proof before test exit.</div>
        <div class="cp-item">Any platform delay compresses all downstream migration windows.</div>
      </div>
      <div class="cp-cell">
        <div class="cp-title">SAP Critical Path</div>
        <div class="cp-item">System copy, role redesign, and interface rewiring drive the highest schedule exposure.</div>
        <div class="cp-item">Mock cutovers and rollback evidence must close before QG4.</div>
        <div class="cp-item">SAP remains the single most material path to GoLive risk.</div>
      </div>
      <div class="cp-cell">
        <div class="cp-title">User and Site Migration</div>
        <div class="cp-item">All four user waves and all 48 sites must complete before final gate approval.</div>
        <div class="cp-item">Local readiness, legal consultation, and support coverage affect wave throughput.</div>
        <div class="cp-item">Post-GoLive work is stabilisation only, not deferred migration.</div>
      </div>
      <div class="cp-cell">
        <div class="cp-title">Programme Principles</div>
        <div class="cp-item">GoLive on {html_escape(fmt_date(golive))} remains the hard business commitment.</div>
        <div class="cp-item">No new transformation scope enters after QG4.</div>
        <div class="cp-item">The approved TSA extension reduces pressure before migration start; formal TSA exit and Bosch handover still complete during hypercare, not later.</div>
      </div>
    </div>
  </div>
</div>

<div class="footer">{html_escape(PROJECT_NAME)} Executive Dashboard | {html_escape(today_label)} | Data sources: {html_escape(SCHEDULE_PATH.name)}, {html_escape(RISK_PATH.name)}, {html_escape(COST_PATH.name)} | CONFIDENTIAL</div>
</div>
</body>
</html>
"""


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(build_html(), encoding="utf-8")
    print(f"[{PROJECT_NAME}] Executive Dashboard: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()