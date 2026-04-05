#!/usr/bin/env python3
"""Generate the Gamma management KPI dashboard as self-contained HTML."""

from __future__ import annotations

import base64
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


HERE = Path(__file__).parent
PROJECT_NAME = "Gamma"
OUTPUT_FOLDER_NAME = "Gamma v1.0"
SELLER = "Robert Bosch China"
BUYER = "Alibaba"
MODEL = "Combination"
REPORT_DATE = datetime(2026, 4, 5)
SCHEDULE_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Project_Schedule.xlsx"
RISK_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Risk_Register.xlsx"
COST_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Cost_Plan.xlsx"
OUTPUT_PATH = HERE / "active-projects" / OUTPUT_FOLDER_NAME / f"{PROJECT_NAME}_Management_KPI_Dashboard.html"
LOGO_PATH = HERE / "Bosch.png"

IMPACT_VALUES = {"Very Low": 1, "Low": 2, "Moderate": 3, "High": 4, "Very High": 5}
PROBABILITY_VALUES = {"10%": 1, "30%": 2, "50%": 3, "70%": 4, "90%": 5}


def days_until(date_str: str) -> int:
    return (datetime.strptime(date_str, "%Y-%m-%d") - REPORT_DATE).days


def fmt_date(value) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%d %b %Y")
    return str(value)


def load_schedule_metrics():
    wb = load_workbook(SCHEDULE_PATH, data_only=False)
    ws = wb["Schedule"]
    milestones = []
    resource_groups = set()
    tasks = 0
    total_hours = 0
    for row in range(2, ws.max_row + 1):
        level = ws.cell(row, 2).value
        if level is None:
            continue
        level = int(level)
        tasks += 1
        name = str(ws.cell(row, 3).value or "").strip()
        start = ws.cell(row, 5).value
        if level >= 3 and ws.cell(row, 8).value:
            days = int(str(ws.cell(row, 4).value).split()[0])
            tokens = [token.strip() for token in str(ws.cell(row, 8).value).split("+") if token.strip()]
            total_hours += days * 8 * len(tokens)
            resource_groups.update(tokens)
        if str(ws.cell(row, 10).value) == "Yes" and ("QG" in name or "GoLive" in name or "Closure" in name):
            iso_date = start.strftime("%Y-%m-%d")
            milestones.append((name, fmt_date(start), iso_date, days_until(iso_date)))
    return milestones, tasks, len(resource_groups), total_hours


def load_risk_metrics():
    wb = load_workbook(RISK_PATH, data_only=False)
    ws = wb["Risk Register"]
    risks = []
    category_counts = defaultdict(int)
    high = 0
    for row in range(5, 140):
        risk_id = ws.cell(row, 2).value
        if risk_id is None:
            continue
        category = ws.cell(row, 4).value
        impact = ws.cell(row, 12).value
        probability = ws.cell(row, 14).value
        risk_type = ws.cell(row, 16).value
        score = IMPACT_VALUES.get(impact, 0) * PROBABILITY_VALUES.get(probability, 0)
        if risk_type == "threat" and score >= 12:
            high += 1
        category_counts[category] += 1
        risks.append(
            {
                "id": int(risk_id),
                "category": category,
                "event": ws.cell(row, 6).value,
                "owner": ws.cell(row, 9).value,
                "score": score,
                "type": risk_type,
            }
        )
    risks.sort(key=lambda item: item["score"], reverse=True)
    return high, category_counts, risks[:6]


def load_cost_metrics():
    wb = load_workbook(COST_PATH, data_only=False)
    ws = wb["Cost Plan"]
    labour_total = 0
    capex_total = 0
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label == "OVERALL PROJECT TOTAL - LABOUR ONLY":
            labour_total = int(ws.cell(row, 6).value or 0)
        if label == "TOTAL CAPEX / ADDITIONAL COSTS":
            capex_total = int(ws.cell(row, 6).value or 0)
    return labour_total, capex_total


def status_class(value: float, green: float, amber: float) -> str:
    if value >= green:
        return "green"
    if value >= amber:
        return "amber"
    return "red"


def main() -> None:
    milestones, task_count, resource_group_count, total_hours = load_schedule_metrics()
    high_risks, category_counts, top_risks = load_risk_metrics()
    labour_total, capex_total = load_cost_metrics()
    logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode()

    spi = 1.00
    cpi = 1.00
    day1_readiness = max(35, 88 - high_risks * 3)
    transition_confidence = max(40, 90 - (category_counts.get("Legal & Compliance", 0) + category_counts.get("Strategy & Portfolio", 0) + category_counts.get("Ecosystems & Ethics", 0)) * 8)
    infrastructure_confidence = max(45, 92 - (category_counts.get("Technology, R&D", 0) + category_counts.get("Security & Data Protection", 0)) * 6)

    actions_horizon = REPORT_DATE + timedelta(days=90)
    kickoff_gap = days_until("2026-08-01")
    actions = [
        ("High", f"Freeze the antitrust decision path and clean-team escalation model before {actions_horizon.strftime('%d %b %Y')} to protect QG0 and QG1 assumptions."),
        ("High", "Lock named infrastructure, security, and legal SMEs into the restricted Gamma team so mobilisation can start on 01 Aug 2026 without approval gaps."),
        ("Medium", "Pre-stage shared-service dependency evidence and vendor pre-reads so concept design can accelerate immediately after kickoff despite disclosure limits."),
        ("Medium", "Prepare the controlled communication pack and stakeholder expansion criteria for use once legal confirms the next disclosure step."),
    ]
    if kickoff_gap < 90:
        actions.insert(0, ("High", "Kickoff falls within the next 90 days - confirm QG0 readiness packs, PMO controls, and inventory interview calendar."))

    workstream_confidence = [
        ("Governance & Confidentiality", max(40, 86 - (category_counts.get("Legal & Compliance", 0) * 7)), "red"),
        ("Infrastructure & Cloud", infrastructure_confidence, status_class(infrastructure_confidence, 75, 55)),
        ("Security & Identity", max(45, 84 - category_counts.get("Security & Data Protection", 0) * 8), "amber"),
        ("Applications & Data", 74, "green"),
        ("Testing & Readiness", max(45, 82 - category_counts.get("Quality", 0) * 10), "amber"),
        ("Hypercare & Handover", 70, "green"),
    ]

    milestone_html = "\n".join(
        f"<li><span>{name}</span><span>{date}</span><span class='status {status_class(100 - abs(days), 80, 40)}'>{days:+d}d</span></li>"
        for name, date, _, days in milestones
    )
    top_risk_rows = "\n".join(
        f"<tr><td>#{risk['id']}</td><td>{risk['category']}</td><td>{risk['score']}</td><td>{risk['owner']}</td><td>{risk['event']}</td></tr>"
        for risk in top_risks
    )
    workstream_rows = "\n".join(
        f"<div class='ws-row'><div class='ws-label'>{name}</div><div class='ws-bar'><div class='{status}' style='width:{score}%;'></div></div><div class='ws-score'>{score}%</div></div>"
        for name, score, status in workstream_confidence
    )
    action_rows = "\n".join(
        f"<div class='action-item {priority.lower()}'><strong>{priority}</strong> - {text}</div>"
        for priority, text in actions
    )

    html = f"""<!DOCTYPE html>
<html lang='en'>
<head>
  <meta charset='UTF-8' />
  <meta name='viewport' content='width=device-width, initial-scale=1.0' />
  <title>{PROJECT_NAME} Management KPI Dashboard</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; background: #f4f6f9; color: #1a1a1a; }}
    .dashboard {{ max-width: 1220px; margin: 0 auto; padding: 20px; }}
    .header {{ background: linear-gradient(135deg, #002147 0%, #003b6e 100%); color: #fff; border-radius: 12px; padding: 20px 22px; margin-bottom: 18px; display:flex; justify-content:space-between; align-items:center; }}
    .bosch-logo {{ display:flex; align-items:center; background:#fff; padding:4px 8px; border-radius:4px; width:fit-content; margin-bottom:10px; }}
    .header h1 {{ margin: 0 0 4px; font-size: 28px; }}
    .header p {{ margin: 0; font-size: 12px; opacity: 0.9; }}
    .header-right {{ text-align:right; font-size:12px; }}
    .grid {{ display:grid; grid-template-columns: repeat(12, 1fr); gap:12px; }}
    .card {{ background:#fff; border-radius:10px; padding:16px; box-shadow:0 2px 6px rgba(0,0,0,0.08); }}
    .span-12 {{ grid-column: span 12; }} .span-8 {{ grid-column: span 8; }} .span-6 {{ grid-column: span 6; }} .span-4 {{ grid-column: span 4; }} .span-3 {{ grid-column: span 3; }}
    .card-header {{ font-weight:700; font-size:13px; color:#003b6e; margin-bottom:10px; padding-bottom:8px; border-bottom:1px solid #e7edf6; }}
    .kpi-value {{ font-size:30px; font-weight:700; color:#003b6e; margin:8px 0 4px; }}
    .subtle {{ color:#5d6d83; font-size:12px; }}
    .status {{ display:inline-block; padding:3px 8px; border-radius:12px; font-size:11px; font-weight:700; color:#fff; text-transform:uppercase; }}
    .status.green {{ background:#007A33; }}
    .status.amber {{ background:#E8A000; color:#1a1a1a; }}
    .status.red {{ background:#CC0000; }}
    .progress {{ height:12px; background:#e1e7f0; border-radius:999px; overflow:hidden; margin:8px 0 10px; }}
    .progress div {{ height:100%; background:linear-gradient(90deg, #0066CC, #003b6e); }}
    .milestones {{ list-style:none; padding:0; margin:0; }}
    .milestones li {{ display:grid; grid-template-columns: 1.6fr .8fr .4fr; gap:10px; align-items:center; padding:8px 0; border-bottom:1px solid #edf1f7; font-size:12px; }}
    .ws-row {{ display:grid; grid-template-columns: 1.4fr 2fr .5fr; gap:10px; align-items:center; margin-bottom:10px; font-size:12px; }}
    .ws-bar {{ background:#e1e7f0; border-radius:999px; overflow:hidden; height:12px; }}
    .ws-bar div {{ height:100%; }}
    .ws-bar .green {{ background:#007A33; }}
    .ws-bar .amber {{ background:#E8A000; }}
    .ws-bar .red {{ background:#CC0000; }}
    table {{ width:100%; border-collapse:collapse; font-size:12px; }}
    th, td {{ padding:8px 10px; text-align:left; border-bottom:1px solid #edf1f7; vertical-align:top; }}
    th {{ background:#eff4fb; color:#003b6e; }}
    .action-item {{ border-left:4px solid #0066CC; background:#f9fbfe; border-radius:8px; padding:12px; margin-bottom:10px; font-size:12px; }}
    .action-item.high {{ border-left-color:#CC0000; }}
    .action-item.medium {{ border-left-color:#E8A000; }}
    .action-item.low {{ border-left-color:#007A33; }}
    .model-box {{ background:#eff4fb; border-radius:10px; padding:12px; font-size:12px; line-height:1.6; }}
    .mini-grid {{ display:grid; grid-template-columns: repeat(2, 1fr); gap:10px; }}
    .mini {{ background:#f8fbff; border:1px solid #dfe8f5; border-radius:8px; padding:10px; }}
  </style>
</head>
<body>
  <div class='dashboard'>
    <div class='header'>
      <div>
        <div class='bosch-logo'><img src='data:image/png;base64,{logo_b64}' alt='Bosch - Invented for Life' style='height:36px;display:block;' /></div>
        <h1>{PROJECT_NAME}</h1>
        <p>Management KPI Dashboard | Pre-start baseline view for {SELLER} to {BUYER} cloud carve-out</p>
      </div>
      <div class='header-right'>
        <div><strong>Report Date:</strong> {REPORT_DATE.strftime('%d %b %Y')}</div>
        <div><strong>Overall Status:</strong> <span class='status amber'>Pre-start baseline</span></div>
        <div><strong>GoLive:</strong> 01 Feb 2027</div>
      </div>
    </div>

    <div class='grid'>
      <div class='card span-3'>
        <div class='card-header'>SCHEDULE PERFORMANCE</div>
        <div class='kpi-value'>{spi:.2f}</div>
        <div class='progress'><div style='width:{spi * 100:.0f}%;'></div></div>
        <div class='subtle'>Baseline SPI held at 1.00 because Gamma is still pre-kickoff and no execution slippage is booked yet.</div>
        <div style='margin-top:8px;'><span class='status green'>Baseline aligned</span></div>
      </div>

      <div class='card span-3'>
        <div class='card-header'>COST PERFORMANCE</div>
        <div class='kpi-value'>{cpi:.2f}</div>
        <div class='progress'><div style='width:{cpi * 100:.0f}%;'></div></div>
        <div class='subtle'>Cost baseline is established at {labour_total:,.0f} EUR labour plus {capex_total:,.0f} EUR contingency and CAPEX.</div>
        <div style='margin-top:8px;'><span class='status green'>Budget aligned</span></div>
      </div>

      <div class='card span-3'>
        <div class='card-header'>DAY 1 READINESS CONFIDENCE</div>
        <div class='kpi-value'>{day1_readiness}%</div>
        <div class='progress'><div style='width:{day1_readiness}%;'></div></div>
        <div class='subtle'>Calculated from the current Gamma high-risk burden, especially legal timing, infrastructure separation, and QG4 readiness exposure.</div>
        <div style='margin-top:8px;'><span class='status {status_class(day1_readiness, 75, 55)}'>{status_class(day1_readiness, 75, 55)}</span></div>
      </div>

      <div class='card span-3'>
        <div class='card-header'>JV TRANSITION CONFIDENCE</div>
        <div class='kpi-value'>{transition_confidence}%</div>
        <div class='progress'><div style='width:{transition_confidence}%;'></div></div>
        <div class='subtle'>Combination-model confidence reflects legal, strategy, and governance exposure around the 50/50 JV decision model.</div>
        <div style='margin-top:8px;'><span class='status {status_class(transition_confidence, 75, 55)}'>{status_class(transition_confidence, 75, 55)}</span></div>
      </div>

      <div class='card span-8'>
        <div class='card-header'>MILESTONE GATE CONTROL TIMELINE</div>
        <ul class='milestones'>
          {milestone_html}
        </ul>
      </div>

      <div class='card span-4'>
        <div class='card-header'>PROGRAMME SNAPSHOT</div>
        <div class='mini-grid'>
          <div class='mini'><strong>{task_count}</strong><br/>Total tasks</div>
          <div class='mini'><strong>{resource_group_count}</strong><br/>Resource groups</div>
          <div class='mini'><strong>{total_hours:,.0f}</strong><br/>Planned person-hours</div>
          <div class='mini'><strong>{high_risks}</strong><br/>High threat risks</div>
        </div>
      </div>

      <div class='card span-6'>
        <div class='card-header'>WORKSTREAM CONFIDENCE</div>
        {workstream_rows}
      </div>

      <div class='card span-6'>
        <div class='card-header'>TOP RISK TABLE</div>
        <table>
          <tr><th>Risk</th><th>Category</th><th>Score</th><th>Owner</th><th>Event</th></tr>
          {top_risk_rows}
        </table>
      </div>

      <div class='card span-4'>
        <div class='card-header'>MODEL KEY DIFFERENCES</div>
        <div class='model-box'>
          <p><strong>Combination model:</strong> Gamma uses a transitional operating path, but the target state is a jointly managed JV rather than a long-term shared environment.</p>
          <p><strong>No SAP:</strong> The technical burden is concentrated in infrastructure, identity, and operations rather than ERP separation.</p>
          <p><strong>Confidentiality:</strong> Clean-team restrictions are a delivery mechanic, not only a communications issue.</p>
        </div>
      </div>

      <div class='card span-8'>
        <div class='card-header'>NEXT 90 DAYS ACTION FORECAST</div>
        {action_rows}
      </div>
    </div>
  </div>
</body>
</html>
"""

    OUTPUT_PATH.write_text(html, encoding="utf-8")
    print(f"[{PROJECT_NAME}] KPI dashboard written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()